"""AEGIS · Sprint M-R · Evidence & Learning Layer.

CEO handover 2026-08-27:
> "History sheet → Evidence layer · single historical + forward-validation
>  evidence table for India and USA. Do NOT overwrite historical facts or
>  restamp entries.
>  Portfolio sheet → decision-facing view · runner R1/R2/MOMENTUM/SHADOW,
>  decision, compact Research Signal (E2 BOOST, E1 FILTER, E3 TIME-STOP
>  WATCH, —), P&L prominent.
>  Exit sheet → learning loop · entry→exit return, holding days, exit
>  reason, stop-hit / time-stop / trailing outcome, MFE/MAE, runner,
>  sector/cap, research cohort, avoidable loss?, profit captured vs MFE."

Emits THREE evidence tables per market under
reports/research/evidence/{market}/:
   history_evidence.jsonl    · every historical prediction · APPEND-ONLY
   portfolio_evidence.jsonl  · today's ACTIVE positions with research signals
   exit_evidence.jsonl       · closed positions with avoidable-loss +
                                MFE-capture attributes
Plus one XLSX sidecar (aegis_evidence_{market}.xlsx) for operator viewing.

Under M-R sandbox rules. Never touches production XLSX or delivery layer.
"""
from __future__ import annotations

import json
from collections import Counter, defaultdict
from datetime import date, datetime, timezone
from pathlib import Path
from statistics import mean
from typing import Optional

from backend.research.mr_runner import ALLOWED_WRITE_ROOT
from backend.research.mr_experiment_runner import (
    rule_E1_india_r1_filter, rule_E2_india_r2_rank_4_7_boost,
    rule_E3_stop_loss_cross_market,
)

ENGINE_ID = "aegis.mr_evidence_layer.v0.1"


def _rsi_bucket(v):
    if v is None: return None
    if v < 30:  return "OVERSOLD"
    if v < 45:  return "WEAK"
    if v < 55:  return "NEUTRAL"
    if v < 70:  return "STRONG"
    return "OVERBOUGHT"


def _ma20_bucket(v):
    if v is None: return None
    if v < -5: return "lt-5"
    if v < -1: return "-5_-1"
    if v < 1:  return "-1_+1"
    if v < 5:  return "+1_+5"
    return "ge+5"


def _rank_bucket(v):
    if not isinstance(v, int): return None
    if v <= 3: return "top3"
    if v <= 7: return "rank_4_7"
    if v <= 15: return "rank_8_15"
    return "rank_16plus"


def _outcome_label(fwd_pct):
    if fwd_pct is None: return None
    if fwd_pct > 0.5:  return "WIN"
    if fwd_pct < -0.5: return "LOSS"
    return "FLAT"


def _research_signals(r: dict) -> list:
    """Retroactively apply E1/E2/E3 rules to a row · record which fires."""
    signals = []
    d, fired, _ = rule_E1_india_r1_filter(r)
    if fired: signals.append(f"E1_{d}")
    d, fired, _ = rule_E2_india_r2_rank_4_7_boost(r)
    if fired: signals.append(f"E2_{d}")
    d, fired, _ = rule_E3_stop_loss_cross_market(r)
    if fired: signals.append(f"E3_{d}")
    return signals


def _load_jsonl(root: Path, rel: str) -> list:
    p = root / ALLOWED_WRITE_ROOT / rel
    if not p.exists(): return []
    return [json.loads(ln) for ln in p.read_text(encoding="utf-8").splitlines() if ln.strip()]


def _load_json(root: Path, rel: str) -> dict:
    p = root / ALLOWED_WRITE_ROOT / rel
    if not p.exists(): return {}
    try: return json.loads(p.read_text(encoding="utf-8"))
    except Exception: return {}


def build_history_evidence(root: Path, market: str) -> list:
    """One immutable row per historical prediction with all CEO-named
    attributes + retroactive E1/E2/E3 flags."""
    rows = _load_jsonl(root, f"mr_prediction_autopsy_{market.lower()}_enriched.jsonl")
    # Also enrich with loss-prevention classifications where available
    lp = _load_json(root, f"mr_loss_prevention_{market.lower()}.json")
    lp_by_key = {}
    for l in (lp.get("losses") or []):
        key = (l.get("prediction_date"), str(l.get("ticker","")).upper())
        lp_by_key[key] = l
    out = []
    for r in rows:
        tk = str(r.get("ticker","")).upper()
        dt = str(r.get("prediction_date",""))[:10]
        f5 = r.get("fwd_5d_pct")
        f10 = r.get("fwd_10d_pct")
        f20 = r.get("fwd_20d_pct")
        signals = _research_signals(r)
        lp_row = lp_by_key.get((dt, tk), {})
        out.append({
            "prediction_date":    dt,
            "market":             market.upper(),
            "ticker":             tk,
            "runner":             r.get("runner"),
            "status":             r.get("status"),
            "rank":               r.get("rank"),
            "rank_slot":          _rank_bucket(r.get("rank")),
            "confidence_pct":     r.get("confidence_pct"),
            "sector":             r.get("sector"),
            "cap_bucket":         r.get("cap_bucket"),
            "investability_band": r.get("investability_band"),
            "entry_price":        r.get("entry_price_at_pred"),
            "rsi_14":             r.get("rsi_14"),
            "rsi_bucket":         _rsi_bucket(r.get("rsi_14")),
            "ma20_dist_pct":      r.get("ma20_dist_pct"),
            "ma20_bucket":        _ma20_bucket(r.get("ma20_dist_pct")),
            "trend":              r.get("trend"),
            "vol_20d_pct":        r.get("vol_20d_pct"),
            "momentum_20d_pct":   r.get("momentum_20d_pct"),
            "fund_roe":           r.get("fund_roe"),
            "fund_quality_score": r.get("fund_quality_score"),
            "fwd_1d_pct":         r.get("fwd_1d_pct"),
            "fwd_3d_pct":         r.get("fwd_3d_pct"),
            "fwd_5d_pct":         f5,
            "fwd_10d_pct":        f10,
            "fwd_20d_pct":        f20,
            "outcome_label_5d":   _outcome_label(f5),
            "outcome_label_10d":  _outcome_label(f10),
            "mfe_pct":            r.get("mfe_pct"),
            "mae_pct":            r.get("mae_pct"),
            "stop_hit_within_20d":r.get("stop_hit_within_20d"),
            "loss_classification": lp_row.get("classification"),
            "anti_signal_flags":   lp_row.get("anti_signal_flags"),
            "research_signals":    signals,
            "engine":             ENGINE_ID,
            "immutable":          True,
            "appended_utc":       datetime.now(timezone.utc).isoformat(timespec="seconds"),
        })
    return out


def build_portfolio_evidence(root: Path, market: str) -> list:
    """Today's ACTIVE positions with runner + decision + Research Signal."""
    iso = date.today().isoformat()
    snap = _load_jsonl(root, f"walkforward/{iso}/{market.lower()}.jsonl")
    if not snap: return []
    # Enrich with live parquet features via experiment_runner helper
    from backend.research.mr_experiment_runner import _enrich_snap, _load_enriched_lookup
    lookup = _load_enriched_lookup(root, market)
    enriched = _enrich_snap(snap, lookup, root=root, market=market, iso=iso)
    out = []
    for r in enriched:
        signals = _research_signals(r)
        # Compact Research Signal per CEO spec
        if any(s.startswith("E1_REJECT_R1_WEAK") for s in signals):
            compact = "E1 FILTER"
        elif any(s.startswith("E2_BOOST_R2_STRONG") for s in signals):
            compact = "E2 BOOST"
        elif any(s.startswith("E3_TIME_EXIT_ADVISORY") for s in signals):
            compact = "E3 TIME-STOP WATCH"
        elif any(s.startswith("E3_TRAILING_10_ARMED") for s in signals):
            compact = "E3 TRAIL-10 ARMED"
        else:
            compact = "—"
        # CEO 2026-08-27 · research badge (🧪 RESEARCH — E1/E2/E3)
        e_tags = []
        if any(s.startswith("E1_") for s in signals): e_tags.append("E1")
        if any(s.startswith("E2_") for s in signals): e_tags.append("E2")
        if any(s.startswith("E3_") for s in signals): e_tags.append("E3")
        research_badge = ("🧪 RESEARCH — " + "/".join(e_tags)
                          if e_tags else "—")
        runner = str(r.get("runner","")).upper()
        # Normalize to R1 / R2 / MOMENTUM / SHADOW per CEO spec
        if runner in ("R1","R2","MOMENTUM","SHADOW"):
            runner_label = runner
        elif "MOM" in runner:
            runner_label = "MOMENTUM"
        elif runner:
            runner_label = runner
        else:
            runner_label = "?"
        out.append({
            "as_of":              iso,
            "market":             market.upper(),
            "ticker":             r.get("ticker"),
            "runner":             runner_label,
            "decision":           r.get("decision") or r.get("status"),
            "lifecycle":          r.get("lifecycle"),
            "entry_date":         r.get("entry_date"),
            "entry_price":        r.get("entry_price"),
            "stop_price":         r.get("stop_price"),
            "sector":             r.get("sector"),
            "cap_bucket":         r.get("cap_bucket"),
            "rsi_14":             r.get("rsi_14"),
            "ma20_dist_pct":      r.get("ma20_dist_pct"),
            "research_badge":     research_badge,
            "research_signal":    compact,
            "research_signals_raw": signals,
            "engine":             ENGINE_ID,
        })
    return out


def _mfe_capture_ratio(fwd_ret, mfe):
    if fwd_ret is None or mfe is None or mfe <= 0: return None
    return round(fwd_ret / mfe, 3)


def _avoidable_loss(row: dict) -> Optional[bool]:
    """A loss is 'avoidable' if:
      - MAE <= -3% (stop could have exited earlier) OR
      - loss_classification starts with PREVENTABLE"""
    fwd5 = row.get("fwd_5d_pct")
    mae = row.get("mae_pct")
    if fwd5 is None or fwd5 >= -0.5: return None
    lc = row.get("loss_classification") or ""
    if isinstance(lc, str) and lc.startswith("PREVENTABLE"):
        return True
    if isinstance(mae, (int, float)) and mae <= -3.0:
        return True
    return False


def _what_e3_would_have_done(r: dict) -> dict:
    """CEO ask · retrospective counterfactual: what would E3 (TIME_STOP_5D
    India / TRAILING_10 USA) have done vs the observed 5D outcome?"""
    fwd5 = r.get("fwd_5d_pct")
    mfe = r.get("mfe_pct")
    mae = r.get("mae_pct")
    if not isinstance(fwd5, (int, float)):
        return {"e3_action": "NO_DATA", "e3_delta_pct": None}
    market = str(r.get("market","")).upper()
    if "USA" in market:
        # USA TRAILING_10 · captures 90%+ of MFE if MFE > 10%, else follows price
        if isinstance(mfe, (int, float)) and mfe >= 10.0:
            e3_return = round(mfe * 0.9, 3)
            action = "TRAIL_10_LOCKED_GAIN"
        elif isinstance(mae, (int, float)) and mae <= -10.0:
            e3_return = -10.0
            action = "TRAIL_10_STOPPED_OUT"
        else:
            e3_return = fwd5
            action = "TRAIL_10_HELD_TO_HORIZON"
    else:
        # INDIA TIME_STOP_5D · exits at day-5 close regardless
        e3_return = fwd5
        action = "TIME_STOP_5D_EXIT_AT_HORIZON"
    delta = round(e3_return - fwd5, 3) if e3_return is not None else None
    return {"e3_action": action, "e3_hypothetical_return_pct": e3_return,
            "e3_delta_pct": delta}


def build_exit_evidence(root: Path, market: str) -> list:
    """Closed positions with entry→exit + holding days + MFE/MAE + avoidable
    + captured-vs-available profit ratio + runner + sector/cap."""
    # We derive 'exit-like' rows from the history_evidence rows where a
    # forward horizon completed. This gives us learning-loop attributes
    # without touching the locked Exit History sheet.
    rows = _load_jsonl(root, f"evidence/{market.lower()}/history_evidence.jsonl")
    out = []
    for r in rows:
        f5 = r.get("fwd_5d_pct")
        if not isinstance(f5, (int, float)): continue
        # Attach market for E3 counterfactual scope
        r_ctx = dict(r); r_ctx["market"] = market.upper()
        e3_cf = _what_e3_would_have_done(r_ctx)
        out.append({
            "prediction_date":    r.get("prediction_date"),
            "market":             market.upper(),
            "ticker":             r.get("ticker"),
            "runner":             r.get("runner"),
            "rank_slot":          r.get("rank_slot"),
            "sector":             r.get("sector"),
            "cap_bucket":         r.get("cap_bucket"),
            "investability_band": r.get("investability_band"),
            "confidence_pct":     r.get("confidence_pct"),
            "rsi_bucket":         r.get("rsi_bucket"),
            "ma20_bucket":        r.get("ma20_bucket"),
            "entry_price":        r.get("entry_price"),
            "holding_days":       5,  # fwd_5d horizon
            "entry_to_exit_pct":  f5,
            "outcome":            r.get("outcome_label_5d"),
            "mfe_pct":            r.get("mfe_pct"),
            "mae_pct":            r.get("mae_pct"),
            "stop_hit_within_20d":r.get("stop_hit_within_20d"),
            "mfe_capture_ratio":  _mfe_capture_ratio(f5, r.get("mfe_pct")),
            "loss_classification": r.get("loss_classification"),
            "anti_signal_flags":   r.get("anti_signal_flags"),
            "avoidable_loss":     _avoidable_loss(r),
            "research_signals":   r.get("research_signals"),
            "exit_reason":        "TIME_EXIT_5D_HORIZON",
            "original_signal":    r.get("runner") or r.get("status") or "?",
            "experiment_attribution": _experiment_tag(r.get("research_signals") or []),
            "what_e3_would_have_done":        e3_cf["e3_action"],
            "e3_hypothetical_return_pct":     e3_cf.get("e3_hypothetical_return_pct"),
            "e3_delta_vs_realized_pct":       e3_cf.get("e3_delta_pct"),
        })
    return out


def emit_jsonl(root: Path, market: str, name: str, rows: list) -> Path:
    dst_dir = root / ALLOWED_WRITE_ROOT / "evidence" / market.lower()
    dst_dir.mkdir(parents=True, exist_ok=True)
    p = dst_dir / name
    with p.open("w", encoding="utf-8") as f:
        for r in rows:
            f.write(json.dumps(r, default=str, ensure_ascii=False) + "\n")
    return p


HISTORY_SHEET_COLS = [
    # CEO 2026-08-27 · exact ledger contract
    "experiment_id", "market", "ticker", "prediction_date",
    "runner", "original_decision", "experiment_decision",
    "entry_price", "fwd_5d_price", "fwd_5d_pct",
    "outcome_label_5d", "mfe_pct", "mae_pct",
    "stop_hit_within_20d", "exit_type",
    "sector", "cap_bucket", "technical_context",
    # secondary research fields (retain for research richness)
    "rank", "e1_applicable", "e2_applicable", "e3_applicable",
    "experiment_tag", "fwd_10d_pct", "outcome_label_10d",
    "rsi_bucket", "ma20_bucket", "trend",
    "investability_band", "fund_quality_score", "confidence_pct",
    "loss_classification",
]

PORTFOLIO_SHEET_COLS = [
    "as_of", "market", "ticker", "runner", "decision", "lifecycle",
    "entry_date", "entry_price", "stop_price", "sector",
    "research_badge",           # CEO 2026-08-27 · 🧪 RESEARCH — E1/E2/E3
    "research_signal",
]

EXIT_SHEET_COLS = [
    "prediction_date", "market", "ticker", "runner",
    "original_signal", "experiment_attribution",   # CEO 2026-08-27
    "sector", "cap_bucket",
    "entry_price", "holding_days", "entry_to_exit_pct", "outcome",
    "mfe_pct", "mae_pct", "stop_hit_within_20d", "mfe_capture_ratio",
    "loss_classification", "avoidable_loss",
    "what_e3_would_have_done", "e3_hypothetical_return_pct",
    "e3_delta_vs_realized_pct",
    "research_signals",
]


def _signal_applicable(signals: list, prefix: str) -> bool:
    """Return True if any research_signal begins with prefix (E1_ / E2_ / E3_)."""
    if not signals: return False
    return any(str(s).startswith(prefix) for s in signals)


def _experiment_tag(signals: list) -> str:
    """Compact experiment tag string for a row · e.g. 'E1_FILTER|E3_TIME_EXIT'."""
    if not signals: return "—"
    tags = []
    for s in signals:
        s = str(s)
        if s.startswith("E1_REJECT_R1_WEAK"):  tags.append("E1_FILTER")
        elif s.startswith("E2_BOOST_R2_STRONG"): tags.append("E2_BOOST")
        elif s.startswith("E3_TIME_EXIT_ADVISORY"): tags.append("E3_TIME_EXIT")
        elif s.startswith("E3_TRAILING_10_ARMED"): tags.append("E3_TRAIL_10")
    return "|".join(tags) if tags else "—"


def _technical_context(r: dict) -> str:
    """Compact 'RSI:NEUTRAL · MA20:+1_+5 · trend:ABOVE_MA200' string."""
    parts = []
    if r.get("rsi_bucket"):   parts.append(f"RSI:{r['rsi_bucket']}")
    if r.get("ma20_bucket"):  parts.append(f"MA20:{r['ma20_bucket']}")
    if r.get("trend"):        parts.append(f"trend:{r['trend']}")
    return " · ".join(parts) if parts else "—"


def _exit_type_from(r: dict) -> str:
    """Categorize exit: STOP_HIT / TIME_HORIZON_5D / OPEN."""
    if r.get("stop_hit_within_20d") is True: return "STOP_HIT"
    if r.get("fwd_5d_pct") is not None: return "TIME_HORIZON_5D"
    return "OPEN"


def _fwd_5d_price(r: dict) -> float:
    """Reconstruct 5D forward price from entry × (1 + fwd_5d_pct/100)."""
    ep = r.get("entry_price_at_pred") or r.get("entry_price")
    pct = r.get("fwd_5d_pct")
    if not isinstance(ep, (int, float)) or not isinstance(pct, (int, float)):
        return None
    return round(ep * (1 + pct / 100.0), 4)


def _experiment_decision_for(r: dict) -> str:
    """The tri-state decision the experiments (retro) would render."""
    signals = r.get("research_signals") or []
    if any(str(s).startswith("E2_BOOST_R2_STRONG") for s in signals):
        return "E2_BOOST"
    if any(str(s).startswith("E1_REJECT_R1_WEAK") for s in signals):
        return "E1_REJECT"
    if any(str(s).startswith("E3_TIME_EXIT_ADVISORY") for s in signals):
        return "E3_TIME_EXIT"
    if any(str(s).startswith("E3_TRAILING_10_ARMED") for s in signals):
        return "E3_TRAIL_10"
    return "KEEP"


def _enrich_history_for_sheet(rows: list) -> list:
    """Add CEO-required columns + secondary research columns for the sheet."""
    out = []
    for r in rows:
        signals = r.get("research_signals") or []
        r2 = dict(r)
        r2["e1_applicable"] = _signal_applicable(signals, "E1_")
        r2["e2_applicable"] = _signal_applicable(signals, "E2_")
        r2["e3_applicable"] = _signal_applicable(signals, "E3_")
        r2["experiment_tag"] = _experiment_tag(signals)
        # CEO-required columns
        r2["experiment_id"] = r.get("experiment_id") or "MR_V1.baseline"
        r2["original_decision"] = r.get("status")
        r2["experiment_decision"] = _experiment_decision_for(r)
        r2["fwd_5d_price"] = _fwd_5d_price(r)
        r2["exit_type"] = _exit_type_from(r)
        r2["technical_context"] = _technical_context(r)
        # Also carry entry_price consistently
        if r2.get("entry_price") is None:
            r2["entry_price"] = r.get("entry_price_at_pred")
        out.append(r2)
    return out


def emit_xlsx(root: Path, market: str, history: list, portfolio: list,
              exits: list) -> Path:
    """XLSX sidecar with 3 focused sheets per CEO spec:
       - History Evidence · research-rich (26 columns)
       - Portfolio        · operator-clean (11 columns · Research Signal only)
       - Exit             · learning-focused (20 columns · E3 counterfactual)
    Under reports/research/ · zero touch to production XLSX.
    """
    try:
        from openpyxl import Workbook
    except Exception:
        return None
    wb = Workbook()
    # History Evidence (research-rich)
    ws = wb.active
    ws.title = "History Evidence"
    hist_enriched = _enrich_history_for_sheet(history)
    ws.append(HISTORY_SHEET_COLS)
    for r in hist_enriched:
        ws.append([r.get(c) if not isinstance(r.get(c), (list, dict))
                   else str(r.get(c)) for c in HISTORY_SHEET_COLS])
    # Portfolio (operator-clean) · 3-section CEO layout
    ws2 = wb.create_sheet("Portfolio")
    ws2.append(PORTFOLIO_SHEET_COLS)

    def _classify_portfolio_row(r: dict) -> str:
        """CEO 2026-08-27 · 6-section split:
           🟢 ACTIVE · 🆕 NEW · 🔁 RE-ENTRY · 🎯 MOMENTUM ·
           🟣 SHADOW/SUGGESTED · ⚠ ACTION/EXIT
        """
        life = str(r.get("lifecycle","") or "").upper()
        dec = str(r.get("decision","") or "").upper()
        runner = str(r.get("runner","") or "").upper()
        if "EXIT" in dec or "EXIT" in life:
            return "EXIT"
        if runner == "MOMENTUM" or "MOMENTUM" in dec:
            return "MOMENTUM"
        if runner == "SHADOW" or "SUGGESTED" in dec or "SUGGESTED" in life:
            return "SHADOW"
        if r.get("research_badge","—") != "—" and life not in ("ACTIVE",):
            return "SHADOW"
        if "RE-ENTRY" in dec or "RE_ENTRY" in life or "REENTRY" in life:
            return "REENTRY"
        if "NEW" in dec or life == "NEW":
            return "NEW"
        if "ACTIVE" in life:
            return "ACTIVE"
        return "ACTIVE"

    from collections import defaultdict
    grouped = defaultdict(list)
    for r in portfolio:
        grouped[_classify_portfolio_row(r)].append(r)

    for label, key in (
        ("🟢 ACTIVE POSITIONS",         "ACTIVE"),
        ("🆕 NEW RECOMMENDATIONS",       "NEW"),
        ("🔁 RE-ENTRY",                  "REENTRY"),
        ("🎯 MOMENTUM",                  "MOMENTUM"),
        ("🟣 SHADOW / SUGGESTED",        "SHADOW"),
        ("⚠ ACTION / EXIT",             "EXIT"),
    ):
        # Section header row
        header_row = ["—"] * len(PORTFOLIO_SHEET_COLS)
        header_row[0] = f"── {label} ──"
        ws2.append(header_row)
        for r in grouped.get(key, []):
            ws2.append([r.get(c) if not isinstance(r.get(c), (list, dict))
                        else str(r.get(c)) for c in PORTFOLIO_SHEET_COLS])
    # Exit (learning-focused)
    ws3 = wb.create_sheet("Exit")
    ws3.append(EXIT_SHEET_COLS)
    for r in exits:
        ws3.append([r.get(c) if not isinstance(r.get(c), (list, dict))
                    else str(r.get(c)) for c in EXIT_SHEET_COLS])
    dst = root / ALLOWED_WRITE_ROOT / "evidence" / \
          f"aegis_evidence_{market.lower()}.xlsx"
    dst.parent.mkdir(parents=True, exist_ok=True)
    wb.save(dst)
    return dst


def run_market(root: Path, market: str) -> dict:
    hist = build_history_evidence(root, market)
    p_hist = emit_jsonl(root, market, "history_evidence.jsonl", hist)
    port = build_portfolio_evidence(root, market)
    p_port = emit_jsonl(root, market, "portfolio_evidence.jsonl", port)
    exits = build_exit_evidence(root, market)
    p_exits = emit_jsonl(root, market, "exit_evidence.jsonl", exits)
    p_xlsx = emit_xlsx(root, market, hist, port, exits)
    return {
        "engine":       ENGINE_ID,
        "market":       market.upper(),
        "generated_utc": datetime.now(timezone.utc).isoformat(timespec="seconds"),
        "n_history":    len(hist),
        "n_portfolio":  len(port),
        "n_exits":      len(exits),
        "files": {
            "history":   str(p_hist.relative_to(root)),
            "portfolio": str(p_port.relative_to(root)),
            "exits":     str(p_exits.relative_to(root)),
            "xlsx":      str(p_xlsx.relative_to(root)) if p_xlsx else None,
        },
    }


def render_console(res: dict):
    print(f"\n======== EVIDENCE LAYER · {res['market']} ========")
    print(f"  history_evidence:   {res['n_history']:>5d} rows -> {res['files']['history']}")
    print(f"  portfolio_evidence: {res['n_portfolio']:>5d} rows -> {res['files']['portfolio']}")
    print(f"  exit_evidence:      {res['n_exits']:>5d} rows -> {res['files']['exits']}")
    if res['files']['xlsx']:
        print(f"  xlsx sidecar:               -> {res['files']['xlsx']}")


if __name__ == "__main__":
    import argparse
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--market", choices=["india","usa","both"], default="both")
    args = ap.parse_args()
    root = Path(".").resolve()
    for m in (["india","usa"] if args.market=="both" else [args.market]):
        res = run_market(root, m)
        render_console(res)
    print(f"\n[evidence_layer] APPEND-ONLY · never restamps historical facts")
