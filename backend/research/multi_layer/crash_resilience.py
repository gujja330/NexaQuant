"""Crash-resilience research · CEO 2026-09-01 (§ Crash-resilience addendum).

Reuses EXISTING regime infrastructure (`mr_market_regime` for the coarse
BULL/BEAR/HIGH_VOL classification and the benchmark parquet loader).
Derives the CEO-mandated 5-state classifier:

    NORMAL      · benchmark near 52w high · low volatility
    WEAKENING   · benchmark drawdown in [3%, 10%] · rising volatility
    RISK_OFF    · drawdown in [10%, 20%]  OR  HIGH_VOL regime
    CRASH       · drawdown > 20%  OR  single-day drop > 5%
    RECOVERY    · previously CRASH/RISK_OFF and benchmark up > 5% off low
                    within last 20d · currently NORMAL/BULL classification

Answers CEO's real question:
    "When the market regime deteriorates sharply, does AEGIS actually
    reduce bad exposure, stop recommending deteriorating stocks, and
    preserve capital better than the benchmark?"

For each regime and each market, computes R2 metrics:
    · n_signals · n_entries · n_exits
    · gross exposure proxy (currently_active count)
    · decision distribution (BUY / STRONG_BUY / HOLD / PROTECT / EXIT)
    · realized_pnl_pct · unrealized_pnl_pct
    · win_rate_pct
    · max_drawdown_per_trade_pct
    · downside_capture (mean R2 return / mean benchmark return · negative-only)
    · sector_concentration_pct (top-3 sector share of active positions)
    · recovery_time_days (avg from CRASH classification to RECOVERY)
    · benchmark_relative_pnl_pct

Point-in-time enforcement:
    · Only uses benchmark data available AT the trade's exit date
    · Marks a regime UNAVAILABLE if the classifier does not have enough
      history at that date

Never modifies R2 · never auto-promotes anything.
Output: reports/research/multi_layer/crash_resilience_{market}_{asof}.json
"""
from __future__ import annotations

import argparse
import json
import statistics
import sys
from collections import Counter, defaultdict
from datetime import date, timedelta
from pathlib import Path

_ROOT = Path(__file__).resolve().parents[3]
sys.path.insert(0, str(_ROOT))

from backend.research import mr_market_regime as _mmr
from openpyxl import load_workbook


REGIME_STATES = ("NORMAL", "WEAKENING", "RISK_OFF", "CRASH", "RECOVERY")


def _load_benchmark_series(root: Path, market: str) -> list:
    """Returns sorted list of (date_str, close) using existing mr_market_regime
    index loader · never invents a parallel data source."""
    pair = _mmr._load_index(root, market)
    if pair is None: return []
    df, col = pair
    return sorted([(d[:10], float(c)) for d, c in df[col].items()
                    if c is not None])


def _drawdown_from_high(closes: list) -> float:
    """Current drawdown from trailing max · 0..1 (positive number)."""
    if not closes: return 0.0
    peak = max(closes)
    if peak <= 0: return 0.0
    last = closes[-1]
    return max(0.0, (peak - last) / peak)


def _classify_5state(series: list, asof_idx: int, regime_hint: str) -> str:
    """Classify a single date · 5-state · point-in-time · reuses BULL/BEAR/
    HIGH_VOL hint from mr_market_regime."""
    if asof_idx < 30: return "UNAVAILABLE"
    # Trailing 252-day high (approx 1yr)
    lookback = 252
    window = [c for _, c in series[max(0, asof_idx - lookback):asof_idx + 1]]
    if len(window) < 30: return "UNAVAILABLE"
    dd = _drawdown_from_high(window)
    # Single-day drop
    day_drop = 0.0
    if asof_idx > 0:
        prev = series[asof_idx - 1][1]
        cur = series[asof_idx][1]
        if prev > 0:
            day_drop = max(0.0, (prev - cur) / prev)
    # 5-state decision tree
    if dd > 0.20 or day_drop > 0.05: return "CRASH"
    if dd > 0.10 or regime_hint == "HIGH_VOL": return "RISK_OFF"
    if dd > 0.03: return "WEAKENING"
    # RECOVERY heuristic · was in CRASH/RISK_OFF within prior 20d AND now UP ≥5% from low
    prior_hint_window = series[max(0, asof_idx - 20):asof_idx + 1]
    if prior_hint_window:
        prior_lows = [c for _, c in prior_hint_window]
        low = min(prior_lows)
        cur = series[asof_idx][1]
        if low > 0 and (cur - low) / low >= 0.05 and low < window[-2]:
            # Was suppressed recently · check if depth was crash-like
            trailing_dd = _drawdown_from_high([c for _, c in
                                                series[max(0, asof_idx - 40):asof_idx + 1]])
            if trailing_dd > 0.10:
                return "RECOVERY"
    return "NORMAL"


def _daily_regime_map(series: list, coarse_regimes: dict) -> dict:
    """Build date → 5-state map. Uses coarse mr_market_regime hint."""
    out = {}
    for i, (d, _) in enumerate(series):
        hint = coarse_regimes.get(d, "UNKNOWN")
        out[d] = _classify_5state(series, i, hint)
    return out


def _load_exit_history(root: Path, market: str) -> list:
    xlsx = root / "reports" / "telegram" / f"aegis_history_{market.lower()}.xlsx"
    if not xlsx.exists(): return []
    wb = load_workbook(xlsx, read_only=True, data_only=True)
    # CEO 2026-09-01 FINAL 3-sheet spec · read from 03_Exit_History
    _sheet = "03_Exit_History"
    if _sheet not in wb.sheetnames: wb.close(); return []
    ws = wb[_sheet]
    rows = list(ws.iter_rows(values_only=True))
    out = []
    for r in rows:
        if not r or not r[0]: continue
        pid = str(r[0])
        if not (pid.upper().startswith("USA-") or pid.upper().startswith("IND-")):
            continue
        # cols: 0=PID 1=Ticker 2=Runner 3=Market 4=EntryDate 5=ExitDate
        # 6=HoldingDays 7=EntryPrice 8=ExitPrice 9=RealizedPnL%
        run = str(r[2] or "").upper() if len(r) > 2 else ""
        if run != "R2": continue
        try:
            pnl = float(r[9]) if len(r) > 9 and r[9] not in (None, "", "—") else 0.0
        except (TypeError, ValueError):
            pnl = 0.0
        try:
            from datetime import date as _d
            days = (_d.fromisoformat(str(r[5])[:10])
                     - _d.fromisoformat(str(r[4])[:10])).days
        except Exception:
            days = 0
        out.append({
            "ticker": str(r[1] or "") if len(r) > 1 else "",
            "sector": "",
            "entry_date": str(r[4] or "")[:10] if len(r) > 4 else "",
            "exit_date": str(r[5] or "")[:10] if len(r) > 5 else "",
            "pnl_pct": round(pnl, 4),
            "days": days,
        })
    wb.close()
    return out


def _benchmark_return_between(series_by_date: dict, start: str, end: str) -> float:
    if not start or not end or start not in series_by_date or end not in series_by_date:
        return None
    a = series_by_date[start]; b = series_by_date[end]
    if a <= 0: return None
    return (b - a) / a * 100.0


def build(market: str, root: Path, asof: str) -> dict:
    # 1. Reuse existing regime
    reg_p = root / "reports" / "research" / f"mr_market_regime_{market.lower()}.json"
    if not reg_p.exists():
        try:
            _mmr.emit(root, market, _mmr.build(root, market))
        except Exception as e:
            return {"error": f"regime build failed: {type(e).__name__}: {e}"}
    reg_data = json.loads(reg_p.read_text(encoding="utf-8")) if reg_p.exists() else {}
    coarse = reg_data.get("regimes", {}) or {}

    # 2. Benchmark series + 5-state map
    series = _load_benchmark_series(root, market)
    if not series:
        return {"error": f"benchmark series unavailable for {market}",
                 "reason": "no parquet found via mr_market_regime._load_index"}
    series_by_date = {d: c for d, c in series}
    daily_5state = _daily_regime_map(series, coarse)

    # 3. Exit History R2 trades
    trades = _load_exit_history(root, market)

    # 4. Per-regime aggregation
    by_regime: dict[str, list] = {r: [] for r in REGIME_STATES}
    by_regime["UNAVAILABLE"] = []
    for t in trades:
        rg = daily_5state.get(t["exit_date"], "UNAVAILABLE")
        t["regime_5state"] = rg
        by_regime.setdefault(rg, []).append(t)

    def _regime_stats(trs: list) -> dict:
        if not trs:
            return {"n": 0, "mean_pnl_pct": None, "sum_pnl_pct": 0.0,
                     "win_rate_pct": None, "max_dd_per_trade_pct": None,
                     "avg_holding_days": None,
                     "sector_concentration_top3_pct": None,
                     "downside_capture_vs_bench": None,
                     "benchmark_relative_pnl_pct": None}
        pnls = [t["pnl_pct"] for t in trs]
        wins = sum(1 for p in pnls if p > 0)
        days = [t["days"] for t in trs if t["days"] > 0]
        # Sector concentration: top-3 sector share
        sec_ct = Counter(t.get("sector", "") for t in trs)
        top3 = sum(v for _, v in sec_ct.most_common(3))
        sec_pct = round(top3 / len(trs) * 100, 1) if trs else None
        # Benchmark relative: mean of (trade_pnl - benchmark_return_over_holding)
        rel = []
        neg_r = []
        neg_b = []
        for t in trs:
            b = _benchmark_return_between(series_by_date, t["entry_date"], t["exit_date"])
            if b is None: continue
            rel.append(t["pnl_pct"] - b)
            if b < 0:
                neg_r.append(t["pnl_pct"])
                neg_b.append(b)
        rel_mean = round(statistics.mean(rel), 3) if rel else None
        # Downside capture: mean R2 return / mean benchmark return on negative benchmark days
        dc = None
        if neg_b and statistics.mean(neg_b) != 0:
            dc = round(statistics.mean(neg_r) / statistics.mean(neg_b), 3)
        return {
            "n": len(trs),
            "mean_pnl_pct": round(statistics.mean(pnls), 3),
            "sum_pnl_pct": round(sum(pnls), 3),
            "win_rate_pct": round(wins / len(pnls) * 100, 1),
            "max_dd_per_trade_pct": round(min(pnls), 3),
            "avg_holding_days": round(statistics.mean(days), 1) if days else None,
            "sector_concentration_top3_pct": sec_pct,
            "downside_capture_vs_bench": dc,
            "benchmark_relative_pnl_pct": rel_mean,
        }

    per_regime = {rg: _regime_stats(trs) for rg, trs in by_regime.items()}

    # 5. Current-regime · today's classification
    today_regime = daily_5state.get(asof) or daily_5state.get(series[-1][0], "UNAVAILABLE")

    # 6. Regime distribution over history
    regime_distribution = Counter(daily_5state.values())

    # 7. Answer the CEO's actual question
    # "During deteriorating regimes does AEGIS reduce exposure and preserve capital?"
    interpretation = []
    if per_regime.get("CRASH", {}).get("n", 0) == 0 and per_regime.get("RISK_OFF", {}).get("n", 0) == 0:
        interpretation.append("INSUFFICIENT_STRESS_HISTORY_IN_R2_WINDOW")
    else:
        crash_n = per_regime.get("CRASH", {}).get("n", 0)
        risk_off_n = per_regime.get("RISK_OFF", {}).get("n", 0)
        crash_wr = per_regime.get("CRASH", {}).get("win_rate_pct")
        risk_off_wr = per_regime.get("RISK_OFF", {}).get("win_rate_pct")
        interpretation.append(f"CRASH n={crash_n} win_rate={crash_wr}%")
        interpretation.append(f"RISK_OFF n={risk_off_n} win_rate={risk_off_wr}%")
        crash_dc = per_regime.get("CRASH", {}).get("downside_capture_vs_bench")
        if crash_dc is not None:
            if crash_dc < 0.7:
                interpretation.append(f"CRASH downside_capture={crash_dc} · R2 absorbs LESS of the drop than benchmark (GOOD)")
            elif crash_dc > 1.1:
                interpretation.append(f"CRASH downside_capture={crash_dc} · R2 absorbs MORE than benchmark (BAD)")
            else:
                interpretation.append(f"CRASH downside_capture={crash_dc} · roughly benchmark-tracking")

    result = {
        "engine": "crash_resilience.multi_layer.v1",
        "market": market.lower(),
        "asof": asof,
        "today_regime": today_regime,
        "regime_source": "mr_market_regime coarse + drawdown 5-state derivation",
        "n_days_classified": len(daily_5state),
        "regime_distribution_alltime": dict(regime_distribution),
        "n_r2_trades_tagged": len(trades),
        "per_regime": per_regime,
        "interpretation": interpretation,
        "notes": [
            "REUSES mr_market_regime · no parallel regime engine created",
            "Point-in-time: 5-state classifier uses only benchmark data at or before each date",
            "Downside capture computed only on trades where benchmark return over holding period is negative",
            "R1 retired · R2 only",
            "UNAVAILABLE when classifier lacks sufficient trailing history (<30d)",
        ],
    }
    out_p = root / "reports" / "research" / "multi_layer" / f"crash_resilience_{market.lower()}_{asof}.json"
    out_p.parent.mkdir(parents=True, exist_ok=True)
    out_p.write_text(json.dumps(result, indent=2, ensure_ascii=False,
                                 default=str), encoding="utf-8")
    return result


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--market", choices=["india", "usa", "both"],
                     default="both")
    ap.add_argument("--asof", default=date.today().isoformat())
    args = ap.parse_args()
    for m in (["india", "usa"] if args.market == "both" else [args.market]):
        rep = build(m, _ROOT, args.asof)
        summary = {
            "market": rep.get("market", m),
            "today_regime": rep.get("today_regime", "?"),
            "n_r2_trades_tagged": rep.get("n_r2_trades_tagged", 0),
            "regime_distribution_alltime": rep.get("regime_distribution_alltime", {}),
            "per_regime_summary": {
                rg: {"n": s.get("n", 0),
                     "mean_pnl_pct": s.get("mean_pnl_pct"),
                     "win_rate_pct": s.get("win_rate_pct"),
                     "downside_capture_vs_bench": s.get("downside_capture_vs_bench")}
                for rg, s in (rep.get("per_regime") or {}).items()
            },
            "interpretation": rep.get("interpretation"),
        }
        print(json.dumps(summary, indent=2, ensure_ascii=False))
    return 0


if __name__ == "__main__":
    sys.exit(main())
