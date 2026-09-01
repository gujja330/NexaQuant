"""Stress-regime research · CEO 2026-09-01 §8-9-10.

Reuses EXISTING regime infrastructure (`backend/research/mr_market_regime`)
· does NOT invent parallel regime engines.

Question this module answers:
    When the market enters stress, does AEGIS actually behave differently,
    and does that behavior protect capital?

For each historical AEGIS R2 exit, tag the exit date with its regime
(BULL / BEAR / HIGH_VOL / NEUTRAL) and compute per-regime:
    · n_trades
    · win_rate_pct
    · realized_pnl_pct_sum
    · mean_pnl_pct
    · worst_trade_pct        (max realized drawdown per trade)
    · avg_holding_days

CEO §8 mandates comparing:
    · normal market
    · correction / bear
    · high-volatility
    · recovery (post-stress)

Recovery is defined here as: regime = BULL/NEUTRAL at exit AND the
prior 20 days included at least 5 HIGH_VOL or BEAR days.

Never modifies R2. Never auto-promotes anything.
Output: reports/research/multi_layer/stress_regime_{market}_{asof}.json
"""
from __future__ import annotations

import argparse
import json
import statistics
import sys
from collections import defaultdict
from datetime import date, timedelta
from pathlib import Path

_ROOT = Path(__file__).resolve().parents[3]
sys.path.insert(0, str(_ROOT))

from backend.research import mr_market_regime as _mmr  # existing · reuse
from openpyxl import load_workbook


def _find_hdr(rows):
    for i, r in enumerate(rows[:10]):
        if r and sum(1 for c in r if c is not None) >= 5:
            return i
    return 0


def _col(hdr, *names):
    for name in names:
        for i, c in enumerate(hdr):
            if c and str(c).lower() == name.lower(): return i
    return None


def _tag_recovery(regimes_by_date: dict, d: str) -> bool:
    """True if today's regime is BULL/NEUTRAL AND prior 20 days had ≥5 BEAR/HIGH_VOL days."""
    try:
        cur_d = date.fromisoformat(d)
    except ValueError:
        return False
    prior_stress = 0
    for offset in range(1, 21):
        prev = (cur_d - timedelta(days=offset)).isoformat()
        r = regimes_by_date.get(prev)
        if r in ("BEAR", "HIGH_VOL"):
            prior_stress += 1
    return prior_stress >= 5 and regimes_by_date.get(d) in ("BULL", "NEUTRAL")


def build(market: str, root: Path, asof: str) -> dict:
    # 1. Load existing regime classification (reuse · don't duplicate)
    regime_p = root / "reports" / "research" / f"mr_market_regime_{market.lower()}.json"
    regime_status = "USED_CACHED"
    if not regime_p.exists():
        regime_status = "COMPUTED_ON_THE_FLY"
        try:
            res = _mmr.build(root, market)
            _mmr.emit(root, market, res)
        except Exception as e:
            return {"error": f"regime build failed: {type(e).__name__}: {e}",
                     "regime_source": "unavailable"}
    reg_data = json.loads(regime_p.read_text(encoding="utf-8")) if regime_p.exists() else {}
    regimes_by_date = reg_data.get("regimes", {}) or {}

    # 2. Load R2 historical exits from Exit History (90d) sheet
    xlsx = root / "reports" / "telegram" / f"aegis_history_{market.lower()}.xlsx"
    r2_trades = []
    if xlsx.exists() and "Exit History (90d)" in load_workbook(xlsx, read_only=True).sheetnames:
        wb = load_workbook(xlsx, read_only=True, data_only=True)
        ws = wb["Exit History (90d)"]
        rows = list(ws.iter_rows(values_only=True))
        hi = _find_hdr(rows)
        hdr = rows[hi]
        c_tk = _col(hdr, "Stock", "Ticker")
        c_run = _col(hdr, "Runner")
        c_ent = _col(hdr, "Entry Date")
        c_ext = _col(hdr, "Exit Date")
        c_days = _col(hdr, "Days Held")
        c_pnl = _col(hdr, "P&L %")
        for r in rows[hi + 1:]:
            if not r or c_tk is None or not r[c_tk]: continue
            run = str(r[c_run] or "").upper() if c_run is not None else ""
            if run != "R2": continue
            exit_d = str(r[c_ext] or "")[:10] if c_ext is not None else ""
            try:
                pnl = float(r[c_pnl]) if c_pnl is not None and r[c_pnl] not in (None, "") else 0.0
            except (TypeError, ValueError):
                pnl = 0.0
            try:
                days = int(r[c_days]) if c_days is not None and r[c_days] not in (None, "") else 0
            except (TypeError, ValueError):
                days = 0
            # P&L in Exit History is expressed as fraction (-0.1062 = -10.62%)
            r2_trades.append({
                "ticker": str(r[c_tk] or ""),
                "exit_date": exit_d,
                "entry_date": str(r[c_ent] or "")[:10] if c_ent is not None else "",
                "pnl_pct": round(pnl * 100.0, 4) if abs(pnl) < 2.0 else round(pnl, 4),
                "days_held": days,
            })
        wb.close()

    # 3. Tag each trade with regime AND recovery flag
    by_regime = defaultdict(list)
    for t in r2_trades:
        rg = regimes_by_date.get(t["exit_date"], "UNKNOWN")
        t["regime"] = rg
        t["is_recovery"] = _tag_recovery(regimes_by_date, t["exit_date"])
        by_regime[rg].append(t)
        if t["is_recovery"]:
            by_regime["RECOVERY"].append(t)

    def _stats(trades: list) -> dict:
        if not trades:
            return {"n": 0, "win_rate_pct": None, "mean_pnl_pct": None,
                     "median_pnl_pct": None, "sum_pnl_pct": 0.0,
                     "worst_trade_pct": None, "avg_holding_days": None}
        pnls = [t["pnl_pct"] for t in trades]
        wins = sum(1 for p in pnls if p > 0)
        days = [t["days_held"] for t in trades if t["days_held"] > 0]
        return {
            "n": len(trades),
            "win_rate_pct": round(wins / len(pnls) * 100, 1),
            "mean_pnl_pct": round(statistics.mean(pnls), 3),
            "median_pnl_pct": round(statistics.median(pnls), 3),
            "sum_pnl_pct": round(sum(pnls), 3),
            "worst_trade_pct": round(min(pnls), 3),
            "avg_holding_days": round(statistics.mean(days), 1) if days else None,
        }

    per_regime = {rg: _stats(trs) for rg, trs in by_regime.items()}
    overall = _stats(r2_trades)

    result = {
        "engine": "stress_regime.multi_layer.v1",
        "market": market.lower(),
        "asof": asof,
        "regime_source": f"mr_market_regime · status={regime_status} · n_days_classified={len(regimes_by_date)}",
        "n_r2_trades_tagged": len(r2_trades),
        "overall": overall,
        "per_regime": per_regime,
        "regime_distribution_in_period": (reg_data.get("regime_distribution") or {}),
        "notes": [
            "R2 only (R1 retired) · Exit History (90d) window",
            "P&L expressed in % · positive = gain",
            "RECOVERY = current regime BULL/NEUTRAL AND ≥5 prior-20d BEAR/HIGH_VOL days",
            "Reused mr_market_regime engine · no parallel regime infrastructure created",
        ],
    }
    out_p = root / "reports" / "research" / "multi_layer" / f"stress_regime_{market.lower()}_{asof}.json"
    out_p.parent.mkdir(parents=True, exist_ok=True)
    out_p.write_text(json.dumps(result, indent=2, ensure_ascii=False),
                      encoding="utf-8")
    return result


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--market", choices=["india", "usa", "both"],
                     default="both")
    ap.add_argument("--asof", default=date.today().isoformat())
    args = ap.parse_args()
    for m in (["india", "usa"] if args.market == "both" else [args.market]):
        rep = build(m, _ROOT, args.asof)
        summary = {
            "market": rep.get("market", m),
            "regime_source": rep.get("regime_source", ""),
            "n_r2_trades_tagged": rep.get("n_r2_trades_tagged", 0),
            "overall": rep.get("overall", {}),
            "per_regime_summary": {
                rg: {"n": s.get("n", 0),
                     "mean_pnl_pct": s.get("mean_pnl_pct"),
                     "worst_trade_pct": s.get("worst_trade_pct")}
                for rg, s in (rep.get("per_regime") or {}).items()
            },
        }
        print(json.dumps(summary, indent=2, ensure_ascii=False))
    return 0


if __name__ == "__main__":
    sys.exit(main())
