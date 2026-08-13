"""P0 · Canonical Outcome Dataset builder.

One row per Position ID · immutable identity · frozen initial values ·
horizon returns · running extremes · terminal fields · sample-size tier.

Reads from:
  - reports/position_store/{market}/positions.json
  - usa/reports/position_store/usa/positions.json
  - reports/research/rank_history.jsonl (initial rank/conf/score snapshot)
  - reports/investability_{market}.json (initial investability snapshot)
  - reports/sector_cache.json
  - configs/india_universe_tiers.yaml
  - data/raw/india/*_D1.parquet (India price history)
  - usa/data/raw/us/*_D1.parquet (USA price history)
  - reports/telegram/aegis_history.xlsx (exit event log · fallback)

Writes to:
  - reports/research/outcome_dataset.parquet (canonical)
  - reports/research/outcome_dataset_summary.json (rollup + tiers)
"""
from __future__ import annotations

import json
from dataclasses import dataclass, asdict, field
from datetime import date, timedelta
from pathlib import Path


_HORIZONS = [1, 3, 5, 10, 17, 30, 60, 90]


def _load_yaml(p: Path) -> dict:
    import yaml
    if not p.exists(): return {}
    return yaml.safe_load(p.read_text(encoding="utf-8")) or {}


def _load_json(p: Path) -> dict | list:
    if not p.exists(): return {}
    try: return json.loads(p.read_text(encoding="utf-8"))
    except Exception: return {}


def _positions(root: Path, market: str) -> dict:
    """Load position_store for market · returns {canonical_ticker: record}."""
    base = root / ("usa/reports" if market.lower() == "usa" else "reports")
    p = base / "position_store" / market.lower() / "positions.json"
    d = _load_json(p) or {}
    return d.get("positions") or {}


def _rank_history(root: Path) -> list:
    p = root / "reports" / "research" / "rank_history.jsonl"
    if not p.exists(): return []
    out = []
    for line in p.read_text(encoding="utf-8").splitlines():
        if not line.strip(): continue
        try: out.append(json.loads(line))
        except Exception: continue
    return out


def _investability(root: Path, market: str) -> dict:
    p = root / "reports" / f"investability_{market.lower()}.json"
    d = _load_json(p) or {}
    return {r["ticker"]: r for r in (d.get("results") or [])}


def _sector_cache(root: Path) -> dict:
    return _load_json(root / "reports" / "sector_cache.json") or {}


def _cap_class(ticker: str, market: str, largecap_set: set) -> str:
    if market.lower() == "usa":
        return "LargeCap"
    short = str(ticker or "").replace(".NS", "").replace(".BO", "").upper()
    return "LargeCap" if short in largecap_set else "MidCap"


def _parquet_df(root: Path, ticker: str, market: str):
    """Load parquet · returns DataFrame indexed by date string."""
    import pandas as pd
    short = str(ticker or "").replace(".NS", "").replace(".BO", "").upper()
    base = "usa/data/raw/us" if market.lower() == "usa" else "data/raw/india"
    p = root / base / f"{short}_D1.parquet"
    if not p.exists(): return None
    try:
        df = pd.read_parquet(p)
        col = "close" if "close" in df.columns else "Close"
        df.index = pd.to_datetime(df.index).strftime("%Y-%m-%d")
        return df[[col]].rename(columns={col: "close"})
    except Exception:
        return None


def _price_at(df, target_date: str):
    """Return close on target_date · fallback to nearest earlier."""
    if df is None or df.empty: return None
    if target_date in df.index: return float(df.loc[target_date, "close"])
    earlier = [d for d in df.index if d <= target_date]
    return float(df.loc[earlier[-1], "close"]) if earlier else None


def _horizon_date(entry_iso: str, trading_days: int) -> str:
    """Return ISO date entry + trading_days (weekends skipped · holidays not)."""
    try:
        d = date.fromisoformat(entry_iso[:10])
    except (ValueError, TypeError):
        return entry_iso
    added = 0
    while added < trading_days:
        d += timedelta(days=1)
        if d.weekday() < 5: added += 1
    return d.isoformat()


def _find_earliest_rank(rank_hist: list, market: str, runner_key: str,
                                    ticker_bare: str) -> dict:
    """Return earliest rank_history entry for this (market, runner, ticker)."""
    entries = []
    for entry in rank_hist:
        if entry.get("market") != market.lower(): continue
        if entry.get("runner") != runner_key: continue
        t = str(entry.get("ticker") or "").replace(".NS","").replace(".BO","").upper()
        if t == ticker_bare.upper():
            entries.append(entry)
    if not entries: return {}
    entries.sort(key=lambda e: str(e.get("asof") or ""))
    return entries[0]


def _load_exit_events_index(root: Path) -> dict:
    """Load XLSX ONCE · return {(market, runner, ticker): exit_info}."""
    xlsx = root / "reports" / "telegram" / "aegis_history.xlsx"
    if not xlsx.exists(): return {}
    index = {}
    try:
        from openpyxl import load_workbook
        wb = load_workbook(xlsx, read_only=True)
        ws = wb["AEGIS Daily"] if "AEGIS Daily" in wb.sheetnames else wb.active
        h = [c.value for c in ws[1]]
        def col(name):
            try: return h.index(name) + 1
            except ValueError: return None
        c_ctry, c_run, c_tk, c_st, c_dt = col("Country"), col("Run_Type"), col("Ticker"), col("Status"), col("Date")
        c_pnl, c_reason = col("Exit P&L %"), col("Exit Reason")
        if not all([c_ctry, c_run, c_tk, c_st, c_dt]):
            wb.close(); return index
        for r in range(2, ws.max_row + 1):
            if ws.cell(r, c_st).value != "EXIT": continue
            market = str(ws.cell(r, c_ctry).value or "").lower()
            runner = str(ws.cell(r, c_run).value or "")
            ticker = str(ws.cell(r, c_tk).value or "").upper()
            ticker = ticker.replace(".NS","").replace(".BO","")
            key = (market, runner, ticker)
            index[key] = {
                "exit_date":   str(ws.cell(r, c_dt).value or "")[:10],
                "exit_pnl":    ws.cell(r, c_pnl).value if c_pnl else None,
                "exit_reason": ws.cell(r, c_reason).value if c_reason else None,
            }
        wb.close()
    except Exception:
        pass
    return index


def _sample_tier(n: int, tiers: list) -> str:
    for t in tiers:
        if t["min"] <= n <= t["max"]:
            return t["label"]
    return "unknown"


@dataclass
class OutcomeRow:
    position_id:                  str
    country:                      str
    runner:                       str
    ticker:                       str
    sector:                       str
    cap:                          str
    initial_rank:                 int | None
    initial_confidence:           float | None
    initial_model_score:          float | None
    initial_investability:        float | None
    initial_investability_verdict: str
    entry_date:                   str
    entry_price:                  float | None
    current_price:                float | None
    current_perf_pct:             float | None
    max_gain_pct:                 float | None
    max_drawdown_pct:             float | None
    ret_1d_pct:                   float | None
    ret_3d_pct:                   float | None
    ret_5d_pct:                   float | None
    ret_10d_pct:                  float | None
    ret_17d_pct:                  float | None
    ret_30d_pct:                  float | None
    ret_60d_pct:                  float | None
    ret_90d_pct:                  float | None
    exit_date:                    str
    exit_price:                   float | None
    exit_pnl_pct:                 float | None
    exit_reason:                  str
    win_flag:                     bool | None   # None if open (per governance rule)
    lifecycle:                    str
    is_closed:                    bool
    sample_tier_note:             str = ""


def build(root: Path) -> list:
    """Build canonical outcome rows from all sources · one per Position ID."""
    tiers_cfg = _load_yaml(root / "configs" / "outcome_dataset_schema.yaml")
    tiers = (tiers_cfg.get("sample_size_tiers") or [])

    india_tiers = _load_yaml(root / "configs" / "india_universe_tiers.yaml")
    largecap_set = set(india_tiers.get("largecap_tickers") or [])
    sector_cache = _sector_cache(root)
    rank_hist = _rank_history(root)

    rows = []
    today_iso = date.today().isoformat()
    exit_index = _load_exit_events_index(root)   # load XLSX ONCE

    for market in ("india", "usa"):
        positions = _positions(root, market)
        inv_map = _investability(root, market)
        market_sector = sector_cache.get(market.lower(), {})

        for canonical_ticker, pos in positions.items():
            ticker_bare = canonical_ticker.replace(".NS","").replace(".BO","").upper()
            entry_date = str(pos.get("first_seen_date") or "")[:10]
            entry_price = pos.get("first_seen_price")

            if not entry_date or not entry_price: continue

            df = _parquet_df(root, ticker_bare, market)
            current_price = _price_at(df, today_iso) if df is not None else None
            current_perf = None
            if entry_price and current_price and entry_price > 0:
                current_perf = round((current_price - entry_price) / entry_price * 100, 2)

            # Running max/min between entry and today
            max_gain = max_dd = None
            if df is not None:
                window = df[(df.index >= entry_date) & (df.index <= today_iso)]
                if not window.empty and entry_price:
                    highest = float(window["close"].max())
                    lowest = float(window["close"].min())
                    max_gain = round((highest - entry_price) / entry_price * 100, 2)
                    max_dd = round((lowest - entry_price) / entry_price * 100, 2)

            # Horizon returns
            horizon_returns = {}
            for h in _HORIZONS:
                target = _horizon_date(entry_date, h)
                if target > today_iso:
                    horizon_returns[h] = None
                    continue
                px = _price_at(df, target) if df is not None else None
                if px and entry_price and entry_price > 0:
                    horizon_returns[h] = round((px - entry_price) / entry_price * 100, 2)
                else:
                    horizon_returns[h] = None

            # Determine runner via rank_history · earliest entry sets runner
            # (rank_history stamps runner per emission)
            r1_earliest = _find_earliest_rank(rank_hist, market, "runner1", ticker_bare)
            r2_earliest = _find_earliest_rank(rank_hist, market, "runner2", ticker_bare)

            for runner_key, earliest, runner_label in [
                ("runner1", r1_earliest, "R1"),
                ("runner2", r2_earliest, "R2"),
            ]:
                if not earliest: continue

                inv = inv_map.get(ticker_bare) or inv_map.get(canonical_ticker) or {}
                sector = market_sector.get(ticker_bare, "") or "Unclassified"
                cap = _cap_class(ticker_bare, market, largecap_set)

                # Look for exit event (indexed lookup · O(1))
                exit_info = exit_index.get((market.lower(), runner_label, ticker_bare.upper()), {})
                exit_date = exit_info.get("exit_date") or ""
                exit_price = _price_at(df, exit_date) if exit_date else None
                exit_pnl = exit_info.get("exit_pnl")
                if exit_price and entry_price and entry_price > 0 and not exit_pnl:
                    exit_pnl = round((exit_price - entry_price) / entry_price * 100, 2)

                is_closed = bool(exit_date and exit_price)
                # Governance rule: win_flag NULL for open positions
                win_flag = None if not is_closed else (
                    exit_pnl > 0 if isinstance(exit_pnl, (int, float)) else None
                )

                # Lifecycle
                if is_closed:
                    lifecycle = "CLOSED"
                elif entry_date == today_iso:
                    lifecycle = "NEW"
                else:
                    lifecycle = "ACTIVE"

                # 2026-08-12 CEO P0 integrity fix · include runner in position_id
                # so (position_id) alone is a true primary key. Previously
                # {TICKER}_{MKT}_{DATE} collided when the same position was
                # picked by both R1 and R2 (LUPIN_IND_20260731 · COALINDIA_IND_20260731
                # both showed 2 rows with identical id). Attribution and P1
                # analyses group by position_id · duplicates skewed counts.
                mkt_upper = "IND" if market.lower() == "india" else "USA"
                position_id = f"{ticker_bare}_{mkt_upper}_{entry_date.replace('-','')}_{runner_label}"

                row = OutcomeRow(
                    position_id=position_id,
                    country=market.upper(),
                    runner=runner_label,
                    ticker=ticker_bare,
                    sector=sector,
                    cap=cap,
                    initial_rank=earliest.get("rank"),
                    initial_confidence=earliest.get("confidence"),
                    initial_model_score=earliest.get("model_score"),
                    initial_investability=inv.get("score"),
                    initial_investability_verdict=inv.get("verdict", ""),
                    entry_date=entry_date,
                    entry_price=round(float(entry_price), 2) if entry_price else None,
                    current_price=round(current_price, 2) if current_price else None,
                    current_perf_pct=current_perf,
                    max_gain_pct=max_gain,
                    max_drawdown_pct=max_dd,
                    ret_1d_pct=horizon_returns.get(1),
                    ret_3d_pct=horizon_returns.get(3),
                    ret_5d_pct=horizon_returns.get(5),
                    ret_10d_pct=horizon_returns.get(10),
                    ret_17d_pct=horizon_returns.get(17),
                    ret_30d_pct=horizon_returns.get(30),
                    ret_60d_pct=horizon_returns.get(60),
                    ret_90d_pct=horizon_returns.get(90),
                    exit_date=exit_date,
                    exit_price=round(exit_price, 2) if exit_price else None,
                    exit_pnl_pct=exit_pnl if isinstance(exit_pnl, (int, float)) else None,
                    exit_reason=str(exit_info.get("exit_reason") or ""),
                    win_flag=win_flag,
                    lifecycle=lifecycle,
                    is_closed=is_closed,
                )
                rows.append(row)

    # Attach sample_tier_note per row · based on total count in row's runner
    from collections import Counter
    runner_counts = Counter(r.runner for r in rows)
    for r in rows:
        r.sample_tier_note = _sample_tier(runner_counts[r.runner], tiers)

    return rows


def emit(root: Path, rows: list) -> dict:
    """Write parquet + summary JSON. Returns summary dict."""
    import pandas as pd

    out_dir = root / "reports" / "research"
    out_dir.mkdir(parents=True, exist_ok=True)

    parquet_path = out_dir / "outcome_dataset.parquet"
    summary_path = out_dir / "outcome_dataset_summary.json"

    if not rows:
        return {"n": 0, "note": "no positions found"}

    df = pd.DataFrame([asdict(r) for r in rows])
    df.to_parquet(parquet_path, index=False)

    # Summary rollups
    from collections import Counter
    tiers_cfg = _load_yaml(root / "configs" / "outcome_dataset_schema.yaml")
    tiers = tiers_cfg.get("sample_size_tiers") or []

    by_country = Counter(r.country for r in rows)
    by_runner = Counter(r.runner for r in rows)
    closed_count = sum(1 for r in rows if r.is_closed)
    open_count = sum(1 for r in rows if not r.is_closed)

    # Runner performance (CLOSED positions only · per governance)
    runner_perf = {}
    for runner_label in ("R1", "R2"):
        closed_r = [r for r in rows if r.runner == runner_label and r.is_closed
                            and isinstance(r.exit_pnl_pct, (int, float))]
        if not closed_r:
            runner_perf[runner_label] = {"n_closed": 0, "tier": _sample_tier(0, tiers)}
            continue
        n = len(closed_r)
        wins = sum(1 for r in closed_r if r.win_flag)
        avg = round(sum(r.exit_pnl_pct for r in closed_r) / n, 3)
        runner_perf[runner_label] = {
            "n_closed": n,
            "win_rate_pct": round(wins / n * 100, 1),
            "avg_pnl_pct": avg,
            "median_pnl_pct": round(sorted([r.exit_pnl_pct for r in closed_r])[n//2], 3),
            "tier": _sample_tier(n, tiers),
        }

    summary = {
        "engine":       "outcome_dataset.v1",
        "generated_utc": __import__("datetime").datetime.now(__import__("datetime").timezone.utc).isoformat(timespec="seconds"),
        "n_positions":  len(rows),
        "n_closed":     closed_count,
        "n_open":       open_count,
        "by_country":   dict(by_country),
        "by_runner":    dict(by_runner),
        "runner_perf_closed_only": runner_perf,
        "parquet_path": str(parquet_path.relative_to(root)),
        "governance":   "win_rate uses CLOSED only · open positions excluded",
    }
    summary_path.write_text(
        json.dumps(summary, indent=2, default=str, ensure_ascii=False),
        encoding="utf-8"
    )
    return summary
