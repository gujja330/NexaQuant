# backend/research/lifecycle_stabilization.py
"""AEGIS · Sprint M · Phase A · Lifecycle Stabilization Engine.

CEO directive 2026-08-25 v2.0: "next phase = AEGIS Validation & Opportunity
Quality · fix the canonical lifecycle first".

Runs the 10 Phase-A audits against the Registry + delivered XLSX +
sender output and emits ONE consolidated verdict per audit. Nothing
ships (Phases B-E) until every audit here is GREEN.

Locks respected (never touched):
  LOCK 1 · Excel sheet format is FROZEN
  LOCK 2 · Position lifecycle is NEW → ACTIVE → ACTIVE+ → EXIT/CLOSED

Audits (10):
  A1  · Position ID immutable across state transitions
  A2  · one-position lifecycle (no duplicate active for same ticker×runner)
  A3  · NEW / EXISTING / RE-ENTRY classifier is correct
  A4  · no duplicate active positions
  A5  · historical CLOSED rows never surface as NEW without new Position ID
  A6  · SKIP tickers never appear in investor Portfolio sheet
  A7  · SKIP tickers never contribute to P&L calculations
  A8  · Active P&L formula uses (Current/Entry - 1) · OPEN only
  A9  · Exit P&L formula uses (Exit/Entry - 1) · CLOSED only
  A10 · entry/current/exit prices reconcile to parquet closes (±0.5%)

Every audit returns one of PASS / WARN / FAIL. Report is written to
reports/context/lifecycle_stabilization_{market}.json.
"""
from __future__ import annotations

import json
from dataclasses import dataclass, field, asdict
from datetime import date, datetime, timezone, timedelta
from pathlib import Path
from typing import Optional


SCHEMA_FINGERPRINT = "aegis.lifecycle_stabilization.v1.20260825"

# LOCK 2 · CEO tightened 2026-08-25 (2nd pass · "exit or closed use only
# one word plz, dont confuse") · terminal state is ONE word: EXIT.
# Canonical lifecycle is exactly 4 states:  NEW / ACTIVE / ACTIVE+ / EXIT
# Registry may persist "CLOSED" for historical reasons · canonicalize()
# maps it to EXIT so display + audits only ever show the 4.
VALID_STATES = {"NEW", "ACTIVE", "ACTIVE+", "EXIT"}
LEGACY_STATES = {"CLOSED", "HOLD", "ROTATED_SAMEDAY"}
ALL_TOLERATED_STATES = VALID_STATES | LEGACY_STATES


def canonicalize_state(state: str) -> str:
    """Map any tolerated state to one of the 4 canonical CEO-locked states.
    CLOSED → EXIT · HOLD → ACTIVE · ROTATED_SAMEDAY → EXIT."""
    s = str(state or "").upper().strip()
    if s in VALID_STATES: return s
    if s in ("CLOSED", "ROTATED_SAMEDAY"): return "EXIT"
    if s == "HOLD": return "ACTIVE"
    return s   # unknown · leave as-is so audits catch it

# LOCK 2 · legitimate transitions (both directions allowed for RE-ENTRY)
# but re-entry MUST get a NEW Position ID (audit A5)


@dataclass
class AuditCheck:
    code: str                    # A1..A10
    name: str
    status: str                  # PASS / WARN / FAIL
    detail: str
    violations: list = field(default_factory=list)


@dataclass
class StabilizationReport:
    market: str
    asof: str
    generated_utc: str
    engine: str = SCHEMA_FINGERPRINT
    verdict: str = "PASS"
    n_audits: int = 10
    n_pass: int = 0
    n_warn: int = 0
    n_fail: int = 0
    audits: list = field(default_factory=list)

    def add(self, check: AuditCheck) -> None:
        self.audits.append(check)
        if check.status == "FAIL":
            self.n_fail += 1
        elif check.status == "WARN":
            self.n_warn += 1
        else:
            self.n_pass += 1
        rank = {"PASS": 0, "WARN": 1, "FAIL": 2}
        if rank[check.status] > rank[self.verdict]:
            self.verdict = check.status


# ─────────────────────────────────────────────────────────────────
# A1 · Position ID immutable across state transitions
# ─────────────────────────────────────────────────────────────────
def audit_a1_position_id_immutable(root: Path, market: str) -> AuditCheck:
    """Every Position ID (opportunity_id in Registry) must appear at most
    once per (created_date, ticker, runner). Reuse = LOCK violation."""
    try:
        from backend.research import opportunity_registry as _oreg
        reg = _oreg.load_all(root)
    except Exception as e:
        return AuditCheck("A1", "Position ID immutable", "WARN",
                          f"Registry not readable · {type(e).__name__}: {e}")
    seen: dict = {}
    violations = []
    for opps in reg.values():
        for o in opps:
            if o.market.lower() != market.lower(): continue
            _pid = getattr(o, "opportunity_id", None) or \
                   f"{o.ticker}_{o.runner}_{o.created_date}"
            key = _pid
            if key in seen:
                violations.append({
                    "position_id": key,
                    "duplicate_ticker": o.ticker,
                    "duplicate_runner": o.runner,
                    "original": str(seen[key]),
                    "reason": "same Position ID used twice · lifecycle violation",
                })
            else:
                seen[key] = (o.ticker, o.runner, o.created_date)
    status = "PASS" if not violations else "FAIL"
    detail = (f"{len(seen)} Position IDs · {len(violations)} duplicated")
    return AuditCheck("A1", "Position ID immutable", status, detail, violations)


# ─────────────────────────────────────────────────────────────────
# A2 · one-position lifecycle · no duplicate active for same ticker×runner
# ─────────────────────────────────────────────────────────────────
def audit_a2_one_active_per_ticker_runner(root: Path, market: str) -> AuditCheck:
    try:
        from backend.research import opportunity_registry as _oreg
        reg = _oreg.load_all(root)
    except Exception as e:
        return AuditCheck("A2", "One active per (ticker, runner)", "WARN",
                          f"Registry not readable · {type(e).__name__}: {e}")
    active_pairs: dict = {}
    violations = []
    for opps in reg.values():
        for o in opps:
            if o.market.lower() != market.lower(): continue
            if not o.is_active(): continue
            key = (o.ticker.upper(), o.runner.upper().replace("_NEW", ""))
            if key in active_pairs:
                violations.append({
                    "ticker": o.ticker, "runner": o.runner,
                    "existing_position_id": str(active_pairs[key]),
                    "new_position_id": getattr(o, "opportunity_id",
                                               f"{o.ticker}_{o.runner}_{o.created_date}"),
                    "reason": "two active positions for same (ticker, runner)",
                })
            else:
                active_pairs[key] = getattr(o, "opportunity_id",
                                            f"{o.ticker}_{o.runner}_{o.created_date}")
    status = "PASS" if not violations else "FAIL"
    detail = (f"{len(active_pairs)} unique active (ticker, runner) pairs · "
              f"{len(violations)} duplicates")
    return AuditCheck("A2", "One active per (ticker, runner)",
                      status, detail, violations)


# ─────────────────────────────────────────────────────────────────
# A3 · NEW / EXISTING / RE-ENTRY classifier (Sprint M Phase A · CEO Part 6)
# ─────────────────────────────────────────────────────────────────
def classify_opportunity_state(
    *, ticker: str, market: str, runner: str, rec_date: str, asof: str,
    registry_history: list,
) -> str:
    """Correct classification per LOCK 2:
       NEW      · first-time ticker+runner OR previous closed + this
                  entry has different Position ID (RE-ENTRY subtype)
       EXISTING · currently active in Registry
       RE-ENTRY · previously CLOSED and now new Position ID
       CLOSED   · Registry says CLOSED
    """
    tk = str(ticker).upper().replace(".NS","").replace(".BO","")
    rn = str(runner).upper().replace("_NEW", "")
    prior = [h for h in registry_history
             if h.get("ticker","").upper() == tk
             and h.get("runner","").upper().replace("_NEW","") == rn]
    if not prior:
        return "NEW"
    has_active = any(h.get("status") in ("NEW","ACTIVE","ACTIVE+","HOLD")
                     for h in prior)
    if has_active:
        return "EXISTING"
    all_closed = all(h.get("status") in ("EXIT","CLOSED","ROTATED_SAMEDAY")
                     for h in prior)
    if all_closed and rec_date == asof:
        return "RE-ENTRY"
    return "CLOSED"


def audit_a3_new_existing_reentry_classifier(root: Path, market: str) -> AuditCheck:
    """A3 audit · run the classifier on Registry's own entries and verify
    outputs are consistent with the state machine."""
    try:
        from backend.research import opportunity_registry as _oreg
        reg = _oreg.load_all(root)
    except Exception as e:
        return AuditCheck("A3", "NEW/EXISTING/RE-ENTRY classifier", "WARN",
                          f"Registry not readable · {type(e).__name__}: {e}")
    asof = date.today().isoformat()
    all_entries = []
    for opps in reg.values():
        for o in opps:
            if o.market.lower() != market.lower(): continue
            all_entries.append({
                "ticker": o.ticker, "runner": o.runner,
                "status": o.status, "created_date": o.created_date,
                "closed_date": o.closed_date,
            })
    violations = []
    n_ok = 0
    for e in all_entries:
        history_before = [h for h in all_entries
                          if (h.get("created_date","") or "") < (e.get("created_date","") or "")]
        classified = classify_opportunity_state(
            ticker=e["ticker"], market=market, runner=e["runner"],
            rec_date=e["created_date"], asof=asof,
            registry_history=history_before,
        )
        # Sanity · NEW state should have no prior active for same (ticker, runner)
        if classified == "NEW" and any(
            (h.get("status") in ("NEW","ACTIVE","ACTIVE+"))
            and h.get("ticker","").upper() == e["ticker"].upper()
            and h.get("runner","").upper().replace("_NEW","") == e["runner"].upper().replace("_NEW","")
            for h in history_before
        ):
            violations.append({
                "ticker": e["ticker"], "runner": e["runner"],
                "date": e["created_date"],
                "classified": classified,
                "reason": "classified NEW but active prior exists · lifecycle violation",
            })
        else:
            n_ok += 1
    status = "PASS" if not violations else "FAIL"
    detail = f"{n_ok} entries classified cleanly · {len(violations)} violations"
    return AuditCheck("A3", "NEW/EXISTING/RE-ENTRY classifier",
                      status, detail, violations)


# ─────────────────────────────────────────────────────────────────
# A4 · Prevent duplicate active positions
# (already covered by A2 · this is a delivered-XLSX cross-check)
# ─────────────────────────────────────────────────────────────────
def audit_a4_no_dup_active_in_portfolio(root: Path, market: str) -> AuditCheck:
    xp = root / "reports" / "telegram" / f"aegis_history_{market.lower()}.xlsx"
    if not xp.exists():
        return AuditCheck("A4", "No dup active in Portfolio", "WARN",
                          "per-market XLSX missing")
    try:
        from openpyxl import load_workbook
        wb = load_workbook(xp, read_only=True)
        if "Portfolio" not in wb.sheetnames:
            wb.close()
            return AuditCheck("A4", "No dup active in Portfolio", "WARN",
                              "Portfolio sheet missing")
        seen = {}
        violations = []
        for row in wb["Portfolio"].iter_rows(min_row=6, values_only=True):
            _tk = str(row[0] or "").upper().replace(".NS","").replace(".BO","")
            if not _tk: continue
            if _tk.startswith(("🟢","🔴","🆕","🟣","AEGIS","📊","🩺","✅","❌")):
                continue
            # Runner is at col index 8 (0-based) in current layout
            _rn = str(row[8] or "").upper().replace("_NEW","").replace("🔹 INDIA ONLY","INDIA")
            key = (_tk, _rn)
            if key in seen:
                violations.append({"ticker": _tk, "runner": _rn})
            else:
                seen[key] = True
        wb.close()
        status = "PASS" if not violations else "FAIL"
        return AuditCheck("A4", "No dup active in Portfolio", status,
                          f"{len(seen)} unique (ticker,runner) · {len(violations)} dups",
                          violations)
    except Exception as e:
        return AuditCheck("A4", "No dup active in Portfolio", "WARN",
                          f"{type(e).__name__}: {e}")


# ─────────────────────────────────────────────────────────────────
# A5 · Historical CLOSED rows never surface as NEW without new Position ID
# ─────────────────────────────────────────────────────────────────
def audit_a5_no_closed_as_new(root: Path, market: str) -> AuditCheck:
    """For every ticker+runner in Registry with a CLOSED entry, verify
    that any newer NEW entry has a DIFFERENT Position ID."""
    try:
        from backend.research import opportunity_registry as _oreg
        reg = _oreg.load_all(root)
    except Exception as e:
        return AuditCheck("A5", "No CLOSED → NEW leaks", "WARN",
                          f"Registry not readable · {type(e).__name__}: {e}")
    by_pair: dict = {}
    for opps in reg.values():
        for o in opps:
            if o.market.lower() != market.lower(): continue
            key = (o.ticker.upper(),
                   o.runner.upper().replace("_NEW",""))
            by_pair.setdefault(key, []).append(o)
    violations = []
    for (tk, rn), items in by_pair.items():
        items_sorted = sorted(
            items, key=lambda x: str(x.created_date or ""))
        closed_ids = set()
        for i in items_sorted:
            _pid = getattr(i, "opportunity_id", None) or \
                   f"{i.ticker}_{i.runner}_{i.created_date}"
            if i.status in ("CLOSED",):
                closed_ids.add(_pid)
            elif i.status in ("NEW","ACTIVE","ACTIVE+","HOLD"):
                if _pid in closed_ids:
                    violations.append({
                        "ticker": tk, "runner": rn,
                        "position_id_reused": _pid,
                        "reason": "closed Position ID surfaces as active/NEW · LOCK 2 violation",
                    })
    status = "PASS" if not violations else "FAIL"
    detail = f"{len(by_pair)} (ticker,runner) pairs · {len(violations)} leaks"
    return AuditCheck("A5", "No CLOSED → NEW leaks",
                      status, detail, violations)


# ─────────────────────────────────────────────────────────────────
# A6 · SKIP tickers never in Portfolio
# ─────────────────────────────────────────────────────────────────
def audit_a6_no_skip_in_portfolio(root: Path, market: str) -> AuditCheck:
    xp = root / "reports" / "telegram" / f"aegis_history_{market.lower()}.xlsx"
    if not xp.exists():
        return AuditCheck("A6", "No SKIP in Portfolio", "WARN",
                          "per-market XLSX missing")
    try:
        from openpyxl import load_workbook
        wb = load_workbook(xp, read_only=True)
        if "Portfolio" not in wb.sheetnames:
            wb.close()
            return AuditCheck("A6", "No SKIP in Portfolio", "WARN",
                              "Portfolio sheet missing")
        violations = []
        for row in wb["Portfolio"].iter_rows(min_row=6, values_only=True):
            # Action column (col 2) shouldn't say SKIP · Status hidden col 20
            _action = str(row[1] or "").upper()
            _decision = str(row[2] or "").upper()
            _tk = str(row[0] or "")
            if "SKIP" in _action or _decision == "SKIP":
                if _tk and not _tk.startswith(("🟢","🔴","🆕","🟣","AEGIS")):
                    violations.append({
                        "ticker": _tk, "action": _action[:40],
                        "decision": _decision[:40],
                    })
        wb.close()
        status = "PASS" if not violations else "FAIL"
        return AuditCheck("A6", "No SKIP in Portfolio", status,
                          f"{len(violations)} SKIP rows found",
                          violations)
    except Exception as e:
        return AuditCheck("A6", "No SKIP in Portfolio", "WARN",
                          f"{type(e).__name__}: {e}")


# ─────────────────────────────────────────────────────────────────
# A7 · SKIP never in P&L
# ─────────────────────────────────────────────────────────────────
def audit_a7_no_skip_in_pnl(root: Path, market: str) -> AuditCheck:
    """Cross-check: any row with Status=SKIP in the unified XLSX must NOT
    have a non-zero P&L computed."""
    xp = root / "reports" / "telegram" / "aegis_history.xlsx"
    if not xp.exists():
        return AuditCheck("A7", "No SKIP in P&L", "WARN",
                          "unified XLSX missing")
    try:
        from openpyxl import load_workbook
        wb = load_workbook(xp, read_only=True)
        sh = wb["AEGIS Daily"]
        h = [c.value for c in sh[1]]
        try:
            c_country = h.index("Country") + 1
            c_st = h.index("Status") + 1
        except ValueError:
            wb.close()
            return AuditCheck("A7", "No SKIP in P&L", "WARN",
                              "column headers missing")
        c_perf = (h.index("Current Perf %") + 1
                  if "Current Perf %" in h else None)
        c_exit_pnl = (h.index("Exit P&L %") + 1
                      if "Exit P&L %" in h else None)
        violations = []
        for row in sh.iter_rows(min_row=2, values_only=True):
            if str(row[c_country-1] or "").upper() != market.upper():
                continue
            _st = str(row[c_st-1] or "").upper()
            if _st != "SKIP": continue
            _perf = row[c_perf-1] if c_perf else None
            _exp = row[c_exit_pnl-1] if c_exit_pnl else None
            if (isinstance(_perf, (int, float)) and abs(_perf) > 0.01) or \
               (isinstance(_exp, (int, float)) and abs(_exp) > 0.01):
                violations.append({
                    "ticker": row[h.index("Ticker")] if "Ticker" in h else "?",
                    "status": _st, "perf": _perf, "exit_pnl": _exp,
                })
        wb.close()
        status = "PASS" if not violations else "FAIL"
        return AuditCheck("A7", "No SKIP in P&L", status,
                          f"{len(violations)} SKIP rows with non-zero P&L",
                          violations)
    except Exception as e:
        return AuditCheck("A7", "No SKIP in P&L", "WARN",
                          f"{type(e).__name__}: {e}")


# ─────────────────────────────────────────────────────────────────
# A8 · Active P&L formula = (Current/Entry - 1) · OPEN only
# A9 · Exit P&L formula = (Exit/Entry - 1) · CLOSED only
# Combined into one audit (they share the source rows)
# ─────────────────────────────────────────────────────────────────
def audit_a8_a9_pnl_formulas(root: Path, market: str) -> tuple:
    xp = root / "reports" / "telegram" / "aegis_history.xlsx"
    if not xp.exists():
        return (AuditCheck("A8", "Active P&L formula", "WARN", "unified XLSX missing"),
                AuditCheck("A9", "Exit P&L formula", "WARN", "unified XLSX missing"))
    try:
        from openpyxl import load_workbook
        wb = load_workbook(xp, read_only=True)
        sh = wb["AEGIS Daily"]
        h = [c.value for c in sh[1]]
        c_country = h.index("Country") + 1
        c_st = h.index("Status") + 1
        c_entry = h.index("Entry Price") + 1 if "Entry Price" in h else None
        c_curr = h.index("Current Price") + 1 if "Current Price" in h else None
        c_perf = h.index("Current Perf %") + 1 if "Current Perf %" in h else None
        c_exit_pnl = h.index("Exit P&L %") + 1 if "Exit P&L %" in h else None
        active_viol = []; exit_viol = []
        tol_pct = 1.0     # 1% tolerance
        n_active = 0; n_exit = 0
        for row in sh.iter_rows(min_row=2, values_only=True):
            if str(row[c_country-1] or "").upper() != market.upper():
                continue
            _st = str(row[c_st-1] or "").upper()
            _e = row[c_entry-1] if c_entry else None
            _c = row[c_curr-1] if c_curr else None
            _perf = row[c_perf-1] if c_perf else None
            _xp_l = row[c_exit_pnl-1] if c_exit_pnl else None
            if _st == "EXIT":
                n_exit += 1
                if (isinstance(_e, (int, float)) and _e > 0
                    and isinstance(_c, (int, float))
                    and isinstance(_xp_l, (int, float))):
                    expected = (_c - _e) / _e * 100
                    if abs(_xp_l - expected) > tol_pct:
                        exit_viol.append({
                            "ticker": row[h.index("Ticker")] if "Ticker" in h else "?",
                            "entry": _e, "exit": _c,
                            "quoted_pnl": _xp_l, "expected": round(expected, 2),
                        })
                # A9 additional: ACTIVE-perf cell must be blank for EXIT rows
                if isinstance(_perf, (int, float)) and abs(_perf) > 0.01:
                    exit_viol.append({
                        "ticker": row[h.index("Ticker")] if "Ticker" in h else "?",
                        "reason": "Active P&L populated on EXIT row (should be blank)",
                    })
            else:
                n_active += 1
                if (isinstance(_e, (int, float)) and _e > 0
                    and isinstance(_c, (int, float))
                    and isinstance(_perf, (int, float))):
                    expected = (_c - _e) / _e * 100
                    if abs(_perf - expected) > tol_pct:
                        active_viol.append({
                            "ticker": row[h.index("Ticker")] if "Ticker" in h else "?",
                            "entry": _e, "current": _c,
                            "quoted_pnl": _perf, "expected": round(expected, 2),
                        })
                # A8 additional: Exit P&L cell must be blank for ACTIVE rows
                if isinstance(_xp_l, (int, float)) and abs(_xp_l) > 0.01:
                    active_viol.append({
                        "ticker": row[h.index("Ticker")] if "Ticker" in h else "?",
                        "reason": "Exit P&L populated on ACTIVE row (should be blank)",
                    })
        wb.close()
        a8 = AuditCheck(
            "A8", "Active P&L formula", "PASS" if not active_viol else "FAIL",
            f"{n_active} active rows · {len(active_viol)} violations",
            active_viol[:10])
        a9 = AuditCheck(
            "A9", "Exit P&L formula", "PASS" if not exit_viol else "FAIL",
            f"{n_exit} exit rows · {len(exit_viol)} violations",
            exit_viol[:10])
        return (a8, a9)
    except Exception as e:
        return (AuditCheck("A8", "Active P&L formula", "WARN",
                           f"{type(e).__name__}: {e}"),
                AuditCheck("A9", "Exit P&L formula", "WARN",
                           f"{type(e).__name__}: {e}"))


# ─────────────────────────────────────────────────────────────────
# A10 · Prices reconcile to parquet closes (delegate to price_integrity)
# ─────────────────────────────────────────────────────────────────
def audit_a10_price_alignment(root: Path, market: str) -> AuditCheck:
    p = root / "reports" / "context" / f"price_integrity_{market.lower()}.json"
    if not p.exists():
        return AuditCheck("A10", "Prices reconcile to parquet", "WARN",
                          "price_integrity JSON not present · run sender first")
    try:
        d = json.loads(p.read_text(encoding="utf-8"))
        _verdict = d.get("verdict", "?")
        _fails = [c for c in d.get("checks", [])
                  if c.get("status") == "FAIL"
                  and c.get("code") in ("PI1", "PI2")]
        n_viol = sum(len(c.get("violations", [])) for c in _fails)
        status = "PASS" if _verdict == "PASS" else ("WARN" if _verdict == "WARN" else "FAIL")
        detail = (f"price_integrity verdict={_verdict} · "
                  f"{len(_fails)} PI1/PI2 fails · "
                  f"{n_viol} price mismatches")
        return AuditCheck("A10", "Prices reconcile to parquet",
                          status, detail, _fails[:3])
    except Exception as e:
        return AuditCheck("A10", "Prices reconcile to parquet", "WARN",
                          f"{type(e).__name__}: {e}")


# ─────────────────────────────────────────────────────────────────
# PUBLIC · compute + emit
# ─────────────────────────────────────────────────────────────────
def compute(root: Path, market: str) -> StabilizationReport:
    rep = StabilizationReport(
        market=market.lower(),
        asof=date.today().isoformat(),
        generated_utc=datetime.now(timezone.utc).isoformat(timespec="seconds"),
    )
    rep.add(audit_a1_position_id_immutable(root, market))
    rep.add(audit_a2_one_active_per_ticker_runner(root, market))
    rep.add(audit_a3_new_existing_reentry_classifier(root, market))
    rep.add(audit_a4_no_dup_active_in_portfolio(root, market))
    rep.add(audit_a5_no_closed_as_new(root, market))
    rep.add(audit_a6_no_skip_in_portfolio(root, market))
    rep.add(audit_a7_no_skip_in_pnl(root, market))
    a8, a9 = audit_a8_a9_pnl_formulas(root, market)
    rep.add(a8)
    rep.add(a9)
    rep.add(audit_a10_price_alignment(root, market))
    return rep


def emit(root: Path, report: StabilizationReport) -> Path:
    p = (root / "reports" / "context"
         / f"lifecycle_stabilization_{report.market}.json")
    p.parent.mkdir(parents=True, exist_ok=True)
    p.write_text(json.dumps(asdict(report), indent=2, default=str,
                            ensure_ascii=False), encoding="utf-8")
    return p


def summary_line(rep: StabilizationReport) -> str:
    return (f"lifecycle_stabilization · verdict={rep.verdict} · "
            f"PASS={rep.n_pass} · WARN={rep.n_warn} · FAIL={rep.n_fail}")


def render_markdown(rep: StabilizationReport) -> str:
    lines = [
        f"# Sprint M Phase A · Lifecycle Stabilization · {rep.market.upper()}",
        f"## Asof {rep.asof} · Verdict {rep.verdict}",
        "",
        f"{rep.n_pass} PASS · {rep.n_warn} WARN · {rep.n_fail} FAIL of {rep.n_audits}",
        "",
    ]
    for a in rep.audits:
        _icon = {"PASS":"✅", "WARN":"⚠️", "FAIL":"❌"}.get(a.status, "?")
        lines.append(f"- {_icon} **{a.code}** · {a.name} · {a.detail}")
        if a.status == "FAIL" and a.violations:
            for v in a.violations[:3]:
                lines.append(f"    - {v}")
    return "\n".join(lines)
