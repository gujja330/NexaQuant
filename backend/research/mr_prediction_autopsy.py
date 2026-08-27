"""AEGIS · M-R · Prediction Autopsy · Sprint M Phase A.

CEO handover 2026-08-27:
> "Historical daily predictions → forward outcomes. For every daily
>  recommendation over the last month, create one immutable research row:
>  Prediction date → ticker → runner → rank → decision → all available
>  features → then what actually happened 1/3/5/10/20 trading days later."

Scope: EVERY row in aegis_history.xlsx AEGIS Daily sheet (~597 India rows),
NOT just currently-held positions. Answers "what did AEGIS believe every
day and what happened afterward" · not "what is currently held".

Emits:
  reports/research/mr_prediction_autopsy_{market}.jsonl  (one row per pred)
  reports/research/mr_prediction_autopsy_{market}_summary.json (aggregate)

Under M-R sandbox rules. Reads locked canonical inputs · writes only to
reports/research/. No coupling to sender / xlsx_validator / xlsx_contract.

Per prediction row we record:
  Prediction date · Ticker · Runner · Rank · Confidence · Status ·
  Sector · Entry price at prediction · Investability band (current proxy)
  Then: +1D, +3D, +5D, +10D, +20D forward returns from prediction close
  Plus: MFE / MAE within 20-day window
  Plus: whether stop distance would have been hit
  Plus: winner / loser classification at each horizon
"""
from __future__ import annotations

import json
from collections import Counter, defaultdict
from dataclasses import dataclass, asdict, field
from datetime import date, datetime, timezone, timedelta
from pathlib import Path
from typing import Optional

from backend.research.mr_runner import EXPERIMENT_ID, ALLOWED_WRITE_ROOT


ENGINE_ID = "aegis.mr_prediction_autopsy.v0.1"
SCHEMA_FINGERPRINT = "aegis.mr_prediction_autopsy.v0.1.20260827"
FORWARD_HORIZONS = [1, 3, 5, 10, 20]


@dataclass
class PredictionRow:
    """One immutable research row per historical prediction."""
    prediction_date:       str
    ticker:                str
    runner:                str           # R1 / R2
    status:                str           # Status at prediction time
    rank:                  Optional[int]
    confidence_pct:        Optional[float]
    sector:                Optional[str]
    recommended_date:      Optional[str]
    entry_price_at_pred:   Optional[float]
    stop_at_pred:          Optional[float]
    investability_band:    str           # QUALITY/OK/MARGINAL/AVOID/PENDING (current proxy)
    # Forward returns from prediction_date's close
    fwd_1d_pct:            Optional[float] = None
    fwd_3d_pct:            Optional[float] = None
    fwd_5d_pct:            Optional[float] = None
    fwd_10d_pct:           Optional[float] = None
    fwd_20d_pct:           Optional[float] = None
    # Trajectory within 20-day window
    mfe_pct:               Optional[float] = None
    mae_pct:               Optional[float] = None
    stop_hit_within_20d:   Optional[bool] = None
    # Classification
    winner_5d:             Optional[bool] = None
    winner_20d:            Optional[bool] = None


def _load_parquet(root: Path, ticker: str, market: str):
    import pandas as pd
    clean = ticker.upper().replace(".NS","").replace(".BO","")
    base = "usa/data/raw/us" if market.lower()=="usa" else "data/raw/india"
    p = root / base / f"{clean}_D1.parquet"
    if not p.exists(): return None
    try:
        df = pd.read_parquet(p)
        col = "close" if "close" in df.columns else "Close"
        df.index = pd.to_datetime(df.index).strftime("%Y-%m-%d")
        return (df, col)
    except Exception:
        return None


def _fwd_return(pair, iso: str, horizon: int) -> Optional[float]:
    if pair is None: return None
    df, col = pair
    dates = sorted(df.index)
    try:
        if iso in df.index:
            i = dates.index(iso)
        else:
            earlier = [d for d in dates if d <= iso]
            if not earlier: return None
            i = dates.index(earlier[-1])
        fi = i + horizon
        if fi >= len(dates): return None
        p0 = float(df.loc[dates[i], col])
        p1 = float(df.loc[dates[fi], col])
        if p0 <= 0: return None
        return round((p1 - p0) / p0 * 100, 3)
    except Exception:
        return None


def _mfe_mae(pair, iso: str, horizon: int) -> tuple:
    """MFE / MAE + stop-hit check within N trading days."""
    if pair is None: return (None, None, None)
    df, col = pair
    dates = sorted(df.index)
    try:
        if iso in df.index:
            i = dates.index(iso)
        else:
            earlier = [d for d in dates if d <= iso]
            if not earlier: return (None, None, None)
            i = dates.index(earlier[-1])
        end = min(i + horizon, len(dates) - 1)
        window = dates[i:end + 1]
        p0 = float(df.loc[dates[i], col])
        if p0 <= 0: return (None, None, None)
        closes = [float(df.loc[d, col]) for d in window]
        mfe = round((max(closes) - p0) / p0 * 100, 3)
        mae = round((min(closes) - p0) / p0 * 100, 3)
        return (mfe, mae, None)
    except Exception:
        return (None, None, None)


_INV_CACHE: dict = {}


def _investability_band(root: Path, market: str, ticker: str) -> str:
    key = market.lower()
    if key not in _INV_CACHE:
        d: dict = {}
        for fname in (f"investability_shadow_{key}.json",
                      f"investability_{key}.json"):
            p = root / "reports" / fname
            if not p.exists(): continue
            try:
                dd = json.loads(p.read_text(encoding="utf-8"))
                for r in (dd.get("results") or []):
                    _t = str(r.get("ticker","")).upper() \
                        .replace(".NS","").replace(".BO","")
                    _v = str(r.get("verdict","")).upper()
                    if _t and _t not in d:
                        if "QUALITY" in _v: d[_t] = "QUALITY"
                        elif "OK" in _v: d[_t] = "OK"
                        elif "MARGINAL" in _v: d[_t] = "MARGINAL"
                        elif "AVOID" in _v: d[_t] = "AVOID"
                        else: d[_t] = "PENDING"
            except Exception:
                continue
        _INV_CACHE[key] = d
    tk = ticker.upper().replace(".NS","").replace(".BO","")
    return _INV_CACHE[key].get(tk, "PENDING")


def compute(root: Path, market: str) -> list:
    """Iterate every prediction in AEGIS Daily and build PredictionRow list."""
    from openpyxl import load_workbook
    p = root / "reports" / "telegram" / "aegis_history.xlsx"
    if not p.exists(): return []
    wb = load_workbook(p, read_only=True, data_only=True)
    sh = wb["AEGIS Daily"]
    h = [c.value for c in sh[1]]
    def _c(n):
        try: return h.index(n)
        except ValueError: return None
    i_dt = _c("Date"); i_ct = _c("Country"); i_rt = _c("Run_Type")
    i_tk = _c("Ticker"); i_st = _c("Status")
    i_rank = _c("Rank"); i_conf = _c("Confidence %")
    i_rec = _c("Recommended"); i_ep = _c("Entry Price")
    i_stop = _c("Stop Loss"); i_sec = _c("Sector")
    if None in (i_dt, i_ct, i_rt, i_tk, i_st):
        wb.close(); return []
    rows: list = []
    for r in sh.iter_rows(min_row=2, values_only=True):
        if len(r) <= max(x for x in (i_dt, i_ct, i_rt, i_tk, i_st) if x is not None):
            continue
        ct = str(r[i_ct] or "").upper()
        if ct != market.upper(): continue
        st = str(r[i_st] or "").upper()
        if st in ("EXIT", "CLOSED", "ROTATED_SAMEDAY"): continue
        tk = str(r[i_tk] or "").upper().replace(".NS","").replace(".BO","")
        rn = str(r[i_rt] or "").upper().replace("_NEW","")
        dt = str(r[i_dt] or "")[:10]
        if not (tk and rn and dt): continue
        pair = _load_parquet(root, tk, market)
        if pair is None: continue
        ep = None
        if i_ep is not None and r[i_ep] not in (None, ""):
            try: ep = float(r[i_ep])
            except Exception: pass
        stop = None
        if i_stop is not None and r[i_stop] not in (None, ""):
            try: stop = float(r[i_stop])
            except Exception: pass
        rank = None
        if i_rank is not None and r[i_rank] not in (None, ""):
            try: rank = int(r[i_rank])
            except Exception: pass
        conf = None
        if i_conf is not None and r[i_conf] not in (None, ""):
            try: conf = float(r[i_conf])
            except Exception: pass
        rec_date = None
        if i_rec is not None and r[i_rec] not in (None, ""):
            rec_date = str(r[i_rec])[:10]
        sec = None
        if i_sec is not None and r[i_sec] not in (None, ""):
            sec = str(r[i_sec])
        # Forward returns from prediction_date
        f1 = _fwd_return(pair, dt, 1)
        f3 = _fwd_return(pair, dt, 3)
        f5 = _fwd_return(pair, dt, 5)
        f10 = _fwd_return(pair, dt, 10)
        f20 = _fwd_return(pair, dt, 20)
        mfe, mae, _ = _mfe_mae(pair, dt, 20)
        stop_hit = None
        if stop and ep and stop > 0 and mae is not None:
            # If MAE reached the stop distance from entry
            stop_dist_pct = (ep - stop) / ep * 100 * -1  # negative pct from entry
            stop_hit = mae <= stop_dist_pct
        rows.append(PredictionRow(
            prediction_date=dt, ticker=tk, runner=rn, status=st,
            rank=rank, confidence_pct=conf, sector=sec,
            recommended_date=rec_date,
            entry_price_at_pred=ep, stop_at_pred=stop,
            investability_band=_investability_band(root, market, tk),
            fwd_1d_pct=f1, fwd_3d_pct=f3, fwd_5d_pct=f5,
            fwd_10d_pct=f10, fwd_20d_pct=f20,
            mfe_pct=mfe, mae_pct=mae,
            stop_hit_within_20d=stop_hit,
            winner_5d=(f5 is not None and f5 > 0.5),
            winner_20d=(f20 is not None and f20 > 0.5),
        ))
    wb.close()
    return rows


def summarize(rows: list) -> dict:
    def _agg(subset: list, key: str) -> dict:
        vals = [getattr(o, key) for o in subset if getattr(o, key) is not None]
        if not vals: return {"n": 0}
        wins = sum(1 for v in vals if v > 0.5)
        return {
            "n":            len(vals),
            "avg_pct":      round(sum(vals)/len(vals), 3),
            "median_pct":   round(sorted(vals)[len(vals)//2], 3),
            "win_rate_pct": round(wins/len(vals)*100, 2),
            "best_pct":     round(max(vals), 3),
            "worst_pct":    round(min(vals), 3),
        }

    def _cohort(subset: list) -> dict:
        return {
            "n":         len(subset),
            "fwd_1d":    _agg(subset, "fwd_1d_pct"),
            "fwd_3d":    _agg(subset, "fwd_3d_pct"),
            "fwd_5d":    _agg(subset, "fwd_5d_pct"),
            "fwd_10d":   _agg(subset, "fwd_10d_pct"),
            "fwd_20d":   _agg(subset, "fwd_20d_pct"),
            "avg_mfe_pct": round(sum(o.mfe_pct for o in subset if o.mfe_pct is not None) /
                                 max(1, sum(1 for o in subset if o.mfe_pct is not None)), 3)
                          if any(o.mfe_pct is not None for o in subset) else None,
            "avg_mae_pct": round(sum(o.mae_pct for o in subset if o.mae_pct is not None) /
                                 max(1, sum(1 for o in subset if o.mae_pct is not None)), 3)
                          if any(o.mae_pct is not None for o in subset) else None,
            "stop_hit_rate_pct": round(
                sum(1 for o in subset if o.stop_hit_within_20d) /
                max(1, sum(1 for o in subset if o.stop_hit_within_20d is not None)) * 100, 2)
                if any(o.stop_hit_within_20d is not None for o in subset) else None,
        }

    by_runner = defaultdict(list)
    by_band = defaultdict(list)
    by_sector = defaultdict(list)
    by_rank_bucket = defaultdict(list)
    for r in rows:
        by_runner[r.runner].append(r)
        by_band[r.investability_band].append(r)
        if r.sector: by_sector[r.sector].append(r)
        if r.rank is not None:
            if r.rank <= 3: bucket = "top3"
            elif r.rank <= 7: bucket = "rank_4_7"
            elif r.rank <= 15: bucket = "rank_8_15"
            else: bucket = "rank_16plus"
            by_rank_bucket[bucket].append(r)

    return {
        "engine":            ENGINE_ID,
        "experiment_id":     EXPERIMENT_ID,
        "schema_fingerprint": SCHEMA_FINGERPRINT,
        "generated_utc":     datetime.now(timezone.utc).isoformat(timespec="seconds"),
        "n_predictions":     len(rows),
        "runner_distribution": dict(Counter(r.runner for r in rows)),
        "band_distribution":   dict(Counter(r.investability_band for r in rows)),
        "cohort_ALL":        _cohort(rows),
        "cohort_by_runner":  {k: _cohort(v) for k, v in by_runner.items()},
        "cohort_by_band":    {k: _cohort(v) for k, v in by_band.items()},
        "cohort_by_sector":  {k: _cohort(v) for k, v in by_sector.items()},
        "cohort_by_rank_bucket": {k: _cohort(v) for k, v in by_rank_bucket.items()},
    }


def emit(root: Path, market: str, rows: list, summary: dict) -> tuple:
    p_rows = root / ALLOWED_WRITE_ROOT / f"mr_prediction_autopsy_{market.lower()}.jsonl"
    p_sum  = root / ALLOWED_WRITE_ROOT / f"mr_prediction_autopsy_{market.lower()}_summary.json"
    p_rows.parent.mkdir(parents=True, exist_ok=True)
    with p_rows.open("w", encoding="utf-8") as f:
        for r in rows:
            f.write(json.dumps(asdict(r), default=str, ensure_ascii=False) + "\n")
    p_sum.write_text(json.dumps(summary, indent=2, ensure_ascii=False), encoding="utf-8")
    return (p_rows, p_sum)


def render_console(market: str, summary: dict):
    print(f"\n== {market.upper()} · n_predictions={summary['n_predictions']} ==")
    print(f"   runners: {summary['runner_distribution']}")
    print(f"   bands: {summary['band_distribution']}")
    a = summary["cohort_ALL"]
    print(f"\n   ALL cohort:")
    for hzn in ("fwd_1d","fwd_3d","fwd_5d","fwd_10d","fwd_20d"):
        m = a[hzn]
        if m.get("n"):
            print(f"     {hzn} · n={m['n']:4d} · WR={m['win_rate_pct']:5.2f}% · "
                  f"avg={m['avg_pct']:+.2f}% · med={m['median_pct']:+.2f}%")
    print(f"     avg_MFE={a['avg_mfe_pct']}% · avg_MAE={a['avg_mae_pct']}% · "
          f"stop_hit_rate={a['stop_hit_rate_pct']}%")
    print(f"\n   By runner:")
    for r, m in summary["cohort_by_runner"].items():
        f5 = m["fwd_5d"]
        if not f5.get("n"): continue
        print(f"     {r:8s} · n={m['n']:4d} · fwd_5d WR={f5['win_rate_pct']:5.2f}% "
              f"avg={f5['avg_pct']:+.2f}% · avg_MAE={m['avg_mae_pct']}% · "
              f"stop_hit={m['stop_hit_rate_pct']}%")
    print(f"\n   By investability band:")
    for b, m in summary["cohort_by_band"].items():
        f5 = m["fwd_5d"]
        if not f5.get("n"): continue
        print(f"     {b:10s} · n={m['n']:4d} · fwd_5d WR={f5['win_rate_pct']:5.2f}% "
              f"avg={f5['avg_pct']:+.2f}% · avg_MAE={m['avg_mae_pct']}%")
    print(f"\n   By rank bucket:")
    for k in ("top3","rank_4_7","rank_8_15","rank_16plus"):
        m = summary["cohort_by_rank_bucket"].get(k, {})
        if not m: continue
        f5 = m["fwd_5d"]
        if not f5.get("n"): continue
        print(f"     {k:12s} · n={m['n']:4d} · fwd_5d WR={f5['win_rate_pct']:5.2f}% "
              f"avg={f5['avg_pct']:+.2f}%")


if __name__ == "__main__":
    import argparse
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--market", choices=["india","usa","both"], default="both")
    args = ap.parse_args()
    root = Path(".").resolve()
    for m in (["india","usa"] if args.market=="both" else [args.market]):
        rows = compute(root, m)
        summary = summarize(rows)
        p_rows, p_sum = emit(root, m, rows, summary)
        print(f"[autopsy:{m}] rows={len(rows)} · {p_rows.name} · {p_sum.name}")
        render_console(m, summary)
