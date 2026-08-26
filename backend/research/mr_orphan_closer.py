"""AEGIS · M-R · Registry orphan closer (pre-flight maintenance).

USA operator-flow defect audit 2026-08-26 · CEO handover:

  Registry has 508 active USA PIDs (470 R2 + 38 R1) while today's
  usa/reports/recommendations.json has only 15 fresh recs. The 493
  ACTIVE positions with no fresh signal in 12+ days inflate the
  Portfolio "Active" count and mislead the operator.

Root cause: Registry `get_or_create` opens opportunities as ACTIVE
when a ticker appears in aegis_history rows. There is no auto-CLOSE
rule when the ticker STOPS appearing. Positions accumulate silently.

## Rule (documented + testable)

For each Registry ACTIVE (ticker, runner, market):

  1. Check if the ticker appears in TODAY's canonical INVESTMENT_ACTIVE
     JSON (`reports/context/portfolio_canonical_{market}.json`).
  2. If YES → leave ACTIVE.
  3. If NO → check age since created_date · if age >= STALE_ORPHAN_DAYS
     AND ticker has never appeared in the last STALE_ORPHAN_DAYS days
     of aegis_history rows for that market · CLOSE it with reason
     `ORPHAN_AUTO_CLOSE`.
  4. Emit a maintenance report to
     `reports/research/mr_orphan_closer_{market}.json` for audit.

## Isolation

- reads canonical Portfolio JSON + Registry + aegis_history (read-only)
- writes only reports/research/ (maintenance report) + Registry
  CLOSED events via `opportunity_registry.close()` (which is the
  canonical closure API)
- does NOT modify aegis_history.xlsx or per-market XLSX
- does NOT touch R1/R2 signals or thresholds
- ONE-way transition: ACTIVE → CLOSED · never re-opens

Threshold: STALE_ORPHAN_DAYS = 10 (adjustable · currently conservative)
"""
from __future__ import annotations

import json
from dataclasses import dataclass, asdict
from datetime import date, datetime, timedelta, timezone
from pathlib import Path
from typing import Optional


STALE_ORPHAN_DAYS = 10
ENGINE_ID = "aegis.mr_orphan_closer.v0.1"
SCHEMA_FINGERPRINT = "aegis.mr_orphan_closer.v0.1.20260826"


@dataclass
class ClosureRecord:
    ticker:              str
    runner:              str
    position_id:         Optional[str]
    created_date:        str
    age_days:            int
    last_seen_in_history: Optional[str]
    action:              str            # CLOSED / KEPT / SKIPPED
    reason:              str


def _canonical_active_tickers(root: Path, market: str) -> set:
    """Today's canonical INVESTMENT_ACTIVE list · single source of truth."""
    p = (root / "reports" / "context"
         / f"portfolio_canonical_{market.lower()}.json")
    if not p.exists():
        return set()
    try:
        d = json.loads(p.read_text(encoding="utf-8"))
        return {(str(o.get("ticker", "")).upper()
                 .replace(".NS", "").replace(".BO", ""),
                 str(o.get("runner", "")).upper().replace("_NEW", ""))
                for o in (d.get("investment_active", []) or [])}
    except Exception:
        return set()


def _last_seen_in_history(root: Path, market: str,
                          within_days: int) -> dict:
    """(ticker, runner) → most recent asof this pair appeared in aegis_
    history within `within_days` days. Returns {} if history unreadable."""
    from datetime import date as _d, timedelta as _td
    p = root / "reports" / "telegram" / "aegis_history.xlsx"
    if not p.exists():
        return {}
    cutoff = (_d.today() - _td(days=within_days)).isoformat()
    try:
        from openpyxl import load_workbook
        wb = load_workbook(p, read_only=True, data_only=True)
        sh = wb["AEGIS Daily"]
        h = [c.value for c in sh[1]]
        i_ct = h.index("Country") if "Country" in h else -1
        i_rt = h.index("Run_Type") if "Run_Type" in h else -1
        i_tk = h.index("Ticker") if "Ticker" in h else -1
        i_dt = h.index("Date") if "Date" in h else -1
        if -1 in (i_ct, i_rt, i_tk, i_dt):
            wb.close()
            return {}
        seen: dict = {}
        for row in sh.iter_rows(min_row=2, values_only=True):
            if len(row) <= max(i_ct, i_rt, i_tk, i_dt): continue
            if str(row[i_ct] or "").upper() != market.upper(): continue
            dt = str(row[i_dt])[:10] if row[i_dt] else ""
            if dt < cutoff: continue
            tk = str(row[i_tk] or "").upper().replace(".NS","").replace(".BO","")
            rn = str(row[i_rt] or "").upper().replace("_NEW","")
            key = (tk, rn)
            if key not in seen or dt > seen[key]:
                seen[key] = dt
        wb.close()
        return seen
    except Exception:
        return {}


def close_orphans(root: Path, market: str,
                  asof: Optional[str] = None,
                  dry_run: bool = False,
                  stale_days: int = STALE_ORPHAN_DAYS) -> list:
    """Scan Registry ACTIVE positions · CLOSE any (ticker, runner) that
    is NOT in today's canonical INVESTMENT_ACTIVE AND hasn't appeared in
    the last `stale_days` days of aegis_history.

    Returns list[ClosureRecord]. When dry_run=True, no Registry mutation
    occurs · used for auditing before enabling in production."""
    from backend.research import opportunity_registry as _oreg
    asof_iso = str(asof or date.today().isoformat())[:10]
    canonical = _canonical_active_tickers(root, market)
    last_seen = _last_seen_in_history(root, market, stale_days)
    reg = _oreg.load_all(root)
    records = []
    seen_pids: set = set()
    for opps in reg.values():
        for o in opps:
            if o.market.lower() != market.lower(): continue
            if not o.is_active(): continue
            pid = getattr(o, "opportunity_id", None)
            if pid in seen_pids: continue
            seen_pids.add(pid)
            tk_norm = o.ticker.upper().replace(".NS","").replace(".BO","")
            rn_norm = o.runner.upper().replace("_NEW","")
            key = (tk_norm, rn_norm)
            cd = str(o.created_date or "")[:10]
            try:
                age = (date.fromisoformat(asof_iso)
                       - date.fromisoformat(cd)).days
            except Exception:
                age = 0
            ls = last_seen.get(key)
            if key in canonical:
                records.append(ClosureRecord(
                    ticker=tk_norm, runner=rn_norm, position_id=pid,
                    created_date=cd, age_days=age,
                    last_seen_in_history=ls, action="KEPT",
                    reason="in today's canonical INVESTMENT_ACTIVE",
                ))
                continue
            if age < stale_days:
                records.append(ClosureRecord(
                    ticker=tk_norm, runner=rn_norm, position_id=pid,
                    created_date=cd, age_days=age,
                    last_seen_in_history=ls, action="KEPT",
                    reason=(f"age {age}d < stale_days threshold "
                            f"({stale_days}d) · give recent entries time"),
                ))
                continue
            if ls and ls >= (
                    date.fromisoformat(asof_iso)
                    - timedelta(days=stale_days)).isoformat():
                records.append(ClosureRecord(
                    ticker=tk_norm, runner=rn_norm, position_id=pid,
                    created_date=cd, age_days=age,
                    last_seen_in_history=ls, action="KEPT",
                    reason=(f"seen in history within {stale_days}d "
                            f"({ls}) · not orphaned"),
                ))
                continue
            # Orphan · close it
            action = "CLOSED" if not dry_run else "WOULD_CLOSE"
            if not dry_run:
                try:
                    _oreg.close(root, pid, asof_iso,
                                reason="ORPHAN_AUTO_CLOSE")
                except Exception as e:
                    action = f"CLOSE_FAILED · {type(e).__name__}"
            records.append(ClosureRecord(
                ticker=tk_norm, runner=rn_norm, position_id=pid,
                created_date=cd, age_days=age,
                last_seen_in_history=ls, action=action,
                reason=(f"age {age}d >= stale_days ({stale_days}d) · "
                        f"not in canonical + last_seen_in_history={ls or 'never'} · "
                        f"orphaned R{rn_norm}-style bloat pattern"),
            ))
    return records


def emit(root: Path, market: str, records: list) -> Path:
    """Write maintenance report · reports/research/ only per M-R isolation."""
    out = root / "reports" / "research" / f"mr_orphan_closer_{market.lower()}.json"
    out.parent.mkdir(parents=True, exist_ok=True)
    from collections import Counter
    action_counts = Counter(r.action for r in records)
    payload = {
        "engine":            ENGINE_ID,
        "schema_fingerprint": SCHEMA_FINGERPRINT,
        "generated_utc":     datetime.now(timezone.utc).isoformat(timespec="seconds"),
        "market":            market.lower(),
        "stale_days":        STALE_ORPHAN_DAYS,
        "n_scanned":         len(records),
        "action_counts":     dict(action_counts),
        "records":           [asdict(r) for r in records],
    }
    out.write_text(json.dumps(payload, indent=2, ensure_ascii=False),
                    encoding="utf-8")
    return out


def summary_line(records: list) -> str:
    from collections import Counter
    c = Counter(r.action for r in records)
    return (f"mr_orphan_closer · scanned={len(records)} · "
            f"CLOSED={c.get('CLOSED', 0)} · KEPT={c.get('KEPT', 0)} · "
            f"WOULD_CLOSE(dry)={c.get('WOULD_CLOSE', 0)}")


if __name__ == "__main__":
    import argparse
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--market", required=True, choices=["india","usa"])
    ap.add_argument("--asof", default=None)
    ap.add_argument("--dry-run", action="store_true",
                    help="scan but do not mutate Registry (recommended first)")
    ap.add_argument("--stale-days", type=int, default=STALE_ORPHAN_DAYS)
    args = ap.parse_args()
    records = close_orphans(Path("."), args.market, args.asof,
                             dry_run=args.dry_run,
                             stale_days=args.stale_days)
    p = emit(Path("."), args.market, records)
    print(summary_line(records))
    try:
        print(f"  written · {p.relative_to(Path('.').resolve())}")
    except ValueError:
        print(f"  written · {p}")
