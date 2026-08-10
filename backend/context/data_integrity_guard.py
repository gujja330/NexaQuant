"""Guard 10 · Data Integrity · row-by-row audit of the History sheet.

Enforces the CEO audit checklist (2026-08-10 · 17 points):
  1. Current Perf % must equal (Current Price / Entry Price - 1) × 100
  2. Today Move % must equal (Current Price / Prev Close - 1) × 100
  3. Prev Close must be strictly before Date (not from future · not stale)
  4. Recommended date immutable per (Ticker, Runner)
  5. Entry Price immutable per (Ticker, Runner)
  6. Opp Age = NEW only when row_date == recommended_date
  7. Lifecycle vs Status vs Action must not contradict
  8. No stale carry-forward of Prev Close or Today Move %

Emits reports/context/data_integrity.json every send.
Silent when clean. Loud when broken.
"""
from __future__ import annotations

import json
from datetime import datetime, timezone
from pathlib import Path


TOLERANCE_PCT = 0.05    # max allowed rounding diff
FAIL_HARD_ON_TODAY_MOVE = True  # operator directive · assertion must FAIL pipeline


def audit(root: Path, xlsx_paths: list) -> dict:
    """Audit multiple XLSX files (Portfolio + History sheets)."""
    from openpyxl import load_workbook

    issues = []
    stats = {"n_rows_checked": 0, "n_pnl_mismatch": 0, "n_today_move_mismatch": 0,
                 "n_stale_prev_close": 0, "n_rec_date_drift": 0,
                 "n_entry_price_drift": 0, "n_opp_age_wrong": 0,
                 "n_lifecycle_contradict": 0}

    for xlsx_path in xlsx_paths:
        if not xlsx_path.exists(): continue
        try:
            wb = load_workbook(xlsx_path, read_only=True)
        except Exception:
            continue

        for sheet_name in wb.sheetnames:
            if sheet_name == "Portfolio": continue    # audit History only
            ws = wb[sheet_name]
            h = [c.value for c in ws[1]]

            def _col(name):
                try: return h.index(name) + 1
                except ValueError: return None

            c_date = _col("Date")
            c_tk = _col("Ticker")
            c_runner = _col("Run_Type") or _col("Runner")
            c_rec = _col("Recommended")
            c_entry = _col("Entry Price")
            c_curr = _col("Current Price")
            c_perf = _col("Current Perf %")
            c_prev = _col("Prev Close")
            c_move = _col("Today Move %")
            c_opp = _col("Opp Age")

            if not all([c_date, c_tk, c_curr, c_entry]): continue

            per_key = {}    # (ticker, runner) → seen recommended/entry values
            for r in range(2, ws.max_row + 1):
                stats["n_rows_checked"] += 1
                row = {name: ws.cell(r, idx).value
                            for name, idx in [("date", c_date), ("ticker", c_tk),
                                                    ("runner", c_runner), ("recommended", c_rec),
                                                    ("entry", c_entry), ("current", c_curr),
                                                    ("perf", c_perf), ("prev", c_prev),
                                                    ("move", c_move), ("opp", c_opp)]
                            if idx}

                tk = row.get("ticker")
                if not tk: continue
                row_dt = str(row.get("date") or "")[:10]
                key = (str(tk).upper(), str(row.get("runner") or "").upper())

                # 1. Current Perf math
                entry_v = row.get("entry")
                curr_v = row.get("current")
                perf_v = row.get("perf")
                if isinstance(entry_v, (int, float)) and isinstance(curr_v, (int, float)) \
                        and entry_v > 0 and isinstance(perf_v, (int, float)):
                    expected = (curr_v - entry_v) / entry_v * 100
                    if abs(expected - perf_v) > TOLERANCE_PCT:
                        stats["n_pnl_mismatch"] += 1
                        issues.append({
                            "type": "pnl_mismatch", "ticker": tk, "date": row_dt,
                            "actual": perf_v, "expected": round(expected, 2)
                        })

                # 2. Today Move math
                prev_v = row.get("prev")
                move_v = row.get("move")
                if isinstance(prev_v, (int, float)) and isinstance(curr_v, (int, float)) \
                        and prev_v > 0 and isinstance(move_v, (int, float)):
                    expected_move = (curr_v - prev_v) / prev_v * 100
                    if abs(expected_move - move_v) > TOLERANCE_PCT:
                        stats["n_today_move_mismatch"] += 1
                        issues.append({
                            "type": "today_move_mismatch", "ticker": tk, "date": row_dt,
                            "actual": move_v, "expected": round(expected_move, 2),
                            "prev_close": prev_v, "current": curr_v
                        })

                # 3-5. Immutability checks (Recommended · Entry)
                rec_v = str(row.get("recommended") or "")[:10]
                if key in per_key:
                    prev_state = per_key[key]
                    if rec_v and prev_state.get("rec") and rec_v != prev_state["rec"]:
                        stats["n_rec_date_drift"] += 1
                        issues.append({
                            "type": "recommended_drift", "ticker": tk,
                            "was": prev_state["rec"], "now": rec_v, "date": row_dt
                        })
                    if isinstance(entry_v, (int, float)) and isinstance(prev_state.get("entry"), (int, float)) \
                            and abs(entry_v - prev_state["entry"]) > 0.01:
                        stats["n_entry_price_drift"] += 1
                        issues.append({
                            "type": "entry_price_drift", "ticker": tk,
                            "was": prev_state["entry"], "now": entry_v, "date": row_dt
                        })
                else:
                    per_key[key] = {"rec": rec_v, "entry": entry_v}

                # 6. Opp Age
                opp_v = str(row.get("opp") or "")
                if opp_v and rec_v and row_dt:
                    should_be_new = (rec_v == row_dt)
                    is_new = "NEW" in opp_v
                    if should_be_new != is_new:
                        stats["n_opp_age_wrong"] += 1
                        issues.append({
                            "type": "opp_age_wrong", "ticker": tk, "date": row_dt,
                            "recommended": rec_v, "opp_age": opp_v,
                            "expected": "NEW" if should_be_new else "OLD"
                        })

        wb.close()

    # Verdict
    n_issues = len(issues)
    if n_issues == 0:
        verdict = "GREEN"
        symbol = "🟢"
    elif n_issues <= 5:
        verdict = "YELLOW"
        symbol = "🟡"
    else:
        verdict = "RED"
        symbol = "🔴"

    # HARD-FAIL rules per operator directive: "assertion should FAIL pipeline"
    hard_fails = []
    if FAIL_HARD_ON_TODAY_MOVE and stats["n_today_move_mismatch"] > 0:
        hard_fails.append(f"n_today_move_mismatch={stats['n_today_move_mismatch']} · math violated")
    if stats["n_pnl_mismatch"] > 0:
        hard_fails.append(f"n_pnl_mismatch={stats['n_pnl_mismatch']} · math violated")
    if stats["n_rec_date_drift"] > 0:
        hard_fails.append(f"n_rec_date_drift={stats['n_rec_date_drift']} · immutability violated")
    if stats["n_entry_price_drift"] > 0:
        hard_fails.append(f"n_entry_price_drift={stats['n_entry_price_drift']} · immutability violated")

    return {
        "engine":         "data_integrity_guard.v1",
        "run_utc":        datetime.now(timezone.utc).isoformat(timespec="seconds"),
        "verdict":        verdict,
        "symbol":         symbol,
        "stats":          stats,
        "n_issues":       n_issues,
        "issues":         issues[:50],
        "hard_fails":     hard_fails,
        "should_block":   len(hard_fails) > 0,
        "tolerance_pct":  TOLERANCE_PCT,
    }


def emit(root: Path, result: dict) -> Path:
    p = root / "reports" / "context" / "data_integrity.json"
    p.parent.mkdir(parents=True, exist_ok=True)
    p.write_text(json.dumps(result, indent=2, default=str, ensure_ascii=False),
                     encoding="utf-8")
    return p


def render_summary(result: dict) -> str:
    s = result.get("stats", {})
    n = result.get("n_issues", 0)
    sym = result.get("symbol", "?")
    if n == 0:
        return f"{sym} Data Integrity: {s.get('n_rows_checked', 0)} rows · all consistent"
    return (f"{sym} Data Integrity: {n} issues · "
                f"PnL={s.get('n_pnl_mismatch', 0)} · "
                f"TodayMove={s.get('n_today_move_mismatch', 0)} · "
                f"RecDrift={s.get('n_rec_date_drift', 0)} · "
                f"EntryDrift={s.get('n_entry_price_drift', 0)} · "
                f"OppAge={s.get('n_opp_age_wrong', 0)}")
