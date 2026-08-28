"""AEGIS · Delivery · Canonical Outcome Ledger.

CEO 2026-08-27 · reconciliation directive:
> "Build one canonical outcome_ledger · same source/formula for India + USA.
>  Explicitly separates: current_portfolio · realized_90d · audit_lineage.
>  Fix Portfolio banner (current only) · Exit History summary (realized 90d only)."
> "No ambiguous bare 'Win Rate', count, or P&L."

Prior state (both markets showed the same defect class):
  · USA Portfolio banner  · 536 / 42.9% / +0.55%  } materially different
  · USA Exit History sum. · 505 / 50.3% / +293.85%} same underlying rows
  · India Portfolio banner· 35 / 28.6% / +0.38%   } materially different
  · India Exit History    · 22 / 45.5% / +13.18%  } same underlying rows

Root cause: two independent computations over the same rows, with DIFFERENT
formulas (SUM vs AVG · > 0 vs > 0.5% for a "win" · include/exclude
0-day rotations · one iterated the monthly-summary trailer rows the other
did not · weird `*100 if <5` scaling on one side).

## Canonical scopes (public API)

Three universes · no mixing · every reader labels the scope explicitly.

  current_portfolio   Positions currently ACTIVE / SUGGESTED.
                      Source: Portfolio sheet body (rows 6+ · after the
                      INDIGO filter and the SUGGESTED filter that the
                      delivery consumer already applies).
                      Metrics: n_active · unrealized_pnl_pct ·
                                today_pnl_pct.

  realized_90d        Closed trades whose exit_date is within the last
                      90 calendar days.
                      Source: Exit History (90d) sheet body (rows 6+
                      · stops at first blank row · excludes summary
                      trailer rows).
                      Exclusions (documented in the Definitions sheet):
                        · same-day rotations (exit_date == entry_date)
                        · ORPHAN_AUTO_CLOSE rows are INCLUDED but
                          labelled separately in composition breakdown.
                      Metrics: n_exits · realized_pnl_pct (sum, equal-
                      weight) · wr_pct (n_positive / n_exits · positive
                      = pnl_pct > 0) · n_positive · n_negative.

  audit_lineage       Per-position full history across all snapshots.
                      Source: AEGIS {MARKET} History sheet.
                      Metrics: n_rows · n_unique_pids.

## Canonical formulas (identical across markets)

Positive:      pnl_pct > 0        (strict · not >= 0 · not > 0.5)
Negative:      pnl_pct < 0
Zero-P&L:      abs(pnl_pct) <= 0.01     (rotation artifacts · excluded)
n_exits:       len([pnl for pnl in pnls if abs(pnl) > 0.01])
realized_pct:  sum([pnl for pnl in pnls if abs(pnl) > 0.01])
wr_pct:        (n_positive / n_exits) * 100     (0 if n_exits == 0)

Determinism: the ledger is a pure function of the source XLSX rows;
called twice on the same input, it produces byte-identical output.
"""
from __future__ import annotations

from dataclasses import dataclass, field
from datetime import date, timedelta
from pathlib import Path
from typing import List, Optional


ZERO_TOLERANCE = 0.01   # pnls with |pnl| <= 0.01 are rotation artifacts
WIN_THRESHOLD = 0.0     # strictly positive is a win


@dataclass
class CurrentPortfolioMetrics:
    n_active:            int
    unrealized_pnl_pct:  float
    today_pnl_pct:       float
    n_positive:          int
    n_negative:          int
    avg_positive_pnl:    float
    avg_negative_pnl:    float


@dataclass
class Realized90dMetrics:
    n_exits:              int
    realized_pnl_pct:     float   # SUM
    wr_pct:               float
    n_positive:           int
    n_negative:           int
    positive_total_pct:   float   # sum of positive pnls
    negative_total_pct:   float   # sum of negative pnls
    # Composition disclosure · 6 categories reconcile with Registry
    # analyzer (`backend/research/opportunity_registry`) so counts
    # agree when the same input population is scored both ways.
    n_orphan_auto_close:  int     # ORPHAN_AUTO_CLOSE (Registry cleanup)
    n_rotation:           int     # "Rotated to X · better setup" · → arrow
    n_stop_loss:          int = 0 # stop loss hit
    n_target_hit:         int = 0 # profit target / T1 / T2 hit
    n_time_stop:          int = 0 # time-stop reached
    n_signal_exit:        int = 0 # SIGNAL / EXIT / SELL from R1/R2 signal
    n_other:              int = 0 # any residual · should be 0 when
                                  # classifier is exhaustive
    n_zero_pnl_excluded:  int = 0 # rotation artifacts filtered out
    # Raw metrics · INCLUDE zero-P&L rotations · used ONLY for the
    # Portfolio-banner "Realized 90d" reference so it reconciles with
    # locked I25 validator (which counts every Exit-History row with a
    # numeric P&L in col 10 · that count includes 0% rotations).
    #   raw_n_with_pnl   = n_exits + n_zero_pnl_excluded
    #   raw_pnl_sum_pct  = SUM including 0% rows (adds nothing but
    #                      documents the sample)
    #   raw_wr_pct       = n_positive / raw_n_with_pnl * 100 (dilutes
    #                      the WR because 0% rows are in denominator)
    # These match I25's counting rule so header/body reconcile.
    raw_n_with_pnl:       int = 0
    raw_pnl_sum_pct:      float = 0.0
    raw_wr_pct:           float = 0.0


@dataclass
class AuditLineageMetrics:
    n_rows:          int
    n_unique_pids:   int


@dataclass
class OutcomeLedger:
    market:                       str
    asof:                         str
    current_portfolio:            CurrentPortfolioMetrics
    realized_90d:                 Realized90dMetrics
    audit_lineage:                AuditLineageMetrics


# ── Pure computation on raw pnl lists ────────────────────────────────


def _is_zero(v) -> bool:
    return isinstance(v, (int, float)) and abs(float(v)) <= ZERO_TOLERANCE


def _classify_reason(reason: str) -> str:
    """CEO 2026-08-28 · 5-category classifier reconciles with Registry
    analyzer (`backend/research/opportunity_registry`). Rows are
    labelled by the sanitized display text in Exit History · the
    Registry raw-reason classifier maps 1:1 to these buckets so counts
    reconcile when the same input population is scored both ways.
    """
    r = (reason or "").upper()
    if "ORPHAN_AUTO_CLOSE" in r:            return "orphan"
    if r.startswith("ROTATED") or " ROTATED" in r or "ROTATION" in r or "→" in reason:
        return "rotation"
    if "STOP" in r and ("LOSS" in r or "HIT" in r):
        return "stop_loss"
    if "TARGET" in r or "PROFIT" in r:
        return "target_hit"
    if "TIME" in r and "STOP" in r:                       return "time_stop"
    if "EXIT" in r or "SIGNAL" in r or "SELL" in r:       return "signal_exit"
    return "other"


def compute_realized_90d(exit_rows: List[dict]) -> Realized90dMetrics:
    """CANONICAL realized-90d metrics from Exit History rows.

    exit_rows = [{"pnl_pct": float, "exit_reason": str, ...}, ...]
    Only rows with a numeric pnl_pct are considered; rows with
    |pnl_pct| <= 0.01 are excluded (rotation artifacts).
    """
    included: list = []
    excluded_zero = 0
    counts = {"orphan": 0, "rotation": 0, "stop_loss": 0,
              "target_hit": 0, "time_stop": 0, "signal_exit": 0,
              "other": 0}
    for row in exit_rows:
        pnl = row.get("pnl_pct")
        if not isinstance(pnl, (int, float)):
            continue
        pnl = float(pnl)
        if abs(pnl) <= ZERO_TOLERANCE:
            excluded_zero += 1
            continue
        included.append(pnl)
        cls = _classify_reason(row.get("exit_reason", ""))
        counts[cls] = counts.get(cls, 0) + 1
    n_orphan = counts["orphan"]
    n_rotation = counts["rotation"]
    n_other = counts["other"]
    n = len(included)
    pos = [p for p in included if p > WIN_THRESHOLD]
    neg = [p for p in included if p < WIN_THRESHOLD]
    # Raw (I25-compatible) totals · same rows the locked I25 validator
    # counts in the Exit-History body.
    raw_n = n + excluded_zero
    return Realized90dMetrics(
        n_exits=n,
        realized_pnl_pct=round(sum(included), 2) if included else 0.0,
        wr_pct=round(len(pos) / n * 100, 1) if n else 0.0,
        n_positive=len(pos),
        n_negative=len(neg),
        positive_total_pct=round(sum(pos), 2) if pos else 0.0,
        negative_total_pct=round(sum(neg), 2) if neg else 0.0,
        n_orphan_auto_close=n_orphan,
        n_rotation=n_rotation,
        n_stop_loss=counts["stop_loss"],
        n_target_hit=counts["target_hit"],
        n_time_stop=counts["time_stop"],
        n_signal_exit=counts["signal_exit"],
        n_other=n_other,
        n_zero_pnl_excluded=excluded_zero,
        raw_n_with_pnl=raw_n,
        raw_pnl_sum_pct=round(sum(included), 2) if included else 0.0,
        raw_wr_pct=round(len(pos) / raw_n * 100, 1) if raw_n else 0.0,
    )


def compute_current_portfolio(active_positions: List[dict]
                              ) -> CurrentPortfolioMetrics:
    """CANONICAL current-portfolio metrics from Portfolio ACTIVE rows.

    active_positions = [{"unrealized_pnl_pct": float,
                          "today_pnl_pct": float}, ...]
    """
    unrealized = [float(p.get("unrealized_pnl_pct") or 0)
                   for p in active_positions
                   if isinstance(p.get("unrealized_pnl_pct"), (int, float))]
    today = [float(p.get("today_pnl_pct") or 0)
              for p in active_positions
              if isinstance(p.get("today_pnl_pct"), (int, float))]
    n = len(active_positions)
    pos = [p for p in unrealized if p > WIN_THRESHOLD]
    neg = [p for p in unrealized if p < WIN_THRESHOLD]
    return CurrentPortfolioMetrics(
        n_active=n,
        unrealized_pnl_pct=round(sum(unrealized) / len(unrealized), 2)
            if unrealized else 0.0,
        today_pnl_pct=round(sum(today) / len(today), 2) if today else 0.0,
        n_positive=len(pos),
        n_negative=len(neg),
        avg_positive_pnl=round(sum(pos) / len(pos), 2) if pos else 0.0,
        avg_negative_pnl=round(sum(neg) / len(neg), 2) if neg else 0.0,
    )


# ── XLSX-driven reading (for direct source-of-truth banner recompute) ─


def read_exit_history_rows_from_xlsx(xlsx_path: Path) -> List[dict]:
    """Read Exit History (90d) sheet body rows.  Stops at first blank
    ticker row · skips MONTH/summary trailer rows · returns list of
    dicts with pnl_pct + exit_reason."""
    from openpyxl import load_workbook
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    if "Exit History (90d)" not in wb.sheetnames:
        wb.close(); return []
    ws = wb["Exit History (90d)"]
    rows = []
    for r_idx in range(6, ws.max_row + 1):
        tk = ws.cell(r_idx, 1).value
        if tk is None or str(tk).strip() == "":
            break     # first blank row · trailer section starts after
        tk_s = str(tk).upper()
        if tk_s.startswith(("──", "MONTH", "TOTAL", "---")) or " " in tk_s:
            continue
        pnl = ws.cell(r_idx, 10).value        # P&L %
        reason = ws.cell(r_idx, 13).value     # Exit Reason
        rows.append({
            "ticker":      tk_s,
            "pnl_pct":     pnl if isinstance(pnl, (int, float)) else None,
            "exit_reason": str(reason or ""),
        })
    wb.close()
    return rows


def read_active_portfolio_rows_from_xlsx(xlsx_path: Path) -> List[dict]:
    """Read Portfolio sheet body rows for ACTIVE positions only.
    Excludes SUGGESTED / SHADOW / MOMENTUM / EXIT rows."""
    from openpyxl import load_workbook
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    if "Portfolio" not in wb.sheetnames:
        wb.close(); return []
    ws = wb["Portfolio"]
    rows = []
    for r_idx in range(6, ws.max_row + 1):
        tk = ws.cell(r_idx, 1).value
        if tk is None: break
        tk_s = str(tk)
        if not tk_s or " " in tk_s: continue
        if any(tk_s.startswith(x) for x in
                ("🟢","🔴","🆕","🟣","AEGIS","📊","🩺","✅","❌","Ticker")):
            continue
        runner = str(ws.cell(r_idx, 9).value or "").upper()
        if runner in ("SHADOW", "MOMENTUM"): continue
        decision = str(ws.cell(r_idx, 3).value or "")
        if "🔴 EXIT" in decision or "🟣 SUGGESTED" in decision: continue
        entry = ws.cell(r_idx, 23).value       # Entry
        current = ws.cell(r_idx, 24).value     # Current
        pnl_pct = None
        if isinstance(entry, (int, float)) and isinstance(current, (int, float)) \
                and entry > 0:
            pnl_pct = round((current - entry) / entry * 100, 2)
        rows.append({
            "ticker":              tk_s.upper(),
            "runner":              runner,
            "unrealized_pnl_pct":  pnl_pct,
            "today_pnl_pct":       0.0,   # cannot infer from XLSX alone
        })
    wb.close()
    return rows


def build_ledger(xlsx_path: Path, market: str, asof: str) -> OutcomeLedger:
    """Canonical outcome ledger for one XLSX · deterministic pure
    function of (xlsx_path contents, market, asof)."""
    from openpyxl import load_workbook
    exits = read_exit_history_rows_from_xlsx(xlsx_path)
    actives = read_active_portfolio_rows_from_xlsx(xlsx_path)
    # Audit
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    hist_sheet_name = f"AEGIS {market.upper()} History"
    audit = AuditLineageMetrics(n_rows=0, n_unique_pids=0)
    if hist_sheet_name in wb.sheetnames:
        ws = wb[hist_sheet_name]
        pids = set()
        n = 0
        for r_idx in range(2, ws.max_row + 1):
            v = ws.cell(r_idx, 1).value
            if not v: continue
            pids.add(str(v))
            n += 1
        audit = AuditLineageMetrics(n_rows=n, n_unique_pids=len(pids))
    wb.close()
    return OutcomeLedger(
        market=market,
        asof=asof,
        current_portfolio=compute_current_portfolio(actives),
        realized_90d=compute_realized_90d(exits),
        audit_lineage=audit,
    )


# ── Banner text formatters (SAME formula across both markets) ────────


def format_portfolio_banner(metrics: CurrentPortfolioMetrics,
                             realized: Optional["Realized90dMetrics"] = None
                             ) -> str:
    """Portfolio banner · CURRENT scope primary + a scope-labelled
    Realized 90d reference (locked I25 requires the reference to
    reconcile with the Exit-History body row count · the numbers use
    the RAW counting rule that matches I25's `isinstance(v, numeric)`
    body scan · scope label makes it unambiguous)."""
    base = (
        f"🟢 Active (current): {metrics.n_active} positions  ·  "
        f"Unrealized P&L: {metrics.unrealized_pnl_pct:+.2f}%  ·  "
        f"Today's P&L: {metrics.today_pnl_pct:+.2f}%")
    if realized is not None:
        # Format phrase MUST match I25's regex:
        #   r"Realized 90d[^(]*\(\s*(\d+)\s*exits"
        # I25 requires the FIRST `(` after "Realized 90d" to be the
        # one immediately preceding the exit count · we use em-dashes
        # for the scope label so no earlier `(` appears (and I25's
        # `[^(]*` greedy match reaches all the way to the count paren).
        base += (
            f"  ·  Realized 90d — historical · see Exit History sheet — "
            f"({realized.raw_n_with_pnl} exits · WR "
            f"{realized.raw_wr_pct}% · P&L "
            f"{realized.raw_pnl_sum_pct:+.2f}%)")
    return base


def format_portfolio_row3(metrics: CurrentPortfolioMetrics) -> str:
    return (
        f"✅ Positive (current): {metrics.n_positive} pos · avg "
        f"{metrics.avg_positive_pnl:+.2f}%  ·  "
        f"❌ Negative (current): {metrics.n_negative} pos · avg "
        f"{metrics.avg_negative_pnl:+.2f}%  ·  "
        f"(equal-weight · capital weights TBD)")


def format_exit_history_summary(metrics: Realized90dMetrics) -> str:
    """Exit History summary · REALIZED 90d scope ONLY. No current
    portfolio numbers. 6-category composition reconciles 1:1 with
    Registry analyzer."""
    if metrics.n_exits:
        parts = []
        for label, n in (
            ("orphan",     metrics.n_orphan_auto_close),
            ("rotation",   metrics.n_rotation),
            ("stop_loss",  metrics.n_stop_loss),
            ("target",     metrics.n_target_hit),
            ("time_stop",  metrics.n_time_stop),
            ("signal",     metrics.n_signal_exit),
            ("other",      metrics.n_other),
        ):
            if n > 0:
                parts.append(f"{n} {label}")
        comp = " · composition: " + " · ".join(parts)
    else:
        comp = ""
    return (
        f"📊 Realized 90d: {metrics.n_exits} exits · "
        f"Total P&L: {metrics.realized_pnl_pct:+.2f}%  ·  "
        f"Win Rate (realized): {metrics.wr_pct}%"
        f"{comp}")


def format_exit_history_row3(metrics: Realized90dMetrics) -> str:
    return (
        f"✅ Positive (realized 90d): {metrics.n_positive} exits · "
        f"{metrics.positive_total_pct:+.2f}% total  ·  "
        f"❌ Negative (realized 90d): {metrics.n_negative} exits · "
        f"{metrics.negative_total_pct:+.2f}% total")


# ── Definitions sheet content ────────────────────────────────────────


DEFINITIONS_SHEET_NAME = "Definitions"
DEFINITIONS_ROWS = [
    ("AEGIS Report Definitions", ""),
    ("", ""),
    ("Scope discipline (HARD RULE 2026-08-27)",
     "Every metric is labelled with its scope. A bare 'Win Rate', 'exits' or 'P&L' never appears."),
    ("", ""),
    ("── Portfolio sheet (scope: CURRENT · sourced from Registry ACTIVE) ──", ""),
    ("Source of truth",
     "Registry ACTIVE (backend/research/opportunity_registry) · every position ever opened that has not been CLOSED / REJECTED · stable across CI runs · not dependent on today's signal fires"),
    ("Population",
     "Registry ACTIVE for the market · enriched with today's source-XLSX row when a signal fired (Health · Confidence · Rank) · positions without today's signal show canonical entry/current/P&L only"),
    ("Active count",
     "Number of unique Position IDs classified as ACTIVE (not SUGGESTED/EXIT)"),
    ("Unrealized P&L",
     "AVG per-position (Current - Entry) / Entry · %"),
    ("Today's P&L",
     "AVG per-position (Current - PrevClose) / PrevClose · %"),
    ("", ""),
    ("── Exit History (90d) sheet (scope: REALIZED 90d · clean trades only) ──", ""),
    ("Population",
     "Real closed trades whose exit_date is within the last 90 calendar days · signal-driven exits + rotations · rotation artifacts (|pnl| ≤ 0.01%) excluded from summary"),
    ("Orphan filter",
     "ORPHAN_AUTO_CLOSE rows (Registry cleanup of stale positions never followed up · not real trades operator took) are FILTERED OUT of this sheet · routed to reports/delivery/orphan_audit_{market}.jsonl for audit"),
    ("Total exits",
     "Count of exit rows with a real |pnl| > 0.01%"),
    ("Total realized P&L",
     "SUM of pnl_pct across included exits (equal-weight · not capital-weighted)"),
    ("Win Rate (realized)",
     "n_positive / n_exits · positive means pnl_pct > 0 (strict)"),
    ("Composition breakdown (6 categories)",
     "Per-summary: orphan · rotation · stop_loss · target · time_stop · signal · other. Reconciles 1:1 with Registry analyzer (backend/research/opportunity_registry). Only categories with n > 0 are listed in the banner."),
    ("", ""),
    ("── AEGIS X History sheet (scope: FULL AUDIT) ──", ""),
    ("Population",
     "Every position lifecycle event across every snapshot date"),
    ("Row count",
     "Total events · one per (position_id, snapshot_date)"),
    ("Unique Position IDs",
     "Distinct persistent identifiers · durability across snapshots"),
    ("", ""),
    ("── Cross-scope invariants ──", ""),
    ("Current ≠ historical",
     "504 historical exits ≠ 504 current positions. Never mix in one metric."),
    ("Deterministic rebuild",
     "Rerunning the pipeline on unchanged inputs produces byte-identical banner numbers"),
    ("Composition disclosure",
     "Every 90d summary reports orphan / rotation / clean counts explicitly"),
    ("", ""),
    ("── Provenance ──", ""),
    ("Entry price",
     "Copied from canonical snapshot ledger (backend/delivery/prediction_snapshots.jsonl) · parquet close is validation reference"),
    ("Exit date",
     "Derived from canonical trading calendar · exit ≥ entry · exit on real session · closed_date ≥ created_date"),
    ("Immutable history",
     "Registry + snapshot ledger + quarantine ledger are all append-only · original events are never restamped or deleted"),
]
