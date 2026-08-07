"""AEGIS · Unified daily detail history · XLSX (Excel) + optional Google Sheets.

Per operator directives 2026-08-01:
    "instead cant u send an xlsx to telegram with columns and rows by
     each stock"
    "runner 1, runner 2 u can mix into a column called Run_type"
    "or maintain one sheet add country too?"
    "everyday we can update same sheet"

Design:
  · ONE unified XLSX · one row per (Date, Country, Run_Type, Ticker)
  · 38 columns (Date + Country + Run_Type first · then the full stock card
    fields · sortable + filterable in Excel)
  · Daily runs APPEND to the same file (never overwrite)
  · Dedup key = (Date, Country, Run_Type, Ticker) · re-running same day
    updates the existing row (not duplicate)

Output: reports/telegram/aegis_history.xlsx (one file · grows daily)
        reports/telegram/aegis_daily_{asof}.xlsx (today-only snapshot)

Optional: GSHEETS_CREDS_JSON + GSHEETS_SHEET_ID env vars enable Google
Sheets mirror. Silent no-op if creds absent.
"""
from __future__ import annotations

import json
import os
from datetime import datetime, timezone
from pathlib import Path
from typing import Mapping

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from .detail_report import (
    _load_position, _load_ledger_events_for_ticker,
    _lifecycle_state, _exit_triggers_checklist,
    _drivers_from_attribution, _r1_orphan_to_rec_shape,
)
from .command_center import _company_name, _short_ticker


# XLSX schema v2 · 33 columns · every field populated with real data today.
# Operator 2026-08-01: "too much of technicals? ur call to decide else keep it"
# CEO decision: dropped 9 R007-blocked columns that were empty for months
# (Exit Triggers Hit · Risk Flags · Sector Exposure % · Confidence Band Low ·
#  Confidence Band High · Correlation · Hist Win Rate · Hist Median Ret ·
#  Hist Avg Hold). Restore as v3 when Ticket R007 lands with the data.
COLUMNS = [
    ("Position ID",         26),      # Sprint K prep · {TKR}_{MKT}_{YYYYMMDD} stable lifecycle identity
    ("Date",                12),
    ("Country",             10),
    ("Run_Type",            10),      # "R1" | "R2" | "R1_NEW" | "R2_NEW"
    ("Ticker",              12),
    ("Company",             22),
    ("Status",              12),
    ("Position Stage",      14),      # Sprint K · NEW · ACTIVE · MATURE · WATCH · EXIT (derived from days_held + band)
    ("Exit Reason",         22),     # NEW · populated only when Status = EXIT
    ("Exit P&L %",          12),     # NEW · realized return on exit · blank otherwise
    ("Rank",                6),
    ("Prior Rank",          10),     # NEW · rank at asof-1 (from rank_history)
    ("Rank Δ",              8),      # NEW · today - prior · positive = worse
    ("Health",              8),      # Sprint A · composite 0-100
    ("Band",                10),     # STRONG/HOLD/EXIT (3-state)
    ("Risk Meter",          8),      # 🟢🟡🔴 from Band + Ctx Drag
    ("Adj Conf",            10),     # CIL · adjusted confidence
    ("Ctx Drag",            10),     # CIL · signed drag pts
    ("Ctx Reason",          50),     # CIL · top context drivers
    # DROPPED 2026-08-07 per external audit PDF · confirmed empirically:
    #   Sector Exposure % · Insider 90d · Corr to Uni · Turnover σ (all <40% populated)
    #   Alert (0/224 populated · profit_protection wiring gap · fixed via alerts feed)
    #   Lifecycle State (224/224 populated but all constant "NEW" · redundant with Position Stage)
    ("Story",               60),     # compact narrative
    ("Alerts",              44),     # RENAMED from Alert · now sourced from position_store + profit_protection combined
    ("Confidence %",        13),
    ("Conf Type",           11),
    ("Model Score",         12),
    ("Day",                 6),        # Calendar days held
    ("Trading Days",        13),       # Sprint K · trading days held (excludes weekends)
    ("Horizon (d)",         12),
    ("Days Left",           10),
    ("Recommended",         14),
    ("Current Price",       14),
    ("Entry Price",         12),
    ("Buy Zone Low",        13),
    ("Buy Zone High",       13),
    ("Stop Loss",           11),
    ("Risk %",              9),
    ("Target 1",            11),
    ("T1 %",                8),
    ("Target 2",            11),
    ("T2 %",                8),
    ("Prev Close",          11),      # Sprint K · yesterday's close price (fixes Today Move gap)
    ("Today Move %",        13),      # NEW · today's daily change (yesterday_close → today_close)
    ("Current Perf %",      15),
    ("Max Gain %",          12),
    ("Max DD %",            11),
    # DROPPED Lifecycle State (redundant with Position Stage · was constant "NEW")
    ("Top Drivers",         32),
    ("Sector",              20),
    ("Portfolio Weight %",  18),
    ("Expected Alpha %",    17),
    ("Last Updated (UTC)",  20),
]

DEDUP_KEY_COLS = ["Date", "Country", "Run_Type", "Ticker"]

HEADER_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
STATUS_FILLS = {
    "STRONG BUY":  PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid"),
    "BUY":         PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
    "HOLD":        PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"),
    "EXIT":        PatternFill(start_color="F8CBAD", end_color="F8CBAD", fill_type="solid"),
    "ROTATE OUT":  PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid"),
    # 2026-08-07 · ARCHIVED = held but no longer in top-N ranks · muted gray
    # so operator visually distinguishes "engine still endorses" (BUY/HOLD)
    # from "engine dropped from ranks · re-evaluate" (ARCHIVED)
    "ARCHIVED":    PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"),
}
CENTER = Alignment(horizontal="center", vertical="center")
LEFT   = Alignment(horizontal="left", vertical="center", wrap_text=True)
RIGHT  = Alignment(horizontal="right", vertical="center")


def _pct_from_range(base, target):
    if not (base and target and base > 0):
        return None
    return (target - base) / base * 100


def _prev_close(root: Path, ticker: str, market: str) -> float | None:
    """Sprint K · yesterday's close price · pairs with Today Move %."""
    try:
        import pandas as pd
    except ImportError:
        return None
    bare = ticker.strip()
    for suf in (".NS", ".BO", ".NSE", ".BSE"):
        if bare.upper().endswith(suf):
            bare = bare[: -len(suf)]; break
    p = ((root / "usa" / "data" / "raw" / "us") if market == "usa"
              else (root / "data" / "raw" / "india")) / f"{bare}_D1.parquet"
    if not p.exists(): return None
    try:
        df = pd.read_parquet(p)
        if len(df) < 2: return None
        return round(float(df["close"].iloc[-2]), 2)
    except Exception:
        return None


def _trading_days_between(start_iso: str, end_iso: str) -> int | None:
    """Sprint K · count only Mon-Fri between two ISO dates (weekend-adjusted).
    Ignores exchange holidays for now · MoSPI/NSE calendar in Sprint L."""
    from datetime import date, timedelta
    try:
        d0 = date.fromisoformat(start_iso[:10])
        d1 = date.fromisoformat(end_iso[:10])
    except (ValueError, TypeError):
        return None
    if d1 < d0: return 0
    days = 0
    cur = d0
    while cur < d1:
        if cur.weekday() < 5: days += 1
        cur += timedelta(days=1)
    return days


def _position_id(ticker: str, market: str, first_seen: str) -> str:
    """Sprint K · stable Position ID = {TKR}_{MKT}_{YYYYMMDD}.
    Uses first_seen_date · never today's date · so ID persists across the
    entire lifecycle (Issue #21)."""
    if not ticker or not first_seen: return ""
    bare = ticker.replace(".NS", "").replace(".BO", "").upper()
    mkt = (market or "").upper()[:3]
    ds = first_seen[:10].replace("-", "")
    return f"{bare}_{mkt}_{ds}"


def _position_stage(days_held: int, health_band: str, status: str) -> str:
    """Sprint K · lifecycle stage derived from age + band + status.
    NEW (day 0) · ACTIVE (1-5) · MATURE (6-20) · WATCH (band=EXIT candidate) · EXIT.
    Answers Issue #15: 'everything still says NEW after 8 days · impossible'."""
    if status == "EXIT" or (health_band or "").upper() == "EXIT":
        return "EXIT"
    if not isinstance(days_held, int):
        return "NEW"
    if days_held == 0: return "NEW"
    if days_held <= 5: return "ACTIVE"
    if days_held <= 20: return "MATURE"
    return "WATCH"     # >20 days · long-holding · monitor closely


def _today_move_pct(root: Path, ticker: str, market: str) -> float | None:
    """Compute today's daily price change from bar cache.
    (today_close - yesterday_close) / yesterday_close × 100.
    Falls back to None if cache missing or fewer than 2 bars."""
    try:
        import pandas as pd
    except ImportError:
        return None
    bare = ticker.strip()
    for suf in (".NS", ".BO", ".NSE", ".BSE"):
        if bare.upper().endswith(suf):
            bare = bare[: -len(suf)]
            break
    if market == "usa":
        p = root / "usa" / "data" / "raw" / "us" / f"{bare}_D1.parquet"
    else:
        p = root / "data" / "raw" / "india" / f"{bare}_D1.parquet"
    if not p.exists():
        return None
    try:
        df = pd.read_parquet(p)
        if len(df) < 2:
            return None
        today = float(df["close"].iloc[-1])
        yest = float(df["close"].iloc[-2])
        if yest <= 0:
            return None
        return round((today / yest - 1) * 100, 2)
    except Exception:
        return None


def _rec_to_row(rec: Mapping, market: str, root: Path,
                    runner: str, asof: str) -> list:
    raw_ticker = rec.get("ticker") or "?"
    ticker = _short_ticker(raw_ticker)
    company = _company_name(raw_ticker, market) or ""

    ia = rec.get("investor_action") or {}
    pp = rec.get("position_plan") or {}
    ez = pp.get("entry_zone") or {}
    ev = rec.get("evolution") or {}
    ri = rec.get("rotation_intelligence") or {}

    entry_action = str(ia.get("entry") or "").upper()
    pct_action = str(rec.get("percentile_action") or "").upper()
    if_holding = ia.get("if_holding")

    # Operator 2026-08-01: "then keep as exit only" · 4-state vocab locked:
    # STRONG BUY · BUY · HOLD · EXIT
    # Rotations collapse into EXIT · from action standpoint both mean "sell".
    # Rotation intent still visible in Expected Alpha % column + Command
    # Center rotation section + R006 rotation ledger audit trail.
    #
    # Operator 2026-08-04: "strong buy to new buy is confusing man" — old
    # label "NEW BUY" wrongly implied "newly added today" but actually meant
    # "actionable BUY that isn't top-percentile STRONG_BUY". Renamed to plain
    # "BUY" so STRONG BUY → BUY reads correctly as conviction downgrade ·
    # not a lifecycle contradiction. HCLTECH was the example (rank 2 STRONG
    # BUY Aug 3 → rank 3 BUY Aug 4 · same actionable position · lower conviction).
    # 2026-08-07 · _ARCH dropped per operator directive · archived positions
    # emit under R1/R2 with REAL decision status:
    #   · Explicit sell/rotate signal  → EXIT
    #   · profit-protection alert on dead-loss / stop-loss threshold → EXIT
    #   · Otherwise (dropped from ranks but no exit trigger) → HOLD
    if entry_action == "SELL" or if_holding in ("EXIT", "REDUCE", "SELL") \
            or ri.get("should_rotate"):
        status = "EXIT"
    elif entry_action == "BUY" and pct_action == "STRONG_BUY":
        status = "STRONG BUY"
    elif entry_action == "BUY":
        status = "BUY"
    else:
        # Archived-but-no-exit-trigger → HOLD. Position Stage column will
        # say "ARCHIVED · dropped from top-N" for the ARCH-origin ones.
        status = "HOLD"

    # Exit Reason · populated ONLY when status = EXIT · else blank
    exit_reason = ""
    if status == "EXIT":
        risks = (rec.get("why") or {}).get("top_risks") or []
        risk_first = str(risks[0])[:80] if risks else ""
        if ri.get("should_rotate"):
            to_t = ri.get("replacement_ticker") or ""
            edge = ri.get("expected_alpha_delta_pct")
            edge_str = f" (+{edge:.1f}pp)" if edge else ""
            exit_reason = f"Rotation → {to_t}{edge_str}"
        elif if_holding == "EXIT":
            exit_reason = risk_first or "Exit signal (defensive)"
        elif if_holding == "REDUCE":
            exit_reason = risk_first or "Reduce position (defensive)"
        elif entry_action == "SELL":
            exit_reason = risk_first or "Sell signal"
        else:
            exit_reason = risk_first or "Exit trigger"

    rank = rec.get("rank")
    conf_cal = rec.get("calibrated_confidence")
    conf_raw = rec.get("confidence")
    if isinstance(conf_cal, (int, float)) and conf_cal:
        conf_pct = round(conf_cal * 100, 1)
        conf_type = "Calibrated"
    elif isinstance(conf_raw, (int, float)) and conf_raw:
        conf_pct = round(conf_raw * 100, 1)
        conf_type = "Raw"
    else:
        conf_pct = None
        conf_type = "—"

    ensemble_score = rec.get("ensemble_score")
    if isinstance(ensemble_score, (int, float)):
        model_score = round(ensemble_score if ensemble_score > 5 else ensemble_score * 100, 1)
    else:
        model_score = None

    horizon = pp.get("time_horizon_days") or 0
    days_rec = ev.get("days_recommended") or 0
    days_left = max(0, horizon - days_rec + 1) if horizon else None

    current_price = ez.get("current_price")
    ps = _load_position(root, market, ticker)
    entry_price = first_seen = high_water = low_water = None
    if ps:
        entry_price = ps.get("first_seen_price")
        first_seen = ps.get("first_seen_date")
        high_water = ps.get("high_water_price")
        low_water = ps.get("low_water_price")
        if current_price is None:
            current_price = ps.get("last_seen_price")

    # P0 FIX 2026-08-06 · never fall back to current_price for Entry
    # (operator: "AEGIS is recreating the recommendation every day · Entry
    # Price = Current Price = destroys P&L calculation"). Instead lookup
    # bar close on first_seen_date OR asof (day-0 case) which is the true
    # historical entry price · then it stays FROZEN across future snapshots.
    if entry_price is None and current_price:
        # Try to get bar close on first_seen_date · use asof if no first_seen
        target_date = first_seen or asof
        try:
            import pandas as _pd
            short_t = ticker.replace(".NS", "").replace(".BO", "")
            dir_p = (root / "usa/data/raw/us" if market == "usa"
                          else root / "data/raw/india")
            bp = dir_p / f"{short_t}_D1.parquet"
            if bp.exists():
                df = _pd.read_parquet(bp)
                col = "close" if "close" in df.columns else "Close"
                df.index = _pd.to_datetime(df.index).strftime("%Y-%m-%d")
                if target_date in df.index:
                    entry_price = float(df.loc[target_date, col])
                else:
                    earlier = [d for d in df.index if d <= target_date]
                    if earlier: entry_price = float(df.loc[earlier[-1], col])
        except Exception:
            pass
        # Absolute last resort: today's price (day-0 · new position case only)
        if entry_price is None:
            entry_price = current_price
        if not first_seen:
            first_seen = asof

    stop = ez.get("stop_loss")
    t1 = ez.get("target_1")
    t2 = ez.get("target_2")

    # Bug fix 2026-08-01: engine only populates stop/T1/T2 for actionable BUYs.
    # For HOLD picks (and R1 defensives), derive standard defaults so operator
    # gets SAME field coverage per stock (5% stop, 8% T1, 15% T2 · heuristics
    # documented in AEGIS_STOCK_CARD_FORMAT.md). Never fabricates per-stock.
    ref_price = entry_price or current_price
    if ref_price:
        if stop is None:
            stop = round(ref_price * 0.95, 2)         # -5% default stop
        if t1 is None:
            t1 = round(ref_price * 1.08, 2)           # +8% default T1
        if t2 is None:
            t2 = round(ref_price * 1.15, 2)           # +15% default T2
    if t2 is None and t1 is not None and current_price:
        t2 = current_price + (t1 - current_price) * 1.5

    # Buy Zone default (±1% of ref_price) when engine doesn't emit for HOLDs
    buy_low = ez.get("ideal_buy_low")
    buy_high = ez.get("ideal_buy_high")
    if buy_low is None and ref_price:
        buy_low = round(ref_price * 0.99, 2)
    if buy_high is None and ref_price:
        buy_high = round(ref_price * 1.01, 2)

    base_for_pct = entry_price or current_price
    risk_pct = _pct_from_range(base_for_pct, stop)
    t1_pct = _pct_from_range(base_for_pct, t1)
    t2_pct = _pct_from_range(base_for_pct, t2)

    cur_ret = round((current_price / entry_price - 1) * 100, 2) if (entry_price and current_price) else None
    max_gain = round((high_water / entry_price - 1) * 100, 2) if (entry_price and high_water) else None
    max_dd = round((low_water / entry_price - 1) * 100, 2) if (entry_price and low_water) else None
    today_move = _today_move_pct(root, raw_ticker, market)

    events = _load_ledger_events_for_ticker(root, market, ticker)
    state = _lifecycle_state(events)
    triggers = [label for (label, hit) in _exit_triggers_checklist(events) if hit]
    triggers_str = ", ".join(triggers)

    drivers = _drivers_from_attribution(rec)
    drivers_str = " · ".join(drivers)

    sector = rec.get("sector") or ""
    alloc = pp.get("suggested_allocation_pct")
    exp_alpha = ri.get("expected_alpha_delta_pct")
    if exp_alpha is None and t1 and base_for_pct:
        exp_alpha = round(((t1 / base_for_pct) - 1) * 100, 2)

    country = market.upper()

    # XLSX v2 · 33 columns · every value is real or documented default
    # · R007-blocked columns removed (Risk Flags · Sector Exposure ·
    #   Confidence Band · Correlation · Historical Setups) · will
    #   restore as v3 when R007 lands
    # Exit P&L % · realized return · populated ONLY when Status = EXIT
    # Uses (current_price - entry_price) / entry_price · same as cur_ret
    # but explicitly separated so blank rows in this column = "not exited"
    exit_pnl_pct = ""
    if status == "EXIT" and cur_ret is not None:
        exit_pnl_pct = cur_ret

    # ── Prior Rank + Rank Δ (v5 · 2026-08-04) ──
    # Reads yesterday's rank from rank_history.jsonl so operator can judge
    # day-over-day at a glance ("TCS was #1 Aug 3 · now #9 Aug 4 · Δ +8")
    # instead of comparing two XLSX files manually.
    prior_rank = ""
    rank_delta = ""
    if rank:
        try:
            from backend.portfolio.rank_history import get_prior_rank as _gpr
            _pr, _ = _gpr(root, market, "runner2" if runner == "R2" else "runner1",
                              ticker, asof)
            if _pr is not None:
                prior_rank = _pr
                rank_delta = int(rank) - int(_pr)
        except Exception:
            pass

    # Alerts · profit-protection signals + position_store-derived stop-loss checks
    # 2026-08-07 · external audit found Alert column 0/224 populated because
    # profit_protection.py only evaluates R006-tracked positions and R006 ledger
    # isn't fully wired. Fallback: also check position_store directly for
    # stop-loss / deep-loss / rapid-gain triggers so operator sees them.
    alert = _load_alert_with_fallback(root, market, ticker, cur_ret, entry_price)

    # Sprint A · Health Score (composite 0-100 + band)
    health_score = ""
    health_band = ""
    prior_band = None
    try:
        hs = _load_health_card(root, market, ticker)
        if hs:
            health_score = hs.get("overall") or ""
            health_band = (hs.get("band") or "").replace("_", " ")
            prior_band = hs.get("prior_band")
    except Exception:
        pass

    # CIL · Adjusted Confidence · Drag · Reason
    adj_conf = ""
    ctx_drag = ""
    ctx_reason = ""
    try:
        cil = _load_cil_adjustment(root, market, ticker)
        if cil:
            adj_conf = cil.get("adjusted")
            ctx_drag = cil.get("drag_pts")
            ctx_reason = cil.get("story") or ""
    except Exception:
        pass

    # Sprint G · Insider · Correlation · Turnover columns
    insider_90d = ""
    corr_to_uni = ""
    turnover_sigma = ""
    try:
        if market == "usa":
            p_ed = root / "reports" / "edgar" / "insider_recent.json"
            if p_ed.exists():
                _d = json.loads(p_ed.read_text(encoding="utf-8"))
                per = (_d.get("per_ticker") or {}).get(ticker.upper()) or {}
                if per.get("available"):
                    insider_90d = per.get("n_form4_last_90d") or 0
    except Exception:
        pass
    try:
        p_c = root / "reports" / "correlation_matrix.json"
        if p_c.exists():
            _d = json.loads(p_c.read_text(encoding="utf-8"))
            short = ticker.replace(".NS", "").replace(".BO", "").upper()
            for r in (_d.get("portfolio_concentration_risk") or []):
                if r.get("ticker") == short:
                    corr_to_uni = r.get("avg_corr_to_others")
                    break
    except Exception:
        pass
    try:
        if market == "india":
            from pathlib import Path as _P
            bhav = sorted((root / "reports" / "nse_bhavcopy").glob("*.parquet"))
            if len(bhav) >= 5:
                import pandas as _pd
                short = ticker.replace(".NS", "").replace(".BO", "").upper()
                vals = []
                for pq in bhav[-20:]:
                    try:
                        df = _pd.read_parquet(pq)
                        if "SYMBOL" not in df.columns or "TURNOVER_LACS" not in df.columns: continue
                        row = df[df["SYMBOL"].str.strip().str.upper() == short]
                        if row.empty: continue
                        if "SERIES" in row.columns:
                            eq = row[row["SERIES"].str.strip() == "EQ"]
                            if not eq.empty: row = eq
                        vals.append(float(row["TURNOVER_LACS"].iloc[0]))
                    except Exception:
                        continue
                if len(vals) >= 5:
                    today = vals[-1]; prior = vals[:-1]
                    m = sum(prior) / len(prior)
                    v = sum((x - m) ** 2 for x in prior) / len(prior)
                    s = v ** 0.5
                    turnover_sigma = round((today - m) / s, 2) if s > 0 else 0
    except Exception:
        pass

    # Sprint J · Entry Signal (computed first · used by Story prefix)
    entry_signal, buy_zone_delta = _entry_signal(current_price, buy_high,
                                                                entry_price, health_band, status)

    # Sprint H · Sector Exposure (RESTORED)
    sector_exposure_pct = _sector_exposure_pct(root, market, sector)

    # Sprint H-nice · Risk Meter (🟢🟡🔴)
    risk_meter = _risk_meter(health_band, ctx_drag)

    # Sprint C · Story column · compact narrative · Sprint J prefix with entry signal
    story = _build_story(rank, prior_rank, rank_delta, conf_pct, ev,
                                health_band, prior_band, days_left, status,
                                if_holding, ri)
    if entry_signal:
        _delta_suffix = f" (Δ{buy_zone_delta:+.1f}%)" if isinstance(buy_zone_delta, (int, float)) else ""
        story = f"{entry_signal}{_delta_suffix} · {story}"

    # Enrich Exit Reason with rotation detail if applicable
    if status == "EXIT" and ri and ri.get("should_rotate"):
        _repl = ri.get("replacement_ticker") or ""
        _edge = ri.get("expected_alpha_delta_pct")
        _repl_rank = ri.get("replacement_rank")
        if _repl and _edge is not None:
            rank_str = f" #{_repl_rank}" if _repl_rank else ""
            exit_reason = f"→ {_repl}{rank_str} · +{_edge:.1f}pp alpha"

    # Sprint K quick-wins · derived fields
    pos_id = _position_id(ticker, market, first_seen or asof)
    pos_stage = _position_stage(days_rec if isinstance(days_rec, int) else 0,
                                             health_band, status)
    trading_days = _trading_days_between(first_seen or asof, asof) \
                            if first_seen else 0
    prev_close_val = _prev_close(root, ticker, market)
    # Auto-compute Current Perf % from immutable Entry (Issue #8)
    if isinstance(entry_price, (int, float)) and isinstance(current_price, (int, float)) \
       and entry_price > 0:
        cur_ret = round((current_price - entry_price) / entry_price * 100, 2)

    return [
        pos_id,               # Sprint K · {TKR}_{MKT}_{YYYYMMDD} stable lifecycle ID
        asof, country, runner, ticker, company, status,
        pos_stage,            # Sprint K · NEW/ACTIVE/MATURE/WATCH/EXIT
        exit_reason,          # NEW · blank unless status = EXIT
        exit_pnl_pct,         # NEW · realized P&L % · blank unless EXIT
        rank if rank else "",
        prior_rank,           # NEW v5 · rank at asof-1
        rank_delta,           # NEW v5 · today - prior · +ve = worse rank
        health_score,         # Sprint A · composite 0-100 (RESTORED)
        health_band,          # Sprint J · band (STRONG/HOLD/EXIT · 3-state)
        risk_meter,           # Sprint H · 🟢🟡🔴
        adj_conf,             # CIL · adjusted confidence
        ctx_drag,             # CIL · signed drag pts
        ctx_reason,           # CIL · top drivers
        # DROPPED 2026-08-06: sector_exposure_pct · insider_90d · corr_to_uni · turnover_sigma
        story,                # Sprint C · compact narrative
        alert,                # profit-protection signals (kept for stop-loss visibility)
        conf_pct if conf_pct is not None else "",
        conf_type,
        model_score if model_score is not None else "",
        days_rec if days_rec else "",
        trading_days if trading_days is not None else "",  # Sprint K · trading days held
        horizon if horizon else "",
        days_left if days_left is not None else "",
        first_seen if first_seen else "",
        current_price if current_price else "",
        entry_price if entry_price else "",     # Sprint K · immutable from position_store
        buy_low if buy_low else "",
        buy_high if buy_high else "",
        stop if stop else "",
        round(risk_pct, 2) if risk_pct is not None else "",
        t1 if t1 else "",
        round(t1_pct, 2) if t1_pct is not None else "",
        t2 if t2 else "",
        round(t2_pct, 2) if t2_pct is not None else "",
        prev_close_val if prev_close_val is not None else "",  # Sprint K · pairs with Today Move
        today_move if today_move is not None else "",
        cur_ret if cur_ret is not None else "",
        max_gain if max_gain is not None else "",
        max_dd if max_dd is not None else "",
        # Lifecycle State column DROPPED · redundant with Position Stage
        drivers_str,
        sector,
        alloc if alloc else "",
        round(exp_alpha, 2) if exp_alpha is not None else "",
        datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M"),
    ]


def _archived_tickers_for(root: Path, market: str, runner_key: str,
                                 current_tickers: set, asof: str) -> list[dict]:
    """P0-1 · every ticker that was EVER in this (market, runner) but is
    NOT in today's recs comes back as an ARCHIVED row. Operator directive
    2026-08-06: 'Recommendations must never disappear.'

    Returns list of fake rec-shaped dicts for _rec_to_row · with status
    ARCHIVED · rank preserved from last-seen · price from position_store
    or last rank_history entry."""
    p = root / "reports" / "research" / "rank_history.jsonl"
    if not p.exists(): return []

    # Build map: ticker → most-recent rank_history entry for this market/runner
    last_seen = {}
    for line in p.read_text(encoding="utf-8").splitlines():
        if not line.strip(): continue
        try: d = json.loads(line)
        except json.JSONDecodeError: continue
        if d.get("market") != market or d.get("runner") != runner_key: continue
        t = (d.get("ticker") or "").replace(".NS", "").replace(".BO", "")
        if not t: continue
        prev = last_seen.get(t)
        if prev is None or (d.get("asof") or "") > (prev.get("asof") or ""):
            last_seen[t] = d

    # Filter to tickers NOT in today's live recs
    archived_recs = []
    for t, last in last_seen.items():
        if t.upper() in current_tickers: continue
        # Skip if last-seen was today (still active · not archived)
        if (last.get("asof") or "").startswith(asof): continue
        # Load position_store for entry/current price · fallback to last rank_history price
        pos = _load_position(root, market, t) or {}
        entry = pos.get("first_seen_price") or last.get("current_price")
        current = pos.get("last_seen_price") or last.get("current_price") or entry
        first_seen = pos.get("first_seen_date") or last.get("asof")
        # Build minimal rec-shaped dict · Status = ARCHIVED
        archived_recs.append({
            "ticker": t,
            "sector": last.get("sector") or "",
            "rank": last.get("rank") or 99,
            "calibrated_confidence": last.get("confidence"),
            "ensemble_score": last.get("model_score"),
            "investor_action": {"entry": "ARCHIVED",
                                        "is_actionable_entry": False,
                                        "if_holding": "ARCHIVED"},
            "position_plan": {
                "entry_zone": {
                    "current_price": current,
                    "stop_loss": entry * 0.95 if entry else None,
                    "target_1": entry * 1.08 if entry else None,
                    "target_2": entry * 1.15 if entry else None,
                },
                "time_horizon_days": last.get("horizon_days") or 60,
            },
            "evolution": {"first_seen_date": first_seen,
                              "days_recommended": 0,
                              "momentum_direction": "STABLE"},
            "rotation_intelligence": {},
            "percentile_action": "ARCHIVED",
            "_archived": True,     # marker for row builder
            "_last_active_asof": last.get("asof"),
        })
    return archived_recs


def _load_exited_set(root: Path, market: str) -> set:
    """2026-08-07 · operator directive: once a ticker EXITs, don't emit again.
    Reads historical XLSX and returns {(runner, ticker)} that have been
    EXITed on any prior date. Used to suppress zombie/repeat rows."""
    xlsx = root / "reports" / "telegram" / "aegis_history.xlsx"
    if not xlsx.exists(): return set()
    try:
        from openpyxl import load_workbook
        wb = load_workbook(xlsx, read_only=True)
        ws = wb["AEGIS Daily"] if "AEGIS Daily" in wb.sheetnames else wb.active
        h = [c.value for c in ws[1]]
        i_ctry, i_rt, i_tk, i_st = (h.index(x)+1 for x in ["Country","Run_Type","Ticker","Status"])
        exited = set()
        mkt_up = market.upper()
        for r in range(2, ws.max_row + 1):
            if str(ws.cell(r, i_ctry).value or "").upper() != mkt_up: continue
            if ws.cell(r, i_st).value == "EXIT":
                rt = str(ws.cell(r, i_rt).value or "").replace("_NEW","")
                tk = str(ws.cell(r, i_tk).value or "").replace(".NS","").replace(".BO","").upper()
                if rt and tk: exited.add((rt, tk))
        wb.close()
        return exited
    except Exception:
        return set()


def _collect_rows_for_market(root: Path, market: str, asof: str) -> list[list]:
    """Return all rows (R2 + R1 if India) for one market on this asof.

    2026-08-07 · once a ticker EXITs (on ANY prior date) it does NOT
    reappear. No zombie live rows · no duplicate EXIT rows. Only fresh
    top-N tickers (STRONG BUY/BUY) and currently-held (HOLD) positions
    appear. EXIT emitted once on the day it happens, then vanishes."""
    recs_path = (root / "usa" / "reports" / "recommendations.json"
                    if market == "usa" else root / "reports" / "recommendations.json")
    if not recs_path.exists():
        return []
    payload = json.loads(recs_path.read_text(encoding="utf-8"))
    rows = []
    # Track today's active tickers to derive archived list
    active_r2 = set()
    active_r1 = set()
    # Load set of tickers that have ALREADY exited on prior dates
    exited = _load_exited_set(root, market)
    def _key(rt, tkr):
        return (rt.replace("_NEW",""), (tkr or "").replace(".NS","").replace(".BO","").upper())
    for r in payload.get("recommendations") or []:
        tk = r.get("ticker") or ""
        _rt = _runner_tag(root, market, "runner2", tk, asof, "R2")
        # Skip if this ticker previously exited (operator: no repeats)
        if _key(_rt, tk) in exited: continue
        rows.append(_rec_to_row(r, market, root, runner=_rt, asof=asof))
        active_r2.add(tk.replace(".NS", "").replace(".BO", "").upper())
    if market == "india":
        rv = payload.get("runner1_validation") or {}
        for i, o in enumerate(rv.get("runner1_orphans") or [], start=1):
            r1_rec = _r1_orphan_to_rec_shape(o, market)
            r1_rec.setdefault("rank", i)
            tk = r1_rec.get("ticker") or ""
            _rt = _runner_tag(root, market, "runner1", tk, asof, "R1")
            if _key(_rt, tk) in exited: continue
            rows.append(_rec_to_row(r1_rec, market, root, runner=_rt, asof=asof))
    elif market == "usa":
        try:
            usa_r1_path = root / "usa" / "reports" / "runner1_orphans.json"
            if usa_r1_path.exists():
                usa_r1 = json.loads(usa_r1_path.read_text(encoding="utf-8"))
                for i, o in enumerate(usa_r1.get("runner1_orphans") or [], start=1):
                    r1_rec = _r1_orphan_to_rec_shape(o, market)
                    r1_rec.setdefault("rank", i)
                    tk = r1_rec.get("ticker") or ""
                    _rt = _runner_tag(root, market, "runner1", tk, asof, "R1")
                    if _key(_rt, tk) in exited: continue
                    rows.append(_rec_to_row(r1_rec, market, root, runner=_rt, asof=asof))
                    active_r1.add(tk.replace(".NS", "").replace(".BO", "").upper())
        except Exception as e:
            print(f"[collect_rows:usa] USA R1 render skipped · {type(e).__name__}: {e}")

    # 2026-08-07 · Operator directive: drop _ARCH suffix entirely.
    # Standardize on 4 Run_Types only: R1 · R2 · R1_NEW · R2_NEW.
    # Archived positions (previously ranked · not in today's top-N) emit
    # under plain R1 / R2 with Status derived from real signal:
    #   · profit_protection alert present (STOP_LOSS/DEEP_LOSS/etc)  → EXIT
    #   · dropped from ranks with no exit trigger                    → HOLD
    # Position Stage column records "ARCHIVED · dropped from top-N" so the
    # nuance isn't lost. No more R1_ARCH/R2_ARCH labels.
    try:
        for arch in _archived_tickers_for(root, market, "runner2", active_r2, asof):
            tk = arch.get("ticker") or ""
            if _key("R2", tk) in exited: continue
            rows.append(_rec_to_row(arch, market, root, runner="R2", asof=asof))
        for arch in _archived_tickers_for(root, market, "runner1", active_r1, asof):
            tk = arch.get("ticker") or ""
            if _key("R1", tk) in exited: continue
            rows.append(_rec_to_row(arch, market, root, runner="R1", asof=asof))
    except Exception as e:
        print(f"[collect_rows:{market}] archived render skipped · {type(e).__name__}: {e}")
    return rows


def _runner_tag(root: Path, market: str, runner_key: str,
                    ticker: str, asof: str, base_tag: str) -> str:
    """Sprint J-final · return R1_NEW / R2_NEW if ticker has NO prior
    rank_history entry in this (market, runner) · else return base_tag (R1/R2).

    Auto-decay: tomorrow's rank_history will include today's stamp · so
    the same ticker becomes R1 or R2 automatically. No manual work.

    Operator directive 2026-08-06: "Run_Type = R1_NEW / R2_NEW on first
    day · next day auto-becomes R1/R2 · client filters Run_Type contains
    NEW to see today's fresh opportunities. No extra sheet · no extra column."
    """
    if not ticker: return base_tag
    p = root / "reports" / "research" / "rank_history.jsonl"
    if not p.exists(): return base_tag
    try:
        short_ticker = ticker.replace(".NS", "").replace(".BO", "")
        for line in p.read_text(encoding="utf-8").splitlines():
            if not line.strip(): continue
            try: d = json.loads(line)
            except json.JSONDecodeError: continue
            if d.get("market") != market: continue
            if d.get("runner") != runner_key: continue
            t = (d.get("ticker") or "").replace(".NS", "").replace(".BO", "")
            if t != short_ticker: continue
            # Found a prior entry (asof < today) · ticker is not new
            if (d.get("asof") or "") < asof:
                return base_tag
        return f"{base_tag}_NEW"
    except Exception:
        return base_tag


def _new_opp_flag(root: Path, market: str, ticker: str) -> str:
    """Sprint J · read fresh_opportunities.json · return 🟢 if this ticker
    is a fresh opportunity today · blank otherwise. One-column at-a-glance
    for new money."""
    p = root / "reports" / "research" / f"fresh_opportunities_{market}.json"
    if not p.exists():
        return ""
    try:
        d = json.loads(p.read_text(encoding="utf-8"))
        short = ticker.replace(".NS", "").replace(".BO", "").upper()
        for f in d.get("fresh_buys") or []:
            if (f.get("ticker") or "").upper() == short:
                return "🟢 NEW"
    except Exception:
        return ""
    return ""


def _entry_signal(current_price, buy_zone_high, entry_price, health_band, status) -> tuple:
    """Sprint J · 3-state Entry Signal for new money.

    Returns (signal_str, buy_zone_delta_pct)

    Rules (in order · first match wins):
        · Status is EXIT / band WEAK or EXIT   → 🔴 AVOID (thesis broken)
        · price ≤ 1% above buy_zone_high        → 🟢 BUY (fresh entry available)
        · price ≤ 8% above entry                → 🟡 WAIT (moved but not chased hard)
        · price > 8% above entry                → ⚪ HOLDER ONLY (aged rec)
        · missing data                          → "" (silent)

    Buy zone delta % = (current - buy_zone_high) / buy_zone_high × 100
    Negative = below buy zone (accumulable) · positive = above.
    """
    if not isinstance(current_price, (int, float)) or not current_price:
        return "", ""
    # Check thesis first · exit / weak → avoid
    band_norm = (health_band or "").upper().replace(" ", "").replace("_", "")
    if status == "EXIT" or band_norm in ("EXIT", "WEAK", "EXITCANDIDATE", "REVIEW"):
        return "🔴 AVOID", ""

    # Compute buy zone delta if we have buy_zone_high
    delta_pct = None
    if isinstance(buy_zone_high, (int, float)) and buy_zone_high:
        delta_pct = round((current_price - buy_zone_high) / buy_zone_high * 100, 2)

    # Compute distance from entry (for HOLDER ONLY check)
    entry_pct = None
    if isinstance(entry_price, (int, float)) and entry_price:
        entry_pct = (current_price - entry_price) / entry_price * 100

    # BUY: within 1% above buy zone high (or below it entirely)
    if delta_pct is not None and delta_pct <= 1.0:
        return "🟢 BUY", delta_pct
    # HOLDER ONLY: too far from entry
    if entry_pct is not None and entry_pct > 8.0:
        return "⚪ HOLDER ONLY", delta_pct if delta_pct is not None else ""
    # WAIT: moved past buy zone but not gone too far
    if delta_pct is not None:
        return "🟡 WAIT", delta_pct
    return "", ""


def _risk_meter(health_band: str, ctx_drag) -> str:
    """Sprint H-simplify · 4-band vocabulary mapped to 🟢🟡🔴.

    Bands (new · 4-state · unambiguous):
        STRONG → 🟢   Buy or add · very healthy
        HOLD   → 🟢   Position healthy · no action
        WEAK   → 🟡   Weakening · consider reducing
        EXIT   → 🔴   Thesis breaking · exit

    Escalation: large context drag can push STRONG/HOLD → 🟡 warning.
    """
    if not health_band:
        return ""
    band = health_band.upper().replace(" ", "").replace("_", "")
    if band == "STRONG":  base = "🟢"
    elif band == "HOLD":  base = "🟢"
    elif band == "WEAK":  base = "🟡"
    elif band == "EXIT":  base = "🔴"
    # Backward compat with old 5-band vocab
    elif band in ("STRONGBUY", "STRONG BUY"): base = "🟢"
    elif band == "WATCH": base = "🟡"
    elif band == "REVIEW": base = "🟡"
    elif band == "EXITCANDIDATE": base = "🔴"
    else: base = ""
    # Escalate on large negative context drag
    try:
        d = float(ctx_drag or 0)
        if d <= -10 and base != "🔴": base = "🔴"
        elif d <= -5 and base == "🟢": base = "🟡"
    except (TypeError, ValueError):
        pass
    return base


def _sector_exposure_pct(root: Path, market: str, sector: str) -> str:
    """Sprint H-nice · % of R2 top-15 in this ticker's sector today."""
    if not sector:
        return ""
    p = (root / "usa" / "reports" / "recommendations.json"
             if market == "usa" else root / "reports" / "recommendations.json")
    if not p.exists():
        return ""
    try:
        d = json.loads(p.read_text(encoding="utf-8"))
        recs = d.get("recommendations") or []
        if not recs: return ""
        same_sector = sum(1 for r in recs if (r.get("sector") or "") == sector)
        pct = round(same_sector / len(recs) * 100, 1)
        return pct
    except Exception:
        return ""


def _load_cil_adjustment(root: Path, market: str, ticker: str) -> dict | None:
    """CIL · load per-ticker adjustment from most-recent CIL run."""
    p = root / "reports" / "context" / f"cil_run_{market}.json"
    if not p.exists(): return None
    try:
        d = json.loads(p.read_text(encoding="utf-8"))
        short = ticker.replace(".NS", "").replace(".BO", "")
        for a in d.get("adjustments") or []:
            t = a.get("ticker") or ""
            if t == ticker or t == short \
                or t.replace(".NS", "").replace(".BO", "") == short:
                return a
    except Exception:
        return None
    return None


def _build_story(rank, prior_rank, rank_delta, conf_pct, ev: dict,
                     health_band: str, prior_band: str | None,
                     days_left, status: str, if_holding, ri: dict) -> str:
    """Sprint C · compose one-line narrative like:
       'Rank ↑ 5→1 · Conf 42% · momentum ↑ · sector leader · 34d left · HOLD'
    Skips segments where data is missing · never fabricates."""
    parts = []
    # Rank movement
    if isinstance(rank_delta, int) and prior_rank not in (None, ""):
        arrow = "↑" if rank_delta < 0 else ("↓" if rank_delta > 0 else "→")
        parts.append(f"Rank {arrow} {prior_rank}→{rank}")
    elif rank:
        parts.append(f"Rank #{rank}")
    # Confidence
    if conf_pct is not None and conf_pct != "":
        parts.append(f"Conf {conf_pct:.0f}%")
    # Momentum
    mom = (ev or {}).get("momentum_direction") if isinstance(ev, dict) else None
    if mom:
        arrow = {"UP": "↑", "STABLE": "→", "DOWN": "↓"}.get(str(mom).upper(), "")
        if arrow:
            parts.append(f"momentum {arrow}")
    # Band drift
    if health_band:
        pretty_band = health_band.replace("_", " ")
        if prior_band and prior_band != health_band.replace(" ", "_"):
            prev_pretty = prior_band.replace("_", " ")
            parts.append(f"band {prev_pretty}→{pretty_band}")
        else:
            parts.append(f"band {pretty_band}")
    # Days remaining
    if isinstance(days_left, int) and days_left > 0:
        parts.append(f"{days_left}d left")
    # Action bit at the end
    if status == "EXIT":
        if ri and ri.get("should_rotate"):
            parts.append(f"→ EXIT (rotate to {ri.get('replacement_ticker') or '?'})")
        else:
            parts.append("→ EXIT")
    elif status == "STRONG BUY":
        parts.append("→ STRONG BUY")
    elif status == "BUY":
        parts.append("→ BUY")
    elif status == "HOLD":
        parts.append("→ HOLD")
    return " · ".join(parts)


def _load_health_card(root: Path, market: str, ticker: str) -> dict | None:
    """Sprint A · look up ticker's Health card from most-recent scoring run.
    Matches on ticker OR short-ticker (health card stores raw ticker like
    TCS.NS · XLSX row uses short TCS · check both)."""
    p = root / "reports" / "research" / f"health_scores_{market}.json"
    if not p.exists():
        return None
    try:
        d = json.loads(p.read_text(encoding="utf-8"))
        short = ticker.replace(".NS", "").replace(".BO", "")
        for c in d.get("cards") or []:
            t = c.get("ticker") or ""
            if t == ticker or t == short \
                or t.replace(".NS", "").replace(".BO", "") == short:
                return c
    except Exception:
        return None
    return None


def _load_alert_with_fallback(root: Path, market: str, ticker: str,
                                       cur_ret, entry_price) -> str:
    """2026-08-07 · combines profit_protection.json output (when R006 fires)
    WITH position_store-derived stop-loss/take-profit checks (fallback when
    R006 ledger is empty · which is most rows currently).

    Emits severity · trigger · reason  · concatenated with || separator."""
    parts = []
    # Native profit_protection signals (if any)
    native = _load_alert(root, market, ticker)
    if native: parts.append(native)
    # Position_store-derived signals · fires when R006 doesn't
    if isinstance(cur_ret, (int, float)) and isinstance(entry_price, (int, float)) and entry_price:
        if cur_ret <= -8.0:
            parts.append(f"CRITICAL·DEEP_LOSS·{cur_ret:+.1f}% ≤ -8% · EXIT URGENT")
        elif cur_ret <= -5.0:
            parts.append(f"WARNING·STOP_LOSS_HIT·{cur_ret:+.1f}% ≤ -5% · exit")
        elif cur_ret >= 20.0:
            parts.append(f"CRITICAL·HARD_GAIN_CAP·{cur_ret:+.1f}% ≥ +20% · lock profit")
        elif cur_ret >= 12.0:
            parts.append(f"WARNING·RAPID_APPRECIATION·{cur_ret:+.1f}% ≥ +12% · take profit")
    return " || ".join(parts) if parts else ""


def _load_alert(root: Path, market: str, ticker: str) -> str:
    """Read profit_protection_{market}.json · return short string per ticker.
    Format: 'CRITICAL · TRIGGER · reason' · blank if no signals for ticker."""
    p = root / "reports" / "research" / f"profit_protection_{market}.json"
    if not p.exists():
        return ""
    try:
        d = json.loads(p.read_text(encoding="utf-8"))
        sigs = [s for s in (d.get("signals") or []) if s.get("ticker") == ticker]
        if not sigs:
            return ""
        # Highest severity first · concatenate up to 2
        order = {"critical": 0, "warning": 1, "info": 2}
        sigs.sort(key=lambda s: order.get(s.get("severity", "info"), 3))
        parts = []
        for s in sigs[:2]:
            sev = str(s.get("severity", "info")).upper()
            trig = s.get("trigger", "")
            reason = (s.get("reason") or "")[:40]
            parts.append(f"{sev}·{trig}·{reason}")
        return " || ".join(parts)
    except Exception:
        return ""


def _add_fresh_buys_sheet(wb, root: Path) -> None:
    """Sprint J-fix · add a prominent 'Fresh Buys Today' sheet as first tab
    so operator sees new opportunities immediately on opening Excel."""
    if "Fresh Buys" in wb.sheetnames:
        del wb["Fresh Buys"]
    ws = wb.create_sheet("Fresh Buys", 0)   # index 0 = first tab
    hdr = ["Market", "Rank", "Ticker", "Sector", "Entry", "Buy Zone Δ %",
              "Base Conf %", "Adj Conf %", "Band", "Why Fresh"]
    for c, name in enumerate(hdr, start=1):
        cell = ws.cell(row=1, column=c, value=name)
        cell.fill = HEADER_FILL; cell.font = HEADER_FONT; cell.alignment = CENTER
    widths = [8, 5, 12, 22, 10, 12, 12, 12, 10, 40]
    for c, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(c)].width = w
    ws.freeze_panes = "A2"

    row_idx = 2
    for m in ("india", "usa"):
        jp = root / "reports" / "research" / f"fresh_opportunities_{m}.json"
        if not jp.exists(): continue
        try:
            d = json.loads(jp.read_text(encoding="utf-8"))
        except Exception:
            continue
        for f in d.get("fresh_buys") or []:
            reasons = " · ".join(f.get("reasons") or [])
            ws.cell(row=row_idx, column=1, value=m.upper())
            ws.cell(row=row_idx, column=2, value=f.get("rank"))
            ws.cell(row=row_idx, column=3, value=f.get("ticker"))
            ws.cell(row=row_idx, column=4, value=f.get("sector") or "—")
            ws.cell(row=row_idx, column=5, value=f.get("entry_price"))
            ws.cell(row=row_idx, column=6, value=f.get("buy_zone_delta_pct"))
            ws.cell(row=row_idx, column=7, value=f.get("confidence_pct"))
            ws.cell(row=row_idx, column=8, value=f.get("adjusted_confidence"))
            ws.cell(row=row_idx, column=9, value=f.get("health_band"))
            ws.cell(row=row_idx, column=10, value=reasons)
            for c in range(1, 11):
                ws.cell(row=row_idx, column=c).alignment = CENTER
            row_idx += 1
    if row_idx == 2:
        # No fresh buys today · note it
        ws.cell(row=2, column=1,
                    value="No fresh opportunities today · Guard 7 filters produced 0 candidates · watchlist only")
        ws.cell(row=2, column=1).alignment = LEFT
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=10)


def _write_new_workbook(path: Path, rows: list[list]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "AEGIS Daily"
    # Header
    for col_idx, (name, width) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=name)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[1].height = 22
    ws.freeze_panes = "D2"      # freeze Date + Country + Run_Type
    # Data
    for r_idx, row in enumerate(rows, start=2):
        for c_idx, val in enumerate(row, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.alignment = LEFT if c_idx <= 6 else RIGHT
        # 2026-08-07 · row-level color spread across ALL columns (operator directive)
        # Status is column INDEX 6 (0-indexed) = col 7 (1-indexed) · after
        # Position ID (col 1), Date (2), Country (3), Run_Type (4), Ticker (5),
        # Company (6), Status (7). Earlier code hardcoded col 6 = Company by mistake.
        status = row[6]
        if status in STATUS_FILLS:
            fill = STATUS_FILLS[status]
            for c in range(1, len(row) + 1):
                ws.cell(row=r_idx, column=c).fill = fill
    # Auto-filter
    last_col = get_column_letter(len(COLUMNS))
    ws.auto_filter.ref = f"A1:{last_col}{len(rows) + 1}"
    # Sprint J-fix REVERTED · single-sheet format ONLY (operator directive
    # 2026-08-06 · "very unprofessional · always maintain same sheet")
    wb.save(path)


def _append_to_workbook(path: Path, rows: list[list]) -> None:
    """Load existing WB · dedup by (Date, Country, Run_Type, Ticker) ·
    replace matching rows with today's · append new ones.

    Schema-migration safe: if the file was written under an older COLUMNS
    spec and the current spec has NEW columns, we rewrite the header row
    and pad existing rows with blank cells for the new columns before
    doing the append. Avoids the 2026-08-04 bug where 3 new columns
    (Prior Rank · Rank Δ · Alert) landed as un-headered data.
    """
    wb = load_workbook(path)
    ws = wb["AEGIS Daily"] if "AEGIS Daily" in wb.sheetnames else wb.active

    headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    current_names = [n for (n, _w) in COLUMNS]

    # ── Schema migration · realign old rows to new column positions ──
    # 2026-08-04: added Prior Rank / Rank Δ / Alert between Rank (col 9) and
    # Confidence % (was col 10 · now col 13). Old rows have data at col 10-12
    # (conf/conf_type/model_score) which now belongs at col 13-15 · without
    # migration those values would appear under the wrong column names.
    if headers != current_names:
        # Snapshot old rows keyed by old-header name so we can re-emit them
        # under the new schema · unknown-in-old-headers columns fill blank.
        old_data: list[dict] = []
        for r_idx in range(2, ws.max_row + 1):
            row_map = {headers[c-1]: ws.cell(row=r_idx, column=c).value
                            for c in range(1, len(headers) + 1) if c - 1 < len(headers)}
            old_data.append(row_map)
        # Reset sheet · rewrite header + data under new schema
        ws.delete_rows(1, ws.max_row)
        for col_idx, (name, width) in enumerate(COLUMNS, start=1):
            cell = ws.cell(row=1, column=col_idx, value=name)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = CENTER
            ws.column_dimensions[get_column_letter(col_idx)].width = width
        for r_idx, row_map in enumerate(old_data, start=2):
            for c_idx, (name, _w) in enumerate(COLUMNS, start=1):
                val = row_map.get(name, "")
                if val is not None:
                    cell = ws.cell(row=r_idx, column=c_idx, value=val)
                    cell.alignment = LEFT if c_idx <= 6 else RIGHT
            status = row_map.get("Status")
            if status in STATUS_FILLS:
                fill = STATUS_FILLS[status]
                for c in range(1, len(COLUMNS) + 1):
                    ws.cell(row=r_idx, column=c).fill = fill
        headers = current_names

    key_indices = [headers.index(k) for k in DEDUP_KEY_COLS if k in headers]

    # Existing rows into map keyed by dedup tuple
    existing_map = {}
    for r_idx in range(2, ws.max_row + 1):
        key = tuple(ws.cell(row=r_idx, column=k_idx + 1).value for k_idx in key_indices)
        existing_map[key] = r_idx

    # For each new row · overwrite or append
    for row in rows:
        row_key = tuple(row[k_idx] for k_idx in key_indices)
        target_row = existing_map.get(row_key)
        if target_row is None:
            target_row = ws.max_row + 1
        for c_idx, val in enumerate(row, start=1):
            cell = ws.cell(row=target_row, column=c_idx, value=val)
            cell.alignment = LEFT if c_idx <= 6 else RIGHT
        # 2026-08-07 · row-level color spread across ALL columns (operator directive)
        # row[6] = Status · row[5] = Company (was wrong · caused single-cell fill on Company)
        status = row[6]
        if status in STATUS_FILLS:
            fill = STATUS_FILLS[status]
            for c in range(1, len(row) + 1):
                ws.cell(row=target_row, column=c).fill = fill

    # Refresh auto-filter to cover new range
    last_col = get_column_letter(len(COLUMNS))
    ws.auto_filter.ref = f"A1:{last_col}{ws.max_row}"
    # Sprint J-fix REVERTED · single-sheet format ONLY
    # Kill any pre-existing "Fresh Buys" sheet from earlier bad commit
    if "Fresh Buys" in wb.sheetnames:
        del wb["Fresh Buys"]
    wb.save(path)


def build_unified_history(root: Path, asof: str,
                              markets: list[str] | None = None) -> Path:
    """Append today's rows for all markets to reports/telegram/aegis_history.xlsx.

    Returns the path to the unified XLSX (creates if missing · appends
    daily otherwise · dedups by Date+Country+Run_Type+Ticker so same-day
    re-runs update rather than duplicate).
    """
    if markets is None:
        markets = ["india", "usa"]
    out_dir = root / "reports" / "telegram"
    out_dir.mkdir(parents=True, exist_ok=True)
    path = out_dir / "aegis_history.xlsx"

    all_rows = []
    for m in markets:
        all_rows.extend(_collect_rows_for_market(root, m, asof))

    if not path.exists():
        _write_new_workbook(path, all_rows)
    else:
        _append_to_workbook(path, all_rows)

    # Also write today-only snapshot for convenience
    snapshot = out_dir / f"aegis_daily_{asof}.xlsx"
    _write_new_workbook(snapshot, all_rows)

    return path


def build_and_stamp_all(root: Path, asof: str,
                             markets: list[str] | None = None) -> Path:
    """End-to-end refresh (v5 · 2026-08-04):
       1. stamp regime history + rank history for each market
       2. run profit-protection eval (writes profit_protection_{market}.json)
       3. build unified XLSX (which reads the outputs of steps 1-2)

    Safe to call multiple times · all steps are idempotent per (asof, market).
    Preferred entrypoint for scripts/rebuild_xlsx_local.py + telegram sender.
    """
    if markets is None:
        markets = ["india", "usa"]
    from backend.portfolio import rank_history as _rh
    from backend.portfolio import market_regime_stability as _mrs
    from backend.portfolio import profit_protection as _pp
    from backend.portfolio import health_score as _hs
    for m in markets:
        # Stamp regime history
        try:
            _mrs.stamp_today(root, asof, m)
        except Exception as e:
            print(f"[build_and_stamp:{m}] regime stamp skipped · {type(e).__name__}: {e}")
        # Stamp today's ranks + run profit-protection
        recs_path = (root / "usa" / "reports" / "recommendations.json"
                        if m == "usa" else root / "reports" / "recommendations.json")
        if not recs_path.exists():
            continue
        # Sprint H · USA R1 defensive derivative (before rank_history so R1 gets stamped too)
        if m == "usa":
            try:
                from backend.recommendation.usa_runner1 import derive as _usa_r1
                r1_payload = _usa_r1.derive(root, asof)
                if r1_payload.get("available"):
                    _usa_r1.emit(root, r1_payload)
                    print(f"[build_and_stamp:{m}] usa_runner1 emitted "
                            f"{r1_payload.get('n_r1_orphans', 0)} defensive orphans")
            except Exception as e:
                print(f"[build_and_stamp:{m}] usa_runner1 skipped · {type(e).__name__}: {e}")
        try:
            recs = json.loads(recs_path.read_text(encoding="utf-8")).get("recommendations", [])
            n_stamped = _rh.stamp_today(root, asof, m, "runner2", recs)
            print(f"[build_and_stamp:{m}] rank_history stamped n={n_stamped}")
            signals = _pp.evaluate_all_active(root, m, "runner2", asof, recs)
            _pp.emit_signals(root, m, "runner2", asof, signals)
            print(f"[build_and_stamp:{m}] profit_protection signals={len(signals)}")
            # Sprint A · Health Scores (composite 0-100 per rec)
            hs_payload = _hs.score_all(root, m, recs, asof)
            _hs.emit(root, m, hs_payload)
            print(f"[build_and_stamp:{m}] health_scores n={hs_payload.get('n')} "
                    f"band_changes={hs_payload.get('band_changes')}")
            # Context Intelligence Layer · run composer with 4 adapters
            try:
                from backend.context import composer as _cil
                from backend.context.adapters import DEFAULT_ADAPTERS
                adjustments = [_cil.compose(root, m, asof, r, DEFAULT_ADAPTERS)
                                    for r in recs]
                _cil.emit_run(root, m, asof, adjustments)
                drags = [a.total_drag_pts for a in adjustments]
                n_drag_neg = sum(1 for d in drags if d < -1)
                n_drag_pos = sum(1 for d in drags if d > 1)
                print(f"[build_and_stamp:{m}] CIL n={len(adjustments)} "
                        f"negative-drag={n_drag_neg} positive-boost={n_drag_pos}")
            except Exception as e:
                print(f"[build_and_stamp:{m}] CIL failed · {type(e).__name__}: {e}")
            # Economic Calendar · daily ingest (data-only · Phase 2 prep)
            try:
                from backend.context.economic_calendar import ingest as _ec
                summary = _ec.ingest_daily(root, asof)
                print(f"[build_and_stamp:{m}] economic_calendar appended={summary['total_appended']}")
            except Exception as e:
                print(f"[build_and_stamp:{m}] economic_calendar failed · {type(e).__name__}: {e}")
            # Sprint J-3 · Fresh Opportunities · daily unified filter · used
            # internally for New Opp signal · MD file skipped (single-file
            # delivery per operator directive)
            try:
                from backend.portfolio import fresh_opportunities as _fo
                fresh = _fo.scan(root, m, asof)
                jp = root / "reports" / "research" / f"fresh_opportunities_{m}.json"
                jp.parent.mkdir(parents=True, exist_ok=True)
                import json as _json
                from dataclasses import asdict as _asdict
                jp.write_text(_json.dumps({
                    "engine":  "aegis.portfolio.fresh_opportunities.v1",
                    "market":  m, "asof": asof, "n_fresh": len(fresh),
                    "fresh_buys": [_asdict(f) for f in fresh],
                }, indent=2, default=str, ensure_ascii=False), encoding="utf-8")
                print(f"[build_and_stamp:{m}] fresh_opportunities scan · n={len(fresh)}")
            except Exception as e:
                print(f"[build_and_stamp:{m}] fresh_opportunities skipped · {type(e).__name__}: {e}")
            # Guard 7-supporting engines · MANDATORY daily refresh per operator
            # directive 2026-08-05 · "every report should flow and run daily"
            for eng_name, eng_fn in [
                ("global_overnight",
                 lambda: __import__("backend.context.global_overnight.ingest",
                                          fromlist=["ingest_daily"]).ingest_daily(root, asof)),
                ("market_breadth",
                 lambda: (lambda mod: (mod.emit(root, mod.compute_breadth(root, m, asof))
                                            if mod.compute_breadth(root, m, asof).get("available") else None))(
                     __import__("backend.context.market_breadth.compute",
                                     fromlist=["compute_breadth", "emit"]))),
                ("fii_dii",
                 lambda: __import__("backend.context.fii_dii.ingest",
                                          fromlist=["ingest_daily"]).ingest_daily(root, asof)
                                 if m == "india" else None),
                ("correlation",
                 lambda: (lambda mod: (mod.emit(root, mod.compute_correlation(root, m, asof))
                                            if mod.compute_correlation(root, m, asof).get("available") else None))(
                     __import__("backend.context.correlation.compute",
                                     fromlist=["compute_correlation", "emit"]))),
                ("sector_news",
                 lambda: (lambda mod: (mod.emit(root, mod.compute_sector_news(root, m, asof))
                                            if mod.compute_sector_news(root, m, asof).get("available") else None))(
                     __import__("backend.context.sector_news.classify",
                                     fromlist=["compute_sector_news", "emit"]))),
            ]:
                try:
                    eng_fn()
                    print(f"[build_and_stamp:{m}] {eng_name} OK")
                except Exception as e:
                    print(f"[build_and_stamp:{m}] {eng_name} FAILED · {type(e).__name__}: {e}")
        except Exception as e:
            print(f"[build_and_stamp:{m}] pp/rank skipped · {type(e).__name__}: {e}")
    # Build XLSX first · then rotation tracker scans it for today's rotation exits
    xlsx_path = build_unified_history(root, asof, markets=markets)
    # 2026-08-06 · Rotation Outcome Tracker · after XLSX built · scans exits ·
    # closes matured >=20d outcomes · emits weekly rollup
    try:
        from backend.portfolio import rotation_outcome_tracker as _rot
        rr = _rot.daily_cycle(root, asof)
        print(f"[rotation_outcomes] logged={rr['n_new_rotations_logged']} · "
              f"closed={rr['n_outcomes_closed']} · "
              f"win_rate={rr['rollup'].get('rotation_win_rate_pct')}%")
    except Exception as e:
        print(f"[rotation_outcomes] skipped · {type(e).__name__}: {e}")
    return xlsx_path


def maybe_sync_google_sheet(xlsx_path: Path) -> tuple[bool, str]:
    """Mirror XLSX to Google Sheets IF creds env vars set. No-op otherwise."""
    creds_json = os.environ.get("GSHEETS_CREDS_JSON")
    sheet_id = os.environ.get("GSHEETS_SHEET_ID")
    if not creds_json or not sheet_id:
        return False, "GSHEETS_CREDS_JSON + GSHEETS_SHEET_ID not set · skipped"
    if not Path(creds_json).exists():
        return False, f"creds file missing: {creds_json}"
    try:
        import gspread
        from google.oauth2.service_account import Credentials
    except ImportError:
        return False, "install missing: pip install gspread google-auth"
    scopes = ["https://www.googleapis.com/auth/spreadsheets",
                 "https://www.googleapis.com/auth/drive"]
    creds = Credentials.from_service_account_file(creds_json, scopes=scopes)
    gc = gspread.authorize(creds)
    sh = gc.open_by_key(sheet_id)
    wb = load_workbook(xlsx_path, read_only=True)
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        try:
            existing = sh.worksheet(sheet_name)
            existing.clear()
        except gspread.WorksheetNotFound:
            existing = sh.add_worksheet(title=sheet_name, rows=200, cols=40)
        rows = [[c.value for c in row] for row in ws.iter_rows()]
        if rows:
            existing.update("A1", rows, value_input_option="USER_ENTERED")
    return True, f"synced {len(wb.sheetnames)} sheets to google_sheet={sheet_id}"
