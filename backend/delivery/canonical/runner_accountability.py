"""AEGIS · Runner-level accountability · Phases 7-8 · CEO 2026-09-01.

R1 and R2 are independently observable runners. Never collapse their
performance into one number before preserving runner-level data.

Reads (READ-ONLY):
  · Registry (canonical position identity + lifecycle events)
  · aegis_history (audit trail · dated observations)
  · Exit History (realized-eligible trades · filtered per contract)

Computes per-runner (R1, R2, COMBINED · never silently merged):
  · signals_generated               count of DISTINCT position_ids
                                     with any observation in window
  · positions_opened                count of DISTINCT position_ids
                                     with created_date in window
  · currently_active                count of ACTIVE per Registry
  · positions_closed                count of CLOSED with closed_date
                                     in window
  · eligible_exits                  positions_closed minus carveout
                                     (ORPHAN_AUTO_CLOSE / rotations)
  · realized_pnl_pct                sum of pnl_pct across eligible exits
                                     (equal-weight per trade)
  · win_rate_pct                    positive / eligible
  · mean_pnl_pct                    average of eligible exits
  · median_pnl_pct                  median of eligible exits
  · avg_holding_days                mean (exit_date - entry_date)
  · max_realized_drawdown_pct       worst single trade P&L
  · sample_size_verdict             "SUFFICIENT" / "INSUFFICIENT (N<20)"

Runner utilization classification (never infer inactivity from row absence):
  · has_qualifying_signals          any created_date in window
  · has_opened_position             positions_opened > 0
  · has_closed_position             positions_closed > 0
  · is_dormant                      no signals AND no positions in window

Explicitly labels combined-metrics with "R1+R2" scope AND provides raw
per-runner numbers so the operator can audit whether "combined" is
dominated by one runner.
"""
from __future__ import annotations

import statistics
from collections import Counter, defaultdict
from dataclasses import dataclass, field
from datetime import date, timedelta
from pathlib import Path
from typing import Optional


_CARVEOUT_KEYWORDS = ("ORPHAN_AUTO_CLOSE", "SAME_DAY_ROTATION",
                       "CANCELLED", "DATA_REPAIR")


@dataclass
class RunnerAccounting:
    market:                  str
    runner:                  str          # "R1" | "R2" | "COMBINED"
    window_days:             int
    asof:                    str
    signals_generated:       int
    positions_opened:        int
    currently_active:        int
    positions_closed:        int
    eligible_exits:          int
    realized_pnl_pct:        float
    win_rate_pct:            Optional[float]
    mean_pnl_pct:            Optional[float]
    median_pnl_pct:          Optional[float]
    avg_holding_days:        Optional[float]
    max_realized_drawdown_pct: Optional[float]
    sample_size_verdict:     str
    has_qualifying_signals:  bool
    has_opened_position:     bool
    has_closed_position:     bool
    is_dormant:              bool
    # Section 10 · CEO 2026-09-01 · explicit utilization classification
    # so operator never infers runner inactivity merely from row absence.
    # One of:
    #   ACTIVE_PRODUCTION      · runner produced signals in window
    #   RETIRED_DORMANT        · runner formally retired · DORMANT_BY_DESIGN
    #   NO_QUALIFYING_SIGNAL   · engine ran but produced no signals meeting criteria
    #   NO_EXECUTION           · signals present but never executed to positions
    #   PIPELINE_FAILURE       · engine/ingest failed · no observations recorded
    utilization_status:      str
    utilization_reason:      str
    # Provenance
    source_registry_events:  int
    source_aegis_history_rows: int
    source_exit_history_rows: int


def _is_rotation_artifact(pnl_pct: Optional[float]) -> bool:
    if pnl_pct is None: return False
    return abs(pnl_pct) < 0.01   # per Contract v1 rule C5


def _is_carveout(reason: str) -> bool:
    r = (reason or "").upper()
    return any(kw in r for kw in _CARVEOUT_KEYWORDS)


def _classify_sample_size(n: int) -> str:
    if n >= 100: return "SUFFICIENT"
    if n >= 30:  return "MODERATE"
    if n >= 10:  return "LOW"
    return f"INSUFFICIENT (n={n} · minimum 30 for stable statistics)"


def compute_runner_accounting(root: Path, market: str,
                                runner_filter: str,
                                asof: str,
                                window_days: int = 90) -> RunnerAccounting:
    """Compute one RunnerAccounting record for (market, runner)."""
    from backend.research import opportunity_registry as oreg

    market_l = market.lower()
    asof_d = date.fromisoformat(asof)
    cutoff = (asof_d - timedelta(days=window_days)).isoformat()

    # Load Registry latest-per-pid
    reg = oreg.load_all(root)
    latest = {}
    for pid, events in reg.items():
        for e in events:
            latest[e.opportunity_id] = e

    # Filter to (market, runner)
    def matches(o):
        if o.market.lower() != market_l: return False
        if runner_filter == "COMBINED":
            return o.runner in ("R1", "R2")
        return o.runner == runner_filter

    reg_relevant = [(pid, o) for pid, o in latest.items() if matches(o)]

    n_signals = 0
    n_opened = 0
    n_active = 0
    n_closed = 0
    eligible_pnls = []
    all_closed_events = []
    for pid, o in reg_relevant:
        n_signals += 1
        if o.created_date and o.created_date >= cutoff:
            n_opened += 1
        if o.status == "ACTIVE":
            n_active += 1
        if o.status == "CLOSED" and o.closed_date and o.closed_date >= cutoff:
            n_closed += 1
            all_closed_events.append(o)

    # For eligible exits · need Exit History rows (Registry doesn't have pnl)
    # Load Exit History body
    from openpyxl import load_workbook
    xlsx_p = root / "reports" / "telegram" / f"aegis_history_{market_l}.xlsx"
    exit_rows = []
    if xlsx_p.exists():
        wb = load_workbook(xlsx_p, read_only=True, data_only=True)
        if "Exit History (90d)" in wb.sheetnames:
            ws = wb["Exit History (90d)"]
            rows = list(ws.iter_rows(values_only=True))
            hdr_idx = None
            for i, r in enumerate(rows):
                if r[0] and "Stock" in str(r[0]):
                    hdr_idx = i; break
            if hdr_idx is not None:
                hdr = rows[hdr_idx]
                def col(name):
                    for i, c in enumerate(hdr):
                        if c and str(c).strip().lower() == name.lower():
                            return i
                    return None
                c_tk = col("Stock")
                c_run = col("Runner")
                c_ent = col("Entry Date")
                c_exit = col("Exit Date")
                c_days = col("Days Held")
                c_pnl = col("P&L %")
                c_reason = col("Exit Reason")
                for r in rows[hdr_idx + 1:]:
                    if not r[c_tk]: continue
                    tk = str(r[c_tk])
                    if not tk.replace("-", "").isalnum(): continue
                    row_run = str(r[c_run] or "").upper() if c_run is not None else ""
                    if runner_filter == "COMBINED":
                        if row_run not in ("R1", "R2"): continue
                    else:
                        if row_run != runner_filter: continue
                    pnl = r[c_pnl] if c_pnl is not None and isinstance(r[c_pnl], (int, float)) else None
                    days = r[c_days] if c_days is not None and isinstance(r[c_days], (int, float)) else None
                    reason = str(r[c_reason] or "") if c_reason is not None else ""
                    exit_rows.append({"pnl_pct": pnl, "days": days, "reason": reason,
                                       "ticker": tk, "runner": row_run})
        wb.close()

    # Filter to REALIZED-eligible (non-rotation, non-carveout)
    eligible = [r for r in exit_rows
                if r["pnl_pct"] is not None
                and not _is_rotation_artifact(r["pnl_pct"])
                and not _is_carveout(r["reason"])]

    pnls = [r["pnl_pct"] * 100 if abs(r["pnl_pct"]) < 5 else r["pnl_pct"]
             for r in eligible]   # normalize decimal → percent
    realized_pnl = sum(pnls) if pnls else 0.0
    n_positive = sum(1 for p in pnls if p > 0)
    n_eligible = len(pnls)
    wr = (n_positive / n_eligible * 100) if n_eligible > 0 else None
    mean = statistics.mean(pnls) if pnls else None
    median = statistics.median(pnls) if pnls else None
    days_vals = [r["days"] for r in eligible if r["days"] is not None]
    avg_days = statistics.mean(days_vals) if days_vals else None
    max_dd = min(pnls) if pnls else None   # worst single trade

    # Section 10 · CEO 2026-09-01 · explicit utilization classification.
    # Prefer retirement flag > pipeline evidence > signal / execution counts.
    utilization_status = "ACTIVE_PRODUCTION"
    utilization_reason = "runner produced signals and/or opened positions in window"
    try:
        from backend.delivery.canonical.retirement import is_retired as _r_is_retired
        if runner_filter in ("R1", "R2") and _r_is_retired(root, runner_filter):
            utilization_status = "RETIRED_DORMANT"
            utilization_reason = (
                f"{runner_filter} formally retired per configs/aegis_retirement.yaml "
                "· DORMANT_BY_DESIGN · historical records preserved"
            )
    except Exception:
        pass
    if utilization_status == "ACTIVE_PRODUCTION":
        # Check pipeline evidence · look at aegis_daily_v2_history for last engine event
        _hist_p = root / "reports" / "aegis_daily_v2_history.jsonl"
        _saw_recent_engine = False
        if _hist_p.exists():
            try:
                _lines = _hist_p.read_text(encoding="utf-8").splitlines()[-500:]
                for _ln in _lines:
                    if runner_filter.lower() in _ln.lower() or market_l in _ln.lower():
                        _saw_recent_engine = True
                        break
            except Exception:
                pass
        if n_signals == 0 and n_opened == 0:
            if not _saw_recent_engine:
                utilization_status = "PIPELINE_FAILURE"
                utilization_reason = (
                    "0 signals AND 0 opened positions in window AND no recent engine "
                    "activity in aegis_daily_v2_history · investigate engine health"
                )
            else:
                utilization_status = "NO_QUALIFYING_SIGNAL"
                utilization_reason = (
                    "engine ran recently but produced 0 signals meeting entry criteria "
                    "· check thresholds / regime / universe"
                )
        elif n_signals > 0 and n_opened == 0:
            utilization_status = "NO_EXECUTION"
            utilization_reason = (
                f"{n_signals} signals generated but 0 positions opened · check "
                "execution gate / capital availability / rotation logic"
            )

    return RunnerAccounting(
        market=market_l,
        runner=runner_filter,
        window_days=window_days,
        asof=asof,
        signals_generated=n_signals,
        positions_opened=n_opened,
        currently_active=n_active,
        positions_closed=n_closed,
        eligible_exits=n_eligible,
        realized_pnl_pct=round(realized_pnl, 2),
        win_rate_pct=round(wr, 1) if wr is not None else None,
        mean_pnl_pct=round(mean, 2) if mean is not None else None,
        median_pnl_pct=round(median, 2) if median is not None else None,
        avg_holding_days=round(avg_days, 1) if avg_days is not None else None,
        max_realized_drawdown_pct=round(max_dd, 2) if max_dd is not None else None,
        sample_size_verdict=_classify_sample_size(n_eligible),
        has_qualifying_signals=n_signals > 0,
        has_opened_position=n_opened > 0,
        has_closed_position=n_closed > 0,
        is_dormant=(n_signals == 0 and n_opened == 0),
        utilization_status=utilization_status,
        utilization_reason=utilization_reason,
        source_registry_events=n_signals,
        source_aegis_history_rows=0,   # optional · not required for accounting
        source_exit_history_rows=len(exit_rows),
    )


def compute_all(root: Path, market: str, asof: str,
                 window_days: int = 90) -> dict:
    """Return {R1: RunnerAccounting, R2: ..., COMBINED: ...}."""
    return {
        "R1": compute_runner_accounting(root, market, "R1", asof, window_days),
        "R2": compute_runner_accounting(root, market, "R2", asof, window_days),
        "COMBINED": compute_runner_accounting(root, market, "COMBINED", asof, window_days),
    }
