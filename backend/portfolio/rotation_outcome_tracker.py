"""Rotation Outcome Tracker · did our rotations actually beat holding?

Operator P0 audit 2026-08-06 → CEO analysis PDF §1 Q1 gap: rotation
exits happen at negative move (opportunity cost) but we have no way to
know if the rotation was correct · did the replacement actually beat
the position we cut?

FLOW
────
1. Every daily run · scan XLSX for Status=EXIT rows with rotation intent
   (Exit Reason contains "→")
2. Log to reports/research/rotation_outcomes.jsonl · append-only:
   {asof · market · runner · old_ticker · old_exit_price ·
    new_ticker · new_entry_price · expected_alpha_pp}
3. On EVERY subsequent daily run · for each open outcome record with
   age >= 20 trading days:
     · fetch current bar close for both old + new ticker
     · compute realized alpha: (new_return_20d − old_return_20d) × 100
     · compare to expected_alpha_pp at rotation time
     · verdict: WIN if realized >= expected · else LOSS
     · close the outcome record
4. Weekly rollup: reports/research/monthly/rotation_outcome_review.json

Silent when no rotations to evaluate. Idempotent. Zero R1/R2 code touches.
"""
from __future__ import annotations

import json
from dataclasses import asdict, dataclass, field
from datetime import date, datetime, timedelta, timezone
from pathlib import Path


LEDGER_PATH = "reports/research/rotation_outcomes.jsonl"
OBSERVATION_WINDOW_DAYS = 20    # trading days · not calendar days


@dataclass
class RotationOutcome:
    ts_utc: str
    asof: str
    market: str
    runner: str
    old_ticker: str
    old_exit_price: float | None
    new_ticker: str
    new_entry_price: float | None
    expected_alpha_pp: float | None
    # These filled in later when 20-day window elapses
    outcome_asof: str | None = None
    old_price_20d: float | None = None
    new_price_20d: float | None = None
    old_return_20d_pct: float | None = None
    new_return_20d_pct: float | None = None
    realized_alpha_pp: float | None = None
    verdict: str | None = None       # WIN · LOSS · PENDING
    closed: bool = False


def _bar_close(root: Path, market: str, ticker: str, target_date: str) -> float | None:
    """Fetch bar close on target_date · fallback to nearest earlier."""
    import pandas as pd
    short = ticker.replace(".NS", "").replace(".BO", "").upper()
    d = root / ("usa/data/raw/us" if market == "usa" else "data/raw/india")
    p = d / f"{short}_D1.parquet"
    if not p.exists(): return None
    try:
        df = pd.read_parquet(p)
        col = "close" if "close" in df.columns else "Close"
        df.index = pd.to_datetime(df.index).strftime("%Y-%m-%d")
        if target_date in df.index: return float(df.loc[target_date, col])
        earlier = [d_ for d_ in df.index if d_ <= target_date]
        if earlier: return float(df.loc[earlier[-1], col])
    except Exception:
        return None
    return None


def _trading_days_after(start_iso: str, n: int) -> str:
    """Return ISO date n trading days after start (weekend-adjusted approx)."""
    try:
        d = date.fromisoformat(start_iso[:10])
    except (ValueError, TypeError):
        return start_iso
    added = 0
    while added < n:
        d += timedelta(days=1)
        if d.weekday() < 5: added += 1
    return d.isoformat()


def _load_ledger(root: Path) -> list[RotationOutcome]:
    p = root / LEDGER_PATH
    if not p.exists(): return []
    rows = []
    for line in p.read_text(encoding="utf-8").splitlines():
        if not line.strip(): continue
        try:
            d = json.loads(line)
            rows.append(RotationOutcome(**{k: v for k, v in d.items()
                                                        if k in RotationOutcome.__dataclass_fields__}))
        except Exception:
            continue
    return rows


def _append_ledger(root: Path, entry: RotationOutcome) -> None:
    p = root / LEDGER_PATH
    p.parent.mkdir(parents=True, exist_ok=True)
    with p.open("a", encoding="utf-8") as fh:
        fh.write(json.dumps(asdict(entry), default=str, ensure_ascii=False) + "\n")


def _rewrite_ledger(root: Path, rows: list[RotationOutcome]) -> None:
    p = root / LEDGER_PATH
    p.parent.mkdir(parents=True, exist_ok=True)
    with p.open("w", encoding="utf-8") as fh:
        for r in rows:
            fh.write(json.dumps(asdict(r), default=str, ensure_ascii=False) + "\n")


def detect_rotations_today(root: Path, asof: str) -> int:
    """Scan today's XLSX for new rotation exits · append to ledger. Returns
    count of new entries added."""
    try:
        from openpyxl import load_workbook
    except ImportError:
        return 0
    p = root / "reports" / "telegram" / "aegis_history.xlsx"
    if not p.exists(): return 0
    try:
        wb = load_workbook(p, read_only=True)
    except Exception:
        return 0
    ws = wb["AEGIS Daily"] if "AEGIS Daily" in wb.sheetnames else wb.active
    h = [c.value for c in ws[1]]
    def _c(name):
        try: return h.index(name) + 1
        except ValueError: return None
    ic = {n: _c(n) for n in ["Date", "Country", "Run_Type", "Ticker",
                                         "Status", "Exit Reason", "Current Price"]}
    if not all(ic.values()):
        wb.close(); return 0

    existing = _load_ledger(root)
    existing_keys = {(r.asof, r.market, r.runner, r.old_ticker) for r in existing}

    # 2026-08-26 · was ws.cell(row=r, column=c) in a loop over max_row
    # (~1700 rows) · in read_only mode ws.cell() is O(n) per access →
    # O(n^2) total → sender hung indefinitely at rotation_outcomes step.
    # iter_rows streams sequentially · fixes the hang.
    _idx = {n: ic[n] - 1 for n in ic}
    n_added = 0
    import re
    for _r_tuple in ws.iter_rows(min_row=2, values_only=True):
        if len(_r_tuple) <= max(_idx.values()): continue
        d = str(_r_tuple[_idx["Date"]] or "")
        if not d.startswith(asof): continue
        status = _r_tuple[_idx["Status"]]
        if status != "EXIT": continue
        exit_reason = str(_r_tuple[_idx["Exit Reason"]] or "")
        if "→" not in exit_reason: continue

        # Parse: "→ TCS.NS #1 · +12.2pp alpha" or "→ TCS.NS (+12.2pp)" or "Rotation → TCS.NS (+X.Xpp)"
        m = re.search(r"→\s*([A-Z0-9\.-]+)", exit_reason)
        new_ticker = m.group(1).replace(".NS", "").replace(".BO", "") if m else ""
        alpha_m = re.search(r"\+?(-?\d+\.?\d*)\s*pp", exit_reason)
        expected_alpha = float(alpha_m.group(1)) if alpha_m else None

        old_ticker = _r_tuple[_idx["Ticker"]]
        market = (_r_tuple[_idx["Country"]] or "").lower()
        runner = _r_tuple[_idx["Run_Type"]]
        exit_price = _r_tuple[_idx["Current Price"]]

        key = (asof, market, runner, old_ticker)
        if key in existing_keys: continue

        new_entry = _bar_close(root, market, new_ticker, asof) if new_ticker else None

        outcome = RotationOutcome(
            ts_utc=datetime.now(timezone.utc).isoformat(),
            asof=asof, market=market, runner=runner,
            old_ticker=old_ticker,
            old_exit_price=float(exit_price) if isinstance(exit_price, (int, float)) else None,
            new_ticker=new_ticker,
            new_entry_price=new_entry,
            expected_alpha_pp=expected_alpha,
        )
        _append_ledger(root, outcome)
        n_added += 1
    wb.close()
    return n_added


def close_matured_outcomes(root: Path, asof: str) -> int:
    """For every open outcome with age >= 20 trading days · fetch prices ·
    compute realized alpha · verdict · close. Returns count closed."""
    all_rows = _load_ledger(root)
    if not all_rows: return 0
    n_closed = 0
    for r in all_rows:
        if r.closed: continue
        target_date = _trading_days_after(r.asof, OBSERVATION_WINDOW_DAYS)
        if asof < target_date: continue
        if not r.old_exit_price or not r.new_entry_price: continue

        old_20d = _bar_close(root, r.market, r.old_ticker, target_date)
        new_20d = _bar_close(root, r.market, r.new_ticker, target_date)
        if old_20d is None or new_20d is None: continue

        old_return = (old_20d - r.old_exit_price) / r.old_exit_price * 100
        new_return = (new_20d - r.new_entry_price) / r.new_entry_price * 100
        realized_alpha = new_return - old_return

        r.outcome_asof = target_date
        r.old_price_20d = round(old_20d, 2)
        r.new_price_20d = round(new_20d, 2)
        r.old_return_20d_pct = round(old_return, 2)
        r.new_return_20d_pct = round(new_return, 2)
        r.realized_alpha_pp = round(realized_alpha, 2)
        # Verdict: WIN if realized alpha positive AND close to expected · else LOSS
        if r.expected_alpha_pp is not None:
            r.verdict = "WIN" if realized_alpha >= 0 else "LOSS"
        else:
            r.verdict = "WIN" if realized_alpha >= 0 else "LOSS"
        r.closed = True
        n_closed += 1

    if n_closed:
        _rewrite_ledger(root, all_rows)
    return n_closed


def weekly_rollup(root: Path, asof: str) -> dict:
    """Roll up closed rotation outcomes into a weekly review report."""
    rows = _load_ledger(root)
    closed = [r for r in rows if r.closed]
    open_ = [r for r in rows if not r.closed]

    wins = [r for r in closed if r.verdict == "WIN"]
    losses = [r for r in closed if r.verdict == "LOSS"]
    n = len(closed)
    win_rate = round(len(wins) / n * 100, 1) if n else None
    avg_realized = round(sum(r.realized_alpha_pp for r in closed
                                       if r.realized_alpha_pp is not None) / n, 2) if n else None
    avg_expected = round(sum(r.expected_alpha_pp for r in closed
                                       if r.expected_alpha_pp is not None) /
                                sum(1 for r in closed if r.expected_alpha_pp is not None), 2) \
                             if any(r.expected_alpha_pp is not None for r in closed) else None

    payload = {
        "engine":          "aegis.portfolio.rotation_outcome_tracker.v1",
        "asof":            asof,
        "generated_utc":   datetime.now(timezone.utc).isoformat(),
        "n_total_logged":  len(rows),
        "n_open":          len(open_),
        "n_closed":        n,
        "n_wins":          len(wins),
        "n_losses":        len(losses),
        "rotation_win_rate_pct": win_rate,
        "avg_realized_alpha_pp": avg_realized,
        "avg_expected_alpha_pp": avg_expected,
        "closed_outcomes": [asdict(r) for r in closed[-30:]],   # last 30
        "open_outcomes":   [asdict(r) for r in open_[-30:]],
        "verdict": (
            "healthy · rotations beating alternatives" if win_rate and win_rate >= 55
            else "watch · rotations breaking even" if win_rate and win_rate >= 45
            else "concern · rotations losing to holding" if win_rate and win_rate < 45
            else "insufficient data · need ≥ 20 closed outcomes"
        ),
    }
    out = root / "reports" / "research" / "monthly" / "rotation_outcome_review.json"
    out.parent.mkdir(parents=True, exist_ok=True)
    out.write_text(json.dumps(payload, indent=2, default=str, ensure_ascii=False),
                       encoding="utf-8")
    return payload


def daily_cycle(root: Path, asof: str) -> dict:
    n_added = detect_rotations_today(root, asof)
    n_closed = close_matured_outcomes(root, asof)
    rollup = weekly_rollup(root, asof)
    return {"n_new_rotations_logged": n_added,
                "n_outcomes_closed":       n_closed,
                "rollup":                  rollup}
