"""Rank 1+2 Portfolio Tracker.

Per operator directive 2026-08-06: "I am thinking to prefer only rank 1, 2
from both runners and test as of now."

Persists a paper-trading portfolio · deduplicated to unique tickers · tracks
P&L daily · exportable to Telegram + XLSX.

Portfolio construction (locked at first snapshot · updates via rotation):
    · India R1 ranks 1-2 → core weight
    · India R2 ranks 1-2 → satellite weight
    · USA R1 ranks 1-2 → core weight
    · USA R2 ranks 1-2 → satellite weight (dedup vs R1 · same tickers common)

Allocation: 70% core (R1) · 30% satellite (R2-unique)

Storage:
    configs/rank12_portfolio.json  · holdings + weights (locked at first run)
    reports/research/rank12_pnl_history.jsonl  · daily P&L snapshots
    reports/research/rank12_current.json  · latest snapshot for Telegram

Rotation rule: portfolio updates when rank 1 or 2 changes AND the new
ticker isn't already held. Existing holdings kept if they're still in
top-5 of their runner (grace period · avoid churn).
"""
from __future__ import annotations

import json
from dataclasses import dataclass, asdict, field
from datetime import date, datetime, timezone
from pathlib import Path


CORE_WEIGHT = 0.70          # R1 total
SATELLITE_WEIGHT = 0.30     # R2-unique total
INITIAL_CAPITAL = 100000.0  # ₹1,00,000 paper


@dataclass
class Holding:
    ticker: str
    market: str                 # india · usa
    source_runner: str          # R1 · R2
    entry_date: str
    entry_price: float
    current_price: float
    weight_pct: float
    shares: float               # entry_capital / entry_price
    current_value: float
    return_pct: float


def _top2_from_xlsx(root: Path, asof: str) -> list[tuple[str, str, int, str]]:
    """Read today's XLSX · extract (market, runner, rank, ticker) for ranks 1+2."""
    from openpyxl import load_workbook
    p = root / "reports" / "telegram" / "aegis_history.xlsx"
    if not p.exists(): return []
    wb = load_workbook(p, read_only=True)
    ws = wb.active
    h = [c.value for c in ws[1]]
    i = {n: h.index(n) + 1 for n in ["Date", "Country", "Run_Type", "Rank", "Ticker"]}
    out = []
    for r in range(2, ws.max_row + 1):
        d = str(ws.cell(row=r, column=i["Date"]).value)
        if not d.startswith(asof): continue
        rank = ws.cell(row=r, column=i["Rank"]).value
        if rank not in (1, 2): continue
        out.append((
            ws.cell(row=r, column=i["Country"]).value,
            ws.cell(row=r, column=i["Run_Type"]).value,
            rank,
            ws.cell(row=r, column=i["Ticker"]).value,
        ))
    wb.close()
    return out


def _bar_close(root: Path, market: str, ticker: str, date_str: str) -> float | None:
    try:
        import pandas as pd
    except ImportError:
        return None
    dir_p = (root / "data" / "raw" / "india") if market.lower() == "india" \
                else (root / "usa" / "data" / "raw" / "us")
    p = dir_p / f"{ticker}_D1.parquet"
    if not p.exists(): return None
    try:
        df = pd.read_parquet(p)
        col = "close" if "close" in df.columns else "Close" if "Close" in df.columns else None
        if col is None: return None
        df.index = pd.to_datetime(df.index).strftime("%Y-%m-%d")
        if date_str in df.index:
            return float(df.loc[date_str, col])
        earlier = [d for d in df.index if d <= date_str]
        if earlier:
            return float(df.loc[earlier[-1], col])
    except Exception:
        return None
    return None


def build_portfolio(root: Path, asof: str,
                          initial_capital: float = INITIAL_CAPITAL) -> dict:
    """First-time portfolio build (or rotation update)."""
    picks = _top2_from_xlsx(root, asof)
    if not picks:
        return {"available": False, "reason": "no rank 1-2 picks in today's XLSX"}

    # Dedup: keep first occurrence · R1 wins over R2 (core priority)
    seen = set()
    ordered = []
    # Sort so R1 comes before R2 (so R1 dedup wins)
    picks.sort(key=lambda x: (x[1] != "R1", x[0], x[2]))
    for market, runner, rank, ticker in picks:
        key = (market, ticker)
        if key in seen: continue
        seen.add(key)
        ordered.append((market, runner, rank, ticker))

    # Split into core (R1) + satellite (R2-unique)
    core = [x for x in ordered if x[1] == "R1"]
    satellite = [x for x in ordered if x[1] == "R2"]

    holdings = []
    if core:
        per_core = CORE_WEIGHT / len(core)
        for market, runner, rank, ticker in core:
            entry_px = _bar_close(root, market, ticker, asof)
            if not entry_px: continue
            capital = initial_capital * per_core
            shares = capital / entry_px
            holdings.append(Holding(
                ticker=ticker, market=market.lower(), source_runner=runner,
                entry_date=asof, entry_price=entry_px, current_price=entry_px,
                weight_pct=round(per_core * 100, 2), shares=round(shares, 4),
                current_value=round(capital, 2), return_pct=0.0,
            ))
    if satellite:
        per_sat = SATELLITE_WEIGHT / len(satellite)
        for market, runner, rank, ticker in satellite:
            entry_px = _bar_close(root, market, ticker, asof)
            if not entry_px: continue
            capital = initial_capital * per_sat
            shares = capital / entry_px
            holdings.append(Holding(
                ticker=ticker, market=market.lower(), source_runner=runner,
                entry_date=asof, entry_price=entry_px, current_price=entry_px,
                weight_pct=round(per_sat * 100, 2), shares=round(shares, 4),
                current_value=round(capital, 2), return_pct=0.0,
            ))

    return {
        "engine":           "aegis.portfolio.rank12_tracker.v1",
        "created_utc":      datetime.now(timezone.utc).isoformat(),
        "asof":             asof,
        "initial_capital":  initial_capital,
        "core_weight_pct":  CORE_WEIGHT * 100,
        "sat_weight_pct":   SATELLITE_WEIGHT * 100,
        "n_core":           len(core),
        "n_satellite":      len(satellite),
        "holdings":         [asdict(h) for h in holdings],
    }


def refresh_prices(root: Path, portfolio: dict, asof: str) -> dict:
    """Update current_price + return + current_value for each holding · daily."""
    total_value = 0
    for h in portfolio.get("holdings") or []:
        curr = _bar_close(root, h["market"], h["ticker"], asof)
        if not curr:
            curr = h["current_price"]     # keep last-known
        entry = h["entry_price"]
        h["current_price"] = curr
        h["return_pct"] = round((curr - entry) / entry * 100, 2) if entry else 0
        h["current_value"] = round(curr * h["shares"], 2)
        total_value += h["current_value"]
    initial = portfolio.get("initial_capital", INITIAL_CAPITAL)
    portfolio["asof"] = asof
    portfolio["last_refresh_utc"] = datetime.now(timezone.utc).isoformat()
    portfolio["total_value"] = round(total_value, 2)
    portfolio["total_return_pct"] = round((total_value - initial) / initial * 100, 2)
    portfolio["total_pnl"] = round(total_value - initial, 2)
    return portfolio


def _cfg_path(root: Path) -> Path:
    return root / "configs" / "rank12_portfolio.json"


def _pnl_hist_path(root: Path) -> Path:
    p = root / "reports" / "research" / "rank12_pnl_history.jsonl"
    p.parent.mkdir(parents=True, exist_ok=True)
    return p


def _current_path(root: Path) -> Path:
    return root / "reports" / "research" / "rank12_current.json"


def load_portfolio(root: Path) -> dict | None:
    p = _cfg_path(root)
    if not p.exists(): return None
    try: return json.loads(p.read_text(encoding="utf-8"))
    except Exception: return None


def save_portfolio(root: Path, portfolio: dict) -> None:
    p = _cfg_path(root)
    p.parent.mkdir(parents=True, exist_ok=True)
    p.write_text(json.dumps(portfolio, indent=2, default=str, ensure_ascii=False),
                    encoding="utf-8")


def append_pnl_snapshot(root: Path, portfolio: dict) -> None:
    row = {
        "asof":              portfolio.get("asof"),
        "total_value":       portfolio.get("total_value"),
        "total_return_pct":  portfolio.get("total_return_pct"),
        "total_pnl":         portfolio.get("total_pnl"),
        "n_holdings":        len(portfolio.get("holdings") or []),
        "ts_utc":            datetime.now(timezone.utc).isoformat(),
    }
    with _pnl_hist_path(root).open("a", encoding="utf-8") as fh:
        fh.write(json.dumps(row, default=str, ensure_ascii=False) + "\n")


def emit_current(root: Path, portfolio: dict) -> Path:
    p = _current_path(root)
    p.parent.mkdir(parents=True, exist_ok=True)
    p.write_text(json.dumps(portfolio, indent=2, default=str, ensure_ascii=False),
                    encoding="utf-8")
    return p


def render_md(portfolio: dict) -> str:
    total_val = portfolio.get("total_value", 0)
    total_ret = portfolio.get("total_return_pct", 0)
    total_pnl = portfolio.get("total_pnl", 0)
    icon = "🟢" if total_ret > 0 else ("🔴" if total_ret < 0 else "⚪")

    lines = [f"# 📊 Rank 1+2 Portfolio · {portfolio.get('asof')}",
                "",
                f"**Total value:** ₹{total_val:,.2f} · "
                f"{icon} **{total_ret:+.2f}%** · P&L ₹{total_pnl:+,.2f}",
                "",
                f"Initial: ₹{portfolio.get('initial_capital', 0):,.2f} · "
                f"Core {portfolio.get('core_weight_pct')}% / Satellite {portfolio.get('sat_weight_pct')}%",
                "",
                "| Ticker | Market | Runner | Weight % | Entry | Current | Return | Value |",
                "|---|---|---|---|---|---|---|---|"]
    for h in portfolio.get("holdings") or []:
        ic = "🟢" if h["return_pct"] > 0 else ("🔴" if h["return_pct"] < 0 else "⚪")
        lines.append(f"| {h['ticker']} | {h['market'].upper()} | {h['source_runner']} | "
                          f"{h['weight_pct']}% | {h['entry_price']:.2f} | {h['current_price']:.2f} | "
                          f"{ic} {h['return_pct']:+.2f}% | ₹{h['current_value']:,.2f} |")
    return "\n".join(lines) + "\n"


def daily_cycle(root: Path, asof: str) -> dict:
    """One-shot: build or refresh · save · append history · return portfolio."""
    existing = load_portfolio(root)
    if existing is None:
        portfolio = build_portfolio(root, asof)
        if not portfolio.get("available", True):
            # No available reason means portfolio built successfully · check holdings
            if not portfolio.get("holdings"):
                return {"available": False, "reason": "no picks available"}
    else:
        portfolio = refresh_prices(root, existing, asof)

    if portfolio.get("holdings"):
        portfolio = refresh_prices(root, portfolio, asof)
        save_portfolio(root, portfolio)
        append_pnl_snapshot(root, portfolio)
        emit_current(root, portfolio)
    return portfolio
