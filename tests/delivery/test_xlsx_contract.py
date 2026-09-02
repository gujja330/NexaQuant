# tests/delivery/test_xlsx_contract.py
"""AEGIS · Delivery Contract + Validator tests.

Contract definitions must be internally consistent · validator must
correctly detect violations · golden integration tests must pass on a
synthetic XLSX + Registry fixture.
"""
from __future__ import annotations

from pathlib import Path
from openpyxl import Workbook
import pytest


# ═════════════════════════════════════════════════════════════════
# Contract structure tests
# ═════════════════════════════════════════════════════════════════
class TestContract:

    def test_all_invariants_have_unique_codes(self):
        from backend.delivery.xlsx_contract import INVARIANTS
        codes = [i.code for i in INVARIANTS]
        assert len(codes) == len(set(codes))

    def test_all_invariants_have_check_fn(self):
        from backend.delivery.xlsx_contract import INVARIANTS
        from backend.delivery.xlsx_validator import XlsxValidator
        for i in INVARIANTS:
            assert hasattr(XlsxValidator, i.check_fn_name), \
                f"Missing check function {i.check_fn_name} for {i.code}"

    def test_all_invariants_severity_valid(self):
        from backend.delivery.xlsx_contract import INVARIANTS
        for i in INVARIANTS:
            assert i.severity in ("BLOCK", "WARN", "INFO"), \
                f"{i.code} has invalid severity {i.severity}"

    def test_20_plus_invariants(self):
        from backend.delivery.xlsx_contract import INVARIANTS
        # CEO's list required 20+ invariants
        assert len(INVARIANTS) >= 20

    def test_at_least_10_blocking(self):
        from backend.delivery.xlsx_contract import BLOCK_INVARIANTS
        # Should be substantial blocking set
        assert len(BLOCK_INVARIANTS) >= 10

    def test_forbidden_states_defined(self):
        from backend.delivery.xlsx_contract import FORBIDDEN_STATES
        assert "PROTECT" in FORBIDDEN_STATES
        assert "REVIEW" in FORBIDDEN_STATES
        assert "TRAIL" in FORBIDDEN_STATES

    def test_canonical_states_correct(self):
        from backend.delivery.xlsx_contract import CANONICAL_STATES
        for s in ("NEW", "ACTIVE", "ACTIVE+", "EXIT"):
            assert s in CANONICAL_STATES


# ═════════════════════════════════════════════════════════════════
# Golden-file integration test · synthetic XLSX + Registry
# ═════════════════════════════════════════════════════════════════
@pytest.fixture
def synthetic_xlsx(tmp_path):
    """Build a minimal valid Portfolio XLSX matching the contract."""
    p = tmp_path / "reports" / "telegram" / "aegis_history_india.xlsx"
    p.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    # CEO 2026-09-02 · LOCKED 3-sheet contract:
    #   01_Portfolio · 02_Today_Momentum · 03_Exit_History
    # title=row1 · analysis=row2 · blank=row3 · header=row4 · data=row5
    # Portfolio sheet · new name + row offsets
    ws = wb.active
    ws.title = "01_Portfolio"
    ws.cell(1, 1, "AEGIS INDIA · PORTFOLIO · current active holdings as of 2026-08-26")
    ws.cell(2, 1, "🟢 R2 ACTIVE: 2 · production runner is R2")
    # Row 4 · header (was row 5)
    headers = ["Position ID", "Ticker", "Runner", "Entry Date",
               "Entry Price", "Current Price", "Unrealized P&L %",
               "Holding Days", "Dynamic Stop", "Engine Verdict",
               "Would-Have-Exited-On", "As-Of", "Provenance"]
    for c, h in enumerate(headers, start=1):
        ws.cell(4, c, h)
    # Row 5 · valid ACTIVE row (TCS)
    ws.cell(5, 1, "IND-R2-TCS-20260801-abc123")
    ws.cell(5, 2, "TCS")
    ws.cell(5, 3, "R2")
    ws.cell(5, 4, "2026-08-01")
    ws.cell(5, 5, 3500.0)
    ws.cell(5, 6, 3550.0)
    ws.cell(5, 7, 1.43)
    ws.cell(5, 8, 25)
    ws.cell(5, 9, 3300.0)
    ws.cell(5, 10, "HOLD (audit-only)")
    ws.cell(5, 11, "—")
    ws.cell(5, 12, "2026-08-26")
    ws.cell(5, 13, "canonical:Registry+prices")
    # Exit History sheet · new name + row offsets
    eh = wb.create_sheet("03_Exit_History")
    eh.cell(1, 1, "AEGIS INDIA · EXIT HISTORY · realized · as of 2026-08-26")
    eh.cell(2, 1, "📕 Closed positions (last 90d): 1 · latest exit first")
    # Row 4 · header (was row 5)
    eh_hdr = ["Position ID", "Stock", "Sector", "Runner", "Market",
              "Entry Date", "Exit Date", "Holding Days",
              "Entry Price", "Exit Price", "Realized P&L %",
              "Exit Reason", "Provenance"]
    for c, h in enumerate(eh_hdr, start=1):
        eh.cell(4, c, h)
    # Row 5 · HDFC exit
    eh.cell(5, 1, "IND-R2-HDFC-20260715-def456")
    eh.cell(5, 2, "HDFC")
    eh.cell(5, 3, "Financial Services")
    eh.cell(5, 4, "R2")
    eh.cell(5, 5, "INDIA")
    eh.cell(5, 6, "2026-07-15")
    eh.cell(5, 7, "2026-08-10")
    eh.cell(5, 8, 26)
    eh.cell(5, 9, 1500.0)
    eh.cell(5, 10, 1620.0)
    eh.cell(5, 11, 8.0)
    eh.cell(5, 12, "Profit target hit")
    eh.cell(5, 13, "canonical:Registry+prices")
    wb.save(p)
    return tmp_path


class TestValidator:

    def test_validate_returns_report(self, synthetic_xlsx):
        from backend.delivery import xlsx_validator as _val
        rep = _val.validate(
            synthetic_xlsx, "india",
            synthetic_xlsx / "reports" / "telegram" / "aegis_history_india.xlsx")
        assert rep.market == "india"
        assert rep.n_pass + rep.n_warn + rep.n_fail + rep.n_skip == \
               len(rep.invariants)

    def test_should_block_send_false_when_all_pass(self, synthetic_xlsx):
        from backend.delivery import xlsx_validator as _val
        # Mock a report with no BLOCK failures
        rep = _val.ValidationReport(
            market="india", asof="2026-08-26",
            generated_utc="2026-08-26T00:00:00Z",
        )
        assert _val.should_block_send(rep) is False

    def test_should_block_send_true_on_block_fail(self):
        from backend.delivery import xlsx_validator as _val
        rep = _val.ValidationReport(
            market="india", asof="2026-08-26",
            generated_utc="2026-08-26T00:00:00Z",
        )
        rep.add(_val.InvariantResult(
            "I1", "test", "BLOCK", "FAIL", "test fail"))
        assert _val.should_block_send(rep) is True

    def test_render_blocked_alert_readable(self):
        from backend.delivery import xlsx_validator as _val
        rep = _val.ValidationReport(
            market="india", asof="2026-08-26",
            generated_utc="2026-08-26T00:00:00Z",
        )
        rep.add(_val.InvariantResult(
            "I1", "EXIT in ACTIVE", "BLOCK", "FAIL", "2 leaks"))
        alert = _val.render_blocked_alert(rep)
        assert "🚫 AEGIS DELIVERY BLOCKED" in alert
        assert "I1" in alert

    def test_summary_line_format(self):
        from backend.delivery import xlsx_validator as _val
        rep = _val.ValidationReport(
            market="india", asof="2026-08-26",
            generated_utc="2026-08-26T00:00:00Z",
        )
        s = _val.summary_line(rep)
        assert "xlsx_validator" in s
        assert "verdict=" in s

    def test_emit_writes_json(self, synthetic_xlsx):
        from backend.delivery import xlsx_validator as _val
        rep = _val.validate(
            synthetic_xlsx, "india",
            synthetic_xlsx / "reports" / "telegram" / "aegis_history_india.xlsx")
        p = _val.emit(synthetic_xlsx, rep)
        assert p.exists()

    def test_sheet_title_check(self, synthetic_xlsx, monkeypatch):
        from backend.delivery.xlsx_validator import XlsxValidator
        v = XlsxValidator(
            synthetic_xlsx, "india",
            synthetic_xlsx / "reports" / "telegram" / "aegis_history_india.xlsx")
        r = v.check_sheet_title()
        assert r.status == "PASS"

    def test_required_headers_check(self, synthetic_xlsx):
        from backend.delivery.xlsx_validator import XlsxValidator
        v = XlsxValidator(
            synthetic_xlsx, "india",
            synthetic_xlsx / "reports" / "telegram" / "aegis_history_india.xlsx")
        r = v.check_required_headers()
        # Our synthetic has all required headers · should PASS
        assert r.status == "PASS", f"Missing headers: {r.violations}"

    def test_analysis_rows_populated(self, synthetic_xlsx):
        from backend.delivery.xlsx_validator import XlsxValidator
        v = XlsxValidator(
            synthetic_xlsx, "india",
            synthetic_xlsx / "reports" / "telegram" / "aegis_history_india.xlsx")
        r = v.check_analysis_rows_populated()
        assert r.status == "PASS"
