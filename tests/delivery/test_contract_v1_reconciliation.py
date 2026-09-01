"""AEGIS · Delivery Data Contract v1 · reconciliation regression tests.

CEO 2026-08-28 · end-to-end stabilization directive · one coherent
change. These tests encode the invariants that must NEVER be silently
regressed by future work.

Scope of this file:
  · Rule C4 · MISSING must never be silently substituted with LOW /
    PENDING / 0 for holding-no-signal rows.
  · Rule C5 · MONTHLY_SUMMARY must live on its own sheet · never as
    trailer rows inside the Exit History body.
  · Portfolio banner "Active (current)" must reconcile to the visible
    body count (banner ≠ independent worksheet scan).
  · Definitions sheet must exist in every shipped workbook.
  · P&L column format · holding-no-signal rows use the same "%" number
    format as signal rows so the operator sees consistent units.

## Freshness contract (2026-08-31 · post-mortem)

These tests inspect the git-committed shipped XLSX at
`reports/telegram/aegis_history_india.xlsx`.

On the CI runner, the workflow order is:
  · Delivery classifier tests (this file · gates the Telegram send)
  · Telegram Command Center (builds the XLSX + sends)

So at test-time the XLSX is WHATEVER git currently tracks · NOT a
freshly-built artifact. The daily-CI `git add` pattern only includes
`reports/AEGIS_*.xlsx` (uppercase, unified) · NOT the per-market
`reports/telegram/aegis_history_*.xlsx` files. Consequence: the
committed per-market XLSX can be arbitrarily stale, and testing its
contents against post-fix invariants produces false-positive failures
that block Telegram delivery.

Fix (this file): SKIP with a clear reason when the committed artifact
is not from today's asof. The invariants remain enforceable when the
tests are run against a fresh build (local dev via
`python scripts/telegram_command_center_send.py --build-only`, or CI
after the send step). The invariants are NOT lost · they are just not
gating CI on a stale sample.

Do NOT convert these skips into passes. If a fresh build is available,
the tests DO run and MUST pass · that is what protects the fix from
silent regression.
"""
from __future__ import annotations

import pytest
from datetime import date
from pathlib import Path

_XLSX_INDIA = Path("reports/telegram/aegis_history_india.xlsx")
_XLSX_USA = Path("reports/telegram/aegis_history_usa.xlsx")

# CEO 2026-09-01 · FINAL 3-sheet spec SUPERSEDES Contract v1's 8-sheet
# invariants (Portfolio · Exit History (90d) · Monthly Summary ·
# AEGIS History · Definitions as separate sheets · specific column
# layouts). The new workbook has exactly 3 sheets: 01_Portfolio ·
# 02_Decisions_Exit_History · 03_Summary_Definitions. Contract v1
# tests remain in the repo as audit history but are skipped when the
# workbook uses the 3-sheet spec (detected by presence of 01_Portfolio).
def _is_3sheet_workbook() -> bool:
    try:
        from openpyxl import load_workbook
        if not _XLSX_INDIA.exists(): return False
        wb = load_workbook(_XLSX_INDIA, read_only=True)
        result = "01_Portfolio" in wb.sheetnames
        wb.close()
        return result
    except Exception:
        return False

pytestmark = pytest.mark.skipif(
    _is_3sheet_workbook(),
    reason="Contract v1 (8-sheet) superseded by CEO 2026-09-01 3-sheet spec. "
             "See docs/AEGIS/R1_RETIREMENT_2026-09-01.md and "
             "scripts/build_aegis_3sheet_workbook.py. Contract v1 checks are "
             "preserved for audit history but do not apply to the 3-sheet workbook."
)


def _artifact_asof(wb) -> str:
    """Extract the asof-date stamp from Portfolio r1 (title row).
    Format is 'AEGIS INDIA PORTFOLIO · as of YYYY-MM-DD'. Returns
    "YYYY-MM-DD" or "" if unparseable."""
    try:
        ws = wb["Portfolio"]
        title = str(ws.cell(1, 1).value or "")
        # Find YYYY-MM-DD substring
        import re
        m = re.search(r"(\d{4}-\d{2}-\d{2})", title)
        if m:
            return m.group(1)
    except Exception:
        pass
    return ""


def _load_wb(p: Path):
    """Load workbook · skip if artifact absent OR stale.

    "Stale" means the Portfolio title's asof stamp is not today. In CI
    this happens because the delivery-tests gate runs BEFORE the
    build step · the git-committed XLSX may be days old. Skipping is
    the correct behavior · the fix invariants can only be validated
    against a fresh build (see module docstring)."""
    if not p.exists():
        pytest.skip(f"artifact not present · {p}")
    from openpyxl import load_workbook
    wb = load_workbook(p, read_only=True, data_only=True)
    asof = _artifact_asof(wb)
    today = date.today().isoformat()
    if asof and asof != today:
        wb.close()
        pytest.skip(
            f"artifact asof={asof} is not today={today} · this is expected "
            f"on CI where delivery-tests gate runs before the XLSX build. "
            f"Rebuild locally with `python scripts/telegram_command_center_send.py "
            f"--build-only` to validate contract-v1 invariants against a "
            f"fresh artifact."
        )
    return wb


def _portfolio_rows(wb):
    ws = wb["Portfolio"]
    return [tuple(r) for r in ws.iter_rows(values_only=True)
             if any(c not in (None, "") for c in r)]


def _exit_rows(wb):
    if "Exit History (90d)" not in wb.sheetnames:
        return []
    ws = wb["Exit History (90d)"]
    return [tuple(r) for r in ws.iter_rows(values_only=True)
             if any(c not in (None, "") for c in r)]


# ── C4 · MISSING ≠ LOW · MISSING ≠ PENDING ────────────────────────


def test_c4_no_low_urgency_fabrication_in_holding_rows():
    """Path-A holding rows (Registry-ACTIVE · no signal today) MUST NOT
    contain 'LOW' as Urgency. Engine did not evaluate · row must show
    'MISSING' semantics ('—' em-dash)."""
    wb = _load_wb(_XLSX_INDIA)
    rows = _portfolio_rows(wb)
    # Find header row
    hdr_idx = None
    for i, r in enumerate(rows):
        if r[0] and "Ticker" in str(r[0]):
            hdr_idx = i
            break
    assert hdr_idx is not None, "Portfolio header row not found"
    hdr = rows[hdr_idx]
    # Find Urgency column
    urgency_col = None
    for c, name in enumerate(hdr):
        if name and "Urgency" in str(name):
            urgency_col = c
            break
    assert urgency_col is not None, "Urgency column not found in Portfolio header"
    body = rows[hdr_idx + 1:]
    # Any row whose ACTION mentions "holding · no signal" is a Path-A row
    violations = []
    for r in body:
        action = str(r[1]) if len(r) > 1 and r[1] else ""
        if "holding" in action.lower() and "no signal" in action.lower():
            urg = str(r[urgency_col]) if urgency_col < len(r) and r[urgency_col] else ""
            if "LOW" in urg.upper():
                violations.append((r[0], urg))
    assert not violations, (
        f"Rule C4 violation · {len(violations)} holding-no-signal rows "
        f"have fabricated LOW Urgency · engine did not evaluate these · "
        f"first 3: {violations[:3]}"
    )
    wb.close()


def test_c4_no_pending_quality_fabrication_in_holding_rows():
    """Path-A holding rows MUST NOT contain 'PENDING' as Inv Quality.
    Engine did not evaluate · row must show 'MISSING' semantics."""
    wb = _load_wb(_XLSX_INDIA)
    rows = _portfolio_rows(wb)
    hdr_idx = None
    for i, r in enumerate(rows):
        if r[0] and "Ticker" in str(r[0]):
            hdr_idx = i
            break
    assert hdr_idx is not None
    hdr = rows[hdr_idx]
    quality_col = None
    for c, name in enumerate(hdr):
        if name and "Inv Quality" in str(name):
            quality_col = c
            break
    assert quality_col is not None, "Inv Quality column not found"
    body = rows[hdr_idx + 1:]
    violations = []
    for r in body:
        action = str(r[1]) if len(r) > 1 and r[1] else ""
        if "holding" in action.lower() and "no signal" in action.lower():
            q = str(r[quality_col]) if quality_col < len(r) and r[quality_col] else ""
            if "PENDING" in q.upper():
                violations.append((r[0], q))
    assert not violations, (
        f"Rule C4 violation · {len(violations)} holding-no-signal rows "
        f"have fabricated PENDING Inv Quality · first 3: {violations[:3]}"
    )
    wb.close()


def test_c4_pnl_column_uses_percent_format_for_holding_rows():
    """Holding-no-signal rows must apply Excel '%' number format so the
    operator sees '-2.69%' not '-0.0269' · matches signal-row format."""
    if not _XLSX_INDIA.exists():
        pytest.skip("artifact not present")
    from openpyxl import load_workbook
    # Full-workbook load (not read-only) to inspect cell.number_format
    wb = load_workbook(_XLSX_INDIA, data_only=True)
    # Same freshness gate as _load_wb · this test needs the not-read-only
    # workbook to inspect cell.number_format · duplicate the asof check.
    _asof = _artifact_asof(wb)
    _today = date.today().isoformat()
    if _asof and _asof != _today:
        wb.close()
        pytest.skip(
            f"artifact asof={_asof} is not today={_today} · rebuild via "
            f"`python scripts/telegram_command_center_send.py --build-only` "
            f"to validate this invariant"
        )
    ws = wb["Portfolio"]
    hdr_row = None
    pnl_col = None
    for r in range(1, ws.max_row + 1):
        if ws.cell(r, 1).value and "Ticker" in str(ws.cell(r, 1).value):
            hdr_row = r
            for c in range(1, ws.max_column + 1):
                v = ws.cell(hdr_row, c).value
                if v and "P&L" in str(v):
                    pnl_col = c
                    break
            break
    assert hdr_row is not None
    assert pnl_col is not None, "P&L column not found in Portfolio header"
    violations = []
    for r in range(hdr_row + 1, ws.max_row + 1):
        action = ws.cell(r, 2).value or ""
        if "holding" in str(action).lower() and "no signal" in str(action).lower():
            cell = ws.cell(r, pnl_col)
            if cell.value is not None and isinstance(cell.value, (int, float)):
                fmt = cell.number_format or ""
                if "%" not in fmt:
                    violations.append((ws.cell(r, 1).value, cell.value, fmt))
    assert not violations, (
        f"Rule C4 violation · {len(violations)} holding-no-signal rows "
        f"have P&L stored without % number format · Excel will render "
        f"the raw decimal (looks like '-0.03' not '-2.69%') · "
        f"first 3: {violations[:3]}"
    )
    wb.close()


# ── C5 · MONTHLY_SUMMARY separate sheet ───────────────────────────


def test_c5_monthly_summary_is_a_separate_sheet():
    """Monthly Summary MUST be a first-class sheet · never trailer rows
    inside Exit History body. Validators must never need trailer-skip
    logic to survive."""
    wb = _load_wb(_XLSX_INDIA)
    sheets = wb.sheetnames
    assert "Monthly Summary" in sheets, (
        f"Monthly Summary sheet missing · rule C5 violation · "
        f"workbook has {sheets}"
    )
    wb.close()


def test_c5_exit_history_body_ends_at_lineage_row():
    """Last non-empty row of Exit History body MUST be a lineage
    (trade) row · not a decoration/summary/trailer row."""
    wb = _load_wb(_XLSX_INDIA)
    rows = _exit_rows(wb)
    if not rows:
        pytest.skip("Exit History sheet empty")
    # Last row's first cell should be a ticker (all-caps, no ── / MONTH /
    # SUMMARY / spaces).
    last = rows[-1]
    first = str(last[0]) if last[0] else ""
    forbidden = ["──", "MONTHLY", "MONTH", "SUMMARY", "TOTAL"]
    for f in forbidden:
        assert f not in first.upper(), (
            f"Rule C5 violation · Exit History body last row is a "
            f"decoration row · '{first[:50]}' · Monthly summary must be "
            f"a separate sheet"
        )


def test_c5_no_trailer_strings_in_exit_history_body():
    """No row in Exit History body should contain 'MONTHLY P&L SUMMARY'
    · that content belongs on the Monthly Summary sheet."""
    wb = _load_wb(_XLSX_INDIA)
    rows = _exit_rows(wb)
    if not rows:
        pytest.skip("Exit History sheet empty")
    for i, r in enumerate(rows, 1):
        for c_v in r:
            if c_v is None: continue
            s = str(c_v).upper()
            assert "MONTHLY P&L SUMMARY" not in s, (
                f"Rule C5 violation · Exit History body r{i} contains "
                f"trailer text '{s[:80]}' · must move to Monthly Summary sheet"
            )
    wb.close()


# ── C6 · Banner reads canonical count not worksheet scan ─────────


def test_banner_active_current_reconciles_to_visible_body():
    """Portfolio banner ACTIVE count must equal the number of visible
    ACTIVE body rows. Banner must not independently invent a count.

    Supports THREE banner formats (evolution of the delivery contract):
      · 2026-08-27 "Active (current): N positions"
      · 2026-08-31 "Current Portfolio: N ACTIVE · M NEW · K SUGGESTED"
      · 2026-08-31 (post-B2) "Lifecycle: N ACTIVE · M NEW · K SUGGESTED"
    """
    wb = _load_wb(_XLSX_INDIA)
    rows = _portfolio_rows(wb)
    banner = str(rows[1][0]) if len(rows) > 1 and rows[1][0] else ""
    import re as _re
    m = _re.search(
        r"(?:Lifecycle|Current Portfolio):\s*(\d+)\s+ACTIVE", banner)
    if not m:
        m = _re.search(r"Active \(current\):\s*(\d+)", banner)
    if not m:
        pytest.skip(f"banner not in recognized shape · {banner[:80]}")
    stated = int(m.group(1))
    # Count body rows (exclude SUGGESTED/SHADOW · they have own header text)
    hdr_idx = None
    for i, r in enumerate(rows):
        if r[0] and "Ticker" in str(r[0]):
            hdr_idx = i
            break
    body = rows[hdr_idx + 1:]
    # B2 · banner "ACTIVE" count = rows whose Lifecycle column contains
    # "ACTIVE" (not "NEW", not blank/SUGGESTED). Decision column can
    # still say "NEW" for a fresh recommendation on an ACTIVE-lifecycle
    # holding · those count as ACTIVE per lifecycle.
    n_active_lifecycle = 0
    for r in body:
        life = str(r[3]) if len(r) > 3 and r[3] else ""
        dec = str(r[2]) if len(r) > 2 and r[2] else ""
        if "SUGGESTED" in dec.upper():
            continue
        life_up = life.upper()
        # Count ACTIVE lifecycle strictly · exclude NEW (which is its own axis)
        if "ACTIVE" in life_up and "NEW" not in life_up:
            n_active_lifecycle += 1
    assert stated == n_active_lifecycle, (
        f"Banner reconciliation failure · stated={stated} "
        f"lifecycle=ACTIVE={n_active_lifecycle} · C6 violation · "
        f"banner must consume same population as body"
    )
    wb.close()


# ── Definitions sheet MUST exist ──────────────────────────────────


def test_definitions_sheet_exists_in_workbook():
    """Every shipped workbook must include the Definitions sheet · so
    the operator can read scope + formula + missing-data semantics
    without leaving the file."""
    wb = _load_wb(_XLSX_INDIA)
    assert "Definitions" in wb.sheetnames, (
        f"Definitions sheet missing from India workbook · sheets={wb.sheetnames}"
    )
    wb.close()
