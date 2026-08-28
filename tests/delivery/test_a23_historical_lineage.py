"""A23 · Historical-lineage validation for Exit History rows.

CEO 2026-08-28 · exact acceptance fixture from directive:
· historically ACTIVE → now CLOSED → PASS
· historically ACTIVE → rotated → PASS
· historically ACTIVE → legitimate exit → PASS
· never existed in canonical lifecycle → FAIL
· fabricated ticker → FAIL
· duplicate/orphan lifecycle → PASS (orphan-audit JSONL is a legit sink)
"""
import json
import pytest
from pathlib import Path


def _write_registry(root: Path, entries: list):
    """entries = [{ticker, runner, market, status, closed_reason?, ts_utc?}]"""
    p = root / "reports" / "research" / "opportunity_registry.jsonl"
    p.parent.mkdir(parents=True, exist_ok=True)
    from backend.research.opportunity_registry import make_opportunity_id
    with p.open("w", encoding="utf-8") as f:
        for e in entries:
            pid = make_opportunity_id(e["market"], e["runner"], e["ticker"],
                                       e.get("created", "2026-08-10"))
            row = {
                "opportunity_id": pid, "market": e["market"],
                "runner": e["runner"], "ticker": e["ticker"],
                "created_date": e.get("created", "2026-08-10"),
                "initial_signal": "BUY", "initial_rank": 1,
                "initial_score": 0.85, "status": e["status"],
                "closed_date": e.get("closed", ""),
                "closed_reason": e.get("closed_reason", ""),
                "last_seen_date": e.get("closed") or e.get("created", "2026-08-10"),
                "ts_utc": e.get("ts_utc", "2026-08-10T00:00:00+00:00"),
            }
            f.write(json.dumps(row) + "\n")


def _write_orphan_audit(root: Path, market: str, tickers: list):
    p = root / "reports" / "delivery" / f"orphan_audit_{market}.jsonl"
    p.parent.mkdir(parents=True, exist_ok=True)
    with p.open("w", encoding="utf-8") as f:
        for tk in tickers:
            f.write(json.dumps({
                "ticker": tk, "runner": "R2", "closed_reason": "ORPHAN_AUTO_CLOSE",
                "created_date": "2026-08-10", "closed_date": "2026-08-10",
            }) + "\n")


def _write_snapshot(root: Path, ticker: str, market: str, entry_date: str,
                     entry_price: float):
    from backend.delivery.prediction_snapshot import record_snapshot
    record_snapshot(root, market=market.upper(), ticker=ticker,
                    prediction_date=entry_date, entry_date=entry_date,
                    entry_price=entry_price, source_close_date=entry_date,
                    source_dataset_version="test",
                    canonical_signal="R2_BUY")


def _write_exit_history_xlsx(root: Path, market: str, tickers: list):
    """Write a minimal aegis_history_{market}.xlsx with Exit History body
    containing the given tickers."""
    from openpyxl import Workbook
    p = root / "reports" / "telegram" / f"aegis_history_{market}.xlsx"
    p.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "Portfolio"     # A23 doesn't read this
    ws2 = wb.create_sheet("Exit History (90d)")
    ws2.cell(1, 1, "AEGIS X · EXIT HISTORY · last 90 days as of 2026-08-27")
    ws2.cell(2, 1, "banner")
    ws2.cell(3, 1, "positive/negative")
    ws2.cell(5, 1, "Stock")   # header row
    for i, tk in enumerate(tickers, start=6):
        ws2.cell(i, 1, tk)
        ws2.cell(i, 10, 1.5)     # numeric P&L
    wb.save(p)


def _run_a23(root: Path, market: str) -> dict:
    """Extract A23 result from wave_regression report."""
    from backend.research.wave_regression import compute
    from datetime import date
    rep = compute(root, market, asof=date.today().isoformat())
    checks = getattr(rep, "checks", None) or getattr(rep, "results", [])
    for r in checks:
        code = getattr(r, "code", None) or (r.get("code") if isinstance(r, dict) else None)
        if code == "A23":
            status = getattr(r, "status", None) or (r.get("status") if isinstance(r, dict) else None)
            detail = getattr(r, "detail", None) or (r.get("detail") if isinstance(r, dict) else "")
            return {"status": status, "detail": detail}
    return {"status": "MISSING", "detail": "A23 not in report"}


# ── Fixture 1 · historically ACTIVE → now CLOSED → PASS ──


def test_historical_active_now_closed_passes(tmp_path):
    _write_registry(tmp_path, [
        {"ticker": "PLTR", "runner": "R2", "market": "usa",
         "status": "CLOSED", "closed": "2026-08-15",
         "closed_reason": "TARGET_1_HIT"},
    ])
    _write_exit_history_xlsx(tmp_path, "usa", ["PLTR"])
    r = _run_a23(tmp_path, "usa")
    assert r["status"] == "PASS", r


# ── Fixture 2 · historically ACTIVE → rotated → PASS ──


def test_historical_active_rotated_passes(tmp_path):
    _write_registry(tmp_path, [
        {"ticker": "MSFT", "runner": "R2", "market": "usa",
         "status": "CLOSED", "closed": "2026-08-14",
         "closed_reason": "Rotated to PLTR · better setup"},
    ])
    _write_exit_history_xlsx(tmp_path, "usa", ["MSFT"])
    r = _run_a23(tmp_path, "usa")
    assert r["status"] == "PASS", r


# ── Fixture 3 · historically ACTIVE → legitimate exit → PASS ──


def test_historical_active_legitimate_exit_passes(tmp_path):
    _write_registry(tmp_path, [
        {"ticker": "GOOG", "runner": "R1", "market": "usa",
         "status": "CLOSED", "closed": "2026-08-20",
         "closed_reason": "STOP_LOSS_HIT"},
    ])
    _write_exit_history_xlsx(tmp_path, "usa", ["GOOG"])
    r = _run_a23(tmp_path, "usa")
    assert r["status"] == "PASS", r


# ── Fixture 4 · never existed in canonical lifecycle → FAIL ──


def test_fabricated_ticker_never_in_registry_fails(tmp_path):
    _write_registry(tmp_path, [
        {"ticker": "PLTR", "runner": "R2", "market": "usa",
         "status": "CLOSED", "closed": "2026-08-15"},
    ])
    # Exit History has a ticker FAKE_XYZ that never appeared in Registry
    _write_exit_history_xlsx(tmp_path, "usa", ["PLTR", "FAKE_XYZ"])
    r = _run_a23(tmp_path, "usa")
    assert r["status"] == "FAIL", r
    assert "FAKE_XYZ" in r["detail"] or "fabricated" in r["detail"].lower()


# ── Fixture 5 · fabricated ticker (only in EH) → FAIL ──


def test_ticker_only_in_exit_history_fails(tmp_path):
    _write_registry(tmp_path, [])   # empty Registry
    _write_exit_history_xlsx(tmp_path, "usa", ["PHANTOM"])
    r = _run_a23(tmp_path, "usa")
    assert r["status"] == "FAIL", r


# ── Fixture 6 · orphan-audit sink is a legit lifecycle path → PASS ──


def test_orphan_filtered_to_audit_jsonl_still_valid_lineage(tmp_path):
    """Registry-CLOSED ticker filtered out of Exit History and routed
    to orphan_audit_usa.jsonl. A23 must recognize this as a legit
    lineage path · NOT flag it as silently-lost."""
    _write_registry(tmp_path, [
        {"ticker": "ORPHAN1", "runner": "R2", "market": "usa",
         "status": "CLOSED", "closed": "2026-08-10",
         "closed_reason": "ORPHAN_AUTO_CLOSE"},
        {"ticker": "REAL1", "runner": "R2", "market": "usa",
         "status": "CLOSED", "closed": "2026-08-15",
         "closed_reason": "TARGET_1_HIT"},
    ])
    # Exit History has REAL1 but not ORPHAN1 (orphan filtered out)
    _write_exit_history_xlsx(tmp_path, "usa", ["REAL1"])
    # orphan_audit JSONL captures the filtered orphan
    _write_orphan_audit(tmp_path, "usa", ["ORPHAN1"])
    r = _run_a23(tmp_path, "usa")
    assert r["status"] == "PASS", \
        f"orphan in audit JSONL should satisfy lineage · got {r}"


# ── Fixture 7 · Registry-CLOSED silently lost (not in EH, not in audit) → FAIL ──


def test_registry_closed_silently_lost_fails(tmp_path):
    """A ticker CLOSED in Registry but NOT in Exit History AND NOT in
    orphan_audit is silently lost · A23 must catch this."""
    _write_registry(tmp_path, [
        {"ticker": "LOST", "runner": "R2", "market": "usa",
         "status": "CLOSED", "closed": "2026-08-15",
         "closed_reason": "TARGET_1_HIT"},
    ])
    _write_exit_history_xlsx(tmp_path, "usa", [])   # empty EH
    # No orphan audit JSONL
    r = _run_a23(tmp_path, "usa")
    assert r["status"] == "FAIL", r
    assert "silently lost" in r["detail"].lower() or "LOST" in r["detail"]


# ── Fixture 8 · snapshot ledger provides lineage even without Registry ──


def test_snapshot_ledger_provides_lineage(tmp_path):
    """A ticker in the snapshot ledger (canonical entry record) has
    provenance even if not currently in Registry."""
    _write_registry(tmp_path, [])
    _write_snapshot(tmp_path, "SNAPPED", "usa", "2026-08-10", 100.0)
    _write_exit_history_xlsx(tmp_path, "usa", ["SNAPPED"])
    r = _run_a23(tmp_path, "usa")
    assert r["status"] == "PASS", \
        f"snapshot ledger provides lineage · got {r}"


# ── Fixture 9 · idempotent · repeated runs produce same verdict ──


def test_a23_deterministic_across_reruns(tmp_path):
    _write_registry(tmp_path, [
        {"ticker": "PLTR", "runner": "R2", "market": "usa",
         "status": "CLOSED", "closed": "2026-08-15"},
    ])
    _write_exit_history_xlsx(tmp_path, "usa", ["PLTR"])
    verdicts = set()
    for _ in range(5):
        verdicts.add(_run_a23(tmp_path, "usa")["status"])
    assert len(verdicts) == 1
