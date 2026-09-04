"""Tests for 01_Investments primary operator sheet (CEO 2026-09-03)."""
from __future__ import annotations

from pathlib import Path
import sys
sys.path.insert(0, str(Path(__file__).resolve().parents[2]))


def test_investments_sheet_meta():
    from backend.delivery.sheets.investments_sheet import (
        sheet_meta, INVESTMENTS_COLUMNS, INVESTMENTS_BANNER,
    )
    m = sheet_meta()
    assert m["sheet_name"] == "01_Investments"
    # Mandatory columns per CEO 2026-09-03
    for req in ("Score", "Action", "Ticker", "Runner", "Entry Date",
                "Entry Price", "Current Price", "Unrealized P&L %",
                "Holding Days", "Dynamic Stop", "Target"):
        assert req in INVESTMENTS_COLUMNS, f"missing mandatory column: {req}"
    # Banner mentions mandatory Dynamic Stop
    assert "Dynamic Stop mandatory" in INVESTMENTS_BANNER


def test_stop_trichotomy_r2_valid_atr():
    from backend.delivery.sheets.investments_sheet import _stop_cell
    cell, prov = _stop_cell("R2", entry_price=100.0, atr=2.5)
    # Should produce numeric stop, no SUGGESTED tag
    assert "SUGGESTED" not in cell
    assert "N/A" not in cell
    assert "DATA_ERROR" not in cell
    assert float(cell) < 100.0   # Stop below entry (k=2 → 100 − 5 = 95)


def test_stop_trichotomy_r1_gets_suggested_tag():
    from backend.delivery.sheets.investments_sheet import _stop_cell
    cell, prov = _stop_cell("R1", entry_price=100.0, atr=2.5)
    assert "SUGGESTED" in cell
    assert prov == "atr14_suggested_r1_no_auto_exit"


def test_stop_trichotomy_r1_no_atr_returns_advisory_only():
    from backend.delivery.sheets.investments_sheet import _stop_cell
    cell, prov = _stop_cell("R1", entry_price=100.0, atr=None)
    assert "N/A" in cell
    assert "advisory" in cell.lower()


def test_stop_trichotomy_r2_data_error_when_missing_atr():
    from backend.delivery.sheets.investments_sheet import _stop_cell
    cell, prov = _stop_cell("R2", entry_price=100.0, atr=None)
    assert cell.startswith("DATA_ERROR")


def test_stop_trichotomy_r2_data_error_when_missing_entry():
    from backend.delivery.sheets.investments_sheet import _stop_cell
    cell, prov = _stop_cell("R2", entry_price=None, atr=2.5)
    assert cell.startswith("DATA_ERROR")


def test_score_bounded_0_100():
    from backend.delivery.sheets.investments_sheet import (
        _score_for_r2_new, _score_for_r1_new, _score_for_active,
    )
    for score in [
        _score_for_r2_new(1.0, 1.0), _score_for_r2_new(-1.0, 0.0),
        _score_for_r1_new("STRONG BUY", 100), _score_for_r1_new(None, None),
        _score_for_active(20, 50, 90), _score_for_active(-30, 0, 20),
    ]:
        assert 0 <= score <= 100


def test_investments_sheet_never_bare_unavailable_for_stop():
    """Every Dynamic Stop cell must be one of: numeric · SUGGESTED · N/A · DATA_ERROR."""
    import io, sys
    from openpyxl import load_workbook
    for m in ("india", "usa"):
        p = Path(__file__).resolve().parents[2] / "reports" / "telegram" / f"aegis_{m}_2026-09-03.xlsx"
        if not p.exists(): continue
        wb = load_workbook(p)
        if "01_Investments" not in wb.sheetnames: continue
        ws = wb["01_Investments"]
        # Header row 4 · Dynamic Stop is column 11
        for r in range(5, ws.max_row + 1):
            action = ws.cell(r, 2).value
            if action is None or str(action).startswith("("): continue
            # Skip header/section rows · Action cell contains literal "Action" text
            if str(action).strip() == "Action": continue
            stop_val = ws.cell(r, 11).value
            if stop_val is None: continue
            # Skip legend rows (start of Score/Action explanations)
            if str(stop_val).strip() in ("Target", ""): continue
            s = str(stop_val)
            # Must NOT be bare "UNAVAILABLE"
            assert s != "UNAVAILABLE", f"{m} row {r} · bare UNAVAILABLE stop (CEO §36 trichotomy violation)"
            # Must be one of the four allowed states
            is_valid = (
                s.replace(".", "", 1).replace("-", "", 1).isdigit()
                or "SUGGESTED" in s
                or "N/A" in s
                or "DATA_ERROR" in s
            )
            assert is_valid, f"{m} row {r} · stop value {s!r} not in trichotomy"


def test_investments_sheet_is_first_tab():
    from openpyxl import load_workbook
    for m in ("india", "usa"):
        p = Path(__file__).resolve().parents[2] / "reports" / "telegram" / f"aegis_{m}_2026-09-03.xlsx"
        if not p.exists(): continue
        wb = load_workbook(p)
        if "01_Investments" not in wb.sheetnames: continue
        assert wb.sheetnames[0] == "01_Investments", (
            f"{m} · 01_Investments must be FIRST tab · got {wb.sheetnames[0]}"
        )


def test_option1_r1_active_present_in_investments_sheet_c19_preserved_in_portfolio():
    """CEO 2026-09-04 · Option 1 fix regression test.

    R1 ACTIVE opportunities MUST render in 01_Investments ACTIVE section
    (fixes 'R1 daily NEW does not translate into ACTIVE' UX gap AND USA
    'no recommendations' where aegis_today_usa.csv does not exist).

    CRITICAL · C19 workbook contract MUST remain intact on 01_Portfolio
    and Exit History · R1 stays retired-from-Portfolio there.
    """
    from openpyxl import load_workbook
    root = Path(__file__).resolve().parents[2]
    found_any = False
    for m in ("india", "usa"):
        # Pick LATEST workbook for the market · Option 1 only exists in files
        # rebuilt after 2026-09-04 (earlier files pre-date the fix).
        candidates = sorted((root / "reports" / "telegram").glob(f"aegis_{m}_*.xlsx"),
                              reverse=True)
        p = next((c for c in candidates if "2026-09-04" in c.name or "2026-09-05" in c.name), None)
        if p is None: continue
        date_str = p.stem.replace(f"aegis_{m}_", "")
        found_any = True
        for _ in [None]:
            wb = load_workbook(p, data_only=True, read_only=True)
            if "01_Investments" not in wb.sheetnames: continue
            ws = wb["01_Investments"]
            # Walk sheet · identify section · count R1 in ACTIVE
            r1_in_active = 0
            section = None
            for row in ws.iter_rows(values_only=True):
                row_str = " ".join(str(c) for c in row if c is not None).upper()
                if "NEW OPPORTUNITIES" in row_str: section = "NEW"
                elif ("ACTIVE POSITIONS" in row_str or
                      "ACTIVE HOLDINGS" in row_str or "HOLDING" in row_str):
                    section = "ACTIVE"
                elif section == "ACTIVE" and len(row) >= 5 and row[4] == "R1":
                    r1_in_active += 1
            # C19 · 01_Portfolio must contain ZERO R1 rows
            if "01_Portfolio" in wb.sheetnames:
                ws2 = wb["01_Portfolio"]
                r1_in_portfolio = sum(
                    1 for row in ws2.iter_rows(values_only=True)
                    for c in row if str(c) == "R1"
                )
                assert r1_in_portfolio == 0, (
                    f"{m} {date_str} · C19 VIOLATION · 01_Portfolio has "
                    f"{r1_in_portfolio} R1 rows (must be 0)"
                )
            # Option 1 · at least one R1 ACTIVE row in 01_Investments (either
            # market with an aegis_*_YYYY-MM-DD.xlsx has R1 registry data)
            assert r1_in_active > 0, (
                f"{m} {date_str} · Option 1 REGRESSION · expected R1 rows "
                f"in 01_Investments ACTIVE section · got 0"
            )
    # We only assert when at least one workbook was present (CI may not have artifacts)
    if not found_any:
        import pytest
        pytest.skip("no aegis_*_{date}.xlsx present in reports/telegram/")
