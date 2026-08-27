"""CEO 2026-08-27 · Portfolio 6-section split property tests."""
from openpyxl import Workbook

from backend.research.mr_evidence_layer import (
    emit_xlsx, PORTFOLIO_SHEET_COLS,
)


def _has_section(sheet, label):
    for row in sheet.iter_rows(values_only=True):
        if row and isinstance(row[0], str) and label in row[0]:
            return True
    return False


def test_portfolio_has_six_sections(tmp_path):
    portfolio = [
        # Active
        {"as_of":"2026-08-27","market":"INDIA","ticker":"SBIN","runner":"R1",
         "decision":"ACTIVE","lifecycle":"ACTIVE","research_badge":"—",
         "research_signal":"—"},
        # New
        {"as_of":"2026-08-27","market":"INDIA","ticker":"MRF","runner":"R1",
         "decision":"NEW","lifecycle":"NEW","research_badge":"—",
         "research_signal":"—"},
        # Re-entry
        {"as_of":"2026-08-27","market":"INDIA","ticker":"HDFC","runner":"R2",
         "decision":"RE-ENTRY","lifecycle":"ACTIVE","research_badge":"—",
         "research_signal":"—"},
        # Momentum
        {"as_of":"2026-08-27","market":"INDIA","ticker":"BEL","runner":"MOMENTUM",
         "decision":"MOMENTUM WATCH","lifecycle":"ACTIVE","research_badge":"—",
         "research_signal":"—"},
        # Shadow
        {"as_of":"2026-08-27","market":"INDIA","ticker":"ONGC","runner":"SHADOW",
         "decision":"SUGGESTED","lifecycle":"SUGGESTED",
         "research_badge":"🧪 RESEARCH — E3","research_signal":"E3 TIME-STOP WATCH"},
        # Exit
        {"as_of":"2026-08-27","market":"INDIA","ticker":"LUPIN","runner":"R1",
         "decision":"EXIT","lifecycle":"EXIT","research_badge":"—",
         "research_signal":"—"},
    ]
    xlsx = emit_xlsx(tmp_path, "india", [], portfolio, [])
    from openpyxl import load_workbook
    wb = load_workbook(xlsx, read_only=True)
    sheet = wb["Portfolio"]
    for label in ("🟢 ACTIVE POSITIONS", "🆕 NEW RECOMMENDATIONS",
                   "🔁 RE-ENTRY", "🎯 MOMENTUM",
                   "🟣 SHADOW / SUGGESTED", "⚠ ACTION / EXIT"):
        assert _has_section(sheet, label), f"missing section: {label}"


def test_portfolio_columns_include_runner_and_research_badge():
    assert "runner" in PORTFOLIO_SHEET_COLS
    assert "research_badge" in PORTFOLIO_SHEET_COLS
    # 12 columns total · slim operator view
    assert len(PORTFOLIO_SHEET_COLS) == 12
