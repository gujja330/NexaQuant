"""Compute a determinism-safe data-only hash of the AEGIS deliverable XLSX.

A production XLSX contains wall-clock timestamps (`Last Updated` column,
`ts_utc` fields) that legitimately change per run. Byte-hashing the whole
file therefore fails determinism checks that ARE actually deterministic
in every business-relevant way.

This script:
  · Loads the XLSX
  · Reads every row of every sheet
  · Excludes columns whose header matches a wall-clock skiplist
  · Formats numeric values to fixed precision
  · Hashes the concatenated normalized cell values

Output: `reports/reconcile/determinism_hash_{market}.json` with the current
hash + wall-clock exclusions used. Compare across two consecutive runs
to prove determinism.

CEO 2026-09-01. Never modifies the XLSX.
"""
from __future__ import annotations
import argparse
import hashlib
import json
import sys
from datetime import date
from pathlib import Path
from openpyxl import load_workbook

_ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(_ROOT))

# Columns whose values legitimately change per run · exclude from hash
_WALLCLOCK_COLS = {
    "last updated", "asof timestamp", "ts_utc", "generated at",
    "build time", "rendered at",
}
# Sheet name patterns to skip entirely (banner-only sheets that print
# the current date in row 1)
_BANNER_ROW_INDEX = 1  # 1-based · skip banner text row where present


def _normalize_cell(v) -> str:
    if v is None:
        return ""
    if isinstance(v, float):
        # 6-digit precision · avoid 1e-10 noise
        return f"{v:.6f}"
    return str(v)


def hash_xlsx(xlsx_path: Path, exclude_cols: set) -> dict:
    if not xlsx_path.exists():
        return {"error": f"missing: {xlsx_path}", "hash": None}
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    h = hashlib.sha256()
    per_sheet = {}
    for name in wb.sheetnames:
        ws = wb[name]
        rows = list(ws.iter_rows(values_only=True))
        # Find header row · scan first 8 rows for one with 5+ non-None
        hdr_i = 0
        for i, r in enumerate(rows[:8]):
            if r and sum(1 for c in r if c is not None) >= 5:
                hdr_i = i
                break
        hdr = rows[hdr_i] if rows else ()
        skip_cols = set()
        for i, c in enumerate(hdr):
            if c and str(c).strip().lower() in exclude_cols:
                skip_cols.add(i)
        sh = hashlib.sha256()
        for r in rows[hdr_i + 1:]:
            if not r: continue
            for i, v in enumerate(r):
                if i in skip_cols: continue
                sh.update(_normalize_cell(v).encode("utf-8"))
                sh.update(b"\x1f")  # unit separator
            sh.update(b"\n")
        per_sheet[name] = {
            "sha256": sh.hexdigest(),
            "n_rows": max(0, len(rows) - hdr_i - 1),
            "n_skipped_cols": len(skip_cols),
            "skipped_col_names": [str(hdr[i]) for i in skip_cols],
        }
        h.update(name.encode("utf-8"))
        h.update(b"\x1e")  # record separator
        h.update(sh.hexdigest().encode("utf-8"))
    wb.close()
    return {
        "hash": h.hexdigest(),
        "per_sheet": per_sheet,
        "exclude_cols": sorted(exclude_cols),
    }


def compute(market: str, root: Path) -> dict:
    market_l = market.lower()
    xlsx = root / "reports" / "telegram" / f"aegis_history_{market_l}.xlsx"
    result = hash_xlsx(xlsx, _WALLCLOCK_COLS)
    result["market"] = market_l
    result["xlsx"] = str(xlsx.relative_to(root))
    result["asof"] = date.today().isoformat()
    out_p = root / "reports" / "reconcile" / f"determinism_hash_{market_l}.json"
    out_p.parent.mkdir(parents=True, exist_ok=True)
    out_p.write_text(json.dumps(result, indent=2, ensure_ascii=False),
                      encoding="utf-8")
    return result


def compare(baseline_path: Path, current_path: Path) -> dict:
    baseline = json.loads(baseline_path.read_text(encoding="utf-8"))
    current = json.loads(current_path.read_text(encoding="utf-8"))
    match = baseline.get("hash") == current.get("hash")
    per_sheet_diffs = []
    for name in baseline.get("per_sheet", {}):
        b = baseline["per_sheet"].get(name, {})
        c = current.get("per_sheet", {}).get(name, {})
        if b.get("sha256") != c.get("sha256"):
            per_sheet_diffs.append({"sheet": name,
                                     "baseline": b.get("sha256"),
                                     "current": c.get("sha256"),
                                     "baseline_rows": b.get("n_rows"),
                                     "current_rows": c.get("n_rows")})
    return {
        "match": match,
        "baseline_hash": baseline.get("hash"),
        "current_hash": current.get("hash"),
        "per_sheet_diffs": per_sheet_diffs,
    }


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--market", choices=["india", "usa", "both"],
                     default="both")
    ap.add_argument("--compare", metavar="BASELINE_JSON",
                     help="compare against a baseline hash json")
    args = ap.parse_args()
    markets = ["india", "usa"] if args.market == "both" else [args.market]
    for m in markets:
        rep = compute(m, _ROOT)
        summary = {"market": m, "hash": rep.get("hash"),
                    "n_sheets": len(rep.get("per_sheet", {}))}
        print(json.dumps(summary, indent=2, ensure_ascii=False))
        if args.compare:
            baseline_p = Path(args.compare)
            if not baseline_p.is_absolute():
                baseline_p = _ROOT / baseline_p
            if baseline_p.exists():
                cur_p = _ROOT / "reports" / "reconcile" / f"determinism_hash_{m}.json"
                diff = compare(baseline_p, cur_p)
                print("MATCH" if diff["match"] else "DRIFT",
                      "·", len(diff["per_sheet_diffs"]), "sheet diffs")
                for d in diff["per_sheet_diffs"]:
                    _line = (f"  DRIFT {d['sheet']}: "
                              f"{d['baseline'][:8]}.. -> {d['current'][:8]}..")
                    print(_line.encode("ascii", errors="replace").decode("ascii"))
    return 0


if __name__ == "__main__":
    sys.exit(main())
