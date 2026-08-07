"""Backfill script · rewrite historical XLSX rows to use frozen Entry Price.

Operator P0 audit 2026-08-06: "AEGIS is recreating the recommendation
every day · not tracking a persistent recommendation · Entry Price =
Current Price = same day · that's overwriting the entry reference."

FIX
───
For every historical row in the XLSX:
    1. If ticker is in position_store today: use position_store first_seen_price
       as Entry Price (immutable · frozen · from the day recommendation was born)
    2. Current Price for that row's Date = bar close on that Date from parquet
    3. Current Perf % = (Current - Entry) / Entry × 100
    4. Max Gain % = highest_close_between(first_seen_date, row_date) / entry - 1
    5. Max DD % = lowest_close_between(first_seen_date, row_date) / entry - 1
    6. Recommended = position_store first_seen_date (immutable)

Runs standalone · idempotent · rewrites the XLSX in place.

Usage:
    python scripts/backfill_frozen_entry.py
"""
from __future__ import annotations

import argparse
import json
import sys
from pathlib import Path

_ROOT = Path(__file__).resolve().parents[1]


def _load_positions(market: str) -> dict:
    reports = _ROOT / ("usa/reports" if market == "usa" else "reports")
    p = reports / "position_store" / market / "positions.json"
    if not p.exists(): return {}
    try:
        return (json.loads(p.read_text(encoding="utf-8")).get("positions") or {})
    except Exception:
        return {}


def _lookup_position(positions: dict, ticker: str) -> dict | None:
    """Find position by ticker OR short-ticker."""
    short = ticker.replace(".NS", "").replace(".BO", "").upper()
    for k, v in positions.items():
        ks = k.replace(".NS", "").replace(".BO", "").upper()
        if ks == short: return v
    return None


def _bar_prices(market: str, ticker: str) -> "pd.DataFrame | None":
    import pandas as pd
    short = ticker.replace(".NS", "").replace(".BO", "")
    dir_p = _ROOT / ("usa/data/raw/us" if market == "usa" else "data/raw/india")
    p = dir_p / f"{short}_D1.parquet"
    if not p.exists(): return None
    try:
        df = pd.read_parquet(p)
        col = "close" if "close" in df.columns else "Close" if "Close" in df.columns else None
        if col is None: return None
        df = df[[col]].copy()
        df.columns = ["close"]
        df.index = pd.to_datetime(df.index).strftime("%Y-%m-%d")
        return df
    except Exception:
        return None


def _find_col(headers: list, name: str) -> int | None:
    try: return headers.index(name) + 1
    except ValueError: return None


def backfill(xlsx_path: Path) -> dict:
    from openpyxl import load_workbook
    if not xlsx_path.exists():
        return {"error": f"{xlsx_path} missing"}

    india_pos = _load_positions("india")
    usa_pos = _load_positions("usa")

    wb = load_workbook(xlsx_path)
    ws = wb["AEGIS Daily"] if "AEGIS Daily" in wb.sheetnames else wb.active
    h = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]

    c_date = _find_col(h, "Date")
    c_country = _find_col(h, "Country")
    c_ticker = _find_col(h, "Ticker")
    c_run = _find_col(h, "Run_Type")
    c_status = _find_col(h, "Status")
    c_entry = _find_col(h, "Entry Price")
    c_curr = _find_col(h, "Current Price")
    c_perf = _find_col(h, "Current Perf %")
    c_recommended = _find_col(h, "Recommended")
    c_maxg = _find_col(h, "Max Gain %")
    c_maxd = _find_col(h, "Max DD %")
    c_exit_pnl = _find_col(h, "Exit P&L %")
    c_pos_id = _find_col(h, "Position ID")

    if not all([c_date, c_country, c_ticker, c_entry, c_curr, c_perf]):
        return {"error": "required columns missing"}

    # Cache bar data per ticker
    bar_cache = {}
    n_rewritten = 0
    n_skipped = 0

    for r in range(2, ws.max_row + 1):
        dt = str(ws.cell(row=r, column=c_date).value or "")
        ctry = ws.cell(row=r, column=c_country).value
        tk = ws.cell(row=r, column=c_ticker).value
        if not (dt and ctry and tk):
            n_skipped += 1
            continue
        mkt = ctry.lower() if ctry else ""
        positions = india_pos if mkt == "india" else usa_pos
        pos = _lookup_position(positions, tk)
        if not pos:
            n_skipped += 1
            continue

        entry = pos.get("first_seen_price")
        first_seen = str(pos.get("first_seen_date") or "")[:10]
        if not entry or not first_seen:
            n_skipped += 1
            continue

        # Skip rows BEFORE the position was opened
        if dt < first_seen:
            n_skipped += 1
            continue

        # Load bars once per ticker
        cache_key = (mkt, tk)
        if cache_key not in bar_cache:
            bar_cache[cache_key] = _bar_prices(mkt, tk)
        df = bar_cache[cache_key]
        if df is None:
            n_skipped += 1
            continue

        # Current close on this row's date
        curr = None
        if dt in df.index:
            curr = float(df.loc[dt, "close"])
        else:
            earlier = [d for d in df.index if d <= dt]
            if earlier: curr = float(df.loc[earlier[-1], "close"])
        if curr is None:
            n_skipped += 1
            continue

        # Max gain / max DD from first_seen to this date
        window = df[(df.index >= first_seen) & (df.index <= dt)]
        max_gain = max_dd = None
        if not window.empty:
            highest = float(window["close"].max())
            lowest = float(window["close"].min())
            max_gain = round((highest - entry) / entry * 100, 2)
            max_dd = round((lowest - entry) / entry * 100, 2)

        # Overwrite the row
        perf = round((curr - entry) / entry * 100, 2)
        ws.cell(row=r, column=c_entry, value=round(entry, 2))
        ws.cell(row=r, column=c_curr, value=round(curr, 2))
        ws.cell(row=r, column=c_perf, value=perf)
        if c_recommended:
            ws.cell(row=r, column=c_recommended, value=first_seen)
        if c_maxg and max_gain is not None:
            ws.cell(row=r, column=c_maxg, value=max_gain)
        if c_maxd and max_dd is not None:
            ws.cell(row=r, column=c_maxd, value=max_dd)
        # 2026-08-07 · fix Exit P&L% (was 0 for all rotation exits)
        if c_exit_pnl and c_status:
            status_val = ws.cell(row=r, column=c_status).value
            if status_val == "EXIT":
                ws.cell(row=r, column=c_exit_pnl, value=perf)
        # 2026-08-07 · populate Position ID for EVERY row (was 71/224)
        if c_pos_id:
            short = tk.replace(".NS", "").replace(".BO", "").upper()
            mkt_short = ctry.upper()[:3] if ctry else "IND"
            fs_short = first_seen[:10].replace("-", "")
            ws.cell(row=r, column=c_pos_id, value=f"{short}_{mkt_short}_{fs_short}")
        n_rewritten += 1

    wb.save(xlsx_path)
    return {"n_rewritten": n_rewritten, "n_skipped": n_skipped,
                "path": str(xlsx_path)}


def main() -> int:
    import io
    try:
        sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    except Exception:
        sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")

    ap = argparse.ArgumentParser()
    ap.add_argument("--path", default=str(_ROOT / "reports" / "telegram" / "aegis_history.xlsx"))
    args = ap.parse_args()

    result = backfill(Path(args.path))
    if "error" in result:
        print(f"[backfill] ERROR · {result['error']}")
        return 1
    print(f"[backfill] rewrote {result['n_rewritten']} rows · skipped {result['n_skipped']}")
    print(f"[backfill] path: {result['path']}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
