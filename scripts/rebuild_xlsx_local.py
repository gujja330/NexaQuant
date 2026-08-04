"""AEGIS · rebuild the unified daily XLSX locally.

Purpose: operator debug tool. Regenerate `reports/telegram/aegis_history.xlsx`
from whatever is currently in `reports/recommendations.json` (India) and
`usa/reports/recommendations.json` (USA) without re-running the daily pipeline
or sending to Telegram.

Usage:
    # Rebuild with today's date · both markets
    python scripts/rebuild_xlsx_local.py

    # Specific date (useful for backfill / debugging)
    python scripts/rebuild_xlsx_local.py --asof 2026-08-01

    # Single market
    python scripts/rebuild_xlsx_local.py --market india
    python scripts/rebuild_xlsx_local.py --market usa

    # Also open the file in Excel when done (Windows only)
    python scripts/rebuild_xlsx_local.py --open

Output: reports/telegram/aegis_history.xlsx (daily-append semantics preserved).
No Telegram send · no pipeline mutation · read-only WRT source recommendations.
"""
from __future__ import annotations

import argparse
import os
import subprocess
import sys
from datetime import date as _date
from pathlib import Path

_ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(_ROOT))

from backend.delivery.telegram.detail_xlsx import (  # noqa: E402
    build_unified_history, build_and_stamp_all,
)


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--asof", default=_date.today().isoformat(),
                       help="Date stamp on rows (YYYY-MM-DD · default: today)")
    ap.add_argument("--market", choices=["india", "usa", "both"], default="both",
                       help="Which markets to (re)build for asof (default: both)")
    ap.add_argument("--open", action="store_true",
                       help="Open the XLSX in default Excel viewer when done")
    args = ap.parse_args()

    markets = ["india", "usa"] if args.market == "both" else [args.market]
    print(f"[rebuild_xlsx] asof={args.asof} · markets={markets}")

    # v5 · uses stamp-first flow · writes rank_history + regime_history +
    # profit_protection_{market}.json before building the XLSX (so Prior
    # Rank, Rank Δ, and Alert columns populate correctly).
    xlsx_path = build_and_stamp_all(_ROOT, args.asof, markets=markets)
    print(f"[rebuild_xlsx] wrote {xlsx_path}")
    print(f"[rebuild_xlsx] size: {xlsx_path.stat().st_size:,} bytes")

    # Quick row count for sanity check
    try:
        from openpyxl import load_workbook
        wb = load_workbook(xlsx_path, read_only=True)
        ws = wb.active
        n_rows = ws.max_row - 1
        n_cols = ws.max_column
        print(f"[rebuild_xlsx] rows={n_rows} · cols={n_cols}")
        wb.close()
    except Exception as exc:
        print(f"[rebuild_xlsx] sanity check skipped: {exc}")

    if args.open:
        if sys.platform == "win32":
            os.startfile(str(xlsx_path))     # noqa: S606
            print(f"[rebuild_xlsx] opened in default viewer")
        elif sys.platform == "darwin":
            subprocess.run(["open", str(xlsx_path)], check=False)
        else:
            subprocess.run(["xdg-open", str(xlsx_path)], check=False)

    return 0


if __name__ == "__main__":
    sys.exit(main())
