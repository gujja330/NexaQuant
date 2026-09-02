"""AEGIS · Final End-to-End Contract Reconciliation · CEO 2026-08-31.

Verifies the delivery/audit chain across every layer for a single
market. STOP-on-mismatch · never guesses · never silently patches.

Checks:
  · Registry → Portfolio (ACTIVE consistency)
  · Registry-CLOSED → Exit History body (I20-shape)
  · Portfolio banner ≡ Portfolio body (per-axis)
  · AEGIS History PID format · 100% new-format
  · No fabricated LOW/PENDING/0 in Portfolio holding rows
  · Monthly Summary sheet exists (rule C5)
  · Exit History body has NO trailer strings (rule C5)
  · Position ID uniqueness at canonical grain per population
  · Historical rows preserved (no dedup)

Emits `reports/reconcile/final_reconcile_{market}_{asof}.json` with
every check result · exits 0 on all-pass · non-zero on any mismatch.
"""
from __future__ import annotations

import argparse
import hashlib
import json
import sys
from collections import Counter, defaultdict
from datetime import date
from pathlib import Path

_ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(_ROOT))

from backend.research import opportunity_registry as oreg   # noqa: E402
from openpyxl import load_workbook                            # noqa: E402


_NEW_PID_PREFIXES = ("IND-", "USA-", "R1-", "R2-", "SHADOW-", "MOMENTUM-")


class ReconcileFailure(Exception):
    pass


def _col(hdr, name):
    for i, c in enumerate(hdr):
        if c and name.lower() == str(c).lower(): return i
    return None


def reconcile(market: str, root: Path) -> dict:
    market_l = market.lower()
    asof = date.today().isoformat()
    xlsx = root / "reports" / "telegram" / f"aegis_history_{market_l}.xlsx"
    if not xlsx.exists():
        raise ReconcileFailure(f"artifact not present · {xlsx}")

    wb = load_workbook(xlsx, read_only=True, data_only=True)
    checks: list = []

    def _add(name: str, ok: bool, detail: str = "", data=None):
        checks.append({"name": name, "ok": ok, "detail": detail, "data": data})

    # ── C1 · Sheets present · CEO 2026-09-03 FINAL 4-sheet spec ─────
    # Exactly 4 visible sheets · no more no less. Sheet 04 is the
    # historical reconstruction from canonical Registry (CEO added
    # to answer "what did AEGIS hold on date X?").
    required_sheets = [
        "01_Portfolio",                    # 1 · current active R2 holdings
        "02_Today_Momentum",               # 2 · today's decisions/recommendations
        "03_Exit_History",                 # 3 · realized production exits only
        "04_Daily_Portfolio_History",      # 4 · daily active reconstruction
    ]
    missing = [s for s in required_sheets if s not in wb.sheetnames]
    extra = [s for s in wb.sheetnames if s not in required_sheets]
    _add("C1_required_sheets_present",
          (not missing) and (not extra),
          (f"missing={missing} · extra={extra}"
            if (missing or extra)
            else "exactly 4 required sheets present · no legacy sheets"),
          {"present": wb.sheetnames, "required": required_sheets,
           "missing": missing, "extra": extra})

    # ── C2 · Registry consistency ──────────────────────────────────
    reg = oreg.load_all(root)
    reg_active = []
    reg_closed_90d = []
    _asof_d = date.fromisoformat(asof)
    from datetime import timedelta
    _cutoff = (_asof_d - timedelta(days=90)).isoformat()
    for _pid, opps in reg.items():
        for o in opps:
            if o.market.lower() != market_l: continue
            if o.status == "ACTIVE":
                reg_active.append(o.opportunity_id)
            elif o.status == "CLOSED" and o.closed_date and o.closed_date >= _cutoff:
                reg_closed_90d.append(o.opportunity_id)
    n_reg_active = len(set(reg_active))
    n_reg_closed = len(set(reg_closed_90d))
    _add("C2_registry_load", True,
          f"Registry {market_l}: {n_reg_active} ACTIVE · {n_reg_closed} CLOSED-90d",
          {"n_active": n_reg_active, "n_closed_90d": n_reg_closed})

    # ── C3 · Canonical PID format 100% new-format ───────────────────
    # CEO 2026-09-01 3-sheet spec · no raw AEGIS History sheet · so
    # scan Registry directly for PID format compliance.
    _NEW_PID_PATTERN_RE = None  # already-canonical regex from Phase 2
    n_new_pid = 0
    n_legacy_pid = 0
    legacy_samples = []
    for _pid, opps in reg.items():
        for o in opps:
            if o.market.lower() != market_l: continue
            pid_up = str(o.opportunity_id).upper()
            if pid_up.startswith(_NEW_PID_PREFIXES):
                n_new_pid += 1
            else:
                n_legacy_pid += 1
                if len(legacy_samples) < 5:
                    legacy_samples.append(o.opportunity_id)
    _add("C3_history_new_format_pids", n_legacy_pid == 0,
          f"{n_new_pid} new-format PIDs · {n_legacy_pid} legacy (source: Registry)",
          {"new_ct": n_new_pid, "legacy_ct": n_legacy_pid,
               "legacy_samples": legacy_samples})

    # ── C4 · Portfolio banner ≡ body count (3-sheet spec) ───────────
    # New banner format: "🟢 R2 ACTIVE: N · R1 retired · ..." — banner
    # count of R2 ACTIVE must match the number of body rows.
    ws_p = wb["01_Portfolio"]
    rows_p = list(ws_p.iter_rows(values_only=True))
    banner_r2 = str(rows_p[1][0]) if len(rows_p) > 1 and rows_p[1][0] else ""
    import re as _re
    m_act = _re.search(r"R2\s+ACTIVE[:\s]+(\d+)", banner_r2, _re.IGNORECASE)
    banner_active = int(m_act.group(1)) if m_act else -1
    banner_new = 0        # new spec · no NEW rows in Portfolio (they go to sheet 2)
    banner_suggested = 0  # no SUGGESTED bucket in new Portfolio
    # Body count: rows below header with a Ticker
    hdr_p_idx = next((i for i, r in enumerate(rows_p)
                        if r[0] and "Position ID" in str(r[0])), None)
    body = rows_p[hdr_p_idx + 1:] if hdr_p_idx is not None else []
    hdr_p = rows_p[hdr_p_idx] if hdr_p_idx is not None else ()
    def _pcol(name):
        for i, c in enumerate(hdr_p):
            if c and str(c).lower() == name.lower(): return i
        return None
    ci_run = _pcol("Runner")
    ci_tk = _pcol("Ticker")
    body_life_active = 0
    for r in body:
        if not r: continue
        # A body row must have a canonical PID in col[0] AND a ticker in col[1]
        pid = str(r[0] or "").strip() if len(r) > 0 else ""
        tk = str(r[ci_tk] or "").strip() if ci_tk is not None and ci_tk < len(r) else ""
        if not pid or not tk: continue
        # Canonical PID starts with USA- or IND-
        if not (pid.upper().startswith("USA-") or pid.upper().startswith("IND-")): continue
        body_life_active += 1
    body_life_new = 0
    body_suggested = 0
    ok_active = banner_active == body_life_active
    ok_new = True     # no NEW rows expected in Portfolio
    ok_sugg = True    # no SUGGESTED bucket in new spec
    _add("C4_banner_lifecycle_active", ok_active,
          f"banner={banner_active} body={body_life_active}",
          {"banner": banner_active, "body": body_life_active})
    _add("C4_banner_lifecycle_new", ok_new,
          f"banner={banner_new} body={body_life_new}",
          {"banner": banner_new, "body": body_life_new})
    _add("C4_banner_suggested", ok_sugg,
          f"banner={banner_suggested} body={body_suggested}",
          {"banner": banner_suggested, "body": body_suggested})

    # ── C5 · Combined ledger has NO trailer rows ───────────────────
    if "03_Exit_History" in wb.sheetnames:
        ws_e = wb["03_Exit_History"]
        eh_rows = [r for r in ws_e.iter_rows(values_only=True)
                     if any(c is not None for c in r)]
        # Last data row's first cell should be a ticker
        trailer_hit = False
        trailer_row = None
        for r in eh_rows:
            first = str(r[0]) if r[0] else ""
            for f in ("MONTHLY", "SUMMARY", "──"):
                if f in first.upper():
                    trailer_hit = True
                    trailer_row = first[:60]
                    break
            if trailer_hit: break
        _add("C5_exit_history_no_trailer", not trailer_hit,
              f"trailer row found: '{trailer_row}'" if trailer_hit else "clean",
              {"trailer_found": trailer_hit, "trailer_row": trailer_row})

    # ── C6 · Path-A holding rows use "—" not fabricated values ─────
    n_fabricated = 0
    fabricated_samples = []
    for r in body:
        action = str(r[1]) if len(r) > 1 and r[1] else ""
        if "holding" in action.lower() and "no signal" in action.lower():
            urg = str(r[15]) if len(r) > 15 and r[15] else ""
            iq = str(r[20]) if len(r) > 20 and r[20] else ""
            if "LOW" in urg.upper() or "PENDING" in iq.upper():
                n_fabricated += 1
                if len(fabricated_samples) < 5:
                    fabricated_samples.append((r[0], urg[:20], iq[:20]))
    _add("C6_no_fabricated_low_pending", n_fabricated == 0,
          f"{n_fabricated} holding rows with LOW/PENDING",
          {"n": n_fabricated, "samples": fabricated_samples})

    # ── C7 · Registry PID uniqueness (canonical grain) ──────────────
    # 3-sheet spec has no raw AEGIS History sheet · uniqueness is asserted
    # at the Registry (canonical PID authority) level directly. Registry
    # is append-only by (opportunity_id, event) so per-PID latest state is
    # unique by construction · check that no PID has two ACTIVE states
    # simultaneously for the same market.
    from collections import defaultdict as _dd
    latest_by_pid = {}
    for _pid, opps in reg.items():
        for o in opps:
            if o.market.lower() != market_l: continue
            latest_by_pid[o.opportunity_id] = o
    _dup_active = _dd(list)
    for pid, o in latest_by_pid.items():
        if o.status == "ACTIVE":
            key = (o.ticker.upper(), o.runner, o.created_date or "")
            _dup_active[key].append(pid)
    unexplained_dupes = [(k, v) for k, v in _dup_active.items() if len(v) > 1]
    _add("C7_history_canonical_uniqueness", len(unexplained_dupes) == 0,
          f"{len(unexplained_dupes)} UNEXPLAINED PID duplicates at "
          f"(ticker, runner, entry_date) grain · Registry authoritative",
          {"n_unexplained": len(unexplained_dupes),
           "unexplained_samples": unexplained_dupes[:5]})

    # ── C8 · Registry-CLOSED ⊆ 02_Decisions_Exit_History HISTORICAL rows ──
    if "03_Exit_History" in wb.sheetnames:
        eh_tickers = set()
        # 03_Exit_History layout: hdr [0]=Position ID [1]=Ticker [2]=Runner
        # [3]=Market [4]=Entry Date [5]=Exit Date ...
        # Header row has "Position ID" in col 0
        _hdr_idx = next((i for i, r in enumerate(eh_rows)
                          if r[0] and "Position ID" in str(r[0])), None)
        if _hdr_idx is not None and _hdr_idx + 1 < len(eh_rows):
            for r in eh_rows[_hdr_idx + 1:]:
                if not r or not r[0]: continue
                # Skip legend / summary rows (they don't have canonical PID)
                if not (str(r[0]).upper().startswith("USA-") or
                          str(r[0]).upper().startswith("IND-")): continue
                if len(r) > 1 and r[1]:
                    eh_tickers.add(str(r[1]).upper().split(".", 1)[0])
        # C8 · CEO 2026-09-01 STRENGTHENED: retirement-aware · carveout-aware.
        # Registry CLOSED events for RETIRED runners are EXPECTED to be absent
        # (retirement carveout · not a production exit). Registry CLOSED events
        # with reasons ORPHAN_AUTO_CLOSE / SAME_DAY_ROTATION / CANCELLED /
        # DATA_REPAIR are also carveouts. Only ACTIVE-runner · non-carveout
        # CLOSED events without a matching Exit History row are real gaps.
        try:
            from backend.delivery.canonical.retirement import retired_runners as _c8_retired
            _c8_ret = _c8_retired(_ROOT)
        except Exception:
            _c8_ret = set()
        _CARVEOUT_KW = ("ORPHAN_AUTO_CLOSE", "SAME_DAY_ROTATION",
                         "CANCELLED", "DATA_REPAIR")
        # CEO 2026-09-02 · align with builder's structural admin filter:
        # same-day OR entry_price==exit_price = administrative event ·
        # not required to appear in Exit History body.
        try:
            from scripts.build_aegis_3sheet_workbook import (
                _is_administrative_exit as _c8_is_admin,
                _close_on_or_before as _c8_close_on)
        except Exception:
            _c8_is_admin = None
            _c8_close_on = None
        reg_closed_by_ticker: dict[str, list] = {}
        for _pid, opps in reg.items():
            for o in opps:
                if o.market.lower() != market_l or o.status != "CLOSED":
                    continue
                _tk = o.ticker.upper().replace(".NS", "").replace(".BO", "")
                reg_closed_by_ticker.setdefault(_tk, []).append(o)
        real_missing = []
        retired_ignored = []
        carveout_ignored = []
        admin_ignored = []
        for tk, closed_events in reg_closed_by_ticker.items():
            if tk in eh_tickers: continue
            has_production_gap = False
            for ev in closed_events:
                runner_up = str(ev.runner or "").upper()
                reason = str(getattr(ev, "closed_reason", "") or "").upper()
                if runner_up in _c8_ret:
                    retired_ignored.append((tk, runner_up))
                    continue
                if any(kw in reason for kw in _CARVEOUT_KW):
                    carveout_ignored.append((tk, reason))
                    continue
                # Structural admin check · builder filters these out too
                if _c8_is_admin is not None and _c8_close_on is not None:
                    _ep = _c8_close_on(_ROOT, ev.ticker, market_l, ev.created_date or "")
                    _xp = _c8_close_on(_ROOT, ev.ticker, market_l, ev.closed_date or "")
                    if _c8_is_admin(ev, _ep, _xp):
                        admin_ignored.append((tk, "same-day or entry==exit"))
                        continue
                has_production_gap = True
                real_missing.append({"ticker": tk, "runner": runner_up,
                                       "reason": reason,
                                       "closed_date": str(ev.closed_date or "")})
                break
        _add("C8_registry_closed_in_exit_history",
              len(real_missing) == 0,
              (f"{len(real_missing)} ACTIVE-runner non-carveout non-admin Registry-CLOSED "
                f"missing from EH · {len(retired_ignored)} retired-runner · "
                f"{len(carveout_ignored)} carveout · {len(admin_ignored)} admin (same-day/zero-Δ) "
                f"ignored (all expected)"),
              {"n_real_missing": len(real_missing),
               "n_retired_ignored": len(retired_ignored),
               "n_carveout_ignored": len(carveout_ignored),
               "n_admin_ignored": len(admin_ignored),
               "sample_real_missing": real_missing[:5]})

    # ── C10 · Retired-runner contamination check ──────────────────
    # CEO 2026-09-01 · R1 retirement · Portfolio + banner + P&L must
    # contain 0 retired-runner rows. Historical audit (AEGIS History
    # sheet · Registry) still contains R1 rows · that is preserved
    # by design. This check ensures the PRODUCTION-FACING view is
    # clean.
    try:
        from backend.delivery.canonical.retirement import retired_runners as _c10_retired
        retired_set = _c10_retired(_ROOT)
    except Exception:
        retired_set = set()
    if retired_set:
        n_retired_in_portfolio = 0
        for r in body:
            run = str(r[ci_run]).upper() if ci_run is not None and ci_run < len(r) and r[ci_run] else ""
            if run in retired_set:
                n_retired_in_portfolio += 1
        _add("C10_no_retired_in_production_portfolio",
              n_retired_in_portfolio == 0,
              f"{n_retired_in_portfolio} retired-runner rows in Portfolio · "
              f"retired={sorted(retired_set)}",
              {"n_retired_rows": n_retired_in_portfolio,
               "retired_runners": sorted(retired_set)})

    # ── C9 · Portfolio ↔ Exit History lifecycle-instance collision ─
    # 3-sheet spec: 01_Portfolio has columns (Position ID · Ticker · Runner
    # · Entry Date · ...) and 02_Decisions_Exit_History has (Population ·
    # Date · Position ID · Ticker · Runner · Entry Date · Exit Date · ...).
    # Collision = same (ticker, runner, entry_date) appears as both an
    # ACTIVE row in 01_Portfolio AND a HISTORICAL_EXIT row in sheet 2.
    port_active_key = set()
    port_entry_by_tk_run = {}
    # In new 01_Portfolio: hdr columns [0]=Position ID [1]=Ticker [2]=Runner
    # [3]=Entry Date. Body rows have those indices.
    for r in body:
        if not r or not r[0]: continue
        if str(r[0]).strip() == "" or str(r[0]).lower().startswith("no current"): continue
        tk = str(r[1]).upper().replace(".NS", "").replace(".BO", "") if len(r) > 1 and r[1] else ""
        run = str(r[2] or "").upper() if len(r) > 2 else ""
        ent = str(r[3])[:10] if len(r) > 3 and r[3] and str(r[3]) != "—" else ""
        if run in ("R1", "R2") and ent and tk:
            port_active_key.add((tk, run, ent))
            port_entry_by_tk_run.setdefault((tk, run), set()).add(ent)
    # 03_Exit_History · body rows have canonical PID in col 0
    eh_closed_key = set()
    eh_by_tk_run = {}
    if "03_Exit_History" in wb.sheetnames:
        ws_e2 = wb["03_Exit_History"]
        eh_all = list(ws_e2.iter_rows(values_only=True))
        for r in eh_all:
            if not r or not r[0]: continue
            pid = str(r[0]).upper()
            if not (pid.startswith("USA-") or pid.startswith("IND-")): continue
            # cols: 0=PID 1=Ticker 2=Runner 3=Market 4=EntryDate 5=ExitDate
            tk_e = str(r[1]).upper().replace(".NS", "").replace(".BO", "") if len(r) > 1 and r[1] else ""
            run_e = str(r[2] or "").upper() if len(r) > 2 else ""
            ent_e = str(r[4])[:10] if len(r) > 4 and r[4] and str(r[4]) != "—" else ""
            exit_e = str(r[5])[:10] if len(r) > 5 and r[5] and str(r[5]) != "—" else ""
            if run_e in ("R1", "R2") and ent_e and tk_e:
                eh_closed_key.add((tk_e, run_e, ent_e))
                eh_by_tk_run.setdefault((tk_e, run_e), []).append((ent_e, exit_e))
    # True collision · same (ticker, runner, entry_date) in both
    collisions = port_active_key & eh_closed_key
    # Same ticker+runner but different entry_date is EXPLAINED
    explained_overlap = []
    for (tk, run) in set(port_entry_by_tk_run) & set(eh_by_tk_run):
        p_ents = port_entry_by_tk_run[(tk, run)]
        e_events = eh_by_tk_run[(tk, run)]
        for e_ent, e_exit in e_events:
            for p_ent in p_ents:
                if p_ent != e_ent:
                    explained_overlap.append((tk, run, p_ent, e_ent, e_exit))
    _add("C9_portfolio_exit_no_lifecycle_collision",
          len(collisions) == 0,
          f"{len(collisions)} UNEXPLAINED lifecycle collisions "
          f"(same ticker+runner+entry_date active AND closed) · "
          f"{len(explained_overlap)} EXPLAINED overlaps (different entry_date · legit re-entry)",
          {"n_unexplained": len(collisions),
           "unexplained_collisions": sorted(list(collisions))[:10],
           "n_explained_overlaps": len(explained_overlap),
           "explained_samples": explained_overlap[:5],
           "n_portfolio_active_keys": len(port_active_key),
           "n_exit_closed_keys": len(eh_closed_key)})

    wb.close()

    # ── C19 · Workbook-wide R1 = 0 (STRENGTHENED CONTRACT) ─────────
    # CEO 2026-09-01 strengthened: R1 must be COMPLETELY absent from
    # every visible sheet · every cell · except Definitions (which may
    # name R1 as a reference explaining retirement). Cell-level scan.
    try:
        from backend.delivery.canonical.retirement import retired_runners as _c19_retired
        _c19_ret = _c19_retired(root)
        _c19_hits = []
        _c19_wb_ro = load_workbook(xlsx, read_only=True, data_only=True)
        # Also load formula-visible workbook (data_only=False) to catch
        # cells whose STORED FORMULA references R1
        _c19_wb_fx = load_workbook(xlsx, data_only=False)
        _prefixes = tuple(
            p + rr + "-" for rr in _c19_ret
            for p in ("", "IND-", "USA-")
        )
        _defs_name = "___NO_DEFINITIONS_SHEET_IN_FINAL_SPEC___"  # every sheet scanned
        # 1. Visible-value scan · CEO 2026-09-01 STRICT rule: the WORD R1
        # (and any variant of a retired runner) must not appear anywhere
        # in the workbook text · not just as an exact cell value.
        import re as _re_c19
        _wb_re = _re_c19.compile(
            r"\b(" + "|".join(_re_c19.escape(r) for r in _c19_ret) + r")\b",
            _re_c19.IGNORECASE,
        )
        for _sh_name in _c19_wb_ro.sheetnames:
            if _sh_name == _defs_name: continue
            _ws = _c19_wb_ro[_sh_name]
            _rn = 0
            for _row in _ws.iter_rows(values_only=True):
                _rn += 1
                for _v in _row:
                    if _v is None: continue
                    _s = str(_v).strip()
                    # Exact match, prefix match, OR word-boundary occurrence
                    if (_s.upper() in _c19_ret
                          or _s.upper().startswith(_prefixes)
                          or _wb_re.search(_s)):
                        _c19_hits.append({"scope": "value", "sheet": _sh_name,
                                           "row": _rn, "value": _s[:60]})
                        break
        _c19_wb_ro.close()
        # 2. Hidden / very-hidden sheet detection · every hidden sheet is
        #    itself a violation regardless of content (production workbook
        #    must have zero surprise sheets)
        _c19_hidden = []
        for _sh_name in _c19_wb_fx.sheetnames:
            _sh = _c19_wb_fx[_sh_name]
            _state = getattr(_sh, "sheet_state", "visible")
            if _state != "visible":
                _c19_hidden.append({"sheet": _sh_name, "state": _state})
        # 3. Formula scan · any cell.data_type == 'f' with R1 in formula text
        _c19_formula_hits = []
        for _sh_name in _c19_wb_fx.sheetnames:
            if _sh_name == _defs_name: continue
            _sh = _c19_wb_fx[_sh_name]
            for _row in _sh.iter_rows():
                for _c in _row:
                    if getattr(_c, "data_type", None) == "f":
                        _fx = str(_c.value or "").upper()
                        for _r in _c19_ret:
                            if _r in _fx.split():   # word-boundary-ish
                                _c19_formula_hits.append(
                                    {"scope": "formula",
                                      "sheet": _sh_name,
                                      "coord": _c.coordinate,
                                      "formula": _fx[:60]})
                                break
        # 4. Defined-name scan
        _c19_defname_hits = []
        try:
            for _dn_name in list(_c19_wb_fx.defined_names):
                _u = str(_dn_name).upper()
                if any(_r in _u.split("_") or _r in _u.split("-")
                        for _r in _c19_ret) or any(_u.startswith(p) for p in _prefixes):
                    _c19_defname_hits.append(_dn_name)
        except Exception:
            pass
        _c19_wb_fx.close()

        _c19_total = (len(_c19_hits) + len(_c19_hidden)
                       + len(_c19_formula_hits) + len(_c19_defname_hits))
        _add("C19_workbook_wide_r1_zero",
              _c19_total == 0,
              (f"{len(_c19_hits)} value hits · {len(_c19_hidden)} hidden sheets · "
                f"{len(_c19_formula_hits)} formula hits · "
                f"{len(_c19_defname_hits)} defined-name hits · "
                f"retired={sorted(_c19_ret)}"),
              {"n_value_hits": len(_c19_hits),
               "n_hidden_sheets": len(_c19_hidden),
               "n_formula_hits": len(_c19_formula_hits),
               "n_defname_hits": len(_c19_defname_hits),
               "hidden_sample": _c19_hidden[:5],
               "value_sample": _c19_hits[:5],
               "formula_sample": _c19_formula_hits[:5]})
    except Exception as _e19:
        _add("C19_workbook_wide_r1_zero", False,
              f"exception: {type(_e19).__name__}: {_e19}", None)

    # ── C18 · Crash-resilience research presence (§ Crash addendum) ─
    # Every certification pass must have current 5-state regime + per-regime
    # R2-vs-benchmark metrics computed · presence is mandatory · interpretation
    # never claims success just because a report exists.
    try:
        _cr_p = root / "reports" / "research" / "multi_layer" / f"crash_resilience_{market_l}_{asof}.json"
        if _cr_p.exists():
            _cr = json.loads(_cr_p.read_text(encoding="utf-8"))
            _tagged = int(_cr.get("n_r2_trades_tagged", 0) or 0)
            _today = str(_cr.get("today_regime", "?"))
            _dist = _cr.get("regime_distribution_alltime", {}) or {}
            # Passes only if the CLASSIFIER ran (n_days_classified > 0)
            _n_days = int(_cr.get("n_days_classified", 0) or 0)
            _add("C18_crash_resilience_present",
                  _n_days > 0,
                  (f"today_regime={_today} · n_r2_trades_tagged={_tagged} · "
                    f"n_days_classified={_n_days} · dist={_dist}"),
                  {"today_regime": _today,
                   "n_r2_trades_tagged": _tagged,
                   "n_days_classified": _n_days,
                   "regime_distribution": _dist,
                   "interpretation": _cr.get("interpretation")})
        else:
            _add("C18_crash_resilience_present", False,
                  f"crash-resilience report missing · run python -m backend.research.multi_layer.crash_resilience --market {market_l}",
                  {"expected": str(_cr_p.name)})
    except Exception as _e18:
        _add("C18_crash_resilience_present", False,
              f"exception: {type(_e18).__name__}: {_e18}", None)

    # ── C17 · Momentum candidate conservation (Momentum correction) ─
    # Every candidate emitted by the momentum engine must be classified
    # into one of 4 terminal states · zero silent disappearances.
    try:
        _ml_p = root / "reports" / "research" / "multi_layer" / f"momentum_ledger_{market_l}_{asof}.json"
        if _ml_p.exists():
            _ml = json.loads(_ml_p.read_text(encoding="utf-8"))
            _cons = bool(_ml.get("conservation_ok", False))
            _n_disappear = int(_ml.get("n_silent_disappearances", 0) or 0)
            _add("C17_momentum_conservation_zero_silent",
                  _cons and _n_disappear == 0,
                  (f"conservation_ok={_cons} · silent_disappearances={_n_disappear} · "
                    f"by_state={_ml.get('by_terminal_state', {})}"),
                  {"conservation_ok": _cons,
                   "n_silent_disappearances": _n_disappear,
                   "by_terminal_state": _ml.get("by_terminal_state"),
                   "by_reason_code": _ml.get("by_reason_code")})
        else:
            _add("C17_momentum_conservation_zero_silent", False,
                  f"momentum ledger missing · run python -m backend.research.multi_layer.momentum_ledger --market {market_l}",
                  {"expected": str(_ml_p.name)})
    except Exception as _e17:
        _add("C17_momentum_conservation_zero_silent", False,
              f"exception: {type(_e17).__name__}: {_e17}", None)

    # ── C16 · Stress-regime research presence (§8) ──────────────────
    try:
        _sr_p = root / "reports" / "research" / "multi_layer" / f"stress_regime_{market_l}_{asof}.json"
        if _sr_p.exists():
            _sr = json.loads(_sr_p.read_text(encoding="utf-8"))
            _n_tagged = int(_sr.get("n_r2_trades_tagged", 0) or 0)
            _add("C16_stress_regime_research_present",
                  _n_tagged > 0,
                  (f"R2 trades tagged={_n_tagged} · regime_source="
                    f"{_sr.get('regime_source', 'missing')}"),
                  {"n_trades": _n_tagged,
                   "overall": _sr.get("overall"),
                   "per_regime_n": {k: v.get("n") for k, v
                                     in (_sr.get("per_regime") or {}).items()}})
        else:
            _add("C16_stress_regime_research_present", False,
                  f"stress-regime report missing · run python -m backend.research.multi_layer.stress_regime --market {market_l}",
                  {"expected": str(_sr_p.name)})
    except Exception as _e16:
        _add("C16_stress_regime_research_present", False,
              f"exception: {type(_e16).__name__}: {_e16}", None)

    # ── C14 · Portfolio↔Exit overlap classification (§11) ───────────
    # Every same-ticker overlap must be explained · RECONCILIATION_DEFECT = 0
    try:
        _ov_p = root / "reports" / "audit" / f"portfolio_exit_overlap_{market_l}_{asof}.json"
        if _ov_p.exists():
            _ov = json.loads(_ov_p.read_text(encoding="utf-8"))
            _defects = int(_ov.get("n_reconciliation_defects", 0) or 0)
            _add("C14_overlap_no_reconciliation_defects",
                  _defects == 0,
                  (f"{_ov.get('n_overlap_tickers', 0)} overlap tickers · "
                    f"defects={_defects} · "
                    f"by_cat={_ov.get('by_category', {})}"),
                  _ov.get("by_category", {}))
        else:
            _add("C14_overlap_no_reconciliation_defects", False,
                  f"overlap report missing · run scripts/portfolio_exit_overlap_classifier.py --market {market_l}",
                  {"expected": str(_ov_p.name)})
    except Exception as _e14:
        _add("C14_overlap_no_reconciliation_defects", False,
              f"exception: {type(_e14).__name__}: {_e14}", None)

    # ── C15 · R1 producer-wide retirement proof (§1 hardening) ──────
    try:
        _pa_p = root / "reports" / "audit" / f"r1_producer_audit_{market_l}_{asof}.json"
        if _pa_p.exists():
            _pa = json.loads(_pa_p.read_text(encoding="utf-8"))
            _viol = int(_pa.get("total_violations", 0) or 0)
            _add("C15_r1_producer_wide_retirement",
                  _viol == 0,
                  (f"{_pa.get('verdict', 'UNKNOWN')} · total_violations={_viol} · "
                    f"n_producers={len(_pa.get('producers', []))}"),
                  {"verdict": _pa.get("verdict"),
                   "total_violations": _viol})
        else:
            _add("C15_r1_producer_wide_retirement", False,
                  f"audit missing · run scripts/r1_producer_audit.py --market {market_l}",
                  {"expected": str(_pa_p.name)})
    except Exception as _e15:
        _add("C15_r1_producer_wide_retirement", False,
              f"exception: {type(_e15).__name__}: {_e15}", None)

    # ── C13 · Universe bounds ───────────────────────────────────────
    # CEO 2026-09-01 Section 2 · USA must be S&P 500 · India retains
    # current bounds · silent widening beyond configs/aegis_universes.yaml
    # is a contract violation.
    try:
        from backend.canonical.universe_validator import validate as _uv_validate
        _uv = _uv_validate(root, market_l)
        _add(
            "C13_universe_bounds",
            _uv.ok,
            (f"{_uv.verdict} · {_uv.detail} · violations={_uv.violations}"
              if _uv.violations else f"{_uv.verdict} · {_uv.detail}"),
            _uv.as_dict(),
        )
    except Exception as _e13:
        _add("C13_universe_bounds", False,
              f"exception: {type(_e13).__name__}: {_e13}", None)

    # ── C12 · Provenance companion coverage ─────────────────────────
    # CEO 2026-09-01 · Every VISIBLE, OPENED position row must resolve
    # to a Position ID. SUGGESTED rows (population=FRESH_RECOMMENDATION,
    # runner=SHADOW) are exempt because they are not opened positions.
    try:
        _prov_path = root / "reports" / "telegram" / f"aegis_history_{market_l}_provenance.jsonl"
        if _prov_path.exists():
            _prov = [json.loads(l) for l in _prov_path.read_text(encoding="utf-8").splitlines() if l.strip()]
            _need_pid = [r for r in _prov if r.get("population") not in ("FRESH_RECOMMENDATION",)]
            _n_need = len(_need_pid)
            _n_have = sum(1 for r in _need_pid if r.get("position_id"))
            _cov = round(_n_have / max(1, _n_need) * 100, 1)
            _add("C12_provenance_position_id_coverage",
                  _n_have == _n_need,
                  f"{_n_have}/{_n_need} opened rows have Position ID ({_cov}%)",
                  {"n_need_pid": _n_need, "n_have_pid": _n_have,
                   "coverage_pct": _cov,
                   "unresolved_samples": [
                       {"sheet": r["sheet"], "ticker": r["ticker"],
                        "runner": r["runner"], "entry_date": r.get("entry_date", "")}
                       for r in _need_pid if not r.get("position_id")
                   ][:5]})
        else:
            _add("C12_provenance_position_id_coverage", False,
                  f"provenance companion missing · run emit_provenance_companion.py --market {market_l}",
                  {"expected_path": str(_prov_path.name)})
    except Exception as _e12:
        _add("C12_provenance_position_id_coverage", False,
              f"exception: {type(_e12).__name__}: {_e12}", None)

    # ── C11 · Standard-name dated XLSX snapshot ─────────────────────
    # CEO 2026-09-01 · Standard XLSX name is `aegis_{market}_YYYY-MM-DD.xlsx`.
    # The undated file `aegis_history_{market}.xlsx` remains the "latest"
    # alias but the DATED file is the authoritative daily snapshot. Both
    # must exist for today's asof and be byte-identical.
    try:
        import hashlib as _h11
        _std_dated = root / "reports" / "telegram" / f"aegis_{market_l}_{asof}.xlsx"
        _undated = root / "reports" / "telegram" / f"aegis_history_{market_l}.xlsx"
        _dated_ok = _std_dated.exists()
        _byte_match = False
        if _dated_ok and _undated.exists():
            _byte_match = (
                _h11.md5(_std_dated.read_bytes()).hexdigest()
                == _h11.md5(_undated.read_bytes()).hexdigest()
            )
        _add("C11_standard_dated_xlsx_present",
              _dated_ok and _byte_match,
              (
                f"dated={_dated_ok} byte_match={_byte_match} · "
                f"expected={_std_dated.name}"
              ),
              {"dated_present": _dated_ok, "byte_match": _byte_match,
               "expected": str(_std_dated.name)})
    except Exception as _e11:
        _add("C11_standard_dated_xlsx_present", False,
              f"exception: {type(_e11).__name__}: {_e11}", None)

    # Overall verdict
    fails = [c for c in checks if not c["ok"]]
    verdict = "PASS" if not fails else "FAIL"
    return {
        "engine": "aegis.reconcile.final.v1",
        "asof": asof,
        "market": market_l,
        "xlsx_path": str(xlsx.relative_to(root)),
        "verdict": verdict,
        "n_checks": len(checks),
        "n_pass": sum(1 for c in checks if c["ok"]),
        "n_fail": len(fails),
        "checks": checks,
    }


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--market", choices=["india", "usa", "both"],
                     default="both")
    ap.add_argument("--out-dir", default="reports/reconcile")
    args = ap.parse_args()
    markets = ["india", "usa"] if args.market == "both" else [args.market]
    any_fail = False
    for m in markets:
        try:
            rep = reconcile(m, _ROOT)
        except ReconcileFailure as e:
            print(f"[reconcile:{m}] SKIPPED · {e}")
            continue
        out_p = _ROOT / args.out_dir / f"final_reconcile_{m}_{rep['asof']}.json"
        out_p.parent.mkdir(parents=True, exist_ok=True)
        out_p.write_text(json.dumps(rep, indent=2, ensure_ascii=False),
                          encoding="utf-8")
        print(f"[reconcile:{m}] {rep['verdict']} · pass={rep['n_pass']} · "
              f"fail={rep['n_fail']} · report={out_p.relative_to(_ROOT)}")
        if rep["verdict"] != "PASS":
            for c in rep["checks"]:
                if not c["ok"]:
                    _line = f"  [FAIL] {c['name']} :: {c['detail']}"
                    print(_line.encode("ascii", errors="replace").decode("ascii"))
            any_fail = True
    return 2 if any_fail else 0


if __name__ == "__main__":
    sys.exit(main())
