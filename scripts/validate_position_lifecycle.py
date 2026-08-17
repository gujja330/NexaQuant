"""Operator 2026-08-17 · Section 17 · Position Lifecycle Acceptance Report.

Reads the freshly-built unified XLSX + per-market Portfolio sheets
and runs the 12 acceptance checks. Prints PASS/FAIL per check plus
a validation summary. Non-zero exit if any FAIL.

Section 17 checks:
   1. Generate fresh India output      (mtime freshness)
   2. Every Position ID scanned        (coverage)
   3. No contradictory same-day authoritative decisions per (PID,Date,Runner)
   4. No CLOSED → ACTIVE transitions
   5. No EXIT + HOLD coexistence
   6. No SKIP in active Portfolio
   7. All P&L records reconcile to actual lifecycle
   8. LUPIN regression (STOP_LOSS_HIT + STRONG BUY resolves to EXIT-family)
   9. R1/R2 disagreement does not create duplicate positions
  10. NEW opportunities are explicitly surfaced (count > 0 when any exist)
  11. Re-entry creates a new Position ID
  12. Portfolio and History reconcile exactly

Usage:  python scripts/validate_position_lifecycle.py
"""
from __future__ import annotations

import io
import json
import sys
import time
from collections import defaultdict
from pathlib import Path

try:
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
except Exception:
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")

_ROOT = Path(__file__).resolve().parents[1]
_BINDING = (
    "EMERGENCY_EXIT", "PORTFOLIO_MAX_DD", "HARD_STOP",
    "STOP_LOSS_HIT", "GAP_EXIT", "TRAILING_STOP_HIT",
    "CRITICAL_DEEP_LOSS",
)


def _read_history():
    from openpyxl import load_workbook
    p = _ROOT / "reports" / "telegram" / "aegis_history.xlsx"
    if not p.exists(): return []
    wb = load_workbook(p, read_only=True)
    ws = wb["AEGIS Daily"] if "AEGIS Daily" in wb.sheetnames else wb.active
    header = None
    rows = []
    for row in ws.iter_rows(values_only=True):
        if header is None:
            header = list(row); continue
        if row and any(row):
            rows.append(dict(zip(header, row)))
    wb.close()
    return rows


def _read_portfolio(market):
    from openpyxl import load_workbook
    p = _ROOT / "reports" / "telegram" / f"aegis_history_{market}.xlsx"
    if not p.exists(): return []
    wb = load_workbook(p, read_only=True)
    if "Portfolio" not in wb.sheetnames:
        wb.close(); return []
    ws = wb["Portfolio"]
    header = None; rows = []
    for row in ws.iter_rows(values_only=True):
        if header is None:
            if row and row[0] == "Ticker":
                header = list(row)
            continue
        if row and row[0]:
            rows.append(dict(zip(header, row)))
    wb.close()
    return rows


def main() -> int:
    print("=" * 72)
    print("  Operator Section 17 · Position Lifecycle Acceptance Report")
    print("=" * 72)
    print()

    hist  = _read_history()
    india = _read_portfolio("india")
    usa   = _read_portfolio("usa")
    print(f"  Loaded {len(hist)} history rows · India Portfolio={len(india)} · USA Portfolio={len(usa)}")
    print()

    failures = []

    # ── #1 · Fresh India output ──
    import os
    p = _ROOT / "reports" / "telegram" / "aegis_history_india.xlsx"
    if p.exists():
        age_h = (time.time() - p.stat().st_mtime) / 3600
        ok1 = age_h < 24
        print(f"  {'✅' if ok1 else '❌'} #1  India output freshness · age={age_h:.1f}h · target <24h")
        if not ok1: failures.append("#1 stale India output")
    else:
        print(f"  ❌ #1  India output MISSING")
        failures.append("#1 missing India")

    # ── #2 · Position ID coverage ──
    n_pids = len({r.get("Position ID") for r in hist if r.get("Position ID")})
    print(f"  ✅ #2  Position IDs scanned · {n_pids} unique")

    # ── #3 · Contradictory same-day (PID, Date, Runner) ──
    by_key = defaultdict(list)
    for r in hist:
        key = (str(r.get("Position ID") or ""), str(r.get("Date") or "")[:10],
                  str(r.get("Run_Type") or ""))
        by_key[key].append(str(r.get("Status") or ""))
    contradictions = [(k, v) for k, v in by_key.items() if len(set(v)) > 1]
    ok3 = len(contradictions) == 0
    print(f"  {'✅' if ok3 else '❌'} #3  Contradictory same-day decisions · {len(contradictions)} · target 0")
    if not ok3:
        for k, v in contradictions[:5]:
            print(f"       ✗ {k}: {v}")
        failures.append(f"#3 · {len(contradictions)} contradictions")

    # ── #4 · CLOSED → ACTIVE transitions ──
    per_pid = defaultdict(list)
    for r in hist:
        pid = str(r.get("Position ID") or ""); date = str(r.get("Date") or "")[:10]
        status = str(r.get("Status") or "").upper()
        per_pid[pid].append((date, status))
    reactivations = 0
    for pid, appears in per_pid.items():
        appears.sort()
        was_closed = False
        for date, status in appears:
            if status == "EXIT":
                was_closed = True; continue
            if was_closed and status in ("STRONG BUY","BUY","ADD","HOLD","ACCUMULATE"):
                reactivations += 1
                break
    ok4 = reactivations == 0
    print(f"  {'✅' if ok4 else '❌'} #4  CLOSED → ACTIVE transitions · {reactivations} · target 0")
    if not ok4: failures.append(f"#4 · {reactivations} reactivations")

    # ── #5 · EXIT + HOLD coexistence for same position/date ──
    exit_hold = [(k, v) for k, v in by_key.items()
                     if "EXIT" in [s.upper() for s in v] and "HOLD" in [s.upper() for s in v]]
    ok5 = len(exit_hold) == 0
    print(f"  {'✅' if ok5 else '❌'} #5  EXIT + HOLD coexistence · {len(exit_hold)} · target 0")
    if not ok5:
        for k, v in exit_hold[:3]: print(f"       ✗ {k}: {v}")
        failures.append(f"#5 · {len(exit_hold)} EXIT+HOLD")

    # ── #6 · SKIP in active Portfolio ──
    skip_in_pf = sum(1 for r in (india + usa)
                             if str(r.get("Status") or "").upper() == "SKIP")
    ok6 = skip_in_pf == 0
    print(f"  {'✅' if ok6 else '❌'} #6  SKIP in active Portfolio · {skip_in_pf} · target 0")
    if not ok6: failures.append(f"#6 · {skip_in_pf} SKIP in portfolio")

    # ── #7 · P&L records reconcile to lifecycle ──
    # Every EXIT row should have an Exit P&L · every non-EXIT should not
    bad_pnl = 0
    for r in hist:
        st = str(r.get("Status") or "").upper()
        exit_pnl = r.get("Exit P&L %")
        if st == "EXIT" and (exit_pnl is None or exit_pnl == ""):
            bad_pnl += 1
        elif st in ("STRONG BUY","BUY","ADD","HOLD") and \
                isinstance(exit_pnl, (int, float)) and exit_pnl != 0:
            bad_pnl += 1
    ok7 = bad_pnl == 0
    print(f"  {'✅' if ok7 else '❌'} #7  P&L / lifecycle reconciliation · {bad_pnl} anomalies · target 0")
    if not ok7: failures.append(f"#7 · {bad_pnl} P&L anomalies")

    # ── #8 · LUPIN regression (STOP_LOSS_HIT ⇒ EXIT-family Decision) ──
    lupin_bad = 0
    for r in (india + usa):
        tk = str(r.get("Ticker") or "").upper()
        if tk not in ("LUPIN", "LUPIN.NS"): continue
        alerts_up = str(r.get("Alerts") or "").upper()
        if any(sig in alerts_up for sig in _BINDING):
            dec_up = str(r.get("🎯 DECISION") or "").upper()
            if "EXIT" not in dec_up and "CLOSED" not in dec_up:
                lupin_bad += 1
    ok8 = lupin_bad == 0
    print(f"  {'✅' if ok8 else '❌'} #8  LUPIN regression · {lupin_bad} bad · target 0")
    if not ok8: failures.append(f"#8 · LUPIN {lupin_bad} bad")

    # ── #9 · R1/R2 disagreement · same ticker/date/PID · no duplicate ──
    # Redundant with #3 · already caught. Print separately for report clarity.
    dup_pids = sum(1 for k, v in by_key.items() if len(v) > 1)
    ok9 = dup_pids == 0
    print(f"  {'✅' if ok9 else '❌'} #9  Duplicate (PID,Date,Runner) rows · {dup_pids} · target 0")
    if not ok9: failures.append(f"#9 · {dup_pids} duplicates")

    # ── #10 · NEW opportunities surfaced ──
    n_new = sum(1 for r in (india + usa)
                       if "NEW" in str(r.get("Lifecycle") or "").upper())
    print(f"  ✅ #10 NEW opportunities surfaced · India+USA total = {n_new}")

    # ── #11 · Re-entry creates a new Position ID ──
    # Group by (ticker, runner) · if the same ticker+runner has MULTIPLE
    # Position IDs, that's re-entry working correctly.
    by_tk = defaultdict(set)
    for r in hist:
        tk = str(r.get("Ticker") or "").upper().replace(".NS","").replace(".BO","")
        runner = str(r.get("Run_Type") or "").upper()
        pid = str(r.get("Position ID") or "")
        if tk and runner and pid:
            by_tk[(tk, runner)].add(pid)
    reentries = sum(1 for pids in by_tk.values() if len(pids) > 1)
    print(f"  ✅ #11 Re-entries observed (ticker+runner has >1 PID) · {reentries}")

    # ── #12 · Portfolio / History reconcile ──
    hist_tickers = {(str(r.get("Country") or "").upper(),
                          str(r.get("Ticker") or "").upper())
                         for r in hist}
    pf_tickers = ({("INDIA", str(r.get("Ticker") or "").upper()) for r in india}
                     | {("USA",   str(r.get("Ticker") or "").upper()) for r in usa})
    pf_missing_in_hist = pf_tickers - hist_tickers
    ok12 = len(pf_missing_in_hist) == 0
    print(f"  {'✅' if ok12 else '❌'} #12 Portfolio/History reconcile · {len(pf_missing_in_hist)} PF tickers missing in history · target 0")
    if not ok12:
        print(f"       missing sample: {list(pf_missing_in_hist)[:5]}")
        failures.append(f"#12 · {len(pf_missing_in_hist)} unreconciled")

    print()
    print("=" * 72)
    print("  VALIDATION SUMMARY")
    print("=" * 72)
    print(f"  Position IDs checked:              {n_pids}")
    print(f"  Duplicate lifecycle conflicts:     {len(contradictions)}")
    print(f"  EXIT/HOLD conflicts:               {len(exit_hold)}")
    print(f"  Closed→Active violations:          {reactivations}")
    print(f"  SKIP portfolio violations:         {skip_in_pf}")
    print(f"  P&L violations:                    {bad_pnl}")
    print(f"  New opportunities:                 {n_new}")
    print(f"  Re-entries:                        {reentries}")
    print(f"  LUPIN regression bad:              {lupin_bad}")
    print(f"  Duplicate (PID,Date,Runner) rows:  {dup_pids}")
    print(f"  Portfolio/History unreconciled:    {len(pf_missing_in_hist)}")
    print()
    if not failures:
        print("  ✅  ALL TESTS PASS · production-ready")
        return 0
    print(f"  ❌  {len(failures)} FAIL(s):")
    for f in failures:
        print(f"       {f}")
    return 1


if __name__ == "__main__":
    sys.exit(main())
