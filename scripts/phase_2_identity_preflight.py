"""AEGIS · Phase 2 preflight · classify every legacy PID as DETERMINISTIC,
AMBIGUOUS, INVALID or ALREADY_CANONICAL against the canonical Registry
authority.

Reads (READ-ONLY):
  · reports/telegram/aegis_history.xlsx        (source XLSX rows)
  · reports/telegram/aegis_history_india.xlsx  (per-market)
  · reports/telegram/aegis_history_usa.xlsx    (per-market)
  · reports/research/outcome_dataset.parquet   (research canonical dataset)
  · reports/research/opportunity_registry.jsonl (authority)

Writes:
  · reports/migration/pid_migration_manifest_YYYYMMDD.jsonl    (all rows)
  · reports/migration/pid_migration_ambiguous_YYYYMMDD.jsonl  (safety-stop rows only)
  · reports/migration/pid_migration_preflight_summary_YYYYMMDD.json

Exit codes:
  0 · all rows DETERMINISTIC or ALREADY_CANONICAL · safe to execute migration
  2 · one or more AMBIGUOUS/INVALID rows · HARD STOP · migration WILL NOT proceed
  1 · unexpected error
"""
from __future__ import annotations

import hashlib
import json
import re
import sys
from collections import Counter, defaultdict
from datetime import date
from pathlib import Path

_ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(_ROOT))

from backend.research.opportunity_registry import make_opportunity_id  # noqa: E402


_CANONICAL_RE = re.compile(r"^(IND|USA)-(R1|R2|R3|SHADOW|MOMENTUM)-([A-Z0-9]+)-(\d{8})-([0-9a-f]{6})$")
_EMITTER2_RE = re.compile(r"^(R1|R2|R3|SHADOW|MOMENTUM)-([A-Z0-9]+)-(IND|USA)-(\d{8})-([0-9a-f]{6})$")
_LEGACY_RE = re.compile(r"^([A-Z0-9]+)_([A-Z]{3})_(\d{8})$")
_MARKET_FROM_TAG = {"IND": "india", "USA": "usa"}
_VALID_RUNNERS = {"R1", "R2", "R3", "SHADOW", "MOMENTUM"}


def _load_registry_index(root: Path) -> dict:
    """Build (market, runner, ticker) → set of Registry PIDs (with their
    created_date). Registry is authoritative."""
    idx = defaultdict(list)   # (mkt, run, tk) → [(created_date, opportunity_id)]
    p = root / "reports" / "research" / "opportunity_registry.jsonl"
    for line in p.read_text(encoding="utf-8", errors="replace").splitlines():
        if not line.strip(): continue
        try:
            o = json.loads(line)
        except Exception:
            continue
        mkt = str(o.get("market", "")).lower()
        run = str(o.get("runner", "")).upper()
        tk = str(o.get("ticker", "")).upper().replace(".NS", "").replace(".BO", "")
        cd = str(o.get("created_date", ""))[:10]
        pid = str(o.get("opportunity_id", ""))
        if mkt and run and tk and cd and pid:
            idx[(mkt, run, tk)].append((cd, pid))
    # Sort each bucket by created_date; keep latest event's opportunity_id
    # per (market, runner, ticker, created_date) via dedup on (cd, pid).
    out = {}
    for key, entries in idx.items():
        dedup = list({(cd, pid) for cd, pid in entries})
        out[key] = sorted(dedup)
    return out


def _classify_row(pid: str, market: str, runner: str, ticker: str,
                    row_date: str, reg_idx: dict) -> dict:
    """Classify a single row's Position ID.

    Returns dict with keys: classification, canonical_pid, authority,
    reason, registry_created_date.
    """
    tk_bare = str(ticker or "").upper().replace(".NS", "").replace(".BO", "")
    run_u = str(runner or "").upper().replace("_NEW", "").strip()
    mkt_l = str(market or "").lower()

    # Sanity
    if run_u not in _VALID_RUNNERS:
        return {"classification": "AMBIGUOUS",
                "reason": f"runner '{run_u}' not in valid set",
                "canonical_pid": None, "authority": None,
                "registry_created_date": None}
    if mkt_l not in ("india", "usa"):
        return {"classification": "INVALID",
                "reason": f"market '{mkt_l}' not india/usa",
                "canonical_pid": None, "authority": None,
                "registry_created_date": None}
    if not tk_bare:
        return {"classification": "INVALID",
                "reason": "ticker empty",
                "canonical_pid": None, "authority": None,
                "registry_created_date": None}

    # Look up Registry authority
    reg_matches = reg_idx.get((mkt_l, run_u, tk_bare), [])

    # Case A · row already canonical
    m_can = _CANONICAL_RE.match(str(pid or ""))
    if m_can:
        pid_mkt = m_can.group(1)
        pid_run = m_can.group(2)
        pid_tk = m_can.group(3)
        pid_ds = m_can.group(4)
        pid_iso = f"{pid_ds[:4]}-{pid_ds[4:6]}-{pid_ds[6:8]}"
        # Is this actually the Registry-authoritative PID?
        if len(reg_matches) == 0:
            return {"classification": "ALREADY_CANONICAL",
                    "canonical_pid": pid, "authority": "canonical_format · no Registry match",
                    "reason": "Row has canonical format but no Registry entry to confirm",
                    "registry_created_date": None}
        # Registry has entries for this (market, runner, ticker)
        # If EXACTLY one Registry match, it's the authoritative canonical PID
        if len(reg_matches) == 1:
            reg_cd, reg_pid = reg_matches[0]
            if reg_pid == pid:
                return {"classification": "ALREADY_CANONICAL",
                        "canonical_pid": pid, "authority": "registry_match",
                        "reason": "Matches Registry authority exactly",
                        "registry_created_date": reg_cd}
            # PID canonical format but doesn't match Registry authority · MUST migrate
            return {"classification": "DETERMINISTIC",
                    "canonical_pid": reg_pid, "authority": "registry_lookup",
                    "reason": f"Row PID '{pid}' has canonical format but disagrees with "
                              f"Registry authority '{reg_pid}' (created={reg_cd})",
                    "registry_created_date": reg_cd}
        # Multiple Registry entries for same (market, runner, ticker)
        # Match by embedded date if possible
        for reg_cd, reg_pid in reg_matches:
            if reg_pid == pid:
                return {"classification": "ALREADY_CANONICAL",
                        "canonical_pid": pid, "authority": "registry_match_multi",
                        "reason": f"Matches one of {len(reg_matches)} Registry entries",
                        "registry_created_date": reg_cd}
        # None match · ambiguous
        return {"classification": "AMBIGUOUS",
                "canonical_pid": None, "authority": None,
                "reason": f"Row PID '{pid}' does not match any of "
                          f"{len(reg_matches)} Registry entries for this (mkt,run,tk)",
                "registry_created_date": None}

    # Case B · Emitter 2 format (RUNNER-TICKER-MKT-DATE-hash) · non-canonical
    m_e2 = _EMITTER2_RE.match(str(pid or ""))
    if m_e2:
        e2_ds = m_e2.group(4)
        e2_iso = f"{e2_ds[:4]}-{e2_ds[4:6]}-{e2_ds[6:8]}"
        if len(reg_matches) == 0:
            # No Registry authority · but the row itself has a runner + date
            # · we can compute canonical from the row's own data
            return {"classification": "DETERMINISTIC",
                    "canonical_pid": make_opportunity_id(mkt_l, run_u, tk_bare, e2_iso),
                    "authority": "row_self_derived · no Registry match",
                    "reason": f"Emitter-2 format · no Registry entry · derive canonical "
                              f"from row's own date '{e2_iso}'",
                    "registry_created_date": None}
        if len(reg_matches) == 1:
            reg_cd, reg_pid = reg_matches[0]
            return {"classification": "DETERMINISTIC",
                    "canonical_pid": reg_pid, "authority": "registry_lookup",
                    "reason": f"Emitter-2 format · single Registry match · "
                              f"use Registry PID (created={reg_cd})",
                    "registry_created_date": reg_cd}
        # Multiple Registry matches · pick the one closest to row's embedded date
        # (row's date is the OBSERVED date · not necessarily created_date)
        # If exactly one Registry match has created_date <= row's date, use that
        candidates = [(cd, pid_r) for cd, pid_r in reg_matches if cd <= e2_iso]
        if len(candidates) == 1:
            reg_cd, reg_pid = candidates[0]
            return {"classification": "DETERMINISTIC",
                    "canonical_pid": reg_pid, "authority": "registry_lookup_dated",
                    "reason": f"Emitter-2 format · multiple Registry entries · "
                              f"one has created_date <= row observation date",
                    "registry_created_date": reg_cd}
        return {"classification": "AMBIGUOUS",
                "canonical_pid": None, "authority": None,
                "reason": f"Emitter-2 format · {len(reg_matches)} Registry entries · "
                          f"cannot deterministically pick which lifecycle instance",
                "registry_created_date": None}

    # Case C · Legacy TICKER_MKT_DATE format · pre-A1
    m_leg = _LEGACY_RE.match(str(pid or ""))
    if m_leg:
        leg_ds = m_leg.group(3)
        leg_iso = f"{leg_ds[:4]}-{leg_ds[4:6]}-{leg_ds[6:8]}"
        if len(reg_matches) == 0:
            return {"classification": "DETERMINISTIC",
                    "canonical_pid": make_opportunity_id(mkt_l, run_u, tk_bare, leg_iso),
                    "authority": "row_self_derived · no Registry match",
                    "reason": f"Legacy format · no Registry · derive from legacy date '{leg_iso}'",
                    "registry_created_date": None}
        if len(reg_matches) == 1:
            reg_cd, reg_pid = reg_matches[0]
            return {"classification": "DETERMINISTIC",
                    "canonical_pid": reg_pid, "authority": "registry_lookup",
                    "reason": f"Legacy format · single Registry match · use Registry PID",
                    "registry_created_date": reg_cd}
        candidates = [(cd, pid_r) for cd, pid_r in reg_matches if cd <= leg_iso]
        if len(candidates) == 1:
            reg_cd, reg_pid = candidates[0]
            return {"classification": "DETERMINISTIC",
                    "canonical_pid": reg_pid, "authority": "registry_lookup_dated",
                    "reason": f"Legacy format · multi Registry · one predates legacy date",
                    "registry_created_date": reg_cd}
        return {"classification": "AMBIGUOUS",
                "canonical_pid": None, "authority": None,
                "reason": f"Legacy format · {len(reg_matches)} Registry entries · "
                          f"cannot pick lifecycle instance",
                "registry_created_date": None}

    # Case D · unrecognized format
    return {"classification": "INVALID",
            "canonical_pid": None, "authority": None,
            "reason": f"Unrecognized PID format: '{pid}'",
            "registry_created_date": None}


def _scan_xlsx(root: Path, xlsx_rel: str, sheet_name: str,
                reg_idx: dict, market_hint: str = "") -> list:
    """Yield classification dicts for each PID-bearing row in the XLSX."""
    from openpyxl import load_workbook
    xlsx_p = root / xlsx_rel
    if not xlsx_p.exists():
        return []
    wb = load_workbook(xlsx_p, read_only=True, data_only=True)
    if sheet_name not in wb.sheetnames:
        wb.close()
        return []
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    hdr = rows[0]
    def col(name):
        for i, c in enumerate(hdr):
            if c and str(c).strip().lower() == name.lower():
                return i
        return None
    c_pid = col("Position ID")
    c_ctry = col("Country") or col("Market")
    c_run = col("Run_Type") or col("Runner")
    c_tk = col("Ticker")
    c_date = col("Date")
    if c_pid is None:
        wb.close()
        return []
    out = []
    for i, r in enumerate(rows[1:], start=2):
        if not r[c_pid]: continue
        pid = str(r[c_pid])
        mkt = str(r[c_ctry] or market_hint).lower() if c_ctry is not None else market_hint
        run = str(r[c_run] or "").upper() if c_run is not None else ""
        tk = str(r[c_tk] or "").upper() if c_tk is not None else ""
        dt = str(r[c_date])[:10] if c_date is not None and r[c_date] else ""
        cls = _classify_row(pid, mkt, run, tk, dt, reg_idx)
        cls.update({
            "file": xlsx_rel,
            "sheet": sheet_name,
            "row": i,
            "market": mkt,
            "runner": run,
            "ticker": tk,
            "row_date": dt,
            "legacy_pid": pid,
        })
        out.append(cls)
    wb.close()
    return out


def _scan_parquet(root: Path, parquet_rel: str, reg_idx: dict) -> list:
    """Yield classification dicts for each row of outcome_dataset.parquet."""
    parquet_p = root / parquet_rel
    if not parquet_p.exists():
        return []
    import pandas as pd
    df = pd.read_parquet(parquet_p)
    out = []
    for idx, row in df.iterrows():
        pid = str(row.get("position_id", ""))
        mkt = str(row.get("country", "")).lower() or str(row.get("market", "")).lower()
        run = str(row.get("runner", "")).upper()
        tk = str(row.get("ticker", "")).upper()
        dt = str(row.get("entry_date", ""))[:10]
        cls = _classify_row(pid, mkt, run, tk, dt, reg_idx)
        cls.update({
            "file": parquet_rel,
            "sheet": None,
            "row": int(idx) + 1,
            "market": mkt,
            "runner": run,
            "ticker": tk,
            "row_date": dt,
            "legacy_pid": pid,
        })
        out.append(cls)
    return out


def main() -> int:
    print("[preflight] loading Registry authority...")
    reg_idx = _load_registry_index(_ROOT)
    print(f"[preflight] Registry: {len(reg_idx)} distinct (market, runner, ticker) keys")
    print()

    all_rows = []
    print("[preflight] scanning aegis_history.xlsx (unified) · AEGIS Daily sheet...")
    all_rows.extend(_scan_xlsx(_ROOT,
        "reports/telegram/aegis_history.xlsx", "AEGIS Daily", reg_idx))
    print("[preflight] scanning aegis_history_india.xlsx · AEGIS INDIA History...")
    all_rows.extend(_scan_xlsx(_ROOT,
        "reports/telegram/aegis_history_india.xlsx", "AEGIS INDIA History",
        reg_idx, market_hint="india"))
    print("[preflight] scanning aegis_history_usa.xlsx · AEGIS USA History...")
    all_rows.extend(_scan_xlsx(_ROOT,
        "reports/telegram/aegis_history_usa.xlsx", "AEGIS USA History",
        reg_idx, market_hint="usa"))
    print("[preflight] scanning outcome_dataset.parquet...")
    all_rows.extend(_scan_parquet(_ROOT,
        "reports/research/outcome_dataset.parquet", reg_idx))
    print(f"[preflight] total rows scanned: {len(all_rows)}")
    print()

    counts = Counter(r["classification"] for r in all_rows)
    print("[preflight] classification distribution:")
    for k in ("ALREADY_CANONICAL", "DETERMINISTIC", "AMBIGUOUS", "INVALID"):
        print(f"  {k:20s} {counts.get(k, 0)}")

    ambiguous = [r for r in all_rows if r["classification"] in ("AMBIGUOUS", "INVALID")]

    # Write manifest
    out_dir = _ROOT / "reports" / "migration"
    out_dir.mkdir(parents=True, exist_ok=True)
    today = date.today().isoformat().replace("-", "")
    manifest_p = out_dir / f"pid_migration_manifest_{today}.jsonl"
    with manifest_p.open("w", encoding="utf-8") as f:
        for r in all_rows:
            f.write(json.dumps(r, ensure_ascii=False, sort_keys=True, default=str) + "\n")
    print(f"\n[preflight] manifest: {manifest_p.relative_to(_ROOT)}")

    if ambiguous:
        ambig_p = out_dir / f"pid_migration_ambiguous_{today}.jsonl"
        with ambig_p.open("w", encoding="utf-8") as f:
            for r in ambiguous:
                f.write(json.dumps(r, ensure_ascii=False, sort_keys=True, default=str) + "\n")
        print(f"[preflight] ambiguous rows: {ambig_p.relative_to(_ROOT)}")

    # Summary
    summary = {
        "engine": "aegis.migration.phase_2_preflight",
        "asof": date.today().isoformat(),
        "counts": dict(counts),
        "total_rows_scanned": len(all_rows),
        "manifest_path": str(manifest_p.relative_to(_ROOT)),
        "verdict": "SAFE_TO_EXECUTE" if not ambiguous else "HARD_STOP",
    }
    summary_p = out_dir / f"pid_migration_preflight_summary_{today}.json"
    summary_p.write_text(json.dumps(summary, indent=2, ensure_ascii=False),
                          encoding="utf-8")
    print(f"[preflight] summary:  {summary_p.relative_to(_ROOT)}")

    print()
    if ambiguous:
        print(f"[preflight] HARD_STOP · {len(ambiguous)} AMBIGUOUS/INVALID rows · "
              "review ambiguous file before migration")
        return 2
    print("[preflight] SAFE_TO_EXECUTE · no ambiguity · migration may proceed")
    return 0


if __name__ == "__main__":
    sys.exit(main())
