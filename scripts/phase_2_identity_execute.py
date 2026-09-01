"""AEGIS · Phase 2 execute · apply DETERMINISTIC PID migrations from the
preflight manifest to each affected artifact.

Rules (per docs/AEGIS/PHASE_2_IDENTITY_MIGRATION.md):
  · Apply only rows classified DETERMINISTIC · never AMBIGUOUS/INVALID
  · Preserve every historical row · zero deletion · zero dedup
  · Preserve `legacy_position_id` (populate before overwriting position_id)
  · Preserve all non-identity fields
  · Idempotent · rows already at canonical PID are skipped

Reads: reports/migration/pid_migration_manifest_YYYYMMDD.jsonl
Writes: mutations in-place on
  · reports/telegram/aegis_history.xlsx
  · reports/telegram/aegis_history_india.xlsx
  · reports/telegram/aegis_history_usa.xlsx
  · reports/research/outcome_dataset.parquet
  · reports/research/outcome_dataset_summary.json (row-count refresh only)

Emits: reports/migration/pid_migration_execute_YYYYMMDD.json
"""
from __future__ import annotations

import json
import sys
from collections import defaultdict
from datetime import date
from pathlib import Path

_ROOT = Path(__file__).resolve().parents[1]


def _load_manifest(today: str) -> list:
    p = _ROOT / "reports" / "migration" / f"pid_migration_manifest_{today}.jsonl"
    return [json.loads(line) for line in p.read_text(encoding="utf-8").splitlines() if line.strip()]


def _apply_xlsx(root: Path, xlsx_rel: str, sheet_name: str,
                  mappings: list) -> dict:
    """Apply DETERMINISTIC PID rewrites to an XLSX file in place.
    `mappings` = list of preflight-manifest rows for this file.
    Returns per-file stats."""
    from openpyxl import load_workbook
    p = root / xlsx_rel
    if not p.exists() or not mappings:
        return {"file": xlsx_rel, "rows_migrated": 0, "reason": "not-present-or-nothing-todo"}
    wb = load_workbook(p)
    if sheet_name not in wb.sheetnames:
        wb.close()
        return {"file": xlsx_rel, "rows_migrated": 0, "reason": "sheet-missing"}
    ws = wb[sheet_name]
    hdr = [c.value for c in ws[1]]
    def col(name):
        for i, c in enumerate(hdr):
            if c and str(c).strip().lower() == name.lower():
                return i + 1
        return None
    c_pid = col("Position ID")
    c_legacy = col("Legacy Position ID")
    if not c_pid:
        wb.close()
        return {"file": xlsx_rel, "rows_migrated": 0, "reason": "no-Position-ID-col"}
    # Index mappings by (row, legacy_pid) for exact match
    by_row = {m["row"]: m for m in mappings if m["classification"] == "DETERMINISTIC"}
    n_migrated = 0
    n_skipped_already = 0
    for r_idx, m in by_row.items():
        pid_cell = ws.cell(r_idx, c_pid)
        current = str(pid_cell.value or "")
        canonical = m.get("canonical_pid")
        if not canonical:
            continue
        if current == canonical:
            n_skipped_already += 1
            continue
        # Preserve legacy PID in Legacy Position ID column if empty
        if c_legacy and not ws.cell(r_idx, c_legacy).value:
            ws.cell(r_idx, c_legacy).value = current
        pid_cell.value = canonical
        n_migrated += 1
    if n_migrated > 0:
        wb.save(p)
    wb.close()
    return {"file": xlsx_rel, "sheet": sheet_name,
            "rows_migrated": n_migrated, "rows_already_canonical": n_skipped_already,
            "mappings_seen": len(mappings)}


def _apply_parquet(root: Path, parquet_rel: str, mappings: list) -> dict:
    """Apply DETERMINISTIC PID rewrites to outcome_dataset.parquet.
    Preserves legacy_position_id and every other column."""
    import pandas as pd
    p = root / parquet_rel
    if not p.exists() or not mappings:
        return {"file": parquet_rel, "rows_migrated": 0, "reason": "not-present-or-nothing-todo"}
    df = pd.read_parquet(p)
    if "position_id" not in df.columns:
        return {"file": parquet_rel, "rows_migrated": 0, "reason": "no-position_id-col"}
    if "legacy_position_id" not in df.columns:
        df["legacy_position_id"] = ""
    # Map preflight row (1-indexed) → mapping · manifest rows are 1-based
    by_row = {m["row"]: m for m in mappings if m["classification"] == "DETERMINISTIC"}
    n_migrated = 0
    n_skipped_already = 0
    for row_idx, m in by_row.items():
        # DataFrame index = row_idx - 1 (0-based)
        i = row_idx - 1
        if i < 0 or i >= len(df): continue
        current = str(df.at[i, "position_id"] or "")
        canonical = m.get("canonical_pid")
        if not canonical:
            continue
        if current == canonical:
            n_skipped_already += 1
            continue
        # Preserve legacy in legacy_position_id
        if not df.at[i, "legacy_position_id"]:
            df.at[i, "legacy_position_id"] = current
        df.at[i, "position_id"] = canonical
        n_migrated += 1
    if n_migrated > 0:
        df.to_parquet(p, index=False)
    return {"file": parquet_rel, "rows_migrated": n_migrated,
            "rows_already_canonical": n_skipped_already, "mappings_seen": len(mappings)}


def main() -> int:
    today = date.today().isoformat().replace("-", "")
    print(f"[execute] loading manifest for {today}...")
    manifest = _load_manifest(today)
    print(f"[execute] manifest rows: {len(manifest)}")

    ambig = [m for m in manifest if m["classification"] in ("AMBIGUOUS", "INVALID")]
    if ambig:
        print(f"[execute] HARD_STOP · {len(ambig)} ambiguous/invalid rows in manifest · "
              "migration blocked", file=sys.stderr)
        return 2

    # Group by file
    by_file = defaultdict(list)
    for m in manifest:
        by_file[m["file"]].append(m)
    print(f"[execute] files to process: {sorted(by_file.keys())}")
    print()

    stats = []
    stats.append(_apply_xlsx(_ROOT, "reports/telegram/aegis_history.xlsx",
                              "AEGIS Daily", by_file.get("reports/telegram/aegis_history.xlsx", [])))
    stats.append(_apply_xlsx(_ROOT, "reports/telegram/aegis_history_india.xlsx",
                              "AEGIS INDIA History",
                              by_file.get("reports/telegram/aegis_history_india.xlsx", [])))
    stats.append(_apply_xlsx(_ROOT, "reports/telegram/aegis_history_usa.xlsx",
                              "AEGIS USA History",
                              by_file.get("reports/telegram/aegis_history_usa.xlsx", [])))
    stats.append(_apply_parquet(_ROOT, "reports/research/outcome_dataset.parquet",
                                 by_file.get("reports/research/outcome_dataset.parquet", [])))

    for s in stats:
        print(f"  [{s['file']}] migrated={s.get('rows_migrated', 0)} "
              f"already_canonical={s.get('rows_already_canonical', 0)} "
              f"mappings_seen={s.get('mappings_seen', 0)}")

    total_migrated = sum(s.get("rows_migrated", 0) for s in stats)
    total_already = sum(s.get("rows_already_canonical", 0) for s in stats)

    report = {
        "engine": "aegis.migration.phase_2_execute",
        "asof": date.today().isoformat(),
        "manifest_rows": len(manifest),
        "total_migrated": total_migrated,
        "total_already_canonical": total_already,
        "files": stats,
    }
    out_p = _ROOT / "reports" / "migration" / f"pid_migration_execute_{today}.json"
    out_p.write_text(json.dumps(report, indent=2, ensure_ascii=False),
                      encoding="utf-8")
    print(f"\n[execute] report: {out_p.relative_to(_ROOT)}")
    print(f"[execute] SUMMARY · migrated={total_migrated} · already_canonical={total_already}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
