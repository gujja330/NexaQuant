"""Portfolio ↔ Exit History overlap classifier · CEO 2026-09-01 §11.

For every security that appears in BOTH the Portfolio (current) and the
Exit History (90d) sheets, determine which category applies. Overlaps
are not automatically errors · but every one must have a machine-
verifiable explanation.

Categories:
    LEGITIMATE_DIFFERENT_LIFECYCLE
      · same ticker · same runner · DIFFERENT entry_date · closed +
        re-opened. Valid re-entry.
    LEGITIMATE_DIFFERENT_RUNNER
      · same ticker · DIFFERENT runner (e.g. R1 exit and R2 held).
    LEGITIMATE_REENTRY
      · same ticker + runner · exit_date < entry_date of current.
    STALE_DUPLICATE
      · same ticker + runner + entry_date · both rows report the same
        lifecycle instance · one should have been cleared.
    RECONCILIATION_DEFECT
      · same ticker + runner + entry_date · Portfolio shows ACTIVE
        AND Exit History shows CLOSED simultaneously. Contract violation.

Output: reports/audit/portfolio_exit_overlap_{market}_{asof}.json
"""
from __future__ import annotations

import argparse
import json
import sys
from datetime import date
from pathlib import Path
from openpyxl import load_workbook

_ROOT = Path(__file__).resolve().parents[1]


def _col(hdr, *names):
    for name in names:
        for i, c in enumerate(hdr):
            if c and str(c).lower() == name.lower(): return i
    return None


def _find_hdr(rows):
    for i, r in enumerate(rows[:10]):
        if r and sum(1 for c in r if c is not None) >= 5:
            return i
    return 0


def _norm(v: object) -> str:
    return str(v or "").split(".", 1)[0].upper().strip()


def classify(market: str, root: Path, asof: str) -> dict:
    xlsx = root / "reports" / "telegram" / f"aegis_history_{market.lower()}.xlsx"
    if not xlsx.exists():
        return {"error": f"missing xlsx: {xlsx}"}
    wb = load_workbook(xlsx, read_only=True, data_only=True)

    # Load Portfolio rows
    portfolio_rows = []
    if "Portfolio" in wb.sheetnames:
        ws = wb["Portfolio"]
        rows = list(ws.iter_rows(values_only=True))
        hi = _find_hdr(rows)
        hdr = rows[hi]
        c_tk = _col(hdr, "Ticker")
        c_run = _col(hdr, "Runner")
        c_life = _col(hdr, "Lifecycle")
        c_ent = _col(hdr, "Entry Date")
        c_dec = _col(hdr, "🎯 DECISION", "DECISION", "Decision")
        for r in rows[hi + 1:]:
            if not r or c_tk is None or not r[c_tk]: continue
            dec = str(r[c_dec] or "") if c_dec is not None else ""
            if "SUGGESTED" in dec.upper(): continue
            portfolio_rows.append({
                "ticker": _norm(r[c_tk]),
                "runner": _norm(r[c_run]) if c_run is not None else "",
                "lifecycle": str(r[c_life] or "") if c_life is not None else "",
                "entry_date": str(r[c_ent] or "")[:10] if c_ent is not None else "",
            })

    exit_rows = []
    eh_sheet = "Exit History (90d)"
    if eh_sheet in wb.sheetnames:
        ws = wb[eh_sheet]
        rows = list(ws.iter_rows(values_only=True))
        hi = _find_hdr(rows)
        hdr = rows[hi]
        c_tk = _col(hdr, "Stock", "Ticker")
        c_run = _col(hdr, "Runner")
        c_ent = _col(hdr, "Entry Date")
        c_ext = _col(hdr, "Exit Date")
        for r in rows[hi + 1:]:
            if not r or c_tk is None or not r[c_tk]: continue
            exit_rows.append({
                "ticker": _norm(r[c_tk]),
                "runner": _norm(r[c_run]) if c_run is not None else "",
                "entry_date": str(r[c_ent] or "")[:10] if c_ent is not None else "",
                "exit_date": str(r[c_ext] or "")[:10] if c_ext is not None else "",
            })
    wb.close()

    portfolio_by_ticker = {}
    for p in portfolio_rows:
        portfolio_by_ticker.setdefault(p["ticker"], []).append(p)
    exit_by_ticker = {}
    for e in exit_rows:
        exit_by_ticker.setdefault(e["ticker"], []).append(e)

    overlap_tickers = sorted(set(portfolio_by_ticker) & set(exit_by_ticker))
    classifications = []
    counts = {
        "LEGITIMATE_DIFFERENT_LIFECYCLE": 0,
        "LEGITIMATE_DIFFERENT_RUNNER": 0,
        "LEGITIMATE_REENTRY": 0,
        "STALE_DUPLICATE": 0,
        "RECONCILIATION_DEFECT": 0,
    }

    for tk in overlap_tickers:
        p_events = portfolio_by_ticker[tk]
        e_events = exit_by_ticker[tk]
        for p in p_events:
            for e in e_events:
                # Same ticker · classify
                if p["runner"] != e["runner"]:
                    cat = "LEGITIMATE_DIFFERENT_RUNNER"
                    explanation = (f"P.runner={p['runner']} E.runner={e['runner']} "
                                    "· independent runner histories")
                elif p["entry_date"] and e["entry_date"] and p["entry_date"] != e["entry_date"]:
                    if e["exit_date"] and p["entry_date"] > e["exit_date"]:
                        cat = "LEGITIMATE_REENTRY"
                        explanation = (f"E.exit={e['exit_date']} < P.entry={p['entry_date']} "
                                        "· closed then re-opened")
                    else:
                        cat = "LEGITIMATE_DIFFERENT_LIFECYCLE"
                        explanation = (f"P.entry={p['entry_date']} E.entry={e['entry_date']} "
                                        "· distinct lifecycle instances")
                elif p["entry_date"] and e["entry_date"] and p["entry_date"] == e["entry_date"]:
                    if "ACTIVE" in p["lifecycle"].upper() and e["exit_date"]:
                        cat = "RECONCILIATION_DEFECT"
                        explanation = (f"Portfolio ACTIVE + Exit CLOSED for identical "
                                        f"(ticker={tk}, runner={p['runner']}, entry={p['entry_date']}) "
                                        "· contract violation")
                    else:
                        cat = "STALE_DUPLICATE"
                        explanation = ("same instance appears in both · one should have been cleared")
                else:
                    # Missing entry_date on one side · degrade gracefully
                    if not p["entry_date"] and not e["entry_date"]:
                        cat = "STALE_DUPLICATE"
                        explanation = "no entry_date on either row · unable to distinguish"
                    else:
                        cat = "LEGITIMATE_DIFFERENT_LIFECYCLE"
                        explanation = "entry_date missing on one row · defaulting to distinct instance"
                counts[cat] += 1
                classifications.append({
                    "ticker": tk,
                    "portfolio": p,
                    "exit": e,
                    "category": cat,
                    "explanation": explanation,
                })

    result = {
        "engine": "portfolio_exit_overlap_classifier.v1",
        "market": market.lower(),
        "asof": asof,
        "n_portfolio_rows": len(portfolio_rows),
        "n_exit_rows": len(exit_rows),
        "n_overlap_tickers": len(overlap_tickers),
        "n_classifications": len(classifications),
        "by_category": counts,
        "n_reconciliation_defects": counts["RECONCILIATION_DEFECT"],
        "classifications": classifications,
    }
    out_p = root / "reports" / "audit" / f"portfolio_exit_overlap_{market.lower()}_{asof}.json"
    out_p.parent.mkdir(parents=True, exist_ok=True)
    out_p.write_text(json.dumps(result, indent=2, ensure_ascii=False),
                      encoding="utf-8")
    return result


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--market", choices=["india", "usa", "both"],
                     default="both")
    ap.add_argument("--asof", default=date.today().isoformat())
    args = ap.parse_args()
    any_defect = False
    for m in (["india", "usa"] if args.market == "both" else [args.market]):
        r = classify(m, _ROOT, args.asof)
        print(f"[overlap:{m}] tickers={r.get('n_overlap_tickers', 0)} · "
              f"defects={r.get('n_reconciliation_defects', 0)}")
        for cat, n in (r.get("by_category") or {}).items():
            _line = f"    {cat:36s} {n}"
            print(_line.encode("ascii", errors="replace").decode("ascii"))
        if r.get("n_reconciliation_defects", 0) > 0:
            any_defect = True
    return 2 if any_defect else 0


if __name__ == "__main__":
    sys.exit(main())
