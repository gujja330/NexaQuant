"""AEGIS · retroactive fix for Recommended/Entry-Price drift in aegis_history.xlsx

Root cause: `_append_to_workbook` (backend/delivery/telegram/detail_xlsx.py)
overwrote all columns on dedup match. When Registry backfill assigned newer
first_seen/Position IDs, the daily writer rewrote historical `Recommended`
+ `Entry Price` fields · producing 103 recommended_drift + 18 entry_price_
drift alerts from guard10.

Source-side fix landed same commit. This one-shot script rewrites the
historical rows in aegis_history.xlsx so guard10:integrity drops to 0.

Rule: for each (Country, Run_Type, Ticker), find the EARLIEST snapshot's
Recommended date + Entry Price. Backfill later snapshots with those
values when they differ. Never mutate a later Recommended to be LATER
than what's already there · always use the earliest.
"""
from __future__ import annotations

import sys
from collections import defaultdict
from pathlib import Path
try:
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
except Exception:
    pass

from openpyxl import load_workbook


def main():
    p = Path("reports/telegram/aegis_history.xlsx")
    if not p.exists():
        print(f"MISSING: {p}"); return 1
    wb = load_workbook(p)
    ws = wb["AEGIS Daily"] if "AEGIS Daily" in wb.sheetnames else wb.active
    h = [c.value for c in ws[1]]

    def _col(name):
        try: return h.index(name) + 1
        except ValueError: return None

    c_ct = _col("Country")
    c_rt = _col("Run_Type")
    c_tk = _col("Ticker")
    c_dt = _col("Date")
    c_rec = _col("Recommended")
    c_ep = _col("Entry Price")
    c_pid = _col("Position ID")

    if not all([c_ct, c_rt, c_tk, c_dt, c_rec, c_ep]):
        print("required columns missing"); return 1

    # Group rows by (country, runner, ticker) · track earliest values
    by_key: dict = defaultdict(list)
    for r in range(2, ws.max_row + 1):
        ct = str(ws.cell(r, c_ct).value or "").upper()
        rt = str(ws.cell(r, c_rt).value or "").upper().replace("_NEW","")
        tk = str(ws.cell(r, c_tk).value or "").upper().replace(".NS","").replace(".BO","")
        dt = str(ws.cell(r, c_dt).value or "")[:10]
        if not (ct and rt and tk and dt): continue
        by_key[(ct, rt, tk)].append(r)

    fixed_rec = 0
    fixed_ep  = 0
    fixed_pid = 0
    samples   = []

    for key, rows in by_key.items():
        # Sort by row date to find earliest
        rows_dated = []
        for r in rows:
            dt = str(ws.cell(r, c_dt).value or "")[:10]
            rows_dated.append((dt, r))
        rows_dated.sort()

        # Canonical values = earliest non-blank
        canonical_rec = None
        canonical_ep  = None
        canonical_pid = None
        for dt, r in rows_dated:
            _rec = ws.cell(r, c_rec).value
            _ep  = ws.cell(r, c_ep).value
            _pid = ws.cell(r, c_pid).value if c_pid else None
            if canonical_rec is None and _rec not in (None, ""):
                canonical_rec = str(_rec)[:10]
            if canonical_ep is None and isinstance(_ep, (int, float)) and _ep > 0:
                canonical_ep = float(_ep)
            if canonical_pid is None and _pid not in (None, ""):
                canonical_pid = str(_pid)
            if canonical_rec and canonical_ep and canonical_pid: break

        # Backfill later rows with canonical values if they differ
        for dt, r in rows_dated:
            _rec_v = ws.cell(r, c_rec).value
            _ep_v  = ws.cell(r, c_ep).value
            _pid_v = ws.cell(r, c_pid).value if c_pid else None
            if canonical_rec is not None:
                v = str(_rec_v)[:10] if _rec_v else ""
                if v != canonical_rec:
                    ws.cell(r, c_rec).value = canonical_rec
                    fixed_rec += 1
                    if len(samples) < 8:
                        samples.append(
                            f"REC {key[2]}/{key[1]} row {r} dt={dt} · "
                            f"{v!r} → {canonical_rec!r}")
            if canonical_ep is not None:
                if isinstance(_ep_v, (int, float)) \
                        and abs(_ep_v - canonical_ep) > 0.01:
                    ws.cell(r, c_ep).value = canonical_ep
                    fixed_ep += 1
                    if len(samples) < 8:
                        samples.append(
                            f"EP  {key[2]}/{key[1]} row {r} dt={dt} · "
                            f"{_ep_v} → {canonical_ep}")
            if c_pid and canonical_pid is not None:
                if _pid_v not in (None, "") and str(_pid_v) != canonical_pid:
                    ws.cell(r, c_pid).value = canonical_pid
                    fixed_pid += 1

    print(f"Recommended date backfills: {fixed_rec}")
    print(f"Entry Price backfills: {fixed_ep}")
    print(f"Position ID backfills: {fixed_pid}")
    for s in samples: print(f"  · {s}")
    if fixed_rec + fixed_ep + fixed_pid > 0:
        wb.save(p)
        print(f"\nSaved: {p}")
    else:
        print("\nNo drift detected · no changes written")
    wb.close()
    return 0


if __name__ == "__main__":
    sys.exit(main())
