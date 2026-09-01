"""Fixed 8-sheet workbook augmenter · CEO 2026-09-01 (Section 11).

The daily emit produces the 5-sheet market workbook
(`aegis_history_{market}.xlsx`). Section 11 mandates a stable 8-sheet
structure:

    1. Portfolio                     (emitted)
    2. Today Decisions               ← ADD
    3. Exit History                  (emitted as `Exit History (90d)`)
    4. Monthly Summary               (emitted)
    5. AEGIS History                 (emitted)
    6. Definitions                   (emitted)
    7. Runner Performance            ← ADD
    8. Research / Quality Analysis   ← ADD

This script opens the emitted workbook and appends the 3 missing sheets
sourced from canonical data. It never modifies the existing 5 sheets.
Idempotent · re-running replaces the 3 augmented sheets from scratch.

Runs after `telegram_command_center_send.py` · never inside it (keeps
scope contained and reversible).

Runner Performance is derived from
`backend/delivery/canonical/runner_accountability.compute_runner_accounting`
· respects R1 retirement (R1 shown as historical · not current).

Research / Quality Analysis is derived from
`reports/research/multi_layer/evidence_{market}_{asof}.json` when present ·
otherwise a placeholder row `NOT_RUN_TODAY`.
"""
from __future__ import annotations

import argparse
import json
import sys
from datetime import date
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

_ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(_ROOT))

from backend.delivery.canonical.runner_accountability import compute_runner_accounting  # noqa: E402
from backend.delivery.canonical.retirement import retired_runners, active_runners  # noqa: E402

_HDR_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
_HDR_FONT = Font(bold=True, color="FFFFFF", size=11)
_BANNER_FILL = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
_BANNER_FONT = Font(bold=True, color="FFFFFF", size=13)


def _clear_or_create(wb, name: str):
    if name in wb.sheetnames:
        del wb[name]
    return wb.create_sheet(name)


def _banner(ws, text: str, ncols: int):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    c = ws.cell(1, 1, text)
    c.font = _BANNER_FONT
    c.fill = _BANNER_FILL
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28


def _header(ws, cols: list[str], row: int = 3):
    for i, name in enumerate(cols, start=1):
        c = ws.cell(row, i, name)
        c.fill = _HDR_FILL
        c.font = _HDR_FONT
        c.alignment = Alignment(horizontal="center", vertical="center")


# ── Sheet 2 · Today Decisions ────────────────────────────────────────
def _emit_today_decisions(wb, market: str, root: Path, asof: str) -> int:
    ws = _clear_or_create(wb, "Today Decisions")
    _banner(ws, f"AEGIS {market.upper()} · TODAY DECISIONS · as of {asof}", 10)
    cols = ["Position ID", "Ticker", "Runner", "Population",
            "Decision", "Urgency", "Reason",
            "Entry Price", "Current Price", "Provenance"]
    _header(ws, cols)
    # Column widths
    for i, w in enumerate([28, 12, 8, 22, 14, 10, 40, 10, 12, 18], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Source: the emitted Portfolio sheet DECISION column
    if "Portfolio" not in wb.sheetnames:
        ws.cell(4, 1, "NO_PORTFOLIO_SHEET")
        return 0
    p = wb["Portfolio"]
    rows = list(p.iter_rows(values_only=True))
    # Find header row
    hdr_i = 0
    for i, r in enumerate(rows[:8]):
        if r and sum(1 for c in r if c is not None) >= 5:
            hdr_i = i
            break
    hdr = rows[hdr_i]

    def _col(name):
        for i, c in enumerate(hdr):
            if c and str(c).lower() == name.lower(): return i
        return None

    c_tk = _col("Ticker")
    c_dec = None
    for cand in ("🎯 DECISION", "DECISION", "Decision"):
        c_dec = _col(cand)
        if c_dec is not None: break
    c_run = _col("Runner")
    c_urg = _col("Urgency")
    c_rea = _col("Reason")
    c_ent = _col("Entry")
    c_curr = _col("Current")

    # Load provenance companion to map Ticker+Runner+EntryDate -> Position ID
    prov_p = root / "reports" / "telegram" / f"aegis_history_{market.lower()}_provenance.jsonl"
    prov_by_key = {}
    if prov_p.exists():
        for line in prov_p.read_text(encoding="utf-8").splitlines():
            if not line.strip(): continue
            r = json.loads(line)
            key = (r.get("ticker", ""), r.get("runner", ""))
            if key not in prov_by_key:
                prov_by_key[key] = r

    retired = retired_runners(root)
    out_row = 4
    n_written = 0
    for r in rows[hdr_i + 1:]:
        if not r or c_tk is None or not r[c_tk]: continue
        tk = str(r[c_tk] or "").split(".", 1)[0].upper()
        dec = str(r[c_dec] or "") if c_dec is not None else ""
        run = str(r[c_run] or "").upper() if c_run is not None else ""
        # Skip retired-runner rows (already hidden from Portfolio · defensive)
        if run in retired: continue
        # Only rows that have an actual decision
        if not dec.strip() or dec.strip() == "—": continue
        urg = str(r[c_urg] or "") if c_urg is not None else ""
        rea = str(r[c_rea] or "") if c_rea is not None else ""
        ent = r[c_ent] if c_ent is not None else ""
        curr = r[c_curr] if c_curr is not None else ""
        prov = prov_by_key.get((tk, run), {})
        pid = prov.get("position_id", "") or "—"
        pop = prov.get("population", "") or "CURRENT_SIGNAL"

        ws.cell(out_row, 1, pid)
        ws.cell(out_row, 2, tk)
        ws.cell(out_row, 3, run)
        ws.cell(out_row, 4, pop)
        ws.cell(out_row, 5, dec)
        ws.cell(out_row, 6, urg)
        ws.cell(out_row, 7, rea)
        ws.cell(out_row, 8, ent)
        ws.cell(out_row, 9, curr)
        ws.cell(out_row, 10, "canonical:portfolio+prov")
        out_row += 1
        n_written += 1
    return n_written


# ── Sheet 7 · Runner Performance ────────────────────────────────────
def _emit_runner_performance(wb, market: str, root: Path, asof: str) -> int:
    ws = _clear_or_create(wb, "Runner Performance")
    _banner(ws, f"AEGIS {market.upper()} · RUNNER PERFORMANCE · as of {asof}", 12)
    cols = ["Runner", "Utilization Status", "Utilization Reason",
            "Signals Generated", "Positions Opened",
            "Currently Active", "Positions Closed", "Eligible Exits",
            "Realized P&L %", "Win Rate %", "Mean DD %",
            "Sample-Size Verdict", "Provenance"]
    _header(ws, cols)
    for i, w in enumerate([12, 22, 60, 18, 18, 16, 18, 16, 16, 14, 12, 22, 18], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    retired = retired_runners(root)
    active = active_runners(root)
    out_row = 4
    n_written = 0
    # CEO 2026-09-01 STRENGTHENED · retired runners MUST NOT appear in
    # the delivered Runner Performance sheet. R1 stats remain in
    # backend/delivery/canonical/runner_accountability for audit access,
    # but the operator-facing workbook shows R2 only.
    _labels_to_emit = tuple(
        r for r in ("R1", "R2", "COMBINED") if r not in retired
    )
    for runner_label in _labels_to_emit:
        try:
            acc = compute_runner_accounting(
                root, market.lower(),
                runner_filter=runner_label if runner_label != "COMBINED" else None,
                asof=asof,
                window_days=90,
            )
        except Exception as e:
            acc = None
        if acc is None:
            continue
        def _num(v, digits):
            if v is None: return "UNAVAILABLE"
            try: return round(float(v), digits)
            except (TypeError, ValueError): return "UNAVAILABLE"
        # Section 10 · CEO 2026-09-01 · canonical utilization classification
        vals = [
            runner_label,
            getattr(acc, "utilization_status", "UNKNOWN") or "UNKNOWN",
            getattr(acc, "utilization_reason", "") or "",
            getattr(acc, "signals_generated", 0) or 0,
            getattr(acc, "positions_opened", 0) or 0,
            getattr(acc, "currently_active", 0) or 0,
            getattr(acc, "positions_closed", 0) or 0,
            getattr(acc, "eligible_exits", 0) or 0,
            _num(getattr(acc, "realized_pnl_pct", None), 2),
            _num(getattr(acc, "win_rate_pct", None), 1),
            _num(getattr(acc, "max_realized_drawdown_pct", None), 2),
            getattr(acc, "sample_size_verdict", "UNKNOWN") or "UNKNOWN",
            "canonical:runner_accountability",
        ]
        for i, v in enumerate(vals, start=1):
            ws.cell(out_row, i, v)
        out_row += 1
        n_written += 1
    return n_written


# ── Sheet 8 · Research / Quality Analysis ───────────────────────────
def _emit_research_quality(wb, market: str, root: Path, asof: str) -> int:
    ws = _clear_or_create(wb, "Research Quality")
    _banner(ws, f"AEGIS {market.upper()} · RESEARCH / QUALITY ANALYSIS · as of {asof}", 8)
    cols = ["Layer", "Category", "Status", "Fold",
            "Train Window", "Test Window", "Walk-Forward Criterion", "Notes"]
    _header(ws, cols)
    for i, w in enumerate([28, 10, 14, 6, 24, 24, 42, 40], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    evidence_p = root / "reports" / "research" / "multi_layer" / f"evidence_{market.lower()}_{asof}.json"
    if not evidence_p.exists():
        ws.cell(4, 1, "NOT_RUN_TODAY")
        ws.cell(4, 3, "UNAVAILABLE")
        ws.cell(4, 8, f"expected: {evidence_p.name}")
        return 0
    data = json.loads(evidence_p.read_text(encoding="utf-8"))
    records = data.get("records", [])
    out_row = 4
    for r in records[:200]:  # cap at 200 rows for legibility
        vals = [
            r.get("layer_key", ""),
            r.get("category", ""),
            r.get("status", ""),
            r.get("fold", 0),
            f"{r.get('train_start', '')}..{r.get('train_end', '')}",
            f"{r.get('test_start', '')}..{r.get('test_end', '')}",
            r.get("walk_forward_criterion", ""),
            r.get("reason", ""),
        ]
        for i, v in enumerate(vals, start=1):
            ws.cell(out_row, i, v)
        out_row += 1
    return len(records)


def _scrub_retired_from_portfolio(wb, root: Path) -> int:
    """CEO 2026-09-01 §1 hardening · in-place removal of retired-runner
    rows from Portfolio sheet. Idempotent · safe if 0 rows to remove."""
    retired = retired_runners(root)
    if not retired or "Portfolio" not in wb.sheetnames: return 0
    ws = wb["Portfolio"]
    rows = list(ws.iter_rows(values_only=True))
    hdr_i = 0
    for i, r in enumerate(rows[:10]):
        if r and r[0] and "Ticker" in str(r[0]):
            hdr_i = i
            break
    hdr = rows[hdr_i]
    ci_run = None
    for i, c in enumerate(hdr):
        if c and str(c).lower() == "runner":
            ci_run = i
            break
    if ci_run is None: return 0
    n_stripped = 0
    for excel_row in range(ws.max_row, hdr_i + 1, -1):
        cell = ws.cell(excel_row, ci_run + 1).value
        if cell and str(cell).upper() in retired:
            ws.delete_rows(excel_row, 1)
            n_stripped += 1
    return n_stripped


def _scrub_retired_from_all_sheets(wb, root: Path) -> dict:
    """CEO 2026-09-01 STRENGTHENED CONTRACT · R1 must be COMPLETELY
    absent from the delivered workbook · every sheet · every cell.
    Only Definitions may reference R1 in a fixed sentence explaining
    the retirement (metadata · not runner-row data).

    Scrubs by:
      · Removing rows where Runner column value is a retired runner
      · Removing rows where any cell contains R1-* prefixed Position ID
        (canonical PID · e.g. USA-R1-TRV-20260810-6f873c)
      · Removing rows where legacy Position ID (col if present) starts
        with R1- or is XXX_R1_

    Never modifies Definitions sheet · that sheet describes the
    retirement contract itself and may name R1 as a reference.
    """
    retired = retired_runners(root)
    if not retired:
        return {"markets_scrubbed": {}, "total_rows_removed": 0}
    per_sheet: dict[str, int] = {}
    total = 0
    for sh_name in list(wb.sheetnames):
        if sh_name == "Definitions":
            continue  # Definitions may name R1 as a reference
        ws = wb[sh_name]
        max_r = ws.max_row
        max_c = ws.max_column
        rows_to_delete: list[int] = []
        for excel_row in range(1, max_r + 1):
            row_is_retired = False
            for c in range(1, max_c + 1):
                v = ws.cell(excel_row, c).value
                if v is None: continue
                s = str(v).strip().upper()
                # Direct runner value
                if s in retired:
                    row_is_retired = True
                    break
                # Canonical PID startswith
                for retired_r in retired:
                    if any(s.startswith(f"{prefix}-{retired_r}-") for prefix in ("USA", "IND")):
                        row_is_retired = True
                        break
                    if s.startswith(f"{retired_r}-"):
                        row_is_retired = True
                        break
                if row_is_retired: break
            if row_is_retired:
                rows_to_delete.append(excel_row)
        # Delete bottom-up so earlier indices remain valid
        for r in sorted(rows_to_delete, reverse=True):
            ws.delete_rows(r, 1)
        if rows_to_delete:
            per_sheet[sh_name] = len(rows_to_delete)
            total += len(rows_to_delete)
    return {"per_sheet": per_sheet, "total_rows_removed": total}


def _emit_research_timing(wb, market: str, root: Path, asof: str) -> int:
    """Sheet 9 · Research Timing · momentum ledger view.
    Never mixed into Portfolio · every candidate has terminal state + reason."""
    ws = _clear_or_create(wb, "Research Timing")
    _banner(ws, f"AEGIS {market.upper()} · RESEARCH / TIMING (momentum ledger) · as of {asof}", 12)
    cols = ["Ticker", "Sector", "Category", "Quality Band",
            "Engine Verdict", "Terminal State", "Reason Code", "Reason Text",
            "Return 1d %", "Return 5d %", "Return 20d %", "Attribution"]
    _header(ws, cols)
    for i, w in enumerate([12, 18, 16, 12, 18, 14, 22, 60, 12, 12, 12, 12], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ledger_p = root / "reports" / "research" / "multi_layer" / f"momentum_ledger_{market.lower()}_{asof}.json"
    if not ledger_p.exists():
        ws.cell(4, 1, "NOT_RUN_TODAY")
        ws.cell(4, 6, "UNAVAILABLE")
        ws.cell(4, 8, f"expected: {ledger_p.name}")
        return 0
    data = json.loads(ledger_p.read_text(encoding="utf-8"))
    entries = data.get("entries") or []
    out_row = 4
    # Sort · ACCEPTED first · then WATCH · REJECTED · NO_EVIDENCE
    state_order = {"ACCEPTED": 0, "WATCH": 1, "REJECTED": 2, "NO_EVIDENCE": 3}
    entries_sorted = sorted(entries, key=lambda e: (
        state_order.get(e.get("terminal_state", ""), 9),
        e.get("ticker", "") or ""
    ))
    for e in entries_sorted[:250]:  # cap for legibility
        vals = [
            e.get("ticker", ""), e.get("sector", ""), e.get("category", ""),
            e.get("quality_band", ""), e.get("verdict_engine", ""),
            e.get("terminal_state", ""), e.get("reason_code", ""),
            e.get("reason_text", ""),
            e.get("return_1d_pct"), e.get("return_5d_pct"),
            e.get("return_20d_pct"), e.get("attribution", "R2"),
        ]
        for i, v in enumerate(vals, start=1):
            ws.cell(out_row, i, v if v is not None else "—")
        out_row += 1
    # Summary row
    ws.cell(out_row + 1, 1, "TOTALS")
    ws.cell(out_row + 1, 2, f"universe={data.get('n_universe_scanned')}")
    ws.cell(out_row + 1, 3, f"src={data.get('n_candidates_source')}")
    ws.cell(out_row + 1, 6, f"by_state={data.get('by_terminal_state')}")
    ws.cell(out_row + 1, 8, f"conservation_ok={data.get('conservation_ok')}")
    return len(entries)


def augment(market: str, root: Path, asof: str) -> dict:
    xlsx = root / "reports" / "telegram" / f"aegis_history_{market.lower()}.xlsx"
    if not xlsx.exists():
        return {"error": f"missing: {xlsx}"}
    wb = load_workbook(xlsx)
    n_scrubbed = _scrub_retired_from_portfolio(wb, root)
    # CEO 2026-09-01 STRENGTHENED · R1 absent from EVERY sheet · not just Portfolio
    workbook_scrub = _scrub_retired_from_all_sheets(wb, root)
    n_td = _emit_today_decisions(wb, market, root, asof)
    n_rp = _emit_runner_performance(wb, market, root, asof)
    n_rq = _emit_research_quality(wb, market, root, asof)
    n_rt = _emit_research_timing(wb, market, root, asof)
    wb.save(xlsx)
    # Sync dated snapshot
    import shutil
    dated = root / "reports" / "telegram" / f"aegis_{market.lower()}_{asof}.xlsx"
    shutil.copyfile(xlsx, dated)
    return {
        "market": market.lower(),
        "xlsx": str(xlsx.relative_to(root)),
        "dated": str(dated.relative_to(root)),
        "retired_rows_scrubbed_from_portfolio": n_scrubbed,
        "workbook_wide_scrub": workbook_scrub,
        "today_decisions_rows": n_td,
        "runner_performance_rows": n_rp,
        "research_quality_rows": n_rq,
        "research_timing_rows": n_rt,
        "sheets_after": load_workbook(xlsx, read_only=True).sheetnames,
    }


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--market", choices=["india", "usa", "both"],
                     default="both")
    ap.add_argument("--asof", default=date.today().isoformat())
    args = ap.parse_args()
    markets = ["india", "usa"] if args.market == "both" else [args.market]
    for m in markets:
        rep = augment(m, _ROOT, args.asof)
        print(json.dumps(rep, indent=2, ensure_ascii=False, default=str))
    return 0


if __name__ == "__main__":
    sys.exit(main())
