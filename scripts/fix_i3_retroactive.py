"""AEGIS · I3 retroactive cleanup.

Blanks Current Perf % on all EXIT rows in reports/telegram/aegis_history.xlsx.

Root cause: cur_ret was computed for ALL statuses before line 613 transferred
it to exit_pnl_pct; the A9 gate at line 760 only recomputed for non-EXIT and
never nulled EXIT. Fixed at source (2026-08-26) so new rows are clean · this
script blanks the historical residue so the validator's I3 check passes.

Safe · idempotent · only modifies EXIT-status cells that are currently
populated with a numeric Active P&L value.
"""
from __future__ import annotations
import sys
from pathlib import Path
try:
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
except Exception:
    pass

from openpyxl import load_workbook


def main():
    p = Path("reports/telegram/aegis_history.xlsx")
    if not p.exists():
        print(f"MISSING: {p}")
        return 1
    wb = load_workbook(p)
    ws = wb["AEGIS Daily"] if "AEGIS Daily" in wb.sheetnames else wb.active
    h = [c.value for c in ws[1]]
    if "Current Perf %" not in h or "Status" not in h:
        print("Required columns missing"); return 1
    c_perf = h.index("Current Perf %") + 1
    c_st = h.index("Status") + 1
    c_tk = h.index("Ticker") + 1 if "Ticker" in h else None
    c_dt = h.index("Date") + 1 if "Date" in h else None
    n_cleaned = 0
    samples = []
    for r in range(2, ws.max_row + 1):
        if ws.cell(r, c_st).value == "EXIT":
            v = ws.cell(r, c_perf).value
            if isinstance(v, (int, float)) and abs(v) > 0.01:
                tk = ws.cell(r, c_tk).value if c_tk else "?"
                dt = ws.cell(r, c_dt).value if c_dt else "?"
                if len(samples) < 5:
                    samples.append(f"  R{r} · {tk} · {dt} · was {v}")
                ws.cell(r, c_perf).value = None
                n_cleaned += 1
    if n_cleaned:
        wb.save(p)
        print(f"Blanked Current Perf % on {n_cleaned} EXIT rows")
        for s in samples: print(s)
    else:
        print("No EXIT rows required cleanup · already clean")
    wb.close()
    return 0


if __name__ == "__main__":
    sys.exit(main())
