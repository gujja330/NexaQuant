"""AEGIS 3-sheet canonical workbook renderer · CEO 2026-09-01 (FINAL closing spec).

Exactly 3 visible sheets · identical structure India + USA · daily rollover:

    01_Portfolio       · current active R2 holdings only (D snapshot)
    02_Today_Momentum  · today's decisions/recommendations (D snapshot · fresh every day)
    03_Exit_History    · closed lifecycle only (accumulates via Registry CLOSED)

Daily rollover semantics:
    · D+1 rebuild reads canonical state at D+1
    · 02_Today_Momentum is regenerated from scratch for D+1
    · Positions that opened at D and are still active → remain in 01_Portfolio
    · Positions that closed at D → leave 01_Portfolio, appear in 03_Exit_History
    · Nothing is copied from D-1's workbook. Canonical state is the source of truth.

Data comes from CANONICAL sources only:
    · Registry (opportunity_registry.jsonl · sole PID authority)
    · usa/data/raw/us/*.parquet · data/raw/india/*.parquet (canonical prices)
    · reports/research/multi_layer/momentum_ledger_*.json (for Today_Momentum)

Rules:
    · R2 only · R1 excluded workbook-wide
    · Latest date first everywhere
    · "—" for N/A · "UNAVAILABLE" for genuine gaps · 0 never used to mean missing
    · Realized vs Unrealized P&L labelled explicitly
    · Only green (P&L>0) / red (P&L<0) coloring · everything else neutral
    · India and USA get IDENTICAL sheet names, columns, headers, formats

Files produced:
    reports/telegram/aegis_{market}_{asof}.xlsx           (dated · deliverable)
    reports/telegram/aegis_history_{market}.xlsx          (latest alias · same bytes)
"""
from __future__ import annotations

import argparse
import json
import sys
from collections import defaultdict
from datetime import date, timedelta
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

_ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(_ROOT))

# Minimal color contract
FILL_BANNER = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
FILL_HEADER = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
FILL_POS = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
FILL_NEG = PatternFill(start_color="F8CBAD", end_color="F8CBAD", fill_type="solid")

FONT_BANNER = Font(bold=True, color="FFFFFF", size=14)
FONT_SUB = Font(bold=True, color="1F4E78", size=11)
FONT_HEADER = Font(bold=True, color="1F4E78", size=11)
FONT_BODY = Font(size=10)
FONT_LEGEND = Font(size=9, color="808080", italic=True)


def _banner(ws, text, ncols):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    c = ws.cell(1, 1, text); c.font = FONT_BANNER; c.fill = FILL_BANNER
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28


def _sub(ws, text, ncols, row):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    c = ws.cell(row, 1, text); c.font = FONT_SUB
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)


def _header(ws, cols, row):
    for i, name in enumerate(cols, start=1):
        c = ws.cell(row, i, name)
        c.font = FONT_HEADER; c.fill = FILL_HEADER
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _write_row(ws, values, row, pnl_col_idx=None):
    for i, v in enumerate(values, start=1):
        c = ws.cell(row, i, v)
        c.font = FONT_BODY
        c.alignment = Alignment(horizontal="left", vertical="center")
        if pnl_col_idx is not None and i == pnl_col_idx and isinstance(v, (int, float)):
            if v > 0: c.fill = FILL_POS
            elif v < 0: c.fill = FILL_NEG


def _legend(ws, lines, start_row, ncols):
    for line in lines:
        ws.merge_cells(start_row=start_row, start_column=1,
                        end_row=start_row, end_column=ncols)
        c = ws.cell(start_row, 1, line)
        c.font = FONT_LEGEND
        c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        start_row += 1
    return start_row


def _load_registry(root, market, retired):
    """R1 retired workbook-wide: excluded from BOTH Portfolio and
    Exit History. Historical R1 CLOSED go to orphan_audit_{market}.jsonl
    (documented sink) so A23 lineage validation passes without
    violating C19 workbook-wide R1-zero contract.

    CEO 2026-09-02 delivery gate reconciliation:
    - C19 (workbook R1-zero): no R1 rows anywhere in workbook
    - A23 (lineage): every Registry-CLOSED tracked (EH body OR orphan_audit)
    Together: R2 CLOSED → EH body · R1 CLOSED → orphan_audit sink."""
    from backend.research import opportunity_registry as oreg
    reg = oreg.load_all(root)
    cutoff = (date.today() - timedelta(days=90)).isoformat()
    active, closed, closed_retired = [], [], []
    # CEO 2026-09-02 · all admin/retired CLOSED must land in orphan_audit
    # (A23 sink) · otherwise USA delivery blocks daily on ORPHAN_AUTO_CLOSE
    # events my filter drops from EH body but leaves untracked.
    closed_admin = []
    for pid, opps in reg.items():
        for o in opps:
            if o.market.lower() != market.lower(): continue
            if o.status == "ACTIVE":
                if o.runner in retired: continue
                active.append(o)
            elif o.status == "CLOSED" and o.closed_date and o.closed_date >= cutoff:
                if o.runner in retired:
                    closed_retired.append(o)     # → orphan_audit (retired sink)
                    continue
                # Production runner CLOSED · classify structurally
                _ep_r = _close_on_or_before(root, o.ticker, market, o.created_date or "")
                _xp_r = _close_on_or_before(root, o.ticker, market, o.closed_date or "")
                if _is_administrative_exit(o, _ep_r, _xp_r):
                    closed_admin.append(o)       # → orphan_audit (admin sink)
                else:
                    closed.append(o)              # → Exit History body
    return {"active": active, "closed_90d": closed,
             "closed_retired_90d": closed_retired,
             "closed_admin_90d": closed_admin}


def _normalize_exit_reason(raw: str) -> str:
    """CEO 2026-09-02 · I18 jargon-free presentation.
    Translate raw registry event codes/arrows/ticker suffixes into
    plain English suitable for operator display. Semantic categories
    preserved · no signal information lost.

    UPDATE 2026-09-02 (CEO XLSX contract cleanup): the '+X pp' relative-
    opportunity number was being misread as realized P&L. Reason column
    now contains only a plain-English category · relative-opportunity
    is emitted separately via _extract_relative_pp."""
    if not raw or raw == "—": return "—"
    r = str(raw).strip()
    r_l = r.lower()
    if r_l.startswith("rotation"):
        return "Rotation swap"
    if "alpha" in r_l and "pp" in r_l:
        return "Outperformance exit"
    # Category 3: registry sync backfill
    if "registry-sync" in r_l or "registry sync" in r_l:
        return "Historical reconciliation"
    # Category 4: orphan auto-close
    if "orphan_auto" in r_l or "orphan auto" in r_l:
        return "Auto-close · orphaned position"
    # Category 5: stop-loss / target / horizon triggers
    if "stop_loss_hit" in r_l or "stop loss hit" in r_l or "_stop_" in r_l:
        return "Stop-loss triggered"
    if "target_hit" in r_l or "t1_hit" in r_l or "t2_hit" in r_l:
        return "Target hit"
    if "horizon" in r_l:
        return "Holding horizon reached"
    if "trailing" in r_l or "trail_" in r_l:
        return "Trailing stop triggered"
    if "risk_signal" in r_l or "risk signal" in r_l:
        return "Risk signal exit"
    if "missing_from_signals" in r_l:
        return "Signal dropped"
    # Default: strip jargon chars but keep the sentence · truncate to 40 char
    cleaned = (r.replace("→", "·")
                 .replace(".NS", "")
                 .replace(".BO", "")
                 .replace("alpha", "gain"))
    # Collapse repeated middots
    while "· ·" in cleaned:
        cleaned = cleaned.replace("· ·", "·")
    return cleaned.strip("· ").strip()[:40] or "—"


def _extract_relative_pp(raw: str):
    """Extract the '+X.X pp' relative-opportunity number from a raw
    registry reason string · returns float or None. Used to populate
    the separate 'Relative Opportunity pp' column so it can't be
    misread as the trade's own realized P&L."""
    if not raw: return None
    import re
    m = re.search(r"([+-]?\d+\.?\d*)\s*pp", str(raw))
    if not m: return None
    try:
        return float(m.group(1))
    except Exception:
        return None


def _is_administrative_exit(o, entry_p=None, exit_p=None) -> bool:
    """CEO 2026-09-02 · realized-exits-only invariant · filter admin
    events out of the operator-facing Exit History using STRUCTURAL
    signals only (no hardcoded string matches on closed_reason).

    Structural criteria (any = administrative):
      1. same-day entry+exit (created_date == closed_date) · no time
         to accumulate a real market-move P&L
      2. entry_price == exit_price within 0.005% · no realized delta
         (rotation/reconciliation artifacts leave entry=exit unchanged)

    Historical audit trail preserved in Registry JSONL · never
    destroyed · just hidden from operator-facing table."""
    cd = str(getattr(o, "created_date", "") or "")[:10]
    xd = str(getattr(o, "closed_date", "") or "")[:10]
    if cd and xd and cd == xd: return True
    if (entry_p and exit_p and entry_p > 0
        and abs((exit_p - entry_p) / entry_p * 100) < 0.005):
        return True
    return False


def _emit_orphan_audit_for_retired(root, market, reg_data):
    """Write ALL non-body CLOSED events to orphan_audit_{market}.jsonl:
      · retired-runner CLOSED (R1 retirement carveout)
      · admin CLOSED (same-day OR entry==exit · e.g. ORPHAN_AUTO_CLOSE)

    A23 lineage validation reads this sink · anything NOT in Exit History
    body MUST appear here or A23 blocks USA delivery daily. Idempotent."""
    p = root / "reports" / "delivery" / f"orphan_audit_{market.lower()}.jsonl"
    p.parent.mkdir(parents=True, exist_ok=True)
    existing_tickers = set()
    if p.exists():
        try:
            for line in p.read_text(encoding="utf-8", errors="replace").splitlines():
                if not line.strip(): continue
                try:
                    e = json.loads(line)
                    t = str(e.get("ticker","")).upper()
                    if t: existing_tickers.add(t)
                except Exception: pass
        except Exception: pass
    added_retired = 0
    added_admin = 0
    with p.open("a", encoding="utf-8") as f:
        for o in reg_data.get("closed_retired_90d", []):
            tk = o.ticker.upper()
            if tk in existing_tickers: continue
            existing_tickers.add(tk)
            entry = {
                "kind": "RETIRED_RUNNER_CLOSED",
                "ticker": tk, "runner": o.runner,
                "market": market.lower(),
                "opportunity_id": o.opportunity_id,
                "closed_date": o.closed_date,
                "closed_reason": getattr(o, "closed_reason", "") or "",
                "rationale": (f"Runner {o.runner} retired · workbook-wide "
                              f"contract · historical accountability preserved"),
            }
            f.write(json.dumps(entry, ensure_ascii=False) + "\n")
            added_retired += 1
        # CEO 2026-09-02 · production-runner admin events (same-day / zero-Δ)
        # also go to orphan_audit · not silently lost from A23's perspective.
        for o in reg_data.get("closed_admin_90d", []):
            tk = o.ticker.upper()
            if tk in existing_tickers: continue
            existing_tickers.add(tk)
            entry = {
                "kind": "ADMIN_ZERO_DELTA_CLOSED",
                "ticker": tk, "runner": o.runner,
                "market": market.lower(),
                "opportunity_id": o.opportunity_id,
                "created_date": o.created_date,
                "closed_date": o.closed_date,
                "closed_reason": getattr(o, "closed_reason", "") or "",
                "rationale": ("Production-runner administrative event · "
                              "same-day OR entry==exit price · never a real "
                              "market delta · filtered from Exit History body · "
                              "tracked here for A23 lineage completeness"),
            }
            f.write(json.dumps(entry, ensure_ascii=False) + "\n")
            added_admin += 1
    return {"path": str(p.relative_to(root)),
            "added_retired": added_retired, "added_admin": added_admin,
            "total_retired_closed": len(reg_data.get("closed_retired_90d", [])),
            "total_admin_closed": len(reg_data.get("closed_admin_90d", []))}


def _load_sector_cache(root):
    """Read reports/sector_cache.json · same source used by the legacy
    sender's _sector_for lookup. Returns {market: {TICKER: sector}}."""
    p = root / "reports" / "sector_cache.json"
    if not p.exists():
        return {}
    try:
        return json.loads(p.read_text(encoding="utf-8"))
    except Exception:
        return {}


def _sector_for(sector_cache, market, ticker):
    """Ticker → sector lookup · returns '—' if unknown."""
    if not sector_cache:
        return "—"
    market_bucket = sector_cache.get(market.lower()) or {}
    short = str(ticker or "").replace(".NS", "").replace(".BO", "").upper()
    return market_bucket.get(short) or "—"


def _close_on_or_before(root, ticker, market, target_date):
    import pandas as pd
    dir_ = "usa/data/raw/us" if market.lower() == "usa" else "data/raw/india"
    ext = "" if market.lower() == "usa" else ".NS"
    for p in (root / dir_ / f"{ticker.upper()}{ext}_D1.parquet",
               root / dir_ / f"{ticker.upper()}_D1.parquet"):
        if not p.exists(): continue
        try:
            df = pd.read_parquet(p)
            if "close" not in df.columns: continue
            idx = pd.to_datetime(df.index).strftime("%Y-%m-%d")
            df = df.copy(); df.index = idx
            sub = df.loc[df.index <= target_date]
            if sub.empty: continue
            return float(sub.iloc[-1]["close"])
        except Exception:
            continue
    return None


def _canonical_ticker(t):
    return str(t or "").split(".", 1)[0].upper().strip()


# ── SHEET 01 · Portfolio · CURRENT ACTIVE R2 HOLDINGS ONLY ──────────
def _load_dynamic_risk(root, market):
    """Load canonical dynamic-risk output · returns {pid: {stop, type, reason}}
    Source: reports/context/dynamic_risk_{market}.json produced by
    backend.risk.dynamic_risk_v2 · this is the authoritative stop level."""
    p = root / "reports" / "context" / f"dynamic_risk_{market.lower()}.json"
    out = {}
    if not p.exists(): return out
    try:
        d = json.loads(p.read_text(encoding="utf-8"))
        for u in (d.get("updates") or []):
            pid = u.get("opportunity_id")
            if not pid: continue
            out[pid] = {
                "stop":         u.get("new_stop"),
                "type":         u.get("stop_type") or "",
                "reason":       u.get("reason") or "",
            }
    except Exception:
        pass
    return out


def _atr14_at_date(root, market, ticker, asof: str):
    """PIT ATR-14 at asof from parquet · returns float or None.
    Same math the P0 exit-bridge replay uses · single source of truth."""
    try:
        import pandas as pd
        from backend.research._paths import price_parquet_path
        p = price_parquet_path(root, market, str(ticker).upper())
        if not p or not p.exists(): return None
        df = pd.read_parquet(p)
        df.index = pd.to_datetime(df.index)
        target_dt = pd.to_datetime(asof).normalize()
        if target_dt in df.index:
            idx = df.index.get_loc(target_dt)
        else:
            mask = df.index <= target_dt
            if not mask.any(): return None
            idx = int(mask.sum()) - 1
        if isinstance(idx, slice) or hasattr(idx, "__len__"): return None
        if idx < 14: return None
        highs = df["high"].to_numpy(); lows = df["low"].to_numpy(); closes = df["close"].to_numpy()
        trs = []
        for i in range(idx - 13, idx + 1):
            if i <= 0: continue
            trs.append(max(highs[i] - lows[i],
                           abs(highs[i] - closes[i-1]),
                           abs(lows[i] - closes[i-1])))
        if not trs: return None
        return sum(trs) / len(trs)
    except Exception:
        return None


def _target_from_atr_fallback(entry_price, atr, m_target: float = 3.0):
    """P0-consistent ATR-based target · used when Registry has no
    structural target field. m=3.0 matches P0 replay parameters."""
    if not entry_price or not atr or atr <= 0: return None
    return float(entry_price) + m_target * float(atr)


def _target_from_registry(o):
    """Extract T1/T2 target from Registry Opportunity if present in
    initial_signal or via structured fields. Returns None if not
    canonically available (never fabricated)."""
    for attr in ("t1_target", "target_1", "initial_t1"):
        v = getattr(o, attr, None)
        if isinstance(v, (int, float)) and v > 0: return v
    sig = getattr(o, "initial_signal", "") or ""
    if isinstance(sig, str) and "target" in sig.lower():
        import re
        m = re.search(r"target[^\d]*(\d+\.?\d*)", sig, re.IGNORECASE)
        if m:
            try:
                v = float(m.group(1))
                if v > 0: return v
            except Exception: pass
    return None


def _horizon_from_registry(o):
    """Extract exit horizon in days from Registry initial_signal · returns
    int (days) or None. Never fabricated."""
    sig = getattr(o, "initial_signal", "") or ""
    if isinstance(sig, str):
        import re
        m = re.search(r"horizon[^\d]*(\d+)\s*d", sig, re.IGNORECASE)
        if m:
            try:
                v = int(m.group(1))
                if v > 0: return v
            except Exception: pass
    return None


def _emit_portfolio(wb, market, root, asof, reg_data):
    ws = wb.create_sheet("01_Portfolio")
    ncols = 20
    _banner(ws, f"AEGIS {market.upper()} · PORTFOLIO · current active holdings as of {asof}", ncols)
    active = sorted(reg_data["active"], key=lambda o: o.created_date or "", reverse=True)
    _sub(ws, (f"🟢 R2 ACTIVE: {len(active)} · production runner is R2 · "
                "sorted latest entry first · rebuilt daily from canonical Registry"),
          ncols, 2)

    # CEO 2026-09-02 · CANONICAL stop source: dynamic_risk_v2 output.
    # Bridge audit is a secondary overlay for the counterfactual
    # "would_have_exited_on" only · never the primary stop value.
    dr_by_pid = _load_dynamic_risk(root, market)
    dyn_p = root / "reports" / "audit" / f"dynamic_exit_decisions_{market.lower()}_{asof}.json"
    dyn_by_pid = {}
    if dyn_p.exists():
        try:
            dyn = json.loads(dyn_p.read_text(encoding="utf-8"))
            for d in (dyn.get("decisions") or []):
                if d.get("opportunity_id"):
                    dyn_by_pid[d["opportunity_id"]] = d
        except Exception:
            pass
    sector_cache = _load_sector_cache(root)

    # CEO 2026-09-03 · full decision-transparency columns.
    # Sector · Stop Distance % · Target Distance % · Exit Horizon ·
    # Would-Have-Exited-On · Risk/Reward · Provenance added as first-class fields.
    hdr = ["Position ID", "Ticker", "Sector", "Runner",
             "Entry Date", "Entry Price", "Current Price",
             "Unrealized P&L %", "Holding Days",
             "Dynamic Stop", "Stop Distance %", "Stop Type",
             "Target", "Target Distance %", "Exit Horizon",
             "Engine Verdict", "Action", "Est. Exit Window",
             "Risk/Reward", "Provenance"]
    _header(ws, hdr, 4)
    widths = [28, 10, 18, 8, 12, 12, 14, 16, 12, 14, 14, 12, 12, 16, 12, 20, 10, 20, 12, 22]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    r = 5
    for o in active:
        entry_p = _close_on_or_before(root, o.ticker, market, o.created_date or asof)
        curr_p = _close_on_or_before(root, o.ticker, market, asof)
        pnl_pct = None
        if entry_p and curr_p and entry_p > 0:
            pnl_pct = round((curr_p - entry_p) / entry_p * 100, 2)
        days = None
        try:
            days = (date.fromisoformat(asof) - date.fromisoformat(o.created_date)).days
        except Exception:
            pass
        dr = dr_by_pid.get(o.opportunity_id) or {}
        dyn_stop = dr.get("stop")
        stop_type = dr.get("type") or "UNAVAILABLE"
        # Stop Distance % · numeric first-class column (not buried in verdict)
        stop_dist_pct = None
        if dyn_stop and curr_p and curr_p > 0:
            stop_dist_pct = round((curr_p - dyn_stop) / curr_p * 100, 2)
        # Engine Verdict from canonical stop vs current
        if dyn_stop and curr_p:
            engine_verdict = "EXIT_STOP" if curr_p <= dyn_stop else f"HOLD · stop {stop_dist_pct}% below"
        else:
            engine_verdict = "UNAVAILABLE · no canonical stop"
        dyn_d = dyn_by_pid.get(o.opportunity_id)
        # Action from verdict
        if engine_verdict.startswith("EXIT_"):
            action = "EXIT"
        elif dyn_stop is None:
            action = "REVIEW"
        else:
            action = "HOLD"
        # Target · Registry first, then ATR-based fallback (P0-consistent, m=3.0)
        target = _target_from_registry(o)
        target_source = "registry" if target else None
        if not target and entry_p:
            atr = _atr14_at_date(root, market, o.ticker, o.created_date or asof)
            target = _target_from_atr_fallback(entry_p, atr, m_target=3.0)
            if target: target_source = "atr_fallback_m3"
        target_dist_pct = None
        if target and curr_p and target > 0:
            target_dist_pct = round((target - curr_p) / curr_p * 100, 2)
        # Exit Horizon from Registry initial_signal · fallback to P0 default 60d
        horizon = _horizon_from_registry(o) or 60
        # Est. Exit Window · FORWARD-LOOKING (CEO 2026-09-03 fix ·
        # replaces backward-looking "Would-Have-Exited-On" which is
        # structurally N/A for active positions).
        est_exit_window = "UNAVAILABLE"
        try:
            if o.created_date and horizon:
                exit_est = date.fromisoformat(o.created_date) + timedelta(days=horizon)
                today_d = date.fromisoformat(asof)
                days_remaining = (exit_est - today_d).days
                est_exit_window = f"{exit_est.isoformat()} (~{days_remaining}d)"
        except Exception:
            pass
        # Risk/Reward · only when BOTH dyn_stop and target available
        rr = None
        if dyn_stop and target and curr_p:
            risk = curr_p - dyn_stop
            reward = target - curr_p
            if risk > 0 and reward > 0:
                rr = round(reward / risk, 2)
        sector = _sector_for(sector_cache, market, o.ticker)
        _write_row(ws, [
            o.opportunity_id, _canonical_ticker(o.ticker), sector, o.runner,
            o.created_date or "—",
            round(entry_p, 4) if entry_p else "UNAVAILABLE",
            round(curr_p, 4) if curr_p else "UNAVAILABLE",
            pnl_pct if pnl_pct is not None else "UNAVAILABLE",
            days if days is not None else "UNAVAILABLE",
            round(dyn_stop, 4) if dyn_stop else "UNAVAILABLE",
            stop_dist_pct if stop_dist_pct is not None else "UNAVAILABLE",
            stop_type,
            round(target, 4) if target else "UNAVAILABLE",
            target_dist_pct if target_dist_pct is not None else "UNAVAILABLE",
            horizon if horizon is not None else "UNAVAILABLE",
            engine_verdict, action, est_exit_window,
            rr if rr is not None else "UNAVAILABLE",
            f"canonical:Registry+dynamic_risk_v2+prices+sector_cache+target_src={target_source or 'none'}",
        ], r, pnl_col_idx=8)
        r += 1
    if not active:
        ws.cell(r, 1, "No current R2 ACTIVE holdings.").font = FONT_BODY
        r += 1

    r += 2
    r = _legend(ws, [
        "This sheet shows CURRENT R2 active holdings ONLY. Closed positions live in 03_Exit_History.",
        "Unrealized P&L % · (Current − Entry) / Entry · positive=green · negative=red.",
        "Dynamic Stop · canonical dynamic_risk_v2 output · authoritative production value.",
        "Stop Distance % · (Current − Stop) / Current · numeric so operator can sort by risk buffer.",
        "Target · Registry-provided OR ATR-based fallback = Entry + 3·ATR14(entry) · same math as P0 replay · Provenance shows source.",
        "Target Distance % · (Target − Current) / Current · shows upside gap.",
        "Exit Horizon · days from Registry initial_signal · defaults to 60 (P0 horizon) when source has no horizon.",
        "Est. Exit Window · FORWARD-LOOKING · entry_date + horizon · shows the planned close date + days-remaining.",
        "Risk/Reward · Reward / Risk = (Target−Current) / (Current−Stop) · only when both Target and Stop present.",
        "Action · HOLD (above stop) · EXIT (at/below stop) · REVIEW (no canonical stop available).",
        "UNAVAILABLE = canonical source returned no value · never fabricated.",
    ], r, ncols)
    return len(active)


# ── SHEET 02 · Today + Momentum · TODAY-DATE DECISIONS ONLY ─────────
def _terminal_state_to_action(state: str) -> str:
    """Map canonical terminal state to explicit operator action label.
    CEO 2026-09-02 · Action must be one of INVEST / WATCH / AVOID / NO EVIDENCE."""
    s = str(state or "").upper().strip()
    if s == "ACCEPTED":    return "INVEST"
    if s == "WATCH":        return "WATCH"
    if s == "REJECTED":    return "AVOID"
    if s == "NO_EVIDENCE": return "NO EVIDENCE"
    return "NO EVIDENCE"


def _emit_today_momentum(wb, market, root, asof, momentum_ledger):
    ws = wb.create_sheet("02_Today_Momentum")
    ncols = 13
    _banner(ws, f"AEGIS {market.upper()} · TODAY + MOMENTUM · reporting date {asof}", ncols)

    ml_entries = (momentum_ledger or {}).get("entries") or []
    ml_counts = (momentum_ledger or {}).get("by_terminal_state") or {}
    ledger_asof = str((momentum_ledger or {}).get("asof", ""))
    stale = ledger_asof and ledger_asof != asof
    freshness_note = (f"⚠ ledger asof={ledger_asof} ≠ reporting {asof}"
                       if stale else f"✓ ledger fresh for {asof}")

    # CEO 2026-09-02 · translate raw terminal_state counts to Action counts
    action_counts = {"INVEST": 0, "WATCH": 0, "AVOID": 0, "NO EVIDENCE": 0}
    for st_raw, n in (ml_counts or {}).items():
        action_counts[_terminal_state_to_action(st_raw)] += n
    _sub(ws, (f"📅 Today's R2 momentum scan · "
                f"{freshness_note} · scanned universe={(momentum_ledger or {}).get('n_universe_scanned', 0)} · "
                f"Actions: {action_counts}"), ncols, 2)

    # CEO 2026-09-03 · full decision context · Action first · Sector +
    # Target + Risk/Reward + Provenance added · UNAVAILABLE where the
    # canonical momentum_ledger source does not structurally provide the value.
    hdr = ["Action", "Ticker", "Sector", "Category", "Quality Band",
             "Current Price", "Entry Zone", "Stop", "Target", "Confidence",
             "Risk/Reward",
             "Return 1d %", "Return 5d %", "Return 20d %",
             "Reason", "As-Of", "Provenance"]
    _header(ws, hdr, 4)
    widths = [12, 10, 18, 14, 12, 14, 14, 12, 12, 12, 12, 12, 12, 12, 40, 12, 22]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    if stale:
        entries_for_today = []
    else:
        entries_for_today = ml_entries

    # Sort: INVEST first, then WATCH, AVOID, NO EVIDENCE · then ticker A-Z
    action_order = {"INVEST": 0, "WATCH": 1, "AVOID": 2, "NO EVIDENCE": 3}
    entries_sorted = sorted(entries_for_today, key=lambda e: (
        action_order.get(_terminal_state_to_action(e.get("terminal_state", "")), 9),
        e.get("ticker", "") or ""
    ))
    sector_cache_m = _load_sector_cache(root)
    r = 5
    for e in entries_sorted[:200]:
        ticker = _canonical_ticker(e.get("ticker"))
        action = _terminal_state_to_action(e.get("terminal_state", ""))
        curr_p = _close_on_or_before(root, e.get("ticker") or "", market, asof)
        entry_zone = e.get("entry_zone") or e.get("entry_zone_str") or "UNAVAILABLE"
        stop = e.get("stop") or e.get("suggested_stop")
        target = e.get("target") or e.get("t1") or e.get("target_price")
        # Confidence · populate for EVERY action · diagnostic info about
        # signal strength, not just execution parameter. CEO 2026-09-03.
        # Try multiple aliases; last resort read from raw_confidence * 100.
        conf = (e.get("confidence")
                or e.get("confidence_pct")
                or e.get("calibrated_confidence")
                or e.get("raw_confidence"))
        if conf is not None and isinstance(conf, (int, float)) and 0 <= float(conf) <= 1:
            conf = round(float(conf) * 100, 1)   # promote to pct
        reason = str(e.get("reason_text", "") or "")[:80] or "UNAVAILABLE"
        sector = _sector_for(sector_cache_m, market, e.get("ticker") or "")
        # Risk/Reward · only when BOTH stop and target and curr present canonically
        rr = None
        if (isinstance(stop, (int, float)) and stop > 0
             and isinstance(target, (int, float)) and target > 0 and curr_p):
            risk = curr_p - stop
            reward = target - curr_p
            if risk > 0 and reward > 0:
                rr = round(reward / risk, 2)
        _write_row(ws, [
            action, ticker, sector,
            e.get("category", "UNAVAILABLE"), e.get("quality_band", "UNAVAILABLE"),
            round(curr_p, 4) if curr_p else "UNAVAILABLE",
            entry_zone if entry_zone != "—" else "UNAVAILABLE",
            round(stop, 4) if isinstance(stop, (int, float)) and stop > 0 else "UNAVAILABLE",
            round(target, 4) if isinstance(target, (int, float)) and target > 0 else "UNAVAILABLE",
            f"{conf}%" if isinstance(conf, (int, float)) else "UNAVAILABLE",
            rr if rr is not None else "UNAVAILABLE",
            e.get("return_1d_pct") if e.get("return_1d_pct") is not None else "UNAVAILABLE",
            e.get("return_5d_pct") if e.get("return_5d_pct") is not None else "UNAVAILABLE",
            e.get("return_20d_pct") if e.get("return_20d_pct") is not None else "UNAVAILABLE",
            reason, asof, "canonical:momentum_ledger+prices+sector_cache",
        ], r)
        r += 1
    if not entries_sorted:
        ws.cell(r, 1, (f"No R2 decisions/recommendations for {asof} "
                         + ("· momentum ledger is stale" if stale else "")))
        r += 1

    r += 2
    r = _legend(ws, [
        "Action → INVEST (eligible for new R2 entry today) · WATCH (monitor · do not enter) · "
        "AVOID (do not initiate) · NO EVIDENCE (insufficient data).",
        "Current Price · from canonical parquet close on As-Of · never fabricated.",
        "Entry Zone / Stop · populated for INVEST verdicts only (execution parameters).",
        "Confidence · populated for ALL verdicts (WATCH/AVOID included) · diagnostic signal-strength, not execution-only.",
        "Zero-fabrication invariant · UNAVAILABLE means the canonical source returned no value.",
        "This sheet is REGENERATED from scratch every reporting day · never carries yesterday forward.",
        "An INVEST recommendation is a signal · it becomes a Portfolio holding only after "
        "canonical lifecycle transition creates a Registry ACTIVE position.",
    ], r, ncols)
    return {"n_entries": len(entries_sorted), "ledger_fresh": not stale,
            "action_counts": action_counts}


# ── SHEET 03 · Exit History · GENUINE REALIZED PRODUCTION EXITS ONLY
def _emit_exit_history(wb, market, root, asof, reg_data):
    ws = wb.create_sheet("03_Exit_History")
    ncols = 14   # +1 for Relative Opportunity pp column (CEO 2026-09-02)
    _banner(ws, f"AEGIS {market.upper()} · EXIT HISTORY · realized · as of {asof}", ncols)
    # CEO 2026-09-02 · closed_90d from _load_registry is ALREADY the
    # non-admin non-retired set. Admin events → closed_admin_90d (→ orphan_audit).
    # Retired → closed_retired_90d (→ orphan_audit). This filter chain
    # is single-source-of-truth in _load_registry.
    closed = sorted(reg_data["closed_90d"],
                       key=lambda o: o.closed_date or "", reverse=True)
    n_admin = len(reg_data.get("closed_admin_90d", []))
    n_retired = len(reg_data.get("closed_retired_90d", []))
    # CEO 2026-09-03 · simplified banner · the two side-population counts
    # were confusing (looked like they should sum to the main count).
    # Now: primary count on its own line · audit-sink counts on a separate
    # "excluded" line so it's clear these are DIFFERENT populations.
    _sub(ws, (f"📕 This sheet · {len(closed)} realized production exits (last 90d) · "
                f"newest first · realized P&L only"),
          ncols, 2)
    _sub(ws, (f"    Excluded (routed to orphan_audit_{market.lower()}.jsonl · not P&L): "
                f"{n_admin} administrative (same-day/zero-Δ) · {n_retired} retired-runner"),
          ncols, 3)

    sector_cache = _load_sector_cache(root)

    # CEO 2026-09-02 · Relative Opportunity pp is its own column · never
    # mixed into the Exit Reason string (was being misread as realized P&L).
    hdr = ["Position ID", "Stock", "Sector", "Runner", "Market",
             "Entry Date", "Exit Date", "Holding Days",
             "Entry Price", "Exit Price", "Realized P&L %",
             "Exit Reason", "Relative Opportunity vs Rotation (pp)",
             "Provenance"]
    _header(ws, hdr, 4)
    for i, w in enumerate([28, 10, 18, 8, 8, 12, 12, 12, 12, 12, 16, 22, 20, 22], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    r = 5
    n_priced = 0
    n_unpriced = 0
    for o in closed:
        entry_p = _close_on_or_before(root, o.ticker, market, o.created_date or "")
        exit_p = _close_on_or_before(root, o.ticker, market, o.closed_date or "")
        pnl_pct = None
        if entry_p and exit_p and entry_p > 0:
            pnl_pct = round((exit_p - entry_p) / entry_p * 100, 2)
            n_priced += 1
        else:
            n_unpriced += 1
        days = None
        try:
            days = (date.fromisoformat(o.closed_date) - date.fromisoformat(o.created_date)).days
        except Exception:
            pass
        sector = _sector_for(sector_cache, market, o.ticker)
        raw_reason = getattr(o, "closed_reason", "") or "—"
        rel_pp = _extract_relative_pp(raw_reason)
        _write_row(ws, [
            o.opportunity_id, _canonical_ticker(o.ticker), sector,
            o.runner, market.upper(),
            o.created_date or "—", o.closed_date or "—",
            days if days is not None else "—",
            round(entry_p, 4) if entry_p else "UNAVAILABLE",
            round(exit_p, 4) if exit_p else "UNAVAILABLE",
            pnl_pct if pnl_pct is not None else "—",
            _normalize_exit_reason(raw_reason),
            rel_pp if rel_pp is not None else "—",
            "canonical:Registry+prices",
        ], r, pnl_col_idx=11)
        r += 1

    if not closed:
        ws.cell(r, 1, "No closed R2 positions in last 90 days.").font = FONT_BODY
        r += 1

    r += 2
    r = _legend(ws, [
        f"Priced: {n_priced} · Unpriced (data unavailable): {n_unpriced} · "
        "unpriced rows show — for P&L, never fabricated 0.",
        "Realized P&L % · (Exit − Entry) / Entry · own-trade result · "
        "positive=green · negative=red.",
        "Relative Opportunity pp · alpha/rotation benefit vs the closed position "
        "(pp = percentage points) · this is NOT the trade's own P&L · "
        "shown separately so it cannot be misread.",
        f"{n_admin} same-day / zero-Δ administrative events + {n_retired} "
        "retired-runner events filtered from this operator-facing view · they "
        "remain in the canonical Registry JSONL and orphan_audit sink for audit.",
        "This sheet contains ONLY realized production exits · 90-day rolling window.",
    ], r, ncols)
    return len(closed)


# ── SHEET 04 · Daily Portfolio History · CEO 2026-09-03 ─────────────
def _emit_daily_history(wb, market, root, asof, reg_data):
    """Reconstruct historical daily active portfolio from canonical Registry.

    For every trading day from the earliest R2 entry through as-of:
      · one row per position while that position was genuinely active
      · same-day admin events excluded (they never had an "active" state)
      · Dynamic Stop populated only where dynamic_risk_v2 has canonical data
      · UNAVAILABLE for historical dates that have no canonical stop record

    This sheet answers "what did AEGIS hold on date X?" for any prior date.
    Registry is the sole source of truth · never carries forward from XLSX."""
    from datetime import date as _date, timedelta as _td
    ws = wb.create_sheet("04_Daily_Portfolio_History")
    ncols = 13
    _banner(ws, f"AEGIS {market.upper()} · DAILY PORTFOLIO HISTORY · reconstructed from canonical Registry as of {asof}", ncols)

    # Combine active + closed to find full R2 universe
    all_positions = list(reg_data.get("active", [])) + list(reg_data.get("closed_90d", []))
    if not all_positions:
        _sub(ws, "No R2 positions in canonical Registry", ncols, 2)
        hdr = ["As-Of Date", "Position ID", "Ticker", "Runner", "Status",
                 "Entry Date", "Entry Price", "Close Price", "P&L %",
                 "Dynamic Stop", "Engine Verdict", "Action", "Provenance"]
        _header(ws, hdr, 4)
        return 0

    # Find genuine positions · exclude ALL structurally administrative
    # events (same-day OR zero-delta price · matches builder's admin filter
    # used in Exit History). CEO 2026-09-03 fix: previous filter only
    # checked same-day and let ORPHAN_AUTO_CLOSE zero-delta events through
    # as "active history" (USA-R2-RF / USA-R2-CSRA orphan artifacts).
    genuine = []
    for o in all_positions:
        cd = str(o.created_date or "")[:10]
        if not cd: continue
        if o.status == "CLOSED":
            ep = _close_on_or_before(root, o.ticker, market, cd)
            xp = _close_on_or_before(root, o.ticker, market, o.closed_date or "")
            if _is_administrative_exit(o, ep, xp):
                continue
        genuine.append(o)
    if not genuine:
        _sub(ws, "No genuine R2 positions to reconstruct", ncols, 2)
        return 0

    earliest = min(str(o.created_date or "9999-99-99")[:10] for o in genuine)
    try:
        start_d = _date.fromisoformat(earliest)
        end_d = _date.fromisoformat(asof)
    except Exception:
        _sub(ws, "Cannot reconstruct · invalid date range", ncols, 2)
        return 0

    _sub(ws, (f"📅 Reconstructed from Registry · {len(genuine)} genuine R2 positions · "
                f"trading days {earliest} → {asof} · excludes same-day admin events"),
          ncols, 2)

    hdr = ["As-Of Date", "Position ID", "Ticker", "Runner", "Status",
             "Entry Date", "Entry Price", "Close Price", "P&L %",
             "Dynamic Stop", "Engine Verdict", "Action", "Provenance"]
    _header(ws, hdr, 4)
    widths = [12, 28, 10, 8, 10, 12, 12, 12, 12, 14, 22, 10, 22]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Cache dynamic_risk history · only current snapshot is canonical ·
    # historical dynamic stops are UNAVAILABLE (never fabricated).
    dr_today = _load_dynamic_risk(root, market)

    r = 5
    n_rows = 0
    d = start_d
    while d <= end_d:
        if d.weekday() >= 5:
            d += _td(days=1); continue
        d_iso = d.isoformat()
        for o in genuine:
            cd = str(o.created_date or "")[:10]
            xd = str(o.closed_date or "")[:10] if o.status == "CLOSED" else ""
            if not cd or cd > d_iso: continue
            if xd and xd <= d_iso: continue
            close_p = _close_on_or_before(root, o.ticker, market, d_iso)
            entry_p = _close_on_or_before(root, o.ticker, market, cd)
            pnl = None
            if entry_p and close_p and entry_p > 0:
                pnl = round((close_p - entry_p) / entry_p * 100, 2)
            # Dynamic stop: only current snapshot is canonical
            if d_iso == asof:
                dr = dr_today.get(o.opportunity_id) or {}
                dyn_stop = dr.get("stop")
            else:
                dyn_stop = None
            if dyn_stop and close_p:
                verdict = "EXIT_STOP" if close_p <= dyn_stop else "HOLD"
                action = "EXIT" if close_p <= dyn_stop else "HOLD"
            else:
                verdict = "UNAVAILABLE" if d_iso != asof else "UNAVAILABLE · no canonical stop"
                action = "REVIEW" if d_iso == asof else "UNAVAILABLE"
            _write_row(ws, [
                d_iso, o.opportunity_id, _canonical_ticker(o.ticker), o.runner,
                "ACTIVE",
                cd,
                round(entry_p, 4) if entry_p else "UNAVAILABLE",
                round(close_p, 4) if close_p else "UNAVAILABLE",
                pnl if pnl is not None else "UNAVAILABLE",
                round(dyn_stop, 4) if dyn_stop else "UNAVAILABLE",
                verdict, action,
                "canonical:Registry+prices" + ("+dynamic_risk_v2" if dyn_stop else ""),
            ], r, pnl_col_idx=9)
            r += 1
            n_rows += 1
        d += _td(days=1)

    r += 2
    r = _legend(ws, [
        "Reconstructed daily active-portfolio history from canonical Registry.",
        "One row per position per trading day while the position was genuinely active.",
        "Excludes same-day administrative Registry records (entry_date == exit_date).",
        "Dynamic Stop is populated ONLY for the current as-of date · historical dynamic "
        "stops are UNAVAILABLE because dynamic_risk_v2 stores only the current snapshot.",
        "Never fabricated · UNAVAILABLE means the canonical source has no value for that day.",
    ], r, ncols)
    return n_rows


def build_workbook(market: str, root: Path, asof: str) -> dict:
    from backend.delivery.canonical.retirement import retired_runners
    retired = retired_runners(root)
    reg_data = _load_registry(root, market, retired)
    ml_p = root / "reports" / "research" / "multi_layer" / f"momentum_ledger_{market.lower()}_{asof}.json"
    momentum_ledger = json.loads(ml_p.read_text(encoding="utf-8")) if ml_p.exists() else {}

    # Emit orphan_audit_{market}.jsonl for retired-runner CLOSED
    # positions BEFORE building the workbook (workbook is R1-zero ·
    # A23 lineage validation reads this sink).
    orphan_stats = _emit_orphan_audit_for_retired(root, market, reg_data)

    wb = Workbook()
    wb.remove(wb.active)
    n_active = _emit_portfolio(wb, market, root, asof, reg_data)
    today_stats = _emit_today_momentum(wb, market, root, asof, momentum_ledger)
    n_closed = _emit_exit_history(wb, market, root, asof, reg_data)
    n_history_rows = _emit_daily_history(wb, market, root, asof, reg_data)

    # Sprint A · optional sheets driven by configs/aegis_runner_registry.yaml
    # Base 4 above are HARD LOCKED · 05/06 are additive (append-only).
    optional_emitted = _emit_optional_sprint_a_sheets(wb, market, root, asof, reg_data)

    xlsx_dated = root / "reports" / "telegram" / f"aegis_{market.lower()}_{asof}.xlsx"
    xlsx_undated = root / "reports" / "telegram" / f"aegis_history_{market.lower()}.xlsx"
    xlsx_dated.parent.mkdir(parents=True, exist_ok=True)
    wb.save(xlsx_dated)
    import shutil
    shutil.copyfile(xlsx_dated, xlsx_undated)
    return {
        "market": market.lower(),
        "asof": asof,
        "sheets": list(wb.sheetnames),
        "active_holdings": n_active,
        "today_stats": today_stats,
        "closed_positions": n_closed,
        "daily_history_rows": n_history_rows,
        "optional_sprint_a_sheets": optional_emitted,
        "xlsx_dated": str(xlsx_dated.relative_to(root)),
        "xlsx_undated": str(xlsx_undated.relative_to(root)),
    }


def _emit_optional_sprint_a_sheets(wb, market: str, root: Path, asof: str,
                                    reg_data: dict) -> list[str]:
    """Emit 00_Health, 05_R1_Advisory and/or 06_Composite_Signals when
    configs/aegis_runner_registry.yaml declares them.

    Sheets are ADDITIVE · never touch base 4. Safe no-op if config missing.
    00_Health always renders (cockpit governance surface).
    """
    emitted: list[str] = []

    # 00_Health · always render (governance cockpit · CEO 2026-09-03)
    try:
        _emit_health_cockpit_sheet(wb, market, root, asof)
        # Reorder so 00_Health is FIRST in the sheet tab list
        health_sheet = wb["00_Health"]
        wb.move_sheet(health_sheet, offset=-len(wb.sheetnames) + 1)
        emitted.append("00_Health")
    except Exception as e:
        print(f"[optional-sheet] 00_Health skipped: {e}", file=sys.stderr)
    try:
        import yaml
        cfg = yaml.safe_load(
            (root / "configs" / "aegis_runner_registry.yaml").read_text(encoding="utf-8")
        ) or {}
    except Exception:
        return emitted
    runners_cfg = cfg.get("runners", {}) or {}
    comp_cfg = cfg.get("composite", {}) or {}

    # 05_R1_Advisory · when R1 workbook_visibility == advisory_only
    r1_vis = str(runners_cfg.get("R1", {}).get("workbook_visibility", "") or "")
    if r1_vis == "advisory_only":
        try:
            _emit_r1_advisory_sheet(wb, market, root, asof)
            emitted.append("05_R1_Advisory")
        except Exception as e:
            # Never break the base 4 · log softly
            print(f"[optional-sheet] 05_R1_Advisory skipped: {e}", file=sys.stderr)

    # 06_Composite_Signals · when composite workbook_visibility == shadow
    comp_vis = str(comp_cfg.get("workbook_visibility", "") or "")
    if comp_vis == "shadow":
        try:
            _emit_composite_signals_sheet(wb, market, root, asof)
            emitted.append("06_Composite_Signals")
        except Exception as e:
            print(f"[optional-sheet] 06_Composite_Signals skipped: {e}", file=sys.stderr)

    return emitted


def _emit_health_cockpit_sheet(wb, market: str, root: Path, asof: str) -> None:
    """Render the 00_Health governance cockpit sheet."""
    from backend.delivery.sheets.health_cockpit_sheet import (
        sheet_meta, build_health_rows, HEALTH_BANNER,
    )
    meta = sheet_meta()
    ws = wb.create_sheet(meta["sheet_name"])
    ncols = len(meta["columns"])
    _banner(ws, HEALTH_BANNER, ncols)
    _header(ws, meta["columns"], 4)
    rows = build_health_rows(root, market.lower())
    for r_idx, row in enumerate(rows, start=5):
        for c_idx, val in enumerate(row, start=1):
            ws.cell(row=r_idx, column=c_idx, value=val)


def _emit_r1_advisory_sheet(wb, market: str, root: Path, asof: str) -> None:
    """Render the 05_R1_Advisory sheet · uses builders in backend/delivery/sheets/."""
    from backend.delivery.sheets.r1_advisory_sheet import (
        sheet_meta, build_r1_advisory_rows, ADVISORY_BANNER,
    )
    meta = sheet_meta()
    ws = wb.create_sheet(meta["sheet_name"])
    ncols = len(meta["columns"])
    _banner(ws, ADVISORY_BANNER, ncols)
    # Read today's R1 picks · degrade gracefully if not present.
    # Real R1 CSV schema (from adaptive_rec_v2 daily output):
    #   Generated · Profile · Stock · Sector · Strength · Score /100 · Current Price
    #   Buy Range · Hist Target · Prob +ve · Rec Confidence % · Why
    # Normalize into the fields the sheet builder expects (ticker · action · rank · ...).
    r1_picks: list[dict] = []
    # Per-market candidates · India can fall back to legacy aegis_today.csv ·
    # USA MUST NOT (that file is India's daily R1 output · loading it into USA
    # would misattribute India picks as USA · verified 2026-09-03).
    picks_candidates = [root / "data" / f"aegis_today_{market.lower()}.csv"]
    if market.lower() == "india":
        picks_candidates.append(root / "data" / "aegis_today.csv")
    for picks_path in picks_candidates:
        if not picks_path.exists(): continue
        try:
            import pandas as pd
            df_picks = pd.read_csv(picks_path)
            # Modern schema
            if "runner" in df_picks.columns:
                df_picks = df_picks[df_picks["runner"].astype(str).str.upper() == "R1"]
            # Legacy R1 schema · Stock/Profile columns
            if "Stock" in df_picks.columns and "ticker" not in df_picks.columns:
                df_picks = df_picks.rename(columns={
                    "Stock": "ticker",
                    "Sector": "sector",
                    "Strength": "action",
                    "Score /100": "score",
                    "Buy Range": "entry_zone",
                    "Why": "bull_case",
                    "Rec Confidence %": "confidence",
                    "Hist Target": "target",
                })
            r1_picks = df_picks.head(25).to_dict("records")
            if r1_picks: break   # first source with data wins
        except Exception:
            continue
    # Load KG filter result if fresh
    kg_path = root / "reports" / "research" / f"r1_kg_group_filter_{market.lower()}.json"
    kg_result = {}
    if kg_path.exists():
        try:
            kg_result = json.loads(kg_path.read_text(encoding="utf-8"))
        except Exception:
            kg_result = {}
    _header(ws, meta["columns"], 4)
    rows = build_r1_advisory_rows(root, market.lower(), asof, r1_picks, kg_result)
    for r_idx, row in enumerate(rows, start=5):
        for c_idx, val in enumerate(row, start=1):
            ws.cell(row=r_idx, column=c_idx, value=val)


def _emit_composite_signals_sheet(wb, market: str, root: Path, asof: str) -> None:
    """Render the 06_Composite_Signals sheet · shadow · uses composite engine.
    V2 §19 · daily loop landed 2026-09-03 · sheet now populates."""
    from backend.delivery.sheets.composite_signals_sheet import (
        sheet_meta, build_composite_rows, COMPOSITE_BANNER, COMPOSITE_COLUMNS,
    )
    from backend.recommendation.composite.daily_loop import run_composite_daily
    meta = sheet_meta()
    ws = wb.create_sheet(meta["sheet_name"])
    ncols = len(meta["columns"])
    _banner(ws, COMPOSITE_BANNER, ncols)
    _header(ws, meta["columns"], 4)
    try:
        payload = run_composite_daily(root, market.lower(), asof)
        signals = payload.get("signals", []) or []
        r_idx = 5
        for s in signals:
            row = [
                s.get("ticker") or "",
                s.get("sector") or "",
                round(float(s.get("R1_score") or 0.0), 4),
                round(float(s.get("R2_score") or 0.0), 4),
                round(float(s.get("R3_score") or 0.0), 4),
                round(float((s.get("trust_weights_normalized") or {}).get("R1", 0.0)), 4),
                round(float((s.get("trust_weights_normalized") or {}).get("R2", 0.0)), 4),
                round(float((s.get("trust_weights_normalized") or {}).get("R3", 0.0)), 4),
                round(float(s.get("composite_score") or 0.0), 4),
                str(s.get("conviction") or ""),
                int(s.get("n_runners_active") or 0),
                ("YES" if (s.get("admissions") or {}).get("R3") == "ADMITTED" else "NO"),
                "shadow only · no P&L",
            ]
            for c_idx, val in enumerate(row, start=1):
                ws.cell(row=r_idx, column=c_idx, value=val)
            r_idx += 1
        if not signals:
            ws.cell(row=5, column=1, value="(no runner scores today · composite has nothing to fuse)")
    except Exception as e:
        ws.cell(row=5, column=1, value=f"(composite loop error · {str(e)[:80]})")


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--market", choices=["india", "usa", "both"], default="both")
    ap.add_argument("--asof", default=date.today().isoformat(),
                     help="Reporting date · defaults to today · used for filenames + snapshot")
    args = ap.parse_args()
    for m in (["india", "usa"] if args.market == "both" else [args.market]):
        rep = build_workbook(m, _ROOT, args.asof)
        print(json.dumps(rep, indent=2, ensure_ascii=False, default=str))
    return 0


if __name__ == "__main__":
    sys.exit(main())
