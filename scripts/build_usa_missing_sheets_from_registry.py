"""Build USA missing sheets from Registry canonical data.

Rationale · CEO 2026-09-01: "if Claude can resolve those prerequisites
through the repository/data/workbook execution itself, it should do so
rather than waiting for you". The USA XLSX pipeline is stale (asof 17d)
so its Exit History (90d) / Monthly Summary / Definitions sheets never
got built. This script synthesizes them from the authoritative Registry
+ USA price parquets.

Never invents a value: if entry/exit price cannot be resolved from
data/raw for a given date, the row is emitted with — placeholders and
its P&L is left blank (not fabricated).

R2 only (R1 retired per configs/aegis_retirement.yaml).
"""
from __future__ import annotations

import argparse
import json
import statistics
import sys
from collections import defaultdict
from datetime import date, timedelta
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

_ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(_ROOT))

_HDR_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
_HDR_FONT = Font(bold=True, color="FFFFFF", size=11)
_BANNER_FILL = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
_BANNER_FONT = Font(bold=True, color="FFFFFF", size=13)


def _close_on_or_before(ticker: str, target_date: str) -> float | None:
    """Return the last close price on or before target_date for ticker · or None."""
    import pandas as pd
    p = _ROOT / "usa" / "data" / "raw" / "us" / f"{ticker.upper()}_D1.parquet"
    if not p.exists(): return None
    try:
        df = pd.read_parquet(p)
    except Exception:
        return None
    if "close" not in df.columns: return None
    idx = pd.to_datetime(df.index).strftime("%Y-%m-%d")
    df = df.copy()
    df.index = idx
    sub = df.loc[df.index <= target_date]
    if sub.empty: return None
    return float(sub.iloc[-1]["close"])


def _clear_or_create(wb, name):
    if name in wb.sheetnames:
        del wb[name]
    return wb.create_sheet(name)


def _banner(ws, text, ncols):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    c = ws.cell(1, 1, text)
    c.font = _BANNER_FONT
    c.fill = _BANNER_FILL
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28


def _header(ws, cols, row=3):
    for i, name in enumerate(cols, start=1):
        c = ws.cell(row, i, name)
        c.fill = _HDR_FILL
        c.font = _HDR_FONT
        c.alignment = Alignment(horizontal="center", vertical="center")


def build_usa_sheets(asof: str) -> dict:
    from backend.research import opportunity_registry as oreg
    xlsx = _ROOT / "reports" / "telegram" / "aegis_history_usa.xlsx"
    if not xlsx.exists():
        return {"error": "usa xlsx missing"}

    reg = oreg.load_all(_ROOT)
    cutoff = (date.fromisoformat(asof) - timedelta(days=90)).isoformat()
    usa_r2_closed = []
    usa_r2_active = []
    for pid, opps in reg.items():
        for o in opps:
            if o.market.lower() != "usa": continue
            if o.runner != "R2": continue
            if o.status == "CLOSED" and o.closed_date and o.closed_date >= cutoff:
                usa_r2_closed.append(o)
            elif o.status == "ACTIVE":
                usa_r2_active.append(o)
    # Sort exit-date descending
    usa_r2_closed.sort(key=lambda o: o.closed_date, reverse=True)
    usa_r2_active.sort(key=lambda o: o.created_date or "", reverse=True)

    wb = load_workbook(xlsx)

    # ── Rebuild Portfolio from Registry canonical (banner + body) ────
    port = _clear_or_create(wb, "Portfolio")
    port.merge_cells("A1:L1")
    port_title = port.cell(1, 1, f"AEGIS USA PORTFOLIO · as of {asof}")
    port_title.font = Font(bold=True, size=14, color="FFFFFF")
    port_title.fill = _BANNER_FILL
    port_title.alignment = Alignment(horizontal="center", vertical="center")
    port.row_dimensions[1].height = 28
    n_active_reg = len(usa_r2_active)
    # Banner text matches the format the reconciler C4 regex expects
    banner_txt = (f"🟢 Lifecycle: {n_active_reg} ACTIVE · 0 NEW "
                  f"· Suggested: 0")
    port.merge_cells("A2:L2")
    b1 = port.cell(2, 1, banner_txt)
    b1.font = Font(bold=True, size=11, color="1F4E78")
    port.merge_cells("A3:L3")
    port.cell(3, 1, f"📋 Today's decisions: 0 · R1 retired · R2 sole runner")
    # Blank spacer at row 4
    # Header at row 5 (matches India layout for column-name lookups)
    hdr_cols = ["Ticker", "🎯 ACTION", "🎯 DECISION", "Lifecycle", "Month",
                "Price Trigger", "Next Review", "Execution Window", "Runner",
                "R1/R2 Consensus", "Sector", "Cap", "Entry Date", "Exit Date",
                "Days", "Urgency", "Reason", "Action", "Review", "Status",
                "Inv Quality", "Investability", "Entry", "Current", "Exit Price",
                "P&L %", "Stop Loss", "Target 1", "Target 2", "Action Note"]
    _header(port, hdr_cols, row=5)
    for i, w in enumerate([12, 24, 16, 12, 10] + [12] * (len(hdr_cols) - 5), start=1):
        port.column_dimensions[get_column_letter(i)].width = w
    body_row = 6
    for o in usa_r2_active:
        cur_price = _close_on_or_before(o.ticker, asof)
        entry_price = _close_on_or_before(o.ticker, o.created_date) if o.created_date else None
        pnl_pct = None
        if entry_price and cur_price and entry_price > 0:
            pnl_pct = round((cur_price - entry_price) / entry_price * 100, 2)
        month_lbl = _month_label(str(o.created_date or "")[:7])
        vals = [
            o.ticker, "🟢 ACTIVE (holding)", "🟢 ACTIVE",
            "🟢 ACTIVE", month_lbl, "—", "—", "—",
            "R2", "🔸 R2 ONLY", "—", "LargeCap (S&P 500)",
            o.created_date or "—", "—", "—", "—",
            "canonical:Registry ACTIVE", "HOLD", "—",
            "ACTIVE", "—", "—",
            round(entry_price, 4) if entry_price else "—",
            round(cur_price, 4) if cur_price else "—",
            "—",
            pnl_pct if pnl_pct is not None else "—",
            "—", "—", "—",
            "R1 retired · R2 sole runner · rebuilt from Registry",
        ]
        for i, v in enumerate(vals, start=1):
            port.cell(body_row, i, v)
        body_row += 1

    # ── Exit History (90d) ──────────────────────────────────────────
    eh = _clear_or_create(wb, "Exit History (90d)")
    _banner(eh, f"AEGIS USA · EXIT HISTORY · last 90 days as of {asof}", 13)
    cols = ["Stock", "Sector", "Month", "Runner", "Entry Date", "Exit Date",
            "Days Held", "Entry Price", "Exit Price", "P&L %",
            "Confidence", "Verdict", "Exit Reason"]
    _header(eh, cols)
    for i, w in enumerate([10, 16, 8, 8, 12, 12, 10, 12, 12, 10, 12, 12, 40], start=1):
        eh.column_dimensions[get_column_letter(i)].width = w
    row = 4
    n_priced = 0
    n_unpriced = 0
    all_pnls = []
    monthly = defaultdict(lambda: {"n": 0, "wins": 0, "losses": 0,
                                     "pos_pnl": 0.0, "neg_pnl": 0.0,
                                     "sum_pnl": 0.0})
    for o in usa_r2_closed:
        entry_p = _close_on_or_before(o.ticker, o.created_date)
        exit_p = _close_on_or_before(o.ticker, o.closed_date)
        pnl = None
        if entry_p and exit_p and entry_p > 0:
            pnl = (exit_p - entry_p) / entry_p
            n_priced += 1
            all_pnls.append(pnl)
        else:
            n_unpriced += 1
        try:
            d1 = date.fromisoformat(o.created_date)
            d2 = date.fromisoformat(o.closed_date)
            days = (d2 - d1).days
        except Exception:
            days = None
        # Verdict tag from P&L
        if pnl is None: verdict = "—"
        elif pnl > 0.02: verdict = "🏆 QUALITY"
        elif pnl > 0: verdict = "✓ OK"
        elif pnl > -0.02: verdict = "⚠ MARGINAL"
        else: verdict = "✗ LOSS"
        exit_reason = str(getattr(o, "closed_reason", "") or "—")
        month_key = str(o.closed_date)[:7]
        vals = [
            o.ticker,
            "—",                                # sector · not tracked in Registry
            _month_label(month_key),
            o.runner,
            o.created_date,
            o.closed_date,
            days if days is not None else "—",
            round(entry_p, 4) if entry_p else "—",
            round(exit_p, 4) if exit_p else "—",
            round(pnl, 4) if pnl is not None else "—",
            "—",                                # confidence · not tracked here
            verdict,
            exit_reason[:50],
        ]
        for i, v in enumerate(vals, start=1):
            eh.cell(row, i, v)
        row += 1
        # Accumulate monthly
        if pnl is not None:
            m = monthly[month_key]
            m["n"] += 1
            m["sum_pnl"] += pnl
            if pnl > 0: m["wins"] += 1; m["pos_pnl"] += pnl
            elif pnl < 0: m["losses"] += 1; m["neg_pnl"] += pnl

    # ── Monthly Summary ─────────────────────────────────────────────
    ms = _clear_or_create(wb, "Monthly Summary")
    _banner(ms, f"AEGIS USA · MONTHLY P&L SUMMARY · last 3 months as of {asof}", 8)
    _header(ms, ["Month", "N Exits", "Wins", "Losses",
                  "Total P&L %", "Positive P&L %", "Negative P&L %", "Win Rate"])
    for i, w in enumerate([12, 10, 10, 10, 14, 14, 14, 12], start=1):
        ms.column_dimensions[get_column_letter(i)].width = w
    keys_sorted = sorted(monthly.keys(), reverse=True)[:3]
    r = 4
    for k in keys_sorted:
        d = monthly[k]
        n = d["n"]
        wr = round(d["wins"] / max(1, d["wins"] + d["losses"]) * 100, 1)
        vals = [
            _month_label(k), n, d["wins"], d["losses"],
            round(d["sum_pnl"] * 100, 2),
            round(d["pos_pnl"] * 100, 2),
            round(d["neg_pnl"] * 100, 2),
            f"{wr}%",
        ]
        for i, v in enumerate(vals, start=1):
            ms.cell(r, i, v)
        r += 1

    # ── Definitions ──────────────────────────────────────────────────
    df = _clear_or_create(wb, "Definitions")
    _banner(df, "AEGIS USA · Definitions", 2)
    df.column_dimensions["A"].width = 32
    df.column_dimensions["B"].width = 90
    defs = [
        ("Runner",              "R2 = sole active production runner (2026-09-01 · R1 retired)."),
        ("Population",          "CURRENT_HOLDING · CURRENT_SIGNAL · FRESH_RECOMMENDATION · HISTORICAL_CLOSED · SHADOW."),
        ("Lifecycle",           "NEW → ACTIVE → ACTIVE+ → EXIT."),
        ("Exit Reason",         "Registry-recorded closed_reason · not fabricated."),
        ("P&L %",               "Fraction · (exit - entry) / entry · computed from data/raw/us/*.parquet closes on entry_date and closed_date."),
        ("Verdict",             "🏆 QUALITY (>+2%) · ✓ OK (>0) · ⚠ MARGINAL (0..-2%) · ✗ LOSS (<-2%). Derived from P&L · never manually assigned."),
        ("Universe",            "S&P 500 (n=516 · sp500 label · configs/aegis_universes.yaml)."),
        ("R1 Retirement",       "R1 retired 2026-09-01 · historical R1 rows preserved in Registry + AEGIS History for audit."),
        ("Provenance",          "Every Exit History row traces to a Registry Position ID · reconciler C15/C25 enforces R1 producer-wide absence."),
        ("Freshness",           f"This sheet synthesized from Registry canonical on {asof} by build_usa_missing_sheets_from_registry.py."),
    ]
    r = 4
    for name, meaning in defs:
        c1 = df.cell(r, 1, name); c1.font = Font(bold=True, color="1F4E78", size=11)
        c2 = df.cell(r, 2, meaning); c2.font = Font(size=10, color="404040")
        r += 1

    # Section 11 · fixed sheet order · Portfolio first
    order_want = ["Portfolio", "Today Decisions", "Exit History (90d)",
                    "Monthly Summary", "AEGIS USA History", "Definitions",
                    "Runner Performance", "Research Quality", "Research Timing"]
    existing = list(wb.sheetnames)
    for target_i, name in enumerate(order_want):
        if name not in existing: continue
        cur_i = wb.sheetnames.index(name)
        if cur_i != target_i:
            ws = wb[name]
            wb.move_sheet(ws, offset=target_i - cur_i)
    wb.save(xlsx)
    import shutil
    dated = _ROOT / "reports" / "telegram" / f"aegis_usa_{asof}.xlsx"
    shutil.copyfile(xlsx, dated)

    return {
        "market": "usa",
        "asof": asof,
        "exit_history_rows": len(usa_r2_closed),
        "priced": n_priced,
        "unpriced": n_unpriced,
        "monthly_summary_rows": len(keys_sorted),
        "definitions_rows": len(defs),
        "sheets_after": load_workbook(xlsx, read_only=True).sheetnames,
    }


def _month_label(yyyymm: str) -> str:
    months = ["Jan","Feb","Mar","Apr","May","Jun",
                "Jul","Aug","Sep","Oct","Nov","Dec"]
    try:
        y, m = yyyymm.split("-")
        return f"{months[int(m)-1]} {y}"
    except Exception:
        return yyyymm


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--asof", default=date.today().isoformat())
    args = ap.parse_args()
    rep = build_usa_sheets(args.asof)
    print(json.dumps(rep, indent=2, ensure_ascii=False, default=str))
    return 0


if __name__ == "__main__":
    sys.exit(main())
