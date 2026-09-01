"""AEGIS · Phase 2 · C9 Registry-sync reconciliation · CEO 2026-09-01.

Resolves Portfolio↔Exit lifecycle collisions where aegis_history has
explicit EXIT status but Registry never received the corresponding
close event.

Authority: aegis_history is the AUDIT trail · every EXIT row there is
an actual observation of position closure. Registry SHOULD have been
updated at that time via `oreg.close()` · the sync at
`backend/delivery/telegram/detail_xlsx.py:503-505` failed silently
for these 4 (likely never invoked · not a code bug).

This script:
  1. Finds every (market, position_id, runner) that has an EXIT status
     row in aegis_history but no CLOSED event in Registry
  2. Uses the EXIT row's Date as the authoritative exit_date
  3. Calls `oreg.close()` to append the CLOSED event (idempotent)
  4. Verifies · re-reads Registry · confirms new state

Reads:
  · reports/telegram/aegis_history.xlsx        (audit trail)
  · reports/research/opportunity_registry.jsonl (Registry state)

Writes:
  · reports/research/opportunity_registry.jsonl (append CLOSED events)
  · reports/migration/c9_registry_sync_YYYYMMDD.json (audit report)

Idempotent · safe to rerun.
"""
from __future__ import annotations

import json
import sys
from collections import defaultdict
from datetime import date
from pathlib import Path

_ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(_ROOT))

from backend.research import opportunity_registry as oreg  # noqa: E402


def main() -> int:
    from openpyxl import load_workbook

    # Load Registry latest-per-pid
    print("[c9-sync] loading Registry state...")
    reg = oreg.load_all(_ROOT)
    latest_by_pid = {}
    for pid, opps in reg.items():
        for o in opps:
            latest_by_pid[o.opportunity_id] = o
    print(f"[c9-sync] Registry latest-per-pid: {len(latest_by_pid)}")

    # Scan aegis_history for EXIT rows
    print("[c9-sync] scanning aegis_history for EXIT status rows...")
    xlsx_p = _ROOT / "reports" / "telegram" / "aegis_history.xlsx"
    wb = load_workbook(xlsx_p, read_only=True, data_only=True)
    ws = wb["AEGIS Daily"]
    rows = list(ws.iter_rows(values_only=True))
    hdr = rows[0]
    def col(name):
        for i, c in enumerate(hdr):
            if c and str(c).strip().lower() == name.lower():
                return i
        return None
    c_pid = col("Position ID")
    c_ctry = col("Country")
    c_run = col("Run_Type")
    c_tk = col("Ticker")
    c_date = col("Date")
    c_st = col("Status")
    c_reason = col("Exit Reason")

    # Collect EXIT observations per PID · take EARLIEST EXIT as authoritative
    exit_obs = defaultdict(list)   # pid → list of (date, reason)
    for r in rows[1:]:
        if not r[c_pid]: continue
        if str(r[c_st] or "").upper() != "EXIT": continue
        pid = str(r[c_pid])
        dt = str(r[c_date])[:10] if r[c_date] else ""
        reason = str(r[c_reason] or "") if c_reason else ""
        exit_obs[pid].append((dt, reason))
    wb.close()
    print(f"[c9-sync] EXIT observations found for {len(exit_obs)} distinct PIDs")

    # Identify PIDs where Registry says ACTIVE but aegis_history says EXIT
    to_close = []
    for pid, obs in exit_obs.items():
        reg_state = latest_by_pid.get(pid)
        if reg_state is None:
            # Registry doesn't know this PID · skip (out of scope)
            continue
        if reg_state.status != "ACTIVE":
            # Registry already CLOSED · nothing to do (idempotent)
            continue
        earliest_exit = sorted(obs)[0]
        to_close.append({
            "opportunity_id": pid,
            "market": reg_state.market,
            "runner": reg_state.runner,
            "ticker": reg_state.ticker,
            "created_date": reg_state.created_date,
            "authoritative_exit_date": earliest_exit[0],
            "reason": earliest_exit[1] or "Registry-sync · aegis_history had EXIT status without oreg.close() call",
        })
    print(f"[c9-sync] PIDs needing CLOSED event: {len(to_close)}")

    # Print each · dry-run visibility
    for c in to_close:
        print(f"  {c['opportunity_id']} · {c['ticker']} {c['runner']} · "
              f"exit_date={c['authoritative_exit_date']} · reason='{c['reason'][:40]}'")

    # Execute · call oreg.close for each
    n_closed = 0
    n_errors = 0
    errors = []
    for c in to_close:
        try:
            oreg.close(_ROOT, c["opportunity_id"], c["authoritative_exit_date"],
                        reason=c["reason"])
            n_closed += 1
        except Exception as e:
            n_errors += 1
            errors.append({"pid": c["opportunity_id"], "error": f"{type(e).__name__}: {e}"})

    print(f"\n[c9-sync] Registry close() calls: {n_closed} succeeded · {n_errors} errored")
    if errors:
        for e in errors:
            print(f"  ERROR {e['pid']}: {e['error']}")

    # Verify
    reg2 = oreg.load_all(_ROOT)
    latest2 = {}
    for pid, opps in reg2.items():
        for o in opps:
            latest2[o.opportunity_id] = o
    still_active = [c for c in to_close
                     if latest2.get(c["opportunity_id"]) and
                        latest2[c["opportunity_id"]].status == "ACTIVE"]
    print(f"[c9-sync] verification · PIDs still ACTIVE after close(): {len(still_active)}")

    # Report
    today = date.today().isoformat().replace("-", "")
    report = {
        "engine": "aegis.migration.phase_2_c9_registry_sync",
        "asof": date.today().isoformat(),
        "candidates": to_close,
        "closed": n_closed,
        "errors": n_errors,
        "error_detail": errors,
        "verification_still_active": [c["opportunity_id"] for c in still_active],
    }
    out_p = _ROOT / "reports" / "migration" / f"c9_registry_sync_{today}.json"
    out_p.write_text(json.dumps(report, indent=2, ensure_ascii=False),
                      encoding="utf-8")
    print(f"[c9-sync] report: {out_p.relative_to(_ROOT)}")

    if n_errors > 0 or still_active:
        return 2
    return 0


if __name__ == "__main__":
    sys.exit(main())
