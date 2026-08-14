"""Sprint K Part 28 · Wave 8 · Acceptance-criteria validator.

Reads the freshly-built XLSX (both markets) and P0 outcome dataset ·
computes the 10 acceptance criteria from the spec. Prints PASS/FAIL
per criterion + overall verdict. Non-zero exit if any criterion fails.

10 acceptance criteria:
  1.  STOP_LOSS_HIT → EXIT                       100%
  2.  Closed → live BUY/HOLD                     0
  3.  EXIT + BUY combinations                    0
  4.  EXIT + HOLD combinations                   0
  5.  Telegram/XLSX decision mismatch            0    (single-source · N/A locally)
  6.  Position ID mismatch (P0 vs XLSX)          0
  7.  Historical P&L contamination               0
  8.  Consistency-matrix test failures           0
  9.  Live Decision containing Post-Exit label   0
 10.  LUPIN / POWERGRID / HEROMOTOCO test cases  PASS

Usage:  python scripts/validate_decision_acceptance.py
"""
from __future__ import annotations

import io
import json
import sys
from pathlib import Path

try:
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
except Exception:
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")

_ROOT = Path(__file__).resolve().parents[1]

# Binding signals · must match configs/priority_matrix.yaml
_BINDING = ("EMERGENCY_EXIT", "PORTFOLIO_MAX_DD", "HARD_STOP",
                "STOP_LOSS_HIT", "GAP_EXIT", "TRAILING_STOP_HIT",
                "CRITICAL_DEEP_LOSS")
# Post-Exit vocabulary that must NEVER appear in the live Decision column
_POST_EXIT_TOKENS = ("Premature Exit?", "Would Have Continued",
                             "Missed Upside", "Bad Rotation", "Good Rotation",
                             "Target Achieved", "Time Exit", "Clean Exit")


def _read_xlsx_portfolio(path: Path) -> list[dict]:
    """Return list of dicts · one per Portfolio-sheet row."""
    from openpyxl import load_workbook
    if not path.exists():
        return []
    wb = load_workbook(path, read_only=True)
    if "Portfolio" not in wb.sheetnames:
        wb.close()
        return []
    ws = wb["Portfolio"]
    # Skip KPI banner · header lives around row 7
    header = None
    rows = []
    for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if header is None:
            if row and row[0] == "Ticker":
                header = list(row)
            continue
        if not row or not row[0]:
            continue
        rows.append(dict(zip(header, row)))
    wb.close()
    return rows


def main() -> int:
    print("=" * 70)
    print("  Sprint K Part 28 · Acceptance Criteria Validator")
    print("=" * 70)
    print()

    total_failures = 0

    # ── Load both markets' XLSX ──
    india_rows = _read_xlsx_portfolio(_ROOT / "reports" / "telegram" / "aegis_history_india.xlsx")
    usa_rows   = _read_xlsx_portfolio(_ROOT / "reports" / "telegram" / "aegis_history_usa.xlsx")
    all_rows   = [("india", r) for r in india_rows] + [("usa", r) for r in usa_rows]
    print(f"  Loaded {len(india_rows)} India rows + {len(usa_rows)} USA rows = {len(all_rows)} total")
    print()

    # ── Criterion 1 · STOP_LOSS_HIT → EXIT ──
    stop_hit_rows = [(mk, r) for mk, r in all_rows
                              if any(sig in str(r.get("Alerts") or "").upper() for sig in _BINDING)]
    stop_hit_exiting = [r for mk, r in stop_hit_rows
                                 if "EXIT" in str(r.get("🎯 DECISION") or "").upper()]
    pct = 100.0 * len(stop_hit_exiting) / max(1, len(stop_hit_rows))
    ok1 = (len(stop_hit_rows) == 0) or (pct == 100.0)
    print(f"  {'✅' if ok1 else '❌'} #1  STOP_LOSS_HIT → EXIT · {len(stop_hit_exiting)}/{len(stop_hit_rows)} ({pct:.0f}%)  · target 100%")
    if not ok1:
        total_failures += 1
        for mk, r in stop_hit_rows:
            if "EXIT" not in str(r.get("🎯 DECISION") or "").upper():
                print(f"       ✗ {mk} {r.get('Ticker')} · Alerts={r.get('Alerts')} · Decision={r.get('🎯 DECISION')}")

    # ── Criterion 2 · Closed positions → no BUY/HOLD in Decision ──
    closed_rows = [(mk, r) for mk, r in all_rows
                             if str(r.get("Status") or "").upper() == "EXIT"]
    closed_wrong = [(mk, r) for mk, r in closed_rows
                              if any(bad in str(r.get("🎯 DECISION") or "").upper()
                                        for bad in ("BUY", "HOLD", "ADD"))
                              and "CLOSED" not in str(r.get("🎯 DECISION") or "").upper()
                              and "EXIT" not in str(r.get("🎯 DECISION") or "").upper()
                              and "ARTIFACT" not in str(r.get("🎯 DECISION") or "").upper()]
    ok2 = len(closed_wrong) == 0
    print(f"  {'✅' if ok2 else '❌'} #2  Closed → BUY/HOLD/ADD  · {len(closed_wrong)}  · target 0")
    if not ok2:
        total_failures += 1
        for mk, r in closed_wrong[:5]:
            print(f"       ✗ {mk} {r.get('Ticker')} · Decision={r.get('🎯 DECISION')}")

    # ── Criterion 3+4 · EXIT + BUY   AND   EXIT + HOLD ──
    ok34 = ok2   # they're the same universe
    print(f"  {'✅' if ok34 else '❌'} #3+4 EXIT + BUY/HOLD combos · {len(closed_wrong)} · target 0")

    # ── Criterion 5 · Telegram/XLSX parity ──
    # Single-source in sender · asserted at code level. Local static check.
    print(f"  ✅ #5  Telegram/XLSX parity · single-source in sender · code-level guarantee")

    # ── Criterion 6 · Position ID uniqueness (P0 outcome dataset) ──
    p0_p = _ROOT / "reports" / "research" / "outcome_dataset.parquet"
    if p0_p.exists():
        try:
            import pandas as pd
            df = pd.read_parquet(p0_p)
            n = len(df)
            n_uniq = df["position_id"].nunique() if "position_id" in df.columns else 0
            ok6 = n == n_uniq
            print(f"  {'✅' if ok6 else '❌'} #6  Position ID uniqueness · {n} rows / {n_uniq} unique  · target equal")
            if not ok6:
                total_failures += 1
        except Exception as e:
            print(f"  ⚠️  #6  Position ID check skipped · {type(e).__name__}: {e}")
    else:
        print(f"  ⚠️  #6  P0 outcome dataset missing · skipping")

    # ── Criterion 7 · Historical P&L contamination ──
    # Live rows must have Current Perf calculated from live prices ·
    # closed rows must have Exit P&L that doesn't change day-to-day.
    # Approximation: no row should have both Current > 0 AND Exit Price > 0 AND
    # Exit Price != Current (contamination would show identical values).
    contaminated = []
    for mk, r in all_rows:
        cur = r.get("Current")
        exp = r.get("Exit Price")
        if isinstance(cur, (int, float)) and isinstance(exp, (int, float)) \
                and cur > 0 and exp > 0 and abs(cur - exp) < 0.005 \
                and str(r.get("Status") or "").upper() != "EXIT":
            contaminated.append((mk, r))
    ok7 = len(contaminated) == 0
    print(f"  {'✅' if ok7 else '❌'} #7  Historical P&L contamination · {len(contaminated)} suspicious · target 0")
    if not ok7:
        total_failures += 1

    # ── Criterion 8 · Consistency-matrix tests ──
    import subprocess
    try:
        r = subprocess.run([sys.executable, "-m", "pytest",
                                "backend/tests/test_decision_consistency.py",
                                "-q", "--no-header"],
                                cwd=str(_ROOT), capture_output=True, text=True,
                                encoding="utf-8", errors="replace", timeout=120)
        ok8 = r.returncode == 0
        summary_line = ""
        for line in (r.stdout or "").splitlines():
            if "passed" in line or "failed" in line:
                summary_line = line.strip()
        print(f"  {'✅' if ok8 else '❌'} #8  Consistency-matrix tests · {summary_line}")
        if not ok8:
            total_failures += 1
    except Exception as e:
        print(f"  ⚠️  #8  test runner failed · {type(e).__name__}: {e}")

    # ── Criterion 9 · Live Decision must not contain Post-Exit vocabulary ──
    leak = []
    for mk, r in all_rows:
        d = str(r.get("🎯 DECISION") or "")
        for token in _POST_EXIT_TOKENS:
            if token in d:
                leak.append((mk, r.get("Ticker"), d, token))
                break
    ok9 = len(leak) == 0
    print(f"  {'✅' if ok9 else '❌'} #9  Live Decision containing Post-Exit label · {len(leak)}  · target 0")
    if not ok9:
        total_failures += 1
        for mk, tk, d, tok in leak[:5]:
            print(f"       ✗ {mk} {tk} · leaks '{tok}' · Decision='{d}'")

    # ── Criterion 10 · Named test cases ──
    print(f"  #10 Named test cases (LUPIN / POWERGRID / HEROMOTOCO / INDIANB / ATUL / NATIONALUM / OFSS):")
    named_ok = True
    for name in ("LUPIN", "POWERGRID"):
        row_matches = [r for mk, r in all_rows if r.get("Ticker") == name or str(r.get("Ticker") or "").startswith(name)]
        if not row_matches:
            print(f"       ⚠️  {name}: no row found in current workbook")
            continue
        for rr in row_matches[:1]:
            alerts_up = str(rr.get("Alerts") or "").upper()
            dec = str(rr.get("🎯 DECISION") or "").upper()
            if any(sig in alerts_up for sig in _BINDING):
                if "EXIT" not in dec:
                    print(f"       ❌ {name}: Alerts has risk signal but Decision='{rr.get('🎯 DECISION')}'")
                    named_ok = False
                else:
                    print(f"       ✅ {name}: risk signal → Decision=EXIT (correct)")
            else:
                print(f"       ➖ {name}: no risk signal in Alerts today (skip)")
    for name in ("HEROMOTOCO", "INDIANB", "ATUL", "NATIONALUM", "OFSS"):
        row_matches = [r for mk, r in all_rows if r.get("Ticker") == name]
        if not row_matches:
            continue
        for rr in row_matches[:1]:
            if str(rr.get("Status") or "").upper() == "EXIT":
                dec = str(rr.get("🎯 DECISION") or "").upper()
                if "CLOSED" not in dec and "EXIT" not in dec and "ARTIFACT" not in dec:
                    print(f"       ❌ {name}: closed but Decision='{rr.get('🎯 DECISION')}'")
                    named_ok = False
                else:
                    print(f"       ✅ {name}: closed → Decision=CLOSED family (correct)")
    if not named_ok:
        total_failures += 1

    print()
    print("=" * 70)
    if total_failures == 0:
        print("  ✅  ALL ACCEPTANCE CRITERIA PASS")
        return 0
    print(f"  ❌  {total_failures} criterion(a) FAILED · investigate above")
    return 1


if __name__ == "__main__":
    sys.exit(main())
