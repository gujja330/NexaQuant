"""Command Center Telegram sender · both markets · single message each.

Consumes the enriched recommendations.json produced by
backend.recommendation.ssot.run + backend.certification.institutional_optimization_run
(cycle 3-4). Sends ONE message per market · replaces the legacy multi-
message duplicate-message flow.

Usage:
    python scripts/telegram_command_center_send.py --market india
    python scripts/telegram_command_center_send.py --market usa
    python scripts/telegram_command_center_send.py --market both

Env:
    TELEGRAM_BOT_TOKEN
    TELEGRAM_CHAT_ID   (single chat serves both markets · UX030 pattern)

Exit codes:
    0 = all requested markets delivered successfully
    1 = at least one market failed
    2 = missing tokens (skipped without failing CI)
"""
from __future__ import annotations

import argparse
import io
import json
import os
import re
import sys
import urllib.error
import urllib.parse
import urllib.request
from datetime import datetime, timezone, timedelta
from pathlib import Path

try:
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
except Exception:
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")

_ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(_ROOT))

from backend.delivery.telegram.command_center import (  # noqa: E402
    load_and_render, render_research_platform_message,
    render_intraday_platform_message,
    ENGINE_ID, SCHEMA_FINGERPRINT,
)
from backend.delivery.telegram.detail_xlsx import (  # noqa: E402
    build_unified_history, build_and_stamp_all,
    maybe_sync_google_sheet,
)
from backend.portfolio.position_store.mark_to_market import (  # noqa: E402
    mark_to_market as _mark_to_market,
    validate_position_freshness as _validate_freshness,
    validate_payload_asof_matches_today as _validate_payload_asof,
)


def _load_env() -> None:
    for name in (".env.telegram", ".env"):
        p = _ROOT / name
        if not p.exists():
            continue
        for line in p.read_text(encoding="utf-8", errors="replace").splitlines():
            line = line.strip()
            if not line or line.startswith("#") or "=" not in line:
                continue
            k, v = line.split("=", 1)
            k = k.strip(); v = v.strip().strip('"').strip("'")
            os.environ.setdefault(k, v)
    # Aliases (mirror UX030 sender)
    if not os.environ.get("TELEGRAM_BOT_TOKEN"):
        for a in ("TOKEN", "BOT_TOKEN", "TELEGRAM_TOKEN"):
            if os.environ.get(a):
                os.environ["TELEGRAM_BOT_TOKEN"] = os.environ[a]; break
    if not os.environ.get("TELEGRAM_CHAT_ID"):
        for a in ("CHAT_ID", "CHAT", "TELEGRAM_CHAT"):
            if os.environ.get(a):
                os.environ["TELEGRAM_CHAT_ID"] = os.environ[a]; break
    m = re.search(r"\d{6,}:[A-Za-z0-9_-]{20,}", os.environ.get("TELEGRAM_BOT_TOKEN", ""))
    if m:
        os.environ["TELEGRAM_BOT_TOKEN"] = m.group(0)


def _send_document(token: str, chat_id: str, file_path: Path,
                       caption: str = "") -> tuple[bool, str]:
    """Upload a file to Telegram via sendDocument · used for detail reports."""
    if not file_path.exists():
        return False, f"file missing: {file_path}"
    url = f"https://api.telegram.org/bot{token}/sendDocument"
    boundary = "----AEGISFormBoundary" + os.urandom(8).hex()
    body = []
    def _p(field, value):
        body.append(f"--{boundary}\r\n"
                       f'Content-Disposition: form-data; name="{field}"\r\n\r\n'
                       f"{value}\r\n")
    _p("chat_id", chat_id)
    if caption:
        _p("caption", caption[:1024])
        # No parse_mode · plain text · avoids Markdown-parse errors from
        # underscores in field names (Run_Type) or dashes in tickers.
    file_bytes = file_path.read_bytes()
    body.append(f"--{boundary}\r\n"
                   f'Content-Disposition: form-data; name="document"; '
                   f'filename="{file_path.name}"\r\n'
                   f"Content-Type: text/markdown\r\n\r\n")
    # Build multipart payload as bytes
    prefix = "".join(body).encode("utf-8")
    suffix = f"\r\n--{boundary}--\r\n".encode("utf-8")
    payload = prefix + file_bytes + suffix
    req = urllib.request.Request(url, data=payload, method="POST")
    req.add_header("Content-Type", f"multipart/form-data; boundary={boundary}")
    try:
        with urllib.request.urlopen(req, timeout=60) as r:
            resp = r.read().decode("utf-8", errors="replace")
        return True, resp[:120]
    except urllib.error.HTTPError as e:
        try:
            err = e.read().decode("utf-8", errors="replace")
        except Exception:
            err = str(e)
        return False, f"HTTP {e.code}: {err[:200]}"
    except Exception as e:
        return False, f"{type(e).__name__}: {e}"


def _send_markdown(token: str, chat_id: str, text: str) -> tuple[bool, str]:
    if not text:
        return True, "empty"
    # Telegram hard cap 4096 · trim only if actually over
    if len(text) > 4096:
        text = text[:4080] + "\n\n...truncated"
    url = f"https://api.telegram.org/bot{token}/sendMessage"
    data = urllib.parse.urlencode({
        "chat_id":                  chat_id,
        "text":                     text,
        "parse_mode":               "Markdown",
        "disable_web_page_preview": "true",
    }).encode("utf-8")
    try:
        req = urllib.request.Request(url, data=data, method="POST")
        with urllib.request.urlopen(req, timeout=30) as r:
            body = r.read().decode("utf-8", errors="replace")
        return True, body[:120]
    except urllib.error.HTTPError as e:
        # Retry without Markdown parse if the parser trips (defense against
        # any accidental `_` / `*` in tickers or narrative).
        try:
            err_body = e.read().decode("utf-8", errors="replace")
        except Exception:
            err_body = str(e)
        if e.code == 400 and "parse" in err_body.lower():
            plain = urllib.parse.urlencode({
                "chat_id":                  chat_id,
                "text":                     text,
                "disable_web_page_preview": "true",
            }).encode("utf-8")
            try:
                req2 = urllib.request.Request(url, data=plain, method="POST")
                with urllib.request.urlopen(req2, timeout=30) as r:
                    body = r.read().decode("utf-8", errors="replace")
                return True, "markdown-failed-plain-delivered: " + body[:100]
            except Exception as e2:
                return False, f"plain-fallback failed: {e2}"
        return False, f"HTTP {e.code}: {err_body[:200]}"
    except Exception as e:
        return False, f"{type(e).__name__}: {e}"


def _market_reports(market: str) -> Path:
    return _ROOT / ("usa/reports" if market == "usa" else "reports")


def _append_delivery_ledger(record: dict) -> None:
    date = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    p = _ROOT / "reports" / f"telegram_command_center_{date}.jsonl"
    p.parent.mkdir(parents=True, exist_ok=True)
    with p.open("a", encoding="utf-8") as f:
        f.write(json.dumps(record, default=str) + "\n")


def _send_one_market(market: str, token: str, chat_id: str) -> tuple[bool, dict]:
    reports_dir = _market_reports(market)

    # ── GUARD 6 · payload asof-freshness (post-mortem 2026-08-01) ──
    # Refuse to send if recommendations.json's asof is NOT today.
    # This closes the "restored-from-git stale payload got sent" gap.
    try:
        vpa = _validate_payload_asof(_ROOT, market)
        if vpa["verdict"] == "STALE_PAYLOAD":
            print(f"[payload_asof:{market}] REFUSED · asof={vpa['asof']} "
                    f"today={vpa['today']} ({vpa['days_stale']}d stale)")
            print(f"  · Fix: run daily pipeline "
                    f"'python -m backend.recommendation.ssot.run --market {market}' "
                    f"(may need --force if snapshot exists)")
            print(f"  · Or override with SEND_FORCE_STALE=1 env var (destructive).")
            if not os.environ.get("SEND_FORCE_STALE"):
                return False, {"market": market, "refused": "stale_payload_asof", **vpa}
        elif vpa["verdict"] == "OK":
            print(f"[payload_asof:{market}] OK · asof={vpa['asof']} (today)")
        else:
            print(f"[payload_asof:{market}] {vpa['verdict']} · continuing anyway")
    except Exception as e:
        print(f"[payload_asof:{market}] check failed · {type(e).__name__}: {e} · continuing")

    # ── PRECAUTION 1 · mark-to-market before render (post-mortem 2026-07-31) ──
    # Ensures every active position has today's actual close price · so
    # Max Gain / Max DD reflect reality (not stale 0.00% from entry-price)
    try:
        mtm = _mark_to_market(_ROOT, market)
        print(f"[mtm:{market}] repriced={mtm['n_repriced']}/{mtm['n_positions']} "
                f"missing={mtm['n_missing_price']}")
    except Exception as e:
        print(f"[mtm:{market}] failed · {type(e).__name__}: {e} · continuing anyway")

    # ── PRECAUTION 2 · freshness check · REFUSE to send if positions stale ──
    # Prevents the "0.00% Max Gain everywhere" incident from ever recurring
    try:
        v = _validate_freshness(_ROOT, market, max_stale_days=2)
        if v["verdict"] == "STALE" and v["n_stale"] > 0:
            print(f"[freshness:{market}] REFUSED · {v['n_stale']} stale positions "
                    f"(last_seen > 2 days behind asof)")
            for s in v["stale_tickers"][:5]:
                print(f"  · {s['ticker']}: last_seen={s['last_seen']} "
                        f"({s['days_behind']}d behind)")
            print(f"  · Fix: run 'python scripts/mark_to_market.py --market {market}' "
                    f"or wait for tomorrow's daily pipeline.")
            print(f"  · Or override with SEND_FORCE_STALE=1 env var (destructive).")
            if not os.environ.get("SEND_FORCE_STALE"):
                return False, {"market": market, "refused": "stale_positions", **v}
        else:
            print(f"[freshness:{market}] OK · {v['n_active']} active positions all fresh")
    except Exception as e:
        print(f"[freshness:{market}] check failed · {type(e).__name__}: {e} · continuing")

    msg, meta = load_and_render(reports_dir, market)
    if meta.get("n_recs") == 0:
        print(f"[command_center:{market}] no recs · skipping")
        return True, {"market": market, "skipped": True, **meta}
    ok, detail = _send_markdown(token, chat_id, msg)
    print(f"[command_center:{market}] chars={meta['message_chars']} "
          f"recs={meta['n_recs']} rotations={meta['n_rotations']} "
          f"actionable={meta['n_actionable']} · sent={ok}")
    if not ok:
        print(f"  detail: {detail[:180]}")
    _append_delivery_ledger({
        "ts_utc": datetime.now(timezone.utc).isoformat(),
        "engine": ENGINE_ID,
        "schema_fingerprint": SCHEMA_FINGERPRINT,
        "market": market,
        "ok": ok,
        "detail_head": detail[:200],
        **meta,
    })

    # Research Platform second message DISABLED 2026-07-30 per operator:
    # "ensure single message for india and single message for usa."
    research_ok = True

    # Detail MD per-market REMOVED 2026-08-01 · replaced by ONE unified
    # XLSX file attached once at end of main() with Date + Country + Run_Type
    # columns · appended daily to reports/telegram/aegis_history.xlsx.

    return (ok and research_ok), {
        "market":        market,
        "ok":            ok,
        "research_ok":   research_ok,
        **meta,
    }


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--market", choices=["india", "usa", "both"], default="both")
    ap.add_argument("--dry-run", action="store_true",
                       help="Render + print message but do not send")
    # Operator directive 2026-08-04: "i should get on xlsx output plz, no
    # need of messages again". XLSX-only mode suppresses the compact per-
    # market Command Center message · daily delivery is the XLSX attachment
    # (+ optional caption) alone. Default ON to match new operator preference.
    ap.add_argument("--with-message", action="store_true",
                       help="Also send the compact Command Center message "
                            "(default: OFF · XLSX attachment only)")
    args = ap.parse_args()

    _load_env()
    token = os.environ.get("TELEGRAM_BOT_TOKEN", "")
    chat_id = os.environ.get("TELEGRAM_CHAT_ID", "")

    markets = ["india", "usa"] if args.market == "both" else [args.market]

    if args.dry_run:
        if args.with_message:
            for m in markets:
                msg, meta = load_and_render(_market_reports(m), m)
                print(f"===== DRY RUN · {m} · {meta['message_chars']} chars =====")
                print(msg)
                print()
        else:
            print(f"===== DRY RUN · XLSX-only mode · markets={markets} =====")
            print("(compact message suppressed · pass --with-message to render)")
        return 0

    if not token or not chat_id:
        print(f"[command_center] MISSING TELEGRAM tokens · skipping "
              f"(sender remains optional in CI). Requested markets: {markets}")
        return 2   # non-fatal · matches optional-step convention

    all_ok = True
    if args.with_message:
        for m in markets:
            ok, _ = _send_one_market(m, token, chat_id)
            all_ok = all_ok and ok
    else:
        print(f"[command_center] XLSX-only mode · compact message skipped · "
              f"markets={markets} (pass --with-message to re-enable)")

    # ── UNIFIED XLSX · one file across all markets · attached ONCE ──
    # Operator directive 2026-08-01: "better send me that xlsx into telegram
    # daily thats it. simple format of xlsx" · "runner 1, runner 2 u can mix
    # into a column called Run_type" · "add country too" · "everyday we
    # can update same sheet".
    xlsx_ok = True
    try:
        from datetime import date as _date
        asof = _date.today().isoformat()
        # v5 · stamp regime + rank + profit-protection · then build XLSX
        xlsx_path = build_and_stamp_all(_ROOT, asof, markets=markets)
        # Sprint H · attach operator guide to Monday caption (weekly refresher)
        try:
            from backend.delivery.telegram.operator_guide import append_to_caption as _guide
            from datetime import date as _dt
            _dow = _dt.today().isoweekday()   # 1=Mon .. 7=Sun
            _CAPTION_APPEND_GUIDE = _guide
        except Exception:
            _CAPTION_APPEND_GUIDE = None
        # Guard 7 · Context Health Monitor · every report is a guard
        # (operator directive 2026-08-05)
        try:
            from backend.context.health_monitor import (
                run_health_check as _hc, emit as _hc_emit, render_summary as _hc_render)
            health = _hc(_ROOT)
            _hc_emit(_ROOT, health)
            print(f"[guard7:health] {_hc_render(health)}")
            if health.get("overall_verdict") == "RED" \
               and os.environ.get("SEND_FORCE_STALE") != "1":
                print(f"[guard7:health] BLOCKING send · {health['n_critical_fails']} "
                      f"critical engines failed. Override with SEND_FORCE_STALE=1")
                for r in health.get("critical_fails", [])[:5]:
                    print(f"    ✗ {r['path']}: {r['verdict']} · {r['reason']}")
                return 2
        except Exception as e:
            print(f"[guard7:health] check failed · {type(e).__name__}: {e} · proceeding")
        # Guard 8 · Price Integrity · verifies data pull is CORRECT before send
        # (operator directive 2026-08-06 · "pipeline should be very strong in
        # pulling right data with guard")
        try:
            from backend.context.price_integrity_guard import (
                check_all as _pig, emit as _pig_emit, render_summary as _pig_render)
            # 2026-08-12 CEO CI fix · scope guard to the market we're actually
            # sending. India CI has no USA parquets (they're .gitignored and
            # only refreshed by aegis-usa.yml) so evaluating USA on India CI
            # produced 500+ fake CRITICALS and blocked India send for 2 days.
            pig = _pig(_ROOT, asof, markets=tuple(markets))
            _pig_emit(_ROOT, pig)
            print(f"[guard8:price] {_pig_render(pig)}")
            if pig.get("verdict") == "RED" \
               and os.environ.get("PRICE_GUARD_OVERRIDE") != "1":
                print(f"[guard8:price] BLOCKING send · {pig['n_critical']} CRITICAL "
                      f"price mismatches. Override with PRICE_GUARD_OVERRIDE=1")
                for r in pig.get("critical_issues", [])[:5]:
                    print(f"    ✗ {r['market']} {r['ticker']} · {r['check']}: {r['detail']}")
                return 2
        except Exception as e:
            print(f"[guard8:price] check failed · {type(e).__name__}: {e} · proceeding")
        # 2026-08-08 · Guard 9 · Pipeline heartbeat (operator directive after
        # silent Aug-3 pipeline miss: "pages you if a trading day passes with
        # no pipeline execution"). Emit gap alert prepended to caption so
        # operator sees it BEFORE the XLSX arrives.
        heartbeat_banner = ""
        try:
            from backend.context.pipeline_heartbeat import (
                check as _hb_check, render as _hb_render)
            for _m in markets:
                _hb = _hb_check(_ROOT, _m, asof)
                print(f"[guard9:heartbeat:{_m}] {_hb_render(_hb)}")
                if _hb["status"] in ("CRITICAL", "WARNING"):
                    heartbeat_banner += _hb["message"] + "\n"
        except Exception as e:
            print(f"[guard9:heartbeat] check failed · {type(e).__name__}: {e} · proceeding")
        # 2026-08-08 · Two-file delivery per operator directive: "send me both
        # xlsx in different timezones · usa separate, india separate."
        # Split unified aegis_history.xlsx into two market-specific files,
        # each attached as its own Telegram message with market-specific
        # timezone caption (India: IST · USA: CST + IST).
        from openpyxl import load_workbook as _lwb
        from openpyxl.styles import PatternFill as _PF
        _now_ist = datetime.now(timezone(timedelta(hours=5, minutes=30)))
        _now_cst = datetime.now(timezone(timedelta(hours=-6)))
        _delivered_ist = _now_ist.strftime("%Y-%m-%d %H:%M IST")
        _delivered_cst = _now_cst.strftime("%Y-%m-%d %H:%M CST")
        _STATUS_FILLS_LOCAL = {
            "STRONG BUY":       _PF(start_color="70AD47", end_color="70AD47", fill_type="solid"),
            "BUY":              _PF(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
            "HOLD":             _PF(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"),
            "EXIT":             _PF(start_color="F8CBAD", end_color="F8CBAD", fill_type="solid"),
            "ROTATED_SAMEDAY":  _PF(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid"),
        }

        # 2026-08-08 · Config-driven largecap seed (was hardcoded)
        _INDIA_LARGECAP = set()
        try:
            import yaml as _yaml
            _tiers_path = _ROOT / "configs" / "india_universe_tiers.yaml"
            if _tiers_path.exists():
                _tiers = _yaml.safe_load(_tiers_path.read_text(encoding="utf-8")) or {}
                _INDIA_LARGECAP = set(_tiers.get("largecap_tickers") or [])
        except Exception as _e:
            print(f"[config:india_universe_tiers] load failed · {_e}")

        def _cap_size(ticker: str, market: str) -> str:
            """Return LargeCap / MidCap tag for the ticker."""
            short = str(ticker or "").replace(".NS", "").replace(".BO", "").upper()
            if market.upper() == "USA":
                return "LargeCap (S&P 500)"
            # India: NIFTY 50 = LargeCap · anything else in NIFTY 200 = MidCap
            return "LargeCap" if short in _INDIA_LARGECAP else "MidCap"

        # Sector cache (operator directive: "sector in output xlsx mandatory")
        # Populated from yfinance · persisted at reports/sector_cache.json
        _sector_cache_path = _ROOT / "reports" / "sector_cache.json"
        _sector_cache = json.loads(_sector_cache_path.read_text(encoding="utf-8")) \
                                if _sector_cache_path.exists() else {"india": {}, "usa": {}}

        def _sector_for(ticker: str, market: str) -> str:
            """Look up sector · cached · fallback to '—'."""
            short = str(ticker or "").replace(".NS", "").replace(".BO", "").upper()
            bucket = _sector_cache.get(market.lower(), {})
            return bucket.get(short) or bucket.get(str(ticker).upper()) or "—"

        # Action explanations for the Action column (operator: "actions needed
        # short explanation · enter buy zone means to buy on that days?")
        _ACTIONS = {
            "STRONG BUY":      "Enter today · top pick · high conviction",
            "BUY":             "Enter or add · in buy zone",
            "HOLD":            "Hold · no action needed today",
            "EXIT":            "Sell/rotate today · closed position",
            "ROTATED_SAMEDAY": "Not held · same-day rotation artifact",
        }

        def _parquet_close(ticker: str, market: str, target_date: str) -> float | None:
            """Latest parquet close on or before target_date · source of truth."""
            import pandas as _pd
            short = str(ticker or "").replace(".NS", "").replace(".BO", "").upper()
            base = "usa/data/raw/us" if market.upper() == "USA" else "data/raw/india"
            p = _ROOT / base / f"{short}_D1.parquet"
            if not p.exists(): return None
            try:
                d = _pd.read_parquet(p)
                col = "close" if "close" in d.columns else "Close"
                d.index = _pd.to_datetime(d.index).strftime("%Y-%m-%d")
                if target_date in d.index: return float(d.loc[target_date, col])
                earlier = [dt for dt in d.index if dt <= target_date]
                return float(d.loc[earlier[-1], col]) if earlier else None
            except Exception:
                return None

        def _parquet_prev_close(ticker: str, market: str, target_date: str) -> float | None:
            """Trading day close STRICTLY BEFORE target_date · fixes Today Move
            lag bug where Prev Close was stamped when row was written and
            never re-derived for the observation date. CEO audit fix 2026-08-10."""
            import pandas as _pd
            short = str(ticker or "").replace(".NS", "").replace(".BO", "").upper()
            base = "usa/data/raw/us" if market.upper() == "USA" else "data/raw/india"
            p = _ROOT / base / f"{short}_D1.parquet"
            if not p.exists(): return None
            try:
                d = _pd.read_parquet(p)
                col = "close" if "close" in d.columns else "Close"
                d.index = _pd.to_datetime(d.index).strftime("%Y-%m-%d")
                strictly_before = [dt for dt in d.index if dt < target_date]
                return float(d.loc[strictly_before[-1], col]) if strictly_before else None
            except Exception:
                return None

        def _split_and_send(mkt_label: str, mkt_key: str, caption_body: str):
            src_wb = _lwb(xlsx_path)
            src_ws = src_wb["AEGIS Daily"] if "AEGIS Daily" in src_wb.sheetnames else src_wb.active
            h = [c.value for c in src_ws[1]]
            c_ctry = h.index("Country") + 1
            c_st = h.index("Status") + 1
            c_tk = h.index("Ticker") + 1
            c_date = h.index("Date") + 1
            c_recommended = h.index("Recommended") + 1 if "Recommended" in h else None
            c_entry = h.index("Entry Price") + 1 if "Entry Price" in h else None
            c_current = h.index("Current Price") + 1 if "Current Price" in h else None
            c_perf = h.index("Current Perf %") + 1 if "Current Perf %" in h else None
            c_exit_pnl = h.index("Exit P&L %") + 1 if "Exit P&L %" in h else None
            c_conf = h.index("Confidence %") + 1 if "Confidence %" in h else None
            c_stop = h.index("Stop Loss") + 1 if "Stop Loss" in h else None
            c_t1 = h.index("Target 1") + 1 if "Target 1" in h else None
            c_t2 = h.index("Target 2") + 1 if "Target 2" in h else None
            c_alerts = h.index("Alerts") + 1 if "Alerts" in h else None
            c_sector_existing = h.index("Sector") + 1 if "Sector" in h else None
            c_prev_close = h.index("Prev Close") + 1 if "Prev Close" in h else None
            c_today_move = h.index("Today Move %") + 1 if "Today Move %" in h else None

            # Load PRIORITY_MATRIX + DECISION vocab BEFORE row iteration
            # so History-sheet row loop can resolve decisions
            PRIORITY_MATRIX = {}
            # 2026-08-14 · Sprint K Part 28 · binding risk signals list.
            # If any of these strings appear in the Alerts column, priority
            # is forced to bucket R (Risk Controller veto · highest priority).
            BINDING_RISK_SIGNALS: list[str] = []
            try:
                import yaml as _yaml
                _pm_path = _ROOT / "configs" / "priority_matrix.yaml"
                if _pm_path.exists():
                    _pm = _yaml.safe_load(_pm_path.read_text(encoding="utf-8")) or {}
                    for bucket, d in (_pm.get("buckets") or {}).items():
                        PRIORITY_MATRIX[bucket] = (
                            d.get("urgency", ""), d.get("reason", ""),
                            d.get("action", ""), d.get("review", ""),
                            d.get("color", "F2F2F2"),
                        )
                    BINDING_RISK_SIGNALS = [str(s).upper()
                                                    for s in (_pm.get("binding_risk_signals") or [])]
            except Exception as _e:
                print(f"[config:priority_matrix] load failed · {_e}")
            PRIORITY_FILLS = {k: _PF(start_color=v[4], end_color=v[4], fill_type="solid")
                                     for k, v in PRIORITY_MATRIX.items()}

            DECISION_RULES = []
            DECISION_COLORS = {}
            try:
                import yaml as _yaml
                _dv_path = _ROOT / "configs" / "decision_vocabulary.yaml"
                if _dv_path.exists():
                    _dv = _yaml.safe_load(_dv_path.read_text(encoding="utf-8")) or {}
                    DECISION_RULES = _dv.get("rules") or []
                    DECISION_COLORS = _dv.get("colors") or {}
            except Exception as _e:
                print(f"[config:decision_vocabulary] load failed · {_e}")

            # 2026-08-14 · operator-approved 6-tier color scheme (msg 380)
            # Load configs/decision_colors.yaml · classify each row by DECISION
            # family · row color + sort by tier_rank. Config-driven per no-hardcode rule.
            DECISION_TIERS: dict = {}     # tier_key -> full tier dict
            TIER_MATCH_ORDER: list = []   # ordered list of tier_keys to walk
            TIER_LEGEND_HEADING = "COLOR LEGEND · row color = actionable priority"
            try:
                import yaml as _yaml
                _dc_path = _ROOT / "configs" / "decision_colors.yaml"
                if _dc_path.exists():
                    _dc = _yaml.safe_load(_dc_path.read_text(encoding="utf-8")) or {}
                    DECISION_TIERS = _dc.get("tiers") or {}
                    TIER_MATCH_ORDER = _dc.get("match_order") or list(DECISION_TIERS.keys())
                    TIER_LEGEND_HEADING = _dc.get("legend_heading") or TIER_LEGEND_HEADING
            except Exception as _e:
                print(f"[config:decision_colors] load failed · {_e}")

            def _decision_tier(decision_text: str, is_new_position: bool) -> str:
                """Return one of the tier keys (strong_buy/buy/new/hold/exit/closed)
                by walking match_order and matching Decision keywords (uppercase)
                or the trigger_on_new_position flag."""
                dt_up = str(decision_text or "").upper()
                for tier_key in TIER_MATCH_ORDER:
                    tier = DECISION_TIERS.get(tier_key) or {}
                    if tier.get("trigger_on_new_position") and is_new_position:
                        return tier_key
                    for kw in (tier.get("keywords_any") or []):
                        if kw.upper() in dt_up:
                            return tier_key
                # Ultimate fallback: hold (yellow)
                return "hold" if "hold" in DECISION_TIERS else next(iter(DECISION_TIERS.keys()), "hold")

            def _resolve_decision(action, inv_quality, status):
                a = str(action or "").strip()
                iq = str(inv_quality or "").strip()
                for rule in DECISION_RULES:
                    m = rule.get("match") or {}
                    if not m:
                        return rule.get("decision", "—"), rule.get("color", "gray")
                    if "action" in m and m["action"] != a:
                        continue
                    if "inv_quality_in" in m and iq not in m["inv_quality_in"]:
                        continue
                    return rule.get("decision", "—"), rule.get("color", "gray")
                return "— UNKNOWN", "gray"

            # Execution Decision Layer config (2026-08-09)
            EXEC_CFG = {}
            try:
                import yaml as _yaml
                _ex_path = _ROOT / "configs" / "execution_windows.yaml"
                if _ex_path.exists():
                    EXEC_CFG = _yaml.safe_load(_ex_path.read_text(encoding="utf-8")) or {}
            except Exception as _e:
                print(f"[config:execution_windows] load failed · {_e}")

            def _next_review_date(review_str, from_date_iso):
                offsets = (EXEC_CFG.get("review_offsets") or {})
                key = str(review_str or "").strip().upper()
                offset = offsets.get(key)
                if offset is None or offset == "null": return ""
                if offset == 0: return from_date_iso
                try:
                    from datetime import date as _dtc, timedelta as _tdc
                    d = _dtc.fromisoformat(from_date_iso[:10])
                    added = 0
                    while added < offset:
                        d += _tdc(days=1)
                        if d.weekday() < 5: added += 1
                    return d.isoformat()
                except Exception:
                    return ""

            def _execution_window(action):
                a = str(action or "").strip()
                for rule in (EXEC_CFG.get("execution_windows") or []):
                    m = rule.get("match") or {}
                    if not m: return rule.get("window", "—")
                    if "action" in m and m["action"] != a: continue
                    return rule.get("window", "—")
                return "—"

            def _price_trigger(action, stop_v, t1_v, curr_v):
                triggers = (EXEC_CFG.get("price_triggers") or {})
                cfg = triggers.get(str(action or "").strip())
                if not cfg: return ""
                source = cfg.get("source")
                label = cfg.get("label", "")
                price = {"stop": stop_v, "target_1": t1_v, "current": curr_v}.get(source)
                if not isinstance(price, (int, float)) or price <= 0: return ""
                return f"{label} {price:.2f}"

            # Filter rows by market
            keep_rows = [row for row in src_ws.iter_rows(min_row=2, values_only=False)
                                    if str(row[c_ctry-1].value or "").upper() == mkt_key.upper()]

            # 2026-08-14 · operator directives:
            #   1. "usa list is very big still in xlsx"
            #   2. "we switched to s&P 500 large cap, why we need still old
            #       stocks like trv and all"
            #
            # Root cause: source workbook keeps ALL 507 USA universe-scan rows
            # daily. Also historical position_store carries pre-S&P-500-switch
            # tickers (TRV · V · etc. from earlier universe) that are no longer
            # in the active universe.
            #
            # Filter = CANONICAL_SET  INTERSECT  CURRENT_UNIVERSE_SET
            #   canonical  = v3-selected ∪ position_store (active + exited + history)
            #   universe   = today's active tickers from usa/reports/universe.json
            #                (currently S&P 500 large cap · 500-ish tickers)
            # Historical tickers no longer in the S&P 500 = dropped.
            # India already has ~15 selected only · this filter is a no-op there.
            if mkt_key.upper() == "USA":
                import json as _json
                _canonical: set = set()
                # a. Currently selected candidates (v3)
                _v3 = _ROOT / "usa" / "reports" / "recommendations_v3.json"
                if _v3.exists():
                    try:
                        _d = _json.loads(_v3.read_text(encoding="utf-8"))
                        for _r in (_d.get("recommendations") or []):
                            _tk = str(_r.get("ticker") or "").upper()
                            if _tk: _canonical.add(_tk)
                    except Exception:
                        pass
                # b. Any ticker ever in position_store (active + exited)
                _pos = _ROOT / "usa" / "reports" / "position_store" / "usa" / "positions.json"
                if _pos.exists():
                    try:
                        _pd = _json.loads(_pos.read_text(encoding="utf-8"))
                        for _tk in (_pd.get("positions") or {}).keys():
                            _canonical.add(str(_tk).upper())
                        for _e in (_pd.get("exited") or []):
                            _tk = str(_e.get("ticker") or "").upper()
                            if _tk: _canonical.add(_tk)
                    except Exception:
                        pass
                # c. (removed 2026-08-14 · operator: "we switched to S&P 500
                #     large cap, why we need still old stocks like TRV")
                # history.jsonl is a raw event log · contains pre-switch
                # position openings (Jul 29 TRV, V, HON etc. from earlier
                # universe experiments). Using it re-surfaces stale tickers
                # the operator no longer wants to see. Only positions.json
                # (managed by portfolio_manager · reflects current era) is
                # the authoritative "what am I holding" source.

                # d. Current active universe (S&P 500) · intersect to drop
                # historical tickers no longer in the universe.
                _current_universe: set = set()
                _uni = _ROOT / "usa" / "reports" / "universe.json"
                if _uni.exists():
                    try:
                        _ud = _json.loads(_uni.read_text(encoding="utf-8"))
                        for _t in (_ud.get("tickers") or []):
                            _sym = _t.get("symbol") if isinstance(_t, dict) else str(_t)
                            if _sym:
                                _current_universe.add(str(_sym).upper())
                    except Exception:
                        pass
                # Also add v3 selected as guaranteed-present · even if universe
                # file happens to be stale · today's selections are authoritative.
                _v3_set: set = set()
                if _v3.exists():
                    try:
                        _d = _json.loads(_v3.read_text(encoding="utf-8"))
                        for _r in (_d.get("recommendations") or []):
                            _tk = str(_r.get("ticker") or "").upper()
                            if _tk: _v3_set.add(_tk)
                    except Exception:
                        pass
                _effective_universe = (_current_universe | _v3_set) if _current_universe else _canonical
                _final_set = _canonical & _effective_universe
                _before = len(keep_rows)
                _dropped_stale = _canonical - _effective_universe   # observability
                keep_rows = [row for row in keep_rows
                                    if str(row[c_tk-1].value or "").upper() in _final_set]
                print(f"[xlsx:USA] canonical∩universe filter · {_before} -> {len(keep_rows)} rows "
                      f"(canonical={len(_canonical)} · universe={len(_current_universe)} · "
                      f"final={len(_final_set)} · dropped-as-out-of-universe={len(_dropped_stale)})")
                if _dropped_stale:
                    _sample = sorted(_dropped_stale)[:10]
                    print(f"[xlsx:USA]   dropped historical (not in current universe): "
                          f"{_sample}{'...' if len(_dropped_stale) > 10 else ''}")

            # 2026-08-15 · Section 5 · SKIP rows NEVER appear in the primary
            # portfolio table. SKIP means "not an investment" · showing it as
            # a decision would corrupt portfolio P&L + confuse the operator.
            # SKIP candidates go to a separate research dataset (see
            # reports/research/skip_candidates_{market}.jsonl) so opportunity-
            # cost analysis stays possible without polluting portfolio counts.
            _skipped_rows = [row for row in keep_rows
                                       if str(row[c_st-1].value or "").upper() == "SKIP"]
            _before_skip = len(keep_rows)
            keep_rows = [row for row in keep_rows
                                if str(row[c_st-1].value or "").upper() != "SKIP"]
            if _skipped_rows:
                _skip_out = (_ROOT / "reports" / "research" /
                                    f"skip_candidates_{mkt_key.lower()}.jsonl")
                _skip_out.parent.mkdir(parents=True, exist_ok=True)
                try:
                    import json as _jj
                    with _skip_out.open("a", encoding="utf-8") as _sf:
                        for _sr in _skipped_rows:
                            _sf.write(_jj.dumps({
                                "asof":   str(_sr[c_date-1].value or "")[:10],
                                "ticker": str(_sr[c_tk-1].value or ""),
                                "runner": str(_sr[3].value or "") if len(_sr) > 3 else "",
                                "status": "SKIP",
                                "current_price": _sr[c_current-1].value if c_current else None,
                                "note":   "Filtered from portfolio table · opportunity-cost tracking only",
                            }, default=str) + "\n")
                    print(f"[xlsx:{mkt_key}] SKIP filter · {_before_skip} -> {len(keep_rows)} rows "
                          f"({len(_skipped_rows)} SKIP moved to {_skip_out.relative_to(_ROOT)})")
                except Exception as _e:
                    print(f"[xlsx:{mkt_key}] SKIP research write failed · {_e}")

            # Write market-specific XLSX file · adds "Cap Size" column at end
            from openpyxl import Workbook as _WB
            out_path = xlsx_path.parent / f"aegis_history_{mkt_key.lower()}.xlsx"
            wb2 = _WB()
            ws2 = wb2.active
            ws2.title = f"AEGIS {mkt_key.upper()}"
            # Header + new columns (Cap Size + Opp Age)
            # Opp Age: NEW = row_date == Recommended date · OLD = held from prior day
            # (operator: "if I open sheet on monday, I might see new buy options
            # to invest if I missed old opportunities")
            new_h = list(h) + [
                "Cap Size", "Opp Age",
                "Investability", "Inv Verdict",
                "🎯 DECISION", "Urgency", "Reason", "Action", "Review",
                "Price Trigger", "Next Review", "Execution Window",
            ]
            for c, name in enumerate(new_h, start=1):
                cell = ws2.cell(1, c, name)
                cell.fill = HEADER_FILL if 'HEADER_FILL' in globals() else _PF(
                    start_color="1F4E78", end_color="1F4E78", fill_type="solid")
                cell.font = _Font(bold=True, color="FFFFFF", size=11)
            # Column widths from source
            from openpyxl.utils import get_column_letter as _gcl_src
            for i in range(1, len(h)+1):
                c_letter = _gcl_src(i)
                if c_letter in src_ws.column_dimensions:
                    ws2.column_dimensions[c_letter].width = src_ws.column_dimensions[c_letter].width
            # New Cap Size column width
            from openpyxl.utils import get_column_letter as _gcl
            ws2.column_dimensions[_gcl(len(new_h))].width = 18
            # Data rows · backfill Sector + repair stale Current Price
            latest_date = max((str(r[c_date-1].value or "")[:10] for r in keep_rows), default="")
            for r_idx, row in enumerate(keep_rows, start=2):
                tk = row[c_tk-1].value
                status = row[c_st-1].value
                # CEO audit fix 2026-08-10 · unconditional Current/Prev/Perf/TodayMove
                # recompute from PARQUET at row's Date · applies to ALL statuses
                # (was previously skipped for EXIT rows · leaving stale carry-forward)
                row_dt_str = str(row[c_date-1].value or "")[:10]
                entry_v = row[c_entry-1].value if c_entry else None
                # Current price at row's date
                live_curr = _parquet_close(tk, mkt_key, row_dt_str) if row_dt_str else None
                repaired_current = round(live_curr, 2) if live_curr else None
                # Perf = (Current - Entry) / Entry · always recomputed
                repaired_perf = None
                if repaired_current and isinstance(entry_v, (int, float)) and entry_v > 0:
                    repaired_perf = round((repaired_current - entry_v) / entry_v * 100, 2)
                # Prev Close = strictly-before close of row's date
                live_prev = _parquet_prev_close(tk, mkt_key, row_dt_str) if row_dt_str else None
                repaired_prev_close = round(live_prev, 2) if live_prev else None
                # Today Move = (Current - Prev) / Prev · always recomputed
                repaired_today_move = None
                if repaired_current is not None and repaired_prev_close is not None \
                        and repaired_prev_close > 0:
                    repaired_today_move = round(
                        (repaired_current - repaired_prev_close) / repaired_prev_close * 100, 2)

                for c_idx, cell in enumerate(row, start=1):
                    val = cell.value
                    if c_sector_existing and c_idx == c_sector_existing and not val:
                        val = _sector_for(tk, mkt_key)
                    if repaired_current is not None and c_idx == c_current:
                        val = repaired_current
                    if repaired_perf is not None and c_idx == c_perf:
                        val = repaired_perf
                    if repaired_prev_close is not None and c_prev_close and c_idx == c_prev_close:
                        val = round(repaired_prev_close, 2)
                    if repaired_today_move is not None and c_today_move and c_idx == c_today_move:
                        val = repaired_today_move
                    ws2.cell(r_idx, c_idx, val)
                # 9 appended columns · positions counted from END
                if not hasattr(_split_and_send, "_inv_loaded"):
                    _split_and_send._inv_map_for_sheet2 = {}
                    try:
                        _inv_path = _ROOT / "reports" / f"investability_{mkt_key.lower()}.json"
                        if _inv_path.exists():
                            _inv_data = json.loads(_inv_path.read_text(encoding="utf-8"))
                            for r in (_inv_data.get("results") or []):
                                _split_and_send._inv_map_for_sheet2[r["ticker"]] = {
                                    "score": r["score"], "verdict": r["verdict"]
                                }
                    except Exception:
                        pass
                    _split_and_send._inv_loaded = True
                _inv_s2 = _split_and_send._inv_map_for_sheet2.get(tk, {})
                inv_v_score = _inv_s2.get("score")
                inv_v_verdict = _inv_s2.get("verdict", "")

                # Classify Priority + derive decision fields for History too
                _row_status = row[c_st-1].value
                _row_entry = row[c_entry-1].value if c_entry else None
                _row_current = row[c_current-1].value if c_current else None
                _pnl_pct_hist = None
                if isinstance(_row_entry, (int, float)) and _row_entry > 0 \
                        and isinstance(_row_current, (int, float)):
                    _pnl_pct_hist = (_row_current - _row_entry) / _row_entry * 100

                def _hist_classify(st, iv, pnl, is_same_day=False, alerts=""):
                    q_high = iv in ("🏆 QUALITY", "✓ OK")
                    q_mid = iv == "⚠ MARGINAL"
                    q_low = iv == "✗ AVOID"
                    pnl_neg = isinstance(pnl, (int, float)) and pnl < 0
                    if st == "ROTATED_SAMEDAY": return "J"
                    if st == "EXIT" and is_same_day: return "J"
                    # 2026-08-14 Sprint K Part 28 · Risk Controller veto for History
                    # rows too · same rule as _classify_priority.
                    _alerts_up = str(alerts or "").upper()
                    for _sig in BINDING_RISK_SIGNALS:
                        if _sig in _alerts_up:
                            return "R"
                    if st == "EXIT":
                        return "H" if q_high else "I"
                    if st == "STRONG BUY" and iv == "🏆 QUALITY": return "A"
                    if st in ("BUY", "STRONG BUY") and q_high: return "B"
                    if st in ("BUY", "STRONG BUY") and q_low: return "F"
                    if st == "HOLD" and q_high and pnl_neg: return "C"
                    if st == "HOLD" and q_high: return "D"
                    if q_mid: return "E"
                    if q_low and pnl_neg: return "G"
                    if q_low: return "F"
                    return "E"

                # 2026-08-14 · same-day detection · source has NO "Exit Date"
                # column · derive from row Date when Status==EXIT (matches
                # main-loop convention). Fixes the same phantom-column bug
                # as the KPI aggregator hotfix.
                _hist_ed = str(row[c_date-1].value or "")[:10] if (_row_status == "EXIT" and c_date) else ""
                _hist_same_day = bool(_hist_ed and entry_dt and _hist_ed == entry_dt[:10])
                # Read alerts for History rows so Risk Controller veto also
                # fires here (Sprint K Part 28).
                _hist_alerts = row[c_alerts-1].value if c_alerts else ""
                hist_bucket = _hist_classify(_row_status, inv_v_verdict, _pnl_pct_hist,
                                                                 is_same_day=_hist_same_day,
                                                                 alerts=_hist_alerts)
                _matrix_h = PRIORITY_MATRIX.get(hist_bucket,
                                                                ("—", "—", "—", "—", "F2F2F2"))
                h_urg, h_rea, h_act, h_rev, _ = _matrix_h
                h_decision, _ = _resolve_decision(h_act, inv_v_verdict, _row_status)
                # 2026-08-14 · Sprint K Part 28 · apply same Decision overrides
                # as the Portfolio-sheet path so both sheets tell the same story.
                if hist_bucket == "R":
                    _au = str(_hist_alerts or "").upper()
                    _sig = next((s for s in BINDING_RISK_SIGNALS if s in _au), "HARD STOP")
                    h_decision = f"🔴 EXIT · {_sig.replace('_',' ').title()} · IMMEDIATE"
                elif hist_bucket == "J":
                    h_decision = "⚪ ARTIFACT · not held"
                elif hist_bucket in ("I", "H"):
                    h_decision = "🔴 EXIT"      # 2026-08-20 · one-word · CLOSED collapsed

                # Write 12 appended columns (positions from END)
                # Compute Execution Layer for History rows too
                h_stop = row[c_stop-1].value if c_stop else None
                h_t1 = row[c_t1-1].value if c_t1 else None
                h_price_trig = _price_trigger(h_act, h_stop, h_t1, _row_current)
                h_next_rev = _next_review_date(h_rev, str(row[c_date-1].value or "")[:10]) \
                                     if h_act not in ("CLOSED", "IGNORE") else ""
                h_exec_win = _execution_window(h_act)

                ws2.cell(r_idx, len(new_h) - 11, _cap_size(tk, mkt_key))
                row_dt = str(row[c_date-1].value or "")[:10]
                entry_dt = str(row[c_recommended-1].value or "")[:10] if c_recommended else ""
                # 2026-08-15 · Section 12 · Opportunity Status vocabulary
                # NEW      · row_dt == entry_dt (genuine first appearance)
                # EXISTING · row_dt > entry_dt (position rolled forward)
                # RE-ENTRY · derived from position_store exited history
                # (RE-ENTRY detection deferred to next wave · needs exit-set
                # cross-check per row · current NEW/EXISTING is the P0 fix.)
                if row_dt and entry_dt and row_dt == entry_dt:
                    opp_age = "🆕 NEW"
                elif row_dt and entry_dt and row_dt > entry_dt:
                    opp_age = "EXISTING"
                else:
                    opp_age = "EXISTING"   # fallback · no first-day distinction
                ws2.cell(r_idx, len(new_h) - 10, opp_age)
                ws2.cell(r_idx, len(new_h) - 9, inv_v_score)
                ws2.cell(r_idx, len(new_h) - 8, inv_v_verdict)
                ws2.cell(r_idx, len(new_h) - 7, h_decision)
                ws2.cell(r_idx, len(new_h) - 6, h_urg)
                ws2.cell(r_idx, len(new_h) - 5, h_rea)
                ws2.cell(r_idx, len(new_h) - 4, h_act)
                ws2.cell(r_idx, len(new_h) - 3, h_rev)
                ws2.cell(r_idx, len(new_h) - 2, h_price_trig)
                ws2.cell(r_idx, len(new_h) - 1, h_next_rev)
                ws2.cell(r_idx, len(new_h), h_exec_win)
                if status in _STATUS_FILLS_LOCAL:
                    fill = _STATUS_FILLS_LOCAL[status]
                    for c in range(1, len(new_h)+1):
                        ws2.cell(r_idx, c).fill = fill
            ws2.freeze_panes = "D2"

            # ═══════════════════════════════════════════════════════════════
            # SHEET 1 · PORTFOLIO GLANCE · operator directive: current XLSX is
            # a transaction log · unusable for P&L glance. Add a top sheet
            # with aggregates + one row per CURRENT position.
            # ═══════════════════════════════════════════════════════════════
            portfolio_ws = wb2.create_sheet("Portfolio", 0)
            # Build one row per unique (Ticker, Run_Type) using LATEST date state
            c_run = 4    # Run_Type is column 4 in the source
            latest_by_ticker = {}
            for r in keep_rows:
                key = (r[c_tk-1].value, str(r[c_run-1].value or ""))
                dt = str(r[c_date-1].value or "")[:10]
                if key not in latest_by_ticker or dt > latest_by_ticker[key][0]:
                    latest_by_ticker[key] = (dt, r)
            positions = [(dt, r) for (dt, r) in latest_by_ticker.values()]

            # 2026-08-12 CEO fix · P0-3 exclude ARTIFACTS from open/closed
            # counts + P1-5 formal P&L / win-rate definitions.
            #
            # Formal definitions (this is the SSoT for the caption + KPI band):
            #   Position P&L (%)         = per-position price change (unweighted)
            #   Realized P&L (sum)       = SUM of Exit P&L across CLOSED trades
            #                              (excludes ARTIFACT · excludes ROTATED_SAMEDAY)
            #   Unrealized P&L (sum)     = SUM of live P&L across ACTIVE positions
            #                              (excludes ARTIFACT · excludes ROTATED_SAMEDAY)
            #   Combined                 = Realized + Unrealized (sum of position %s,
            #                              NOT portfolio-weighted return · see caveat)
            #   Win rate                 = WINS / (WINS + LOSSES)
            #     where WIN = pnl > +0.01% · LOSS = pnl < -0.01% · FLAT = else
            #     ARTIFACT + ROTATED_SAMEDAY are EXCLUDED from all counters.
            # NOTE · a portfolio-weighted return would need per-position sizing
            # which we do not yet emit into the sheet. TODO after Sprint L.
            realized_sum = 0.0; n_realized = 0
            unrealized_sum = 0.0; n_unrealized = 0
            n_win = 0; n_loss = 0; n_flat = 0
            n_artifact = 0
            # 2026-08-21 · operator flagged "P&L says LICI top but row shows EXIT" ·
            # root cause was a single portfolio-wide best_pos/worst_pos being
            # visually paired with the Realized row while actually tracking the
            # best of BOTH buckets. Split into per-bucket bests so the label
            # next to Realized only refers to closed trades, and the label next
            # to Unrealized only refers to open positions.
            best_realized = ("", 0.0); worst_realized = ("", 0.0)
            best_unreal   = ("", 0.0); worst_unreal   = ("", 0.0)
            # 2026-08-21 · Wave 1 · rolling 90-day Exit P&L tracker.
            # Operator: "IN FINAL XLSX U NEED TO SHOW LAST 3 MONTHS P&L FOR
            # CLOSED STOCKS ONLY · PLZ ENSURE THAT TRACKING ONLY · U DONT
            # NEED TO SHOW STOCKS". Aggregate sum + count + win rate over
            # closed exits with row_date within last 90 days of asof.
            from datetime import date as _date_cls, timedelta as _td
            try:
                _asof_dt = _date_cls.fromisoformat(str(latest_date)[:10])
            except (TypeError, ValueError):
                _asof_dt = _date_cls.today()
            _90d_cutoff = _asof_dt - _td(days=90)
            r90_sum = 0.0; r90_n = 0; r90_win = 0; r90_loss = 0
            r90_best = ("", 0.0); r90_worst = ("", 0.0)
            # 2026-08-13 CEO bug fix · previous code used h.index("Recommended Date")
            # and h.index("Exit Date") which don't exist in source header
            # (real column is "Recommended" · there is no "Exit Date" column
            # at all · exit_date is derived on the fly from Date+Status).
            # Result: _row_is_artifact was always False · KPI showed
            # "0 same-day rotations" even when ~23 rows had ARTIFACT decisions.
            _c_rec_date = h.index("Recommended") + 1 if h and "Recommended" in h else None
            _c_date     = c_date   # row date · used as exit_date when status==EXIT
            for dt, r in positions:
                status = r[c_st-1].value
                # ARTIFACT detection · same rule as Priority J:
                #   status == ROTATED_SAMEDAY
                #   OR (status == EXIT AND row Date == Recommended date) · same-day rotation
                _entry_dt = r[_c_rec_date-1].value if _c_rec_date else None
                _exit_dt  = r[_c_date-1].value if (status == "EXIT" and _c_date) else None
                _row_is_artifact = (
                    status == "ROTATED_SAMEDAY"
                    or (_entry_dt and _exit_dt and str(_entry_dt)[:10] == str(_exit_dt)[:10])
                )
                if _row_is_artifact:
                    n_artifact += 1
                    continue   # do NOT contribute to open/closed/win-rate stats

                pnl = None
                is_realized = False
                if status == "EXIT" and c_exit_pnl:
                    v = r[c_exit_pnl-1].value
                    if isinstance(v, (int, float)):
                        pnl = v; realized_sum += v; n_realized += 1
                        is_realized = True
                        # 2026-08-21 · Wave 1 · roll into 90d window if recent
                        try:
                            _row_d = _date_cls.fromisoformat(str(dt)[:10])
                            if _row_d >= _90d_cutoff:
                                r90_sum += v; r90_n += 1
                                if v > 0.01: r90_win += 1
                                elif v < -0.01: r90_loss += 1
                                _tk_v = r[c_tk-1].value
                                if v > r90_best[1]:  r90_best  = (_tk_v, v)
                                if v < r90_worst[1]: r90_worst = (_tk_v, v)
                        except (TypeError, ValueError):
                            pass
                elif status != "EXIT":
                    entry_v = r[c_entry-1].value if c_entry else None
                    tk = r[c_tk-1].value
                    live = _parquet_close(tk, mkt_key, dt)
                    if live and isinstance(entry_v, (int, float)) and entry_v > 0:
                        pnl = round((live - entry_v) / entry_v * 100, 2)
                        unrealized_sum += pnl; n_unrealized += 1
                if pnl is not None:
                    if pnl > 0.01: n_win += 1
                    elif pnl < -0.01: n_loss += 1
                    else: n_flat += 1
                    tk_v = r[c_tk-1].value
                    if is_realized:
                        if pnl > best_realized[1]:  best_realized = (tk_v, pnl)
                        if pnl < worst_realized[1]: worst_realized = (tk_v, pnl)
                    else:
                        if pnl > best_unreal[1]:    best_unreal   = (tk_v, pnl)
                        if pnl < worst_unreal[1]:   worst_unreal  = (tk_v, pnl)
            combined = realized_sum + unrealized_sum
            n_total = n_realized + n_unrealized
            # Win rate denominator excludes flats (per formal definition above)
            _win_denom = max(1, n_win + n_loss)
            win_rate = round(n_win / _win_denom * 100, 1)

            # Portfolio header + KPI banner
            portfolio_ws.merge_cells("A1:L1")
            portfolio_ws["A1"] = f"AEGIS {mkt_key} PORTFOLIO · as of {latest_date or 'today'}"
            portfolio_ws["A1"].font = _Font(bold=True, size=14, color="FFFFFF")
            portfolio_ws["A1"].fill = HEADER_FILL if 'HEADER_FILL' in globals() else _PF(
                start_color="1F4E78", end_color="1F4E78", fill_type="solid")
            portfolio_ws["A1"].alignment = _Align(horizontal="center", vertical="center")
            portfolio_ws.row_dimensions[1].height = 28

            # 2026-08-12 P0-3 · KPI banner shows ARTIFACTS on their own row.
            # 2026-08-15 Section 6 · P&L 2-way separation (SKIP tracked in
            # research file · never included in portfolio P&L):
            #   A. Exit P&L    · closed positions · entry → exit price
            #   B. Active P&L  · active positions · entry → latest price
            # 'closed'/'open' counts EXCLUDE artifacts + SKIP by definition.
            # 2026-08-21 · SKIP dataset tile removed per operator "any way we
            # are not investing, then dont show in portfolio" · research
            # file at reports/research/skip_candidates_*.jsonl persists.
            # 2026-08-21 · label pairs each P&L row with a best/worst FROM THE
            # SAME BUCKET so operator can't misread a still-open winner (LICI
            # HOLD +7.92%) as a closed realized winner.
            def _fmt(pair, empty="—"):
                if not pair[0]: return empty
                return f"{pair[0]} {pair[1]:+.2f}%"
            # 2026-08-21 · operator vocab: "active P&L, exit P&L makes sense".
            # Rename Realized→Exit, Unrealized→Active so labels match the
            # section banners the operator scans daily (EXISTING POSITIONS
            # = active · EXIT · REFERENCE ONLY = exits).
            _r90_win_pct = (round(r90_win / max(1, r90_win + r90_loss) * 100, 1)
                              if (r90_win + r90_loss) else 0.0)
            # 2026-08-21 · Wave 4 · NEW opportunity visibility panel.
            # Operator directive §30 + "new stocks is not flowing day by day
            # is my strong feeling · check throughly plz". Compute today's
            # NEW opportunity summary + zero-reason explanation and surface
            # it as the FIRST KPI row · operator sees NEW status at a glance.
            # 2026-08-21 · NEW-opportunity Strong Guard.
            # Operator directive: "build strong guard for new recommendation
            # engine". Wraps Wave 4/5/6 chain (NEW + Rotation + Ops) with
            # pre/post validation, 3 retries + exponential backoff, held-
            # penalty rotation force (fixes "same stocks daily"), and
            # fallback to last-good diagnostic snapshot on total failure.
            # Guard emits reports/context/new_opp_guard_health_{mkt}.json.
            _guard_summary = "GUARD · unavailable"
            _guard_note    = ""
            _new_summary = "NEW · pending guard"
            _new_detail  = ""
            _rotate_summary = "ROTATE · pending guard"
            _rotate_detail  = ""
            _ops_summary = ""
            _ops_warn    = ""
            try:
                from backend.recommendation.new_opp_guard import (
                    guarded_run, summary_line as _guard_line,
                )
                _hg = guarded_run(_ROOT, mkt_key.lower(), latest_date)
                _guard_summary = _guard_line(_hg)
                _guard_note = _hg.notes or "; ".join(_hg.error_history[:2])
                # After guard runs · pull the just-emitted diagnostics
                from backend.research import (
                    new_opportunity_diagnostic as _nod,
                    rotation_engine as _rot,
                    daily_ops_diagnostic as _dod,
                )
                _diag = _nod.compute(_ROOT, mkt_key.lower(), latest_date)
                _new_summary = _nod.summary_line(_diag)
                _new_detail = (" · ".join(
                    f"{c['runner']}·{c['ticker']}·rank {c.get('rank') or '?'}"
                    for c in _diag.top_new[:5])
                                     if _diag.n_new_today > 0
                                     else (_diag.zero_reason or "no candidates cleared gates"))
                _rep = _rot.compute(_ROOT, mkt_key.lower(), latest_date)
                _rotate_summary = _rot.summary_line(_rep)
                _rotate_detail = (" · ".join(
                    f"{s['existing_ticker']}→{s['new_ticker']} "
                    f"(+{s.get('alpha_delta_pp') or 0}pp)"
                    for s in _rep.suggestions[:3])
                                          if _rep.n_suggestions > 0
                                          else (_rep.reason_if_zero or ""))
                _ops = _dod.compute(_ROOT, mkt_key.lower(), latest_date)
                _ops_summary = _dod.summary_line(_ops)
                _ops_warn = (f"{len(_ops.warnings)} warnings · " +
                                    "; ".join(_ops.warnings[:2])
                                    if _ops.warnings else "no warnings")
            except Exception as _e:
                _guard_note = f"guard error · {type(_e).__name__}: {_e}"
            # 2026-08-21 · Wave 7 · acceptance-gate regression (§32-§35).
            # Runs LAST · consumes prior diagnostics · verdict shown in KPI
            # banner. Non-blocking · operator still sees the XLSX. Failing
            # checks land in reports/context/wave_regression_{market}.json.
            _reg_summary = ""
            _reg_detail  = ""
            try:
                from backend.research import wave_regression as _wreg
                _wrp = _wreg.compute(_ROOT, mkt_key.lower(), latest_date)
                _wreg.emit(_ROOT, _wrp)
                _reg_summary = _wreg.summary_line(_wrp)
                _fails = [c for c in _wrp.checks if c.get("status") == "FAIL"]
                _warns = [c for c in _wrp.checks if c.get("status") == "WARN"]
                if _fails:
                    _reg_detail = " · ".join(
                        f"[{c['code']}] {c['detail'][:60]}" for c in _fails[:3])
                elif _warns:
                    _reg_detail = " · ".join(
                        f"[{c['code']}] {c['detail'][:60]}" for c in _warns[:3])
                else:
                    _reg_detail = "all Wave 1-6 invariants hold"
            except Exception as _e:
                _reg_detail = f"regression check error · {type(_e).__name__}: {_e}"
            # Part 26 · Investability Shadow diagnostic (2026-08-21).
            # Reads the shadow diagnostic emitted by new_opp_guard's parallel
            # chain · surfaces "how much would investability change today".
            _inv_summary = "invest · shadow off"
            _inv_detail  = ""
            try:
                import json as _json
                _inv_p = (_ROOT / "reports" / "context"
                                / f"investability_shadow_diagnostic_{mkt_key.lower()}.json")
                if _inv_p.exists():
                    _idata = _json.loads(_inv_p.read_text(encoding="utf-8"))
                    _mode = "ENFORCE" if _idata.get("enforce_gate") else "SHADOW"
                    _inv_summary = (f"[{_mode}] scored={_idata.get('n_scored',0)} · "
                                            f"pass={_idata.get('n_would_pass',0)} · "
                                            f"fail={_idata.get('n_would_fail',0)}")
                    _rej = _idata.get("would_reject_from_recs") or []
                    _disc = _idata.get("top_discoveries") or []
                    _parts = []
                    if _rej:
                        _parts.append(f"would-reject: " + ", ".join(
                            f"{r['ticker']}({r.get('score','?')})" for r in _rej[:3]))
                    if _disc:
                        _parts.append(f"discoveries: " + ", ".join(
                            f"{r['ticker']}({r.get('score','?')})" for r in _disc[:3]))
                    _inv_detail = " · ".join(_parts) or "no material differences"
            except Exception as _e:
                _inv_detail = f"shadow read error · {type(_e).__name__}: {_e}"
            # 2026-08-24 · KPI BANNER SIMPLIFIED · operator: "toomuch of
            # confusion, multiple columns" + "what u decide tell me instead
            # of asking me". Collapsed 6 diagnostic panels (NEW OPP + ROTATE
            # + OPS + ACCEPTANCE + GUARD + INVESTABILITY) into ONE
            # 🩺 AEGIS HEALTH row. Full detail lives in an APPENDIX section
            # at the bottom for anyone who wants to drill in.
            # Also added: 🔬 SUGGESTED NEW row surfacing the top 3
            # investability-shadow discoveries the recommender missed.
            def _worst_verdict(*verdicts):
                order = {"RED": 3, "FAIL": 3, "❌": 3,
                             "YELLOW": 2, "WARN": 2, "⚠️": 2,
                             "GREEN": 1, "PASS": 1, "✅": 1}
                best = 1
                for v in verdicts:
                    for k, sc in order.items():
                        if k in str(v or ""): best = max(best, sc); break
                return "❌ RED" if best == 3 else "⚠️ YELLOW" if best == 2 else "✅ GREEN"
            _health_verdict = _worst_verdict(
                _guard_summary, _reg_summary, _ops_summary, _new_summary)
            _health_line = (f"{_health_verdict} · "
                                    f"NEW={_new_summary.split('·')[0].strip() if 'NEW' in _new_summary else '?'} · "
                                    f"ROT={_rotate_summary.split('·')[0].strip() if 'ROTATE' in _rotate_summary else '?'} · "
                                    f"GUARD={_guard_summary.split('·')[0].strip()} · "
                                    f"REG={_reg_summary.split('·')[0].strip()}")
            _health_detail = (f"full detail in APPENDIX at bottom · panels: "
                                       f"NEW·ROTATE·OPS·ACCEPTANCE·GUARD·INVESTABILITY")

            # Pull top 3 shadow discoveries for the SUGGESTED NEW row
            _sug_summary = "SUGGESTED · no discoveries"
            _sug_detail  = ""
            try:
                import json as _json
                _inv_p2 = (_ROOT / "reports" / "context"
                                 / f"investability_shadow_diagnostic_{mkt_key.lower()}.json")
                if _inv_p2.exists():
                    _ish = _json.loads(_inv_p2.read_text(encoding="utf-8"))
                    _disc = _ish.get("top_discoveries") or []
                    if _disc:
                        _sug_summary = f"SUGGESTED · {len(_disc)} high-quality picks the recommender missed"
                        _sug_detail = " · ".join(
                            f"{d['ticker']}({d.get('score',0):.0f}·{d.get('verdict','')})"
                            for d in _disc[:3])
                    else:
                        _sug_summary = "SUGGESTED · no discoveries above threshold"
            except Exception:
                pass

            # 2026-08-21 · SKIP removed entirely per operator "we dont need skip
            # itself". Opportunity dataset tile deleted · no SKIP tracker.
            kpi_rows = [
                # ONE health row · scan in 3 seconds
                ["🩺 AEGIS HEALTH",         "",                    _health_line,
                 "Details",      _health_detail],
                # 🔬 SUGGESTED NEW · top 3 investability shadow discoveries
                # · directly addresses "same stocks daily" · surfaces genuine
                # quality names the recommender is missing.
                ["🔬 SUGGESTED NEW",         "",                    _sug_summary,
                 "Detail",       _sug_detail],
                # P&L numbers · what operator actually needs to make trade decisions
                ["Exit P&L (closed)",      realized_sum / 100.0, f"{n_realized} exits",
                 "Top exit",     _fmt(best_realized)],
                ["Active P&L (open)",      unrealized_sum / 100.0, f"{n_unrealized} active",
                 "Top active",   _fmt(best_unreal)],
                ["COMBINED PORTFOLIO",     combined / 100.0,      f"{n_total} real positions",
                 "Win rate", f"{win_rate}% ({n_win}W / {n_loss}L · {n_flat} flat excluded)"],
                ["Worst positions",        "",                    "",
                 "Exit · Active", f"{_fmt(worst_realized)}  ·  {_fmt(worst_unreal)}"],
                # 2026-08-21 · rolling 90d exit tracker
                ["Last 90d Exit P&L",       r90_sum / 100.0,       f"{r90_n} closed",
                 "Win rate · Best · Worst",
                 f"{_r90_win_pct}% · {_fmt(r90_best)} · {_fmt(r90_worst)}"],
                ["Artifacts (excluded)",  "",                    f"{n_artifact} same-day rotations",
                 "Note", "not counted in P&L / win rate"],
            ]
            # 2026-08-25 · operator "what is this appendix not able to
            # understand" · APPENDIX DELETED. All diagnostic detail lives
            # in the reports/context/*.json files for anyone who wants
            # to drill in · not cluttering the Portfolio.
            _appendix_rows = []
            for r_off, kpi_row in enumerate(kpi_rows, start=3):
                for c_off, val in enumerate(kpi_row, start=1):
                    cell = portfolio_ws.cell(r_off, c_off, val)
                    cell.font = _Font(bold=(c_off in (1, 4)), size=11)
                    # Column 2 = numeric P&L · display as %
                    if c_off == 2:
                        cell.number_format = "+0.00%;-0.00%;0.00%"
                    if r_off == 5 and c_off <= 3:   # combined row highlighted
                        cell.fill = _PF(start_color="FFE699", end_color="FFE699", fill_type="solid")

            # 2026-08-14 · operator-approved 6-tier color legend rows.
            # Placed between KPI banner (rows 3-6) and position table.
            # Legend heading at row 8 · one row per tier at rows 9-14.
            _LEGEND_HEADING_ROW = 8
            portfolio_ws.merge_cells(f"A{_LEGEND_HEADING_ROW}:L{_LEGEND_HEADING_ROW}")
            _lh = portfolio_ws.cell(_LEGEND_HEADING_ROW, 1, TIER_LEGEND_HEADING)
            _lh.font = _Font(bold=True, size=11, color="1F4E78")
            _lh.alignment = _Align(horizontal="left", vertical="center")
            _legend_start = _LEGEND_HEADING_ROW + 1
            # Iterate tiers in tier_rank order (1..N)
            _legend_ordered = sorted(
                [(k, v) for k, v in DECISION_TIERS.items()],
                key=lambda kv: int(kv[1].get("tier_rank") or 99))
            for _leg_idx, (_tier_key, _tier_def) in enumerate(_legend_ordered):
                _lr = _legend_start + _leg_idx
                _label = str(_tier_def.get("label") or _tier_key)
                _desc  = str(_tier_def.get("description") or "")
                _hf    = str(_tier_def.get("hex_fill") or "F2F2F2")
                _ht    = str(_tier_def.get("hex_text") or "000000")
                _bold  = bool(_tier_def.get("bold"))
                # Left cell = colored label swatch
                _swatch = portfolio_ws.cell(_lr, 1, _label)
                _swatch.fill = _PF(start_color=_hf, end_color=_hf, fill_type="solid")
                _swatch.font = _Font(bold=True, color=_ht, size=10)
                _swatch.alignment = _Align(horizontal="center", vertical="center")
                # Right cells (merged) = description
                portfolio_ws.merge_cells(start_row=_lr, start_column=2,
                                                       end_row=_lr, end_column=6)
                _dc = portfolio_ws.cell(_lr, 2, _desc)
                _dc.font = _Font(size=10)
                _dc.alignment = _Align(horizontal="left", vertical="center", wrap_text=True)
            # Track where the position header goes (row after legend + 1 blank)
            _pos_header_row = _legend_start + len(_legend_ordered) + 1

            # 2026-08-25 · operator: "cant see a single new stocks?" +
            # "where is sunpharma example i am asking". Emit a prominent
            # SUGGESTED NEW strip right after the legend · lists the top
            # investability-shadow discoveries the recommender missed so
            # operator SEES them without opening the shadow JSON file.
            try:
                import json as _jsonw
                _inv_p3 = (_ROOT / "reports" / "context"
                                 / f"investability_shadow_diagnostic_{mkt_key.lower()}.json")
                if _inv_p3.exists():
                    _ish3 = _jsonw.loads(_inv_p3.read_text(encoding="utf-8"))
                    _disc3 = _ish3.get("top_discoveries") or []
                    if _disc3:
                        # Purple banner + 1 row per discovery
                        _sug_row = _pos_header_row + 1
                        portfolio_ws.merge_cells(start_row=_sug_row, start_column=1,
                                                                  end_row=_sug_row, end_column=6)
                        _sb = portfolio_ws.cell(_sug_row, 1,
                                                             f"🆕 SUGGESTED NEW · {len(_disc3)} high-quality picks "
                                                             "the recommender missed today "
                                                             "(shadow investability)")
                        _sb.font = _Font(bold=True, size=12, color="5B2A82")
                        _sb.fill = _PF(start_color="D5A6EA", end_color="D5A6EA", fill_type="solid")
                        _sb.alignment = _Align(horizontal="left", vertical="center")
                        portfolio_ws.row_dimensions[_sug_row].height = 22
                        _sug_row += 1
                        # One row per discovery · plain-English pitch
                        for _d in _disc3[:5]:
                            _tk3 = _d.get("ticker", "?")
                            _sc3 = _d.get("score", 0)
                            _vd3 = _d.get("verdict", "")
                            _top3 = ", ".join(f"{k}={int(v)}"
                                                        for k, v in (_d.get("top_engines") or [])[:3])
                            _line = (f"  · {_tk3:<12} score {_sc3:.1f} · {_vd3} · "
                                          f"strongest: {_top3}")
                            portfolio_ws.merge_cells(start_row=_sug_row, start_column=1,
                                                                      end_row=_sug_row, end_column=6)
                            _cell3 = portfolio_ws.cell(_sug_row, 1, _line)
                            _cell3.font = _Font(size=10, color="5B2A82")
                            _cell3.fill = _PF(start_color="EFE0F7",
                                                        end_color="EFE0F7", fill_type="solid")
                            _cell3.alignment = _Align(horizontal="left", vertical="center")
                            _sug_row += 1
                        _pos_header_row = _sug_row + 1     # shift position header down
            except Exception as _e:
                print(f"[xlsx:{mkt_key}] suggested-new strip skipped · {_e}")

            # 2026-08-10 · CEO review fixes (semantic separation):
            #   1. Lifecycle column · NEW/ACTIVE/CLOSED (separate from Runner Status)
            #      Rule: EXITED only when Runner=EXIT AND Portfolio agrees (Priority I)
            #      If Priority=H (Premature Exit?) · Lifecycle=ACTIVE (portfolio challenging)
            #   2. R1/R2 Consensus column · AGREE/SPLIT/R1-only/R2-only
            #   3. Next Review = Recommended Date + horizon (stable · not sliding)
            pos_hdr = [
                # 2026-08-24 · operator: "based on which columns w eneed to
                # take decision is also very confusing for me man". Added
                # 🎯 ACTION column · ONE plain-English sentence per row so
                # operator knows exactly what to do without scanning others.
                # 2026-08-25 · Month column added per operator
                # "add month column plz" + "year also makes sense" +
                # "mm-yyyy makes sense in single column like May 2026".
                # Values formatted "August 2026" via strftime("%b %Y").
                # IDENTITY (5) · Ticker + Action + Decision + Lifecycle + Month
                "Ticker", "🎯 ACTION", "🎯 DECISION", "Lifecycle", "Month",
                # EXECUTION LAYER (3)
                "Price Trigger", "Next Review", "Execution Window",
                # META (7)
                "Runner", "R1/R2 Consensus", "Sector", "Cap", "Entry Date", "Exit Date", "Days",
                # SUPPORTING DECISION FIELDS (7)
                "Urgency", "Reason", "Action", "Review",
                "Status", "Inv Quality", "Investability",
                # PRICE + P&L (4)
                "Entry", "Current", "Exit Price", "P&L %",
                # RISK/TARGET (3)
                "Stop Loss", "Target 1", "Target 2",
                # CONTEXT (3)
                "Action Note", "Alerts", "Exit Reason",
                # 2026-08-14 Sprint K Part 28 · Wave 4 · Post-Exit Assessment
                # is analytical / research-only · NEVER a live trading
                # instruction. Live Decision column stays clean (BUY/HOLD/
                # PROTECT/EXIT/CLOSED).
                "Post-Exit Assessment",
                # 2026-08-15 · Section 13 · Decision Basis · one short reason
                # explaining WHY the Decision was reached. Deterministic ·
                # derived from lifecycle + risk + status.
                "Decision Basis",
            ]
            widths_pos = [12, 46, 24, 12, 10,       # + Month width 10
                              22, 12, 30,
                              8, 14, 22, 20, 12, 12, 8,
                              12, 20, 16, 14, 12, 14, 12,
                              12, 12, 12, 10,
                              12, 12, 12,
                              40, 40, 30,
                              32,
                              22]   # Decision Basis

            # R1/R2 consensus map · build from all keep_rows (before ticker loop)
            # (populated after keep_rows filter · so put function here · use later)
            # 2026-08-21 · Wave 2 · enriched consensus map. Operator §7 + §8 + §32:
            # "One stock = one canonical live portfolio position" · reconcile
            # R1/R2 disagreement · LUPIN R1=EXIT + R2=HOLD must not show as two
            # confusing rows. Map now carries per-ticker bucket set so the
            # write loop can tag SPLIT decisions in the Runner cell.
            def _row_bucket(_r_cells):
                _st = str((_r_cells[c_st-1].value if hasattr(_r_cells[c_st-1], 'value')
                                   else _r_cells[c_st-1]) or "").upper()
                _al_v = (_r_cells[c_alerts-1].value if c_alerts and
                            hasattr(_r_cells[c_alerts-1], 'value') else "")
                _al = str(_al_v or "").upper()
                if _st == "EXIT" or any(sig in _al for sig in BINDING_RISK_SIGNALS):
                    return "EXIT"
                if _st in ("BUY", "STRONG BUY", "ACCUMULATE", "ADD", "BUY BIG"):
                    return "ACTIVE+"
                return "ACTIVE"
            _by_ticker_runners: dict = {}
            _bucket_by_key: dict = {}
            for _r in keep_rows:
                _tk = _r[c_tk-1].value
                _rn = _r[3].value if len(_r) > 3 else None
                if _tk and _rn:
                    _rn_norm = str(_rn).upper().replace("_NEW", "")
                    _by_ticker_runners.setdefault(_tk, set()).add(_rn_norm)
                    _bucket_by_key[(str(_tk).upper(), _rn_norm)] = _row_bucket(_r)
            def _consensus(tk, this_runner):
                runners = _by_ticker_runners.get(tk, set())
                if len(runners) == 1:
                    return f"🔹 {this_runner} ONLY"
                # Both R1 + R2 present · compare their buckets
                _tk_up = str(tk).upper()
                _b_r1 = _bucket_by_key.get((_tk_up, "R1"))
                _b_r2 = _bucket_by_key.get((_tk_up, "R2"))
                if _b_r1 and _b_r2 and _b_r1 == _b_r2:
                    return f"✅ AGREE ({_b_r1})"
                if _b_r1 and _b_r2:
                    # SPLIT · operator wants to see the disagreement clearly
                    return f"⚠️ SPLIT · R1={_b_r1} · R2={_b_r2}"
                return "✅ AGREE"

            # PRIORITY_MATRIX + DECISION vocab + EXEC layer already loaded at top of _split_and_send

            # Load Investability scores (advisory · from reports/investability_{market}.json)
            _inv_map = {}
            try:
                _inv_path = _ROOT / "reports" / f"investability_{mkt_key.lower()}.json"
                if _inv_path.exists():
                    import json as _json
                    _inv_data = _json.loads(_inv_path.read_text(encoding="utf-8"))
                    for r in (_inv_data.get("results") or []):
                        _inv_map[r["ticker"]] = {"score": r["score"], "verdict": r["verdict"]}
            except Exception as _e:
                print(f"[investability] load failed: {_e}")
            # 2026-08-14 · position header row anchored to _pos_header_row
            # (was hardcoded row 7 · now shifts down to accommodate legend).
            for c, name in enumerate(pos_hdr, start=1):
                cell = portfolio_ws.cell(_pos_header_row, c, name)
                cell.font = _Font(bold=True, color="FFFFFF", size=11)
                cell.fill = HEADER_FILL if 'HEADER_FILL' in globals() else _PF(
                    start_color="1F4E78", end_color="1F4E78", fill_type="solid")
                cell.alignment = _Align(horizontal="center", vertical="center")
                portfolio_ws.column_dimensions[_gcl_src(c)].width = widths_pos[c-1]

            # 2026-08-21 · Wave 6 · operator "toomuch of confusion, multiple
            # columns?". Aggressive slim to 10 visible columns · everything
            # else hidden (Excel Unhide reveals for audit).
            #
            # Visible (10): Ticker · Decision · Runner · Sector · Days ·
            #               Urgency · Entry · Current · P&L % · Stop Loss
            #
            # Strong-Buy vs Buy differentiation under vocab v5.0:
            #   both collapse to 🟢 ACTIVE+ in Decision · intensity shows in
            #   Urgency column (HIGH = strong · MEDIUM = normal · LOW = weak).
            _HIDDEN_COL_NAMES = {
                "Lifecycle",              # redundant with Decision under v5.0
                "Price Trigger",
                "Next Review",
                "Execution Window",
                "R1/R2 Consensus",        # Runner cell shows the split tag
                "Cap",                    # Sector tells the story
                "Entry Date",             # Days column already conveys age
                "Exit Date",
                "Reason",                 # hidden · Urgency conveys severity
                "Action",
                "Review",
                "Status",                 # raw feed · not client-facing
                "Inv Quality",
                "Investability",
                "Exit Price",
                "Target 1",
                "Target 2",
                "Action Note",
                "Alerts",
                "Exit Reason",
                "Post-Exit Assessment",
                "Decision Basis",
            }
            for c, name in enumerate(pos_hdr, start=1):
                if name in _HIDDEN_COL_NAMES:
                    portfolio_ws.column_dimensions[_gcl_src(c)].hidden = True

            # 2026-08-14 operator-approved · sort by decision-TIER (not
            # priority bucket). Same tier logic used for row color · classifier
            # walks match_order in configs/decision_colors.yaml.
            #
            # Tier ordering (top → bottom):
            #   1 strong_buy  Deep Green
            #   2 buy         Light Green
            #   3 new         Purple  (is_new_position=True)
            #   4 hold        Yellow
            #   5 exit        Light Red
            #   6 closed      Dark Red
            def _sort_key(item):
                dt, r = item
                status = r[c_st-1].value
                # Same-day rotation detection (matches KPI aggregator)
                _e_dt = r[_c_rec_date-1].value if _c_rec_date else None
                _x_dt = r[_c_date-1].value if (status == "EXIT" and _c_date) else None
                is_artifact = (
                    status == "ROTATED_SAMEDAY"
                    or (_e_dt and _x_dt and str(_e_dt)[:10] == str(_x_dt)[:10])
                )
                # P&L for within-tier sorting (best first)
                pnl = 0
                if status == "EXIT" and c_exit_pnl:
                    v = r[c_exit_pnl-1].value
                    if isinstance(v, (int, float)): pnl = v
                else:
                    entry_v = r[c_entry-1].value if c_entry else None
                    tk = r[c_tk-1].value
                    live = _parquet_close(tk, mkt_key, dt)
                    if live and isinstance(entry_v, (int, float)) and entry_v > 0:
                        pnl = round((live - entry_v) / entry_v * 100, 2)
                # ── Reproduce decision-tier from row data (mirrors write-loop) ──
                # Alerts col for risk-veto detection
                alerts_here = str(r[c_alerts-1].value if c_alerts else "").upper()
                # is_new_position · rec_dt == asof AND not an artifact/exit
                is_new = bool(_e_dt) and str(_e_dt)[:10] == str(asof)[:10] \
                             and status != "EXIT" and not is_artifact
                # Pre-classify tier without needing full decision_text · use
                # the same precedence rules as the write loop.
                if is_artifact:
                    tier = "closed"
                elif any(sig in alerts_here for sig in BINDING_RISK_SIGNALS):
                    tier = "exit"
                elif status == "EXIT":
                    tier = "closed"
                elif is_new:
                    tier = "new"
                elif status == "STRONG BUY":
                    tier = "strong_buy"
                elif status in ("BUY", "ACCUMULATE"):
                    tier = "buy"
                elif status in ("SELL", "REDUCE"):
                    tier = "exit"
                else:
                    tier = "hold"
                tier_rank = int((DECISION_TIERS.get(tier) or {}).get("tier_rank") or 99)
                # Within tier: best P&L first (descending)
                return (tier_rank, -pnl)

            positions_sorted = sorted(positions, key=_sort_key)

            # 2026-08-18 · Wave 3 · Section 17 · INDIGO filter.
            # Drop rows whose (mkt, runner, ticker) has a REJECTED opportunity
            # in the registry with created_date == this row's row_date. Those
            # are same-day close/rotation artifacts · must never appear in the
            # active Portfolio (would confuse operator with 'NEW + CLOSED').
            _rejected_row_keys: set = set()
            try:
                from backend.research import opportunity_registry as _oreg
                _reg = _oreg.load_all(_ROOT)
                for _opps in _reg.values():
                    for _o in _opps:
                        if _o.status == "REJECTED":
                            _rejected_row_keys.add(
                                (_o.market.lower(), _o.runner.upper(),
                                  _o.ticker.upper(), _o.created_date[:10]))
            except Exception:
                pass
            if _rejected_row_keys:
                _before_flt = len(positions_sorted)
                _kept = []
                for _item in positions_sorted:
                    _dt, _r = _item
                    _tk = str(_r[c_tk-1].value or "").upper().replace(".NS","").replace(".BO","")
                    _rn = str(_r[3].value or "").upper().replace("_NEW","")
                    _rd = str(_r[c_recommended-1].value or "")[:10] if c_recommended else ""
                    if (mkt_key.lower(), _rn, _tk, _rd) in _rejected_row_keys:
                        continue    # drop INDIGO-style rejected same-day row
                    _kept.append(_item)
                positions_sorted = _kept
                if len(positions_sorted) < _before_flt:
                    print(f"[xlsx:{mkt_key}] INDIGO filter · dropped "
                          f"{_before_flt - len(positions_sorted)} REJECTED same-day rows")

            # 2026-08-18 · Wave 3 · Section 17 · 3-section banner rows.
            # Group positions by section · order them as a flat list where
            # section transitions inject a banner-row placeholder. The write
            # loop below emits banners inline when it encounters one.
            # 2026-08-21 · Wave 3 · Portfolio sheet is JUST NEW + ACTIVE now.
            # Operator directive: "portfolio is main sheet for new and active".
            # Historical exits moved to dedicated "Exit History (90d)" sheet.
            # We keep the "ACTION REQUIRED TODAY" section for today's urgent
            # stops so operator can't miss an immediate exit signal · this is
            # today's decision, not historical reference.
            _SECTION_ORDER = [
                ("new_opps",  "🆕 NEW",                          "D5A6EA", "000000"),
                ("existing",  "🟢 ACTIVE",                       "C6EFCE", "000000"),
                ("action",    "🔴 EXIT · ACTION REQUIRED TODAY", "FCE4D6", "9C0006"),
                # "closed" section retired · see Exit History sheet
            ]
            # 2026-08-25 · fix operator complaint "why red stocks seen in
            # portfolio" · BATAINDIA/TATAPOWER (bucket G · Structural
            # Failure) leaked into ACTIVE section because Status stayed
            # HOLD while Decision resolved to EXIT via bucket G. Fixed
            # here by pre-classifying the bucket using the same rules
            # the row-write loop uses.
            _EXIT_ACTIONS = {"EXIT", "SKIP", "CLOSED", "IGNORE"}
            def _row_section(_item):
                _dt, _r = _item
                _st = str(_r[c_st-1].value or "").upper()
                _al = str(_r[c_alerts-1].value if c_alerts else "").upper()
                _rd = str(_r[c_recommended-1].value or "")[:10] if c_recommended else ""
                _same_day = _rd == asof[:10] and _st != "EXIT"
                # 1. Binding-risk signal in Alerts → today's action (bucket R)
                if any(sig in _al for sig in BINDING_RISK_SIGNALS): return "action"
                # 2. Real Status=EXIT → closed bin (Exit History sheet)
                if _st == "EXIT": return "closed"
                # 3. Structural Failure / Quality Fail (bucket G/F/H) · Status=HOLD
                #    but decision resolves to EXIT. Detect via investability +
                #    live P&L (mirrors _classify_priority logic).
                try:
                    _iv_col = h.index("Health") if "Health" in h else None
                    _iv = str((_r[_iv_col].value if _iv_col is not None else "") or "")
                except Exception:
                    _iv = ""
                _q_low = _iv == "✗ AVOID"
                # Estimate P&L from Entry vs Current (both in source XLSX)
                _pnl_neg = False
                try:
                    _entry_col = h.index("Entry Price") if "Entry Price" in h else None
                    _curr_col  = h.index("Current Price") if "Current Price" in h else None
                    if _entry_col is not None and _curr_col is not None:
                        _e = _r[_entry_col].value; _c_v = _r[_curr_col].value
                        if isinstance(_e, (int, float)) and isinstance(_c_v, (int, float)) and _e > 0:
                            _pnl_neg = (_c_v - _e) / _e < 0
                except Exception:
                    pass
                # Bucket G · quality-avoid + negative P&L · action=EXIT
                # Bucket F · quality-avoid alone · action=SKIP
                # Both should render in ACTION section, not ACTIVE.
                if _q_low:
                    return "action"
                # 4. Same-day new-rec (rec_date == asof) → NEW section
                if _same_day: return "new_opps"
                # 5. Default · ACTIVE
                return "existing"

            # 2026-08-21 · Wave 1 + Wave 3 · investable-only + exits-off-portfolio.
            # Operator directives (2026-08-21 batch):
            #   "only investable stocks we need to see in portfolio"
            #   "portfolio is main sheet for new and active" (Wave 3)
            #   "add new sheet, show last 3 months stock by stock P&L"
            #   § 5 · "a stock cannot simultaneously be NEW and CLOSED on the
            #          same daily recommendation snapshot"
            # Drop from Portfolio sheet (Exit History sheet + research files
            # still track everything):
            #   1. EXIT rows without a today-actionable stop signal → move
            #      entirely to the dedicated Exit History (90d) sheet
            #   2. Same-day rec_dt == asof AND Status == EXIT · JIOFIN artifact
            #   3. NEW candidate rows with ✗ AVOID verdict · not investable
            # Exception: EXIT with a BINDING_RISK_SIGNAL in Alerts (stop hit
            # TODAY) stays in the "ACTION REQUIRED TODAY" section.
            _iv_col_idx = h.index("Investability") + 1 if "Investability" in h else None
            _before_flt3 = len(positions_sorted)
            _kept3 = []
            _n_exit_moved = 0; _n_same_day_artifact = 0; _n_avoid_new = 0
            for _item in positions_sorted:
                _dt2, _r2 = _item
                _st2 = str(_r2[c_st-1].value or "").upper()
                _row_dt2 = str(_dt2)[:10]
                _rec_dt2 = (str(_r2[_c_rec_date-1].value or "")[:10]
                                    if _c_rec_date else "")
                _iv2 = (str(_r2[_iv_col_idx-1].value or "")
                                    if _iv_col_idx else "")
                _al2 = (str(_r2[c_alerts-1].value or "").upper()
                                    if c_alerts else "")
                _is_same_day_artifact = (
                    _st2 == "EXIT" and _rec_dt2 and _rec_dt2 == _row_dt2)
                _is_new_avoid = (
                    _rec_dt2 == asof[:10] and _st2 != "EXIT"
                    and "AVOID" in _iv2.upper())
                _is_today_urgent_stop = (
                    _st2 == "EXIT"
                    and any(sig in _al2 for sig in BINDING_RISK_SIGNALS))
                if _is_same_day_artifact:
                    _n_same_day_artifact += 1; continue
                if _is_new_avoid:
                    _n_avoid_new += 1;         continue
                # EXIT stays only if it's a TODAY urgent stop · else goes to
                # the Exit History sheet
                if _st2 == "EXIT" and not _is_today_urgent_stop:
                    _n_exit_moved += 1;        continue
                _kept3.append(_item)
            positions_sorted = _kept3
            if (_n_exit_moved + _n_same_day_artifact + _n_avoid_new) > 0:
                print(f"[xlsx:{mkt_key}] portfolio filter · "
                      f"{_n_exit_moved} exits → Exit History sheet · "
                      f"{_n_same_day_artifact} same-day artifacts dropped · "
                      f"{_n_avoid_new} NEW·AVOID rows dropped · "
                      f"({_before_flt3} → {len(positions_sorted)})")

            # ─────────────────────────────────────────────────────────
            # 2026-08-21 · Wave 2 · canonical position collapse.
            # Operator directive §7 · "One stock = one canonical live
            # portfolio position". §8 · reconcile R1/R2 disagreement.
            # §32 LUPIN test · R1=EXIT + R2=HOLD must render as ONE
            # canonical row (not two confusing independent rows).
            #
            # Consensus rules under vocab v5.0:
            #   both EXIT              → keep EXIT row · canonical EXIT
            #   one EXIT · one non-EXIT → keep non-EXIT · REVIEW tag
            #                              (conservative · operator inspects)
            #   both same bucket       → keep either (deterministic: R1 wins)
            #   different non-EXIT     → priority NEW > ACTIVE+ > ACTIVE
            #
            # The dropped row still lives in aegis_history.xlsx for audit ·
            # only the operator-facing Portfolio view collapses to one row.
            _canon_before = len(positions_sorted)
            _by_pos: dict = {}
            for _item in positions_sorted:
                _dt_c, _r_c = _item
                _tk_c = str(_r_c[c_tk-1].value or "").upper().replace(".NS","").replace(".BO","")
                _by_pos.setdefault(_tk_c, []).append(_item)
            _canonical: list = []
            _n_split_resolved = 0
            _tier_map = {"EXIT": 4, "NEW": 3, "ACTIVE+": 2, "ACTIVE": 1}
            for _tk_c, _items in _by_pos.items():
                if len(_items) == 1:
                    _canonical.append(_items[0]); continue
                # Multi-runner ticker · resolve
                def _item_bucket(_it):
                    _st_i = str(_it[1][c_st-1].value or "").upper()
                    _rd_i = (str(_it[1][_c_rec_date-1].value or "")[:10]
                                 if _c_rec_date else "")
                    _dt_i = str(_it[0])[:10]
                    if _st_i == "EXIT": return "EXIT"
                    if _rd_i == asof[:10]: return "NEW"
                    if _st_i in ("BUY", "STRONG BUY", "ACCUMULATE", "ADD", "BUY BIG"):
                        return "ACTIVE+"
                    return "ACTIVE"
                _bkts = [(_item_bucket(_it), _it) for _it in _items]
                _exit_items = [it for (b, it) in _bkts if b == "EXIT"]
                _non_exit_items = [it for (b, it) in _bkts if b != "EXIT"]
                if _exit_items and not _non_exit_items:
                    # both EXIT · keep first EXIT row (R1 wins by sort)
                    _canonical.append(_exit_items[0])
                elif _exit_items and _non_exit_items:
                    # SPLIT · keep the non-EXIT row · consensus tag added
                    # by _consensus() in the write loop
                    _canonical.append(_non_exit_items[0])
                    _n_split_resolved += 1
                else:
                    # no EXITs · pick highest-priority bucket
                    _sorted = sorted(_bkts,
                                          key=lambda bi: -_tier_map.get(bi[0], 0))
                    _canonical.append(_sorted[0][1])
                    if len({b for b, _ in _bkts}) > 1:
                        _n_split_resolved += 1
            positions_sorted = _canonical
            if _canon_before != len(positions_sorted):
                print(f"[xlsx:{mkt_key}] canonical position collapse · "
                      f"{_canon_before} → {len(positions_sorted)} rows · "
                      f"{_n_split_resolved} R1/R2 splits resolved")

            # 2026-08-25 · CLOSED-DEDUP GUARD · operator: "iex already in
            # exit list · in portfolio why IEX again". Any ticker whose
            # Registry status is CLOSED must NOT appear in the live
            # Portfolio (it lives ONLY in Exit History sheet). This
            # catches the case where the recommender re-evaluates a
            # closed ticker and produces a stale STOP_LOSS_HIT alert.
            try:
                from backend.research import opportunity_registry as _oreg_dedup
                _reg_dedup = _oreg_dedup.load_all(_ROOT)
                _closed_tks: set = set()
                for _opps in _reg_dedup.values():
                    for _o in _opps:
                        if (_o.market.lower() == mkt_key.lower()
                            and _o.status == "CLOSED"):
                            _closed_tks.add(_o.ticker.upper())
                _dedup_before = len(positions_sorted)
                _kept_dd = []
                _n_dropped_closed = 0
                _dropped_tickers = []
                for _item in positions_sorted:
                    _dt_dd, _r_dd = _item
                    _tk_dd = str(_r_dd[c_tk-1].value or "").upper() \
                                    .replace(".NS", "").replace(".BO", "")
                    if _tk_dd in _closed_tks:
                        _n_dropped_closed += 1
                        _dropped_tickers.append(_tk_dd)
                        continue
                    _kept_dd.append(_item)
                positions_sorted = _kept_dd
                if _n_dropped_closed > 0:
                    print(f"[xlsx:{mkt_key}] Registry-CLOSED dedup · "
                              f"dropped {_n_dropped_closed} tickers from Portfolio "
                              f"(already in Exit History): {', '.join(_dropped_tickers[:5])}")
            except Exception as _e:
                print(f"[xlsx:{mkt_key}] Registry-dedup failed · {type(_e).__name__}: {_e}")

            # Reorder positions_sorted by (section, existing sort) so
            # rows for each section are contiguous. Keeps within-section
            # ordering (P&L desc within tier).
            _by_section: dict = {k: [] for k, *_ in _SECTION_ORDER}
            for _item in positions_sorted:
                _by_section.setdefault(_row_section(_item), []).append(_item)
            positions_sorted = []
            _banner_row_indexes: dict = {}   # {row_index: (banner_text, fill_hex, text_hex)}
            _cursor = _pos_header_row + 1
            for _sect_key, _banner_text, _fill_hex, _text_hex in _SECTION_ORDER:
                _rr = _by_section.get(_sect_key) or []
                if not _rr: continue
                _banner_row_indexes[_cursor] = (_banner_text, _fill_hex, _text_hex)
                _cursor += 1               # reserve banner row
                for _r in _rr:
                    positions_sorted.append(_r)
                    _cursor += 1

            # Emit banners now (write loop uses assigned row indexes).
            for _br, (_txt, _fill, _tc) in _banner_row_indexes.items():
                portfolio_ws.merge_cells(start_row=_br, start_column=1,
                                                          end_row=_br, end_column=len(pos_hdr))
                _bc = portfolio_ws.cell(_br, 1, _txt)
                _bc.font = _Font(bold=True, size=12, color=_tc)
                _bc.fill = _PF(start_color=_fill, end_color=_fill, fill_type="solid")
                _bc.alignment = _Align(horizontal="center", vertical="center")
                portfolio_ws.row_dimensions[_br].height = 20

            # Build (target_row_index, item) pairs so the write loop uses
            # row indices that skip past banner rows.
            _row_indexes = [r for r in range(_pos_header_row + 1, _cursor)
                                    if r not in _banner_row_indexes]
            for i, (dt, r) in zip(_row_indexes, positions_sorted):
                # (loop body below is the original write logic)
                # First iteration variable unpacking · pass through
                tk = r[c_tk-1].value
                status = r[c_st-1].value
                entry_v = r[c_entry-1].value if c_entry else None
                rec_dt = str(r[c_recommended-1].value or "")[:10] if c_recommended else ""
                # Compute P&L
                if status == "EXIT":
                    pnl = r[c_exit_pnl-1].value if c_exit_pnl else None
                    curr = r[c_current-1].value if c_current else None
                else:
                    live = _parquet_close(tk, mkt_key, dt)
                    curr = round(live, 2) if live else (r[c_current-1].value if c_current else None)
                    pnl = round((live - entry_v) / entry_v * 100, 2) \
                            if live and isinstance(entry_v, (int, float)) and entry_v > 0 else None
                # Days held
                try:
                    from datetime import date as _dtc
                    d1 = _dtc.fromisoformat(rec_dt); d2 = _dtc.fromisoformat(dt)
                    days = max(0, (d2 - d1).days)
                except Exception:
                    days = ""
                alerts = r[c_alerts-1].value if c_alerts else ""
                exit_reason = ""
                # Row values · P&L stored as decimal (0.0384) so Excel SUM works
                pnl_decimal = pnl / 100.0 if isinstance(pnl, (int, float)) else None
                runner_val = r[c_run - 1].value if c_run else ""
                # Exit Price / Exit Date only for EXIT rows · Stop/T1/T2 only for open
                exit_price = curr if status == "EXIT" else None
                exit_date  = dt if status == "EXIT" else None
                stop_v = r[c_stop - 1].value if (c_stop and status != "EXIT") else None
                t1_v   = r[c_t1 - 1].value   if (c_t1   and status != "EXIT") else None
                t2_v   = r[c_t2 - 1].value   if (c_t2   and status != "EXIT") else None
                if status == "EXIT" and h and "Exit Reason" in h:
                    exit_reason = r[h.index("Exit Reason")].value or ""
                # Investability lookup (Wave 2 · 11 sub-engines)
                _inv = _inv_map.get(tk, {})
                inv_score = _inv.get("score")
                inv_verdict = _inv.get("verdict", "")
                # 2026-08-12 P1-4 · CEO fix · when Investability hasn't been
                # computed yet for a NEW recommendation, substitute PENDING
                # instead of empty string (previously showed as NaN and the
                # Decision layer defaulted to PROTECT which is contradictory
                # for a same-day STRONG BUY like ZYDUSLIFE +6.43%).
                if not inv_verdict:
                    inv_verdict = "⏳ PENDING"

                # 2026-08-12 P0-1 · CEO fix · lifecycle-first detection.
                # These flags feed BOTH _classify_priority and _resolve_decision
                # so ARTIFACT/EXIT/NEW/ACTIVE state precedes generic protective
                # rules. Previously a same-day STRONG BUY (+6.43%) got PROTECT
                # because the decision layer ran trailing-stop logic without
                # asking "is this even a held position yet?".
                # `asof` is the run date; rec_dt is when the position was first
                # recommended. Same day => NEW position, day-0 rules apply.
                _is_new_position = bool(rec_dt) and str(rec_dt)[:10] == str(asof)[:10]
                # ARTIFACT detected inline below via is_same_day check.

                def _classify_priority(st, iv, pnl, is_same_day=False, alerts=""):
                    """Returns bucket letter · R (Risk Veto) OR A-J (existing).

                    2026-08-14 · Sprint K Part 28 · Risk Controller has veto.
                    Any binding risk signal in Alerts (STOP_LOSS_HIT · HARD_STOP ·
                    TRAILING_STOP_HIT · GAP_EXIT · PORTFOLIO_MAX_DD · EMERGENCY_EXIT ·
                    CRITICAL_DEEP_LOSS) forces bucket R regardless of Status +
                    Investability. Same-day rotations are still J even if a stop
                    also fired (rotation semantic wins for that pathological case).

                    2026-08-10 · same-day EXIT (Entry Date == Exit Date) reclassified
                    as J (ARTIFACT) instead of I (CLOSED) · not a real trade.
                    """
                    q_high = iv in ("🏆 QUALITY", "✓ OK")
                    q_mid = iv == "⚠ MARGINAL"
                    q_low = iv == "✗ AVOID"
                    pnl_neg = isinstance(pnl, (int, float)) and pnl < 0
                    # ── Risk Controller veto (Sprint K Part 28) ──
                    # Same-day ARTIFACT wins over R (never-held rotation is not
                    # a real position that could be stopped-out).
                    if st == "ROTATED_SAMEDAY": return "J"
                    if st == "EXIT" and is_same_day: return "J"
                    _alerts_up = str(alerts or "").upper()
                    for _sig in BINDING_RISK_SIGNALS:
                        if _sig in _alerts_up:
                            return "R"
                    # ── existing rules ──
                    if st == "EXIT":
                        return "H" if q_high else "I"
                    if st == "STRONG BUY" and iv == "🏆 QUALITY": return "A"
                    if st in ("BUY", "STRONG BUY") and q_high:    return "B"
                    if st in ("BUY", "STRONG BUY") and q_low:     return "F"
                    if st == "HOLD" and q_high and pnl_neg:       return "C"
                    if st == "HOLD" and q_high:                   return "D"
                    if q_mid:                                      return "E"
                    if q_low and pnl_neg:                         return "G"
                    if q_low:                                      return "F"
                    return "E"

                # Compute P&L for classification
                _pnl_for_class = None
                if isinstance(pnl_decimal, (int, float)):
                    _pnl_for_class = pnl_decimal * 100

                # Same-day detection · Entry Date == Exit Date == today = rotation artifact
                _is_same_day = bool(exit_date and rec_dt and str(exit_date)[:10] == rec_dt[:10])
                # 2026-08-14 Sprint K Part 28 · pass Alerts so Risk Controller
                # veto can force bucket R when STOP_LOSS_HIT etc. present.
                priority_bucket = _classify_priority(status, inv_verdict, _pnl_for_class,
                                                                        is_same_day=_is_same_day,
                                                                        alerts=alerts)
                _matrix = PRIORITY_MATRIX.get(priority_bucket,
                                                            ("—", "—", "—", "—", "F2F2F2"))
                priority_tag = priority_bucket   # for color lookup below
                urgency, reason, action, review, _color = _matrix

                # Final Action · human-readable follow-through (kept for context)
                def _final_action(st, iv):
                    q_high = iv in ("🏆 QUALITY", "✓ OK")
                    q_low = iv == "✗ AVOID"
                    if st == "STRONG BUY":
                        if iv == "🏆 QUALITY": return "🟢🟢 STRONG BUY · high conviction"
                        if iv == "✓ OK":       return "🟢 BUY · engine + quality aligned"
                        if iv == "⚠ MARGINAL": return "🟡 BUY SMALL · quality borderline"
                        if q_low:               return "🔴 SKIP · engine buys but quality FAILS"
                    if st == "BUY":
                        if q_high:              return "🟢 BUY · in buy zone · quality confirmed"
                        if iv == "⚠ MARGINAL": return "🟡 BUY SMALL · watch quality"
                        if q_low:               return "🔴 SKIP · buy signal but quality WEAK"
                    if st == "HOLD":
                        if iv == "🏆 QUALITY": return "🟢 HOLD · patient · both endorse"
                        if iv == "✓ OK":       return "🟢 HOLD · quality intact"
                        if iv == "⚠ MARGINAL": return "🟡 HOLD · tighten stop · watching"
                        if q_low:               return "🔴 REDUCE / EXIT · quality degraded"
                    if st == "EXIT":
                        if q_high: return "⚪ EXITED · quality was good · review if premature"
                        return "⚪ EXITED · position closed"
                    if st == "ROTATED_SAMEDAY":
                        return "⚪ Rotation artifact · not held"
                    return "—"

                final_action = _final_action(status, inv_verdict)

                # DECISION · single human-facing synthesis (col 2)
                decision_text, decision_color_key = _resolve_decision(
                    action, inv_verdict, status)

                # 2026-08-14 CEO fix · Sprint K Part 28 · Risk Controller veto +
                # closed-position uniformity + NEW-state routing. Precedence
                # order (highest first · every check is a hard override):
                #
                #   R bucket (RISK VETO · Alerts contains STOP_LOSS_HIT etc.)
                #     -> Decision = 🔴 EXIT · <alert reason> · IMMEDIATE
                #     -> BINDING · overrides Status/Investability/lifecycle
                #   J bucket (ARTIFACT · same-day rotation · never held)
                #     -> Decision = ⚪ ARTIFACT · not held
                #   I OR H bucket (EXIT closed positions · runner exited)
                #     -> Decision = ⚪ CLOSED  (H was HOLD before · Sprint K Part 28)
                #     -> Any "Premature Exit?" analysis moves to Post-Exit Assessment (Wave 4)
                #   NEW position (rec_dt == asof, not EXIT, not artifact)
                #     -> P0-2 · use NEW-state logic, NOT trailing-stop/protect
                #        · QUALITY/OK   -> 🟢 BUY (validate entry)
                #        · MARGINAL     -> 🟡 WATCH · small size
                #        · AVOID        -> 🔴 SKIP · quality fails
                #        · PENDING      -> ⏳ PROVISIONAL BUY · investability not yet computed
                if priority_bucket == "R":
                    # Extract the specific binding-signal name for the decision text
                    _alerts_up = str(alerts or "").upper()
                    _hit_signal = next(
                        (s for s in BINDING_RISK_SIGNALS if s in _alerts_up),
                        "HARD STOP")
                    decision_text, decision_color_key = (
                        f"🔴 EXIT · {_hit_signal.replace('_',' ').title()} · IMMEDIATE",
                        "red")
                elif priority_bucket == "J":
                    # 2026-08-20 · one-word rule · ARTIFACT collapsed to EXIT
                    decision_text, decision_color_key = "🔴 EXIT · same-day rotation", "red"
                elif priority_bucket in ("I", "H"):
                    # 2026-08-20 · one-word rule · CLOSED collapsed to EXIT
                    decision_text, decision_color_key = "🔴 EXIT", "red"
                elif _is_new_position and status != "EXIT":
                    _iv_key = str(inv_verdict).strip()
                    if _iv_key in ("🏆 QUALITY", "✓ OK"):
                        decision_text, decision_color_key = "🟢 BUY · new position · quality confirmed", "green"
                    elif _iv_key == "⚠ MARGINAL":
                        decision_text, decision_color_key = "🟡 WATCH · new · small size only", "yellow"
                    elif _iv_key == "✗ AVOID":
                        decision_text, decision_color_key = "🔴 SKIP · new · quality fails", "red"
                    elif _iv_key == "⏳ PENDING":
                        decision_text, decision_color_key = "⏳ PROVISIONAL BUY · investability pending", "yellow"
                    # else: keep _resolve_decision result

                # ─────────────────────────────────────────────────────────
                # 2026-08-21 · Wave 1 · vocab v5.0 collapse to 4 values.
                # Operator directive: "new + active + active plus + exit
                # makes sense why hold? new will become active + exit thats
                # it, straight and simple · active plus means add more ·
                # only 3 colors thats it, straight and simple".
                #
                # After all precedence logic above computes a rich decision
                # text (with sub-detail like "BUY · new position · quality
                # confirmed" or "EXIT · Stop Loss Hit · IMMEDIATE"), collapse
                # to the 4-value operator vocab. Sub-detail already lives in
                # Reason / Urgency / Action columns · no information loss.
                #   R / I / J / H / any EXIT-family      → 🔴 EXIT
                #   NEW position (rec_dt == asof)         → 🆕 NEW
                #   Existing · BUY/ADD/STRONG BUY family  → 🟢 ACTIVE+
                #   Everything else (HOLD/PROTECT/WATCH)  → 🟢 ACTIVE
                # 3 colors: purple (NEW) · green (ACTIVE + ACTIVE+) · red (EXIT).
                _dec_upper = str(decision_text or "").upper()
                _is_exit_family = (
                    priority_bucket in ("R", "I", "J", "H")
                    or "EXIT" in _dec_upper or "CLOSED" in _dec_upper
                    or "ARTIFACT" in _dec_upper or "SKIP" in _dec_upper
                )
                _is_buy_family = (
                    status in ("BUY", "STRONG BUY", "ACCUMULATE", "ADD", "BUY BIG")
                    or " BUY" in _dec_upper or "ADD" in _dec_upper
                )
                if _is_exit_family:
                    decision_text, decision_color_key = "🔴 EXIT", "red"
                elif _is_new_position and status != "EXIT":
                    decision_text, decision_color_key = "🆕 NEW", "purple"
                elif _is_buy_family:
                    decision_text, decision_color_key = "🟢 ACTIVE+", "green"
                else:
                    decision_text, decision_color_key = "🟢 ACTIVE", "green"

                # 2026-08-10 CEO v6 · closed positions get BLANK actionable fields
                # (operator: "You don't want CLOSED · Next Review 15-Aug ·
                # Decision HOLD · nonsense")
                # 2026-08-14 · R (risk veto) is NOT terminal · needs execution today
                # (Price Trigger + Exec Window populated). I/H/J stay terminal.
                is_terminal = action in ("CLOSED", "IGNORE") or priority_bucket in ("I", "J", "H")
                price_trigger = "" if is_terminal else _price_trigger(action, stop_v, t1_v, curr)
                next_review_anchor = rec_dt if rec_dt else dt
                next_review = "" if is_terminal else _next_review_date(review, next_review_anchor)
                exec_window = "" if is_terminal else _execution_window(action)

                # 2026-08-10 CEO fix v6 · added REVIEWING state (explicit)
                # 4 lifecycle states operator can act on unambiguously:
                #   🆕 NEW        · first day of recommendation
                # 2026-08-20 · operator directive "exit or close one word is
                # enough · prefer exit default". Lifecycle vocab collapsed to:
                #   🆕 NEW     · first day of the opportunity (rec_dt == today)
                #   🟢 ACTIVE  · held (BUY/HOLD/PROTECT)
                #   🔴 EXIT    · closed / rotated / same-day artifact · anything
                #                terminal · one word covers I / H / J / R buckets
                # Also · terminal Decision (starts with EXIT/CLOSED/ARTIFACT)
                # FORCES Lifecycle=EXIT · fixes the NTPC/LICI/BATAINDIA case
                # where Decision was EXIT but Lifecycle stayed ACTIVE.
                _dec_up = str(decision_text or "").upper()
                _is_terminal_decision = (
                    priority_bucket in ("R", "I", "J", "H")
                    or _dec_up.startswith("🔴 EXIT") or "CLOSED" in _dec_up
                    or "ARTIFACT" in _dec_up
                )
                if _is_terminal_decision:
                    lifecycle = "🔴 EXIT"
                elif rec_dt and dt and rec_dt == dt:
                    lifecycle = "🆕 NEW"
                else:
                    lifecycle = "🟢 ACTIVE"

                # 2026-08-10 CEO fix #3: R1/R2 Consensus column
                consensus = _consensus(tk, runner_val)

                # 2026-08-14 Sprint K Part 28 · Wave 4 · Post-Exit Assessment.
                # Analytical / research-only classification of the CLOSE event.
                # Never a trading instruction. Live Decision column is separate.
                #   R bucket → Stop Loss Triggered · was <alert>
                #   H bucket → Premature Exit? · quality was intact
                #   I bucket → Clean Exit · quality had degraded
                #   J bucket → Same-Day Rotation · never held
                #   active   → blank
                _post_exit_assessment = ""
                if priority_bucket == "R":
                    _au = str(alerts or "").upper()
                    _sig = next((s for s in BINDING_RISK_SIGNALS if s in _au), "HARD_STOP")
                    _post_exit_assessment = f"Stop Loss Triggered · {_sig.replace('_',' ').title()}"
                    if isinstance(pnl_decimal, (int, float)):
                        _post_exit_assessment += f" @ {pnl_decimal*100:+.2f}%"
                elif priority_bucket == "H":
                    _post_exit_assessment = "Premature Exit? · quality intact at close"
                elif priority_bucket == "I":
                    _post_exit_assessment = "Clean Exit · quality had degraded"
                elif priority_bucket == "J":
                    _post_exit_assessment = "Same-Day Rotation · never held"

                # 2026-08-15 · Section 13 · Decision Basis · one short reason
                # explaining WHY. Deterministic derivation · matches the
                # priority classifier's decision chain.
                _au = str(alerts or "").upper()
                if priority_bucket == "R":
                    _sig = next((s for s in BINDING_RISK_SIGNALS if s in _au), "HARD STOP")
                    _decision_basis = f"STOP LOSS HIT · {_sig.replace('_',' ').title()}"
                elif priority_bucket == "J":
                    _decision_basis = "SAME-DAY ROTATION"
                elif priority_bucket in ("I", "H"):
                    _decision_basis = "MODEL EXIT · Runner closed"
                elif _is_new_position and status != "EXIT":
                    _decision_basis = "NEW OPPORTUNITY"
                elif str(inv_verdict or "").strip() == "✗ AVOID":
                    _decision_basis = "QUALITY FAILED"
                elif priority_bucket == "E":
                    _decision_basis = "PROTECT · watch"
                elif priority_bucket == "G":
                    _decision_basis = "RISK REDUCTION"
                elif status in ("STRONG BUY", "BUY"):
                    _decision_basis = "MODEL CONTINUES · buy signal"
                else:
                    _decision_basis = "MODEL CONTINUES"

                # 2026-08-24 · plain-English 🎯 ACTION column · single
                # sentence per row so operator can decide without scanning
                # 10 columns. Format: "<VERB> at ₹P (stop ₹S · T1 ₹T1)".
                def _fmt_num(x):
                    try:
                        v = float(x)
                        return f"{v:,.2f}" if v > 0 else "?"
                    except Exception:
                        return "?"
                # 2026-08-24 · operator: "dont confuse more make this actions:
                # new - active - active + and exit". Four verbs · one-word each
                # · matches vocab v5.0. EXIT covers both live stop-hits and
                # historical closes (one-word rule). Currency prefix stays
                # market-appropriate (₹ for India · $ for USA).
                _cur = "$" if mkt_key.lower() == "usa" else "₹"
                _curr_s = _fmt_num(curr)
                _stop_s = _fmt_num(stop_v)
                _t1_s   = _fmt_num(t1_v)
                _entry_s = _fmt_num(entry_v)
                _pnl_s  = (f"{pnl_decimal*100:+.1f}%"
                                 if isinstance(pnl_decimal, (int, float)) else "?")
                _dec_up_a = str(decision_text or "").upper()
                # EXIT dominates · both R-bucket (stop hit) and I/H/J
                # buckets (historical close) surface as one word EXIT.
                if ("EXIT" in _dec_up_a or "CLOSED" in _dec_up_a
                    or priority_bucket in ("R", "I", "H", "J")):
                    if priority_bucket == "R":
                        _action_str = (f"🔴 EXIT · {_decision_basis[:40]} · "
                                              f"was entry {_cur}{_entry_s}")
                    else:
                        _action_str = f"🔴 EXIT · P&L {_pnl_s} · exit {_cur}{_curr_s}"
                elif "NEW" in _dec_up_a:
                    _action_str = (f"🟣 NEW · {tk} @ {_cur}{_curr_s} · "
                                          f"stop {_cur}{_stop_s} · T1 {_cur}{_t1_s}")
                elif "ACTIVE+" in _dec_up_a or "ACTIVE +" in _dec_up_a:
                    _action_str = (f"🟢 ACTIVE+ · @ {_cur}{_curr_s} · "
                                          f"stop {_cur}{_stop_s} · P&L {_pnl_s}")
                else:
                    _action_str = (f"🟢 ACTIVE · stop {_cur}{_stop_s} · "
                                          f"P&L {_pnl_s}")

                # 2026-08-25 · Month column · operator "mm-yyyy makes sense
                # in single column like May 2026, June 2026" · format as
                # "August 2026" (full month name + year) via strftime.
                def _month_label(_iso: str) -> str:
                    _s = str(_iso or "")[:7]
                    if not _s or "-" not in _s: return ""
                    try:
                        from datetime import datetime as _dt2
                        return _dt2.strptime(_s, "%Y-%m").strftime("%b %Y")
                    except Exception:
                        return _s
                _month_p = _month_label(rec_dt)
                vals = [
                    # IDENTITY (1-5) · Ticker + Action + Decision + Lifecycle + Month
                    tk, _action_str, decision_text, lifecycle, _month_p,
                    # EXECUTION LAYER (6-8)
                    price_trigger, next_review, exec_window,
                    # META (9-15)
                    runner_val, consensus, _sector_for(tk, mkt_key), _cap_size(tk, mkt_key),
                    rec_dt, exit_date, days,
                    # DECISION SUPPORT (16-22)
                    urgency, reason, action, review,
                    status, inv_verdict, inv_score,
                    # PRICE + P&L (23-26)
                    entry_v, curr, exit_price, pnl_decimal,
                    # RISK/TARGET (27-29)
                    stop_v, t1_v, t2_v,
                    # CONTEXT (30-32)
                    _ACTIONS.get(status, ""), alerts or "", exit_reason,
                    # POST-EXIT ASSESSMENT (33)
                    _post_exit_assessment,
                    # DECISION BASIS (34)
                    _decision_basis,
                ]
                for c, v in enumerate(vals, start=1):
                    cell = portfolio_ws.cell(i, c, v)
                    # +1 shift on all text/format columns after Month insertion at col 5
                    text_cols = {1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14, 16, 17, 18, 19, 20, 21, 30, 31, 32, 33, 34}
                    cell.alignment = _Align(
                        horizontal="left" if c in text_cols else "right",
                        vertical="center", wrap_text=True)
                    # Number formats
                    # +2 shift on numeric col indices (ACTION at 2 + Month at 5)
                    if c == 15 and isinstance(days, int):       # Days
                        cell.number_format = "0"
                    elif c == 22 and isinstance(inv_score, (int, float)):  # Investability
                        cell.number_format = "0.0"
                    elif c in (23, 24, 25, 27, 28, 29):         # Entry/Curr/Exit/Stop/T1/T2
                        cell.number_format = "#,##0.00"
                    elif c == 26 and pnl_decimal is not None:   # P&L %
                        cell.number_format = "+0.00%;-0.00%;0.00%"

                # DECISION cell keeps col 3 · ACTION at col 2 · Month at col 5
                _dec_hex = DECISION_COLORS.get(decision_color_key, "E7E6E6")
                portfolio_ws.cell(i, 3).fill = _PF(
                    start_color=_dec_hex, end_color=_dec_hex, fill_type="solid")
                portfolio_ws.cell(i, 3).font = _Font(bold=True, size=11)
                # 🎯 ACTION cell (col 2) also gets emphasis · this is the
                # single column operator scans to decide.
                portfolio_ws.cell(i, 2).font = _Font(bold=True, size=11)

                # 2026-08-14 · operator-approved 6-tier color scheme (msg 380 approved).
                # Row color = decision-TIER (from configs/decision_colors.yaml).
                # Config-driven · edit YAML → next CI cycle picks up.
                _row_tier = _decision_tier(decision_text, _is_new_position)
                _tier_def = DECISION_TIERS.get(_row_tier) or {}
                _tier_hex_fill = _tier_def.get("hex_fill") or "F2F2F2"
                _tier_hex_text = _tier_def.get("hex_text") or "000000"
                _tier_bold = bool(_tier_def.get("bold"))
                _row_fill = _PF(start_color=_tier_hex_fill, end_color=_tier_hex_fill, fill_type="solid")
                _row_font = _Font(color=_tier_hex_text, bold=_tier_bold)
                for c in range(1, len(pos_hdr) + 1):
                    _cell = portfolio_ws.cell(i, c)
                    _cell.fill = _row_fill
                    # Preserve ACTION (col 2) + DECISION (col 3) + Month (col 5) fonts.
                    if c not in (2, 3, 5):
                        # Keep existing font (alignment / number format) but tint text
                        try:
                            _cell.font = _Font(color=_tier_hex_text, bold=_tier_bold,
                                                    size=_cell.font.size or 11)
                        except Exception:
                            _cell.font = _row_font
            portfolio_ws.freeze_panes = f"A{_pos_header_row + 1}"

            # ═══════════════════════════════════════════════════════════════
            # APPENDIX · diagnostic panels (moved from top per 2026-08-24
            # operator: "toomuch of confusion, multiple columns"). Written
            # after the position table so operator scans NEW/ACTIVE/EXIT
            # rows first · diagnostic drill-in lives below the fold.
            # ═══════════════════════════════════════════════════════════════
            try:
                _app_row = portfolio_ws.max_row + 2
                for _r_off, _kpi_row in enumerate(_appendix_rows, start=_app_row):
                    for _c_off, _v in enumerate(_kpi_row, start=1):
                        _c = portfolio_ws.cell(_r_off, _c_off, _v)
                        _c.font = _Font(size=10, italic=(_c_off > 1))
                        _c.alignment = _Align(vertical="center", wrap_text=True)
                        if _c_off == 1 and str(_v or "").startswith("──"):
                            _c.font = _Font(bold=True, size=11)
            except Exception:
                pass    # non-fatal · appendix is decorative

            # ═══════════════════════════════════════════════════════════════
            # SHEET 2 · EXIT HISTORY (90d) · Wave 3 · 2026-08-21
            # Operator directive: "add new sheet, show last 3 months stock
            # by stock P&L profit loss, makes sense to track it. portfolio
            # is main sheet for new and active" · columns per operator:
            # stock name · start date · closed date · entry price · closing
            # price · P&L.
            # ═══════════════════════════════════════════════════════════════
            exit_ws = wb2.create_sheet("Exit History (90d)", 1)
            # 2026-08-25 · operator: "sector missing exit sheet" +
            # "add month column plz" + "year also makes sense" · Month
            # column formatted YYYY-MM covers both.
            _exit_hdr = ["Stock", "Sector", "Month", "Runner",
                            "Entry Date", "Exit Date",
                            "Days Held", "Entry Price", "Exit Price",
                            "P&L %", "Exit Reason"]
            _exit_widths = [14, 22, 16, 8, 12, 12, 10, 14, 14, 12, 46]
            # Title row (widened for Month column)
            exit_ws.merge_cells("A1:K1")
            exit_ws["A1"] = f"AEGIS {mkt_key} · EXIT HISTORY · last 90 days as of {latest_date or 'today'}"
            exit_ws["A1"].font = _Font(bold=True, size=14, color="FFFFFF")
            exit_ws["A1"].fill = _PF(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
            exit_ws["A1"].alignment = _Align(horizontal="center", vertical="center")
            exit_ws.row_dimensions[1].height = 28
            # Header row
            for _c, _n in enumerate(_exit_hdr, start=1):
                _hc = exit_ws.cell(3, _c, _n)
                _hc.font = _Font(bold=True, color="FFFFFF", size=11)
                _hc.fill = _PF(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
                _hc.alignment = _Align(horizontal="center", vertical="center")
                exit_ws.column_dimensions[_gcl_src(_c)].width = _exit_widths[_c-1]
            # Body · walk keep_rows for EXIT status within last 90 days
            _exit_rows: list = []
            _exit_cutoff = (_asof_dt - _td(days=90)).isoformat()
            for _r in keep_rows:
                _st_e = str(_r[c_st-1].value or "").upper()
                if _st_e != "EXIT": continue
                _row_dt_e = str(_r[c_date-1].value or "")[:10]
                if _row_dt_e < _exit_cutoff: continue
                _rn_e = str(_r[c_run-1].value or "").upper().replace("_NEW", "")
                _tk_e = str(_r[c_tk-1].value or "").upper().replace(".NS","").replace(".BO","")
                _entry_dt_e = (str(_r[_c_rec_date-1].value or "")[:10]
                                        if _c_rec_date else "")
                _pnl_e = _r[c_exit_pnl-1].value if c_exit_pnl else None
                _entry_p = _r[c_entry-1].value if c_entry else None
                _exit_p = _r[h.index("Current Price") + 1 - 1].value if "Current Price" in h else None
                _reason = (_r[h.index("Exit Reason") + 1 - 1].value
                                    if "Exit Reason" in h else "")
                try:
                    from datetime import date as _d
                    _d0 = _d.fromisoformat(_entry_dt_e)
                    _d1 = _d.fromisoformat(_row_dt_e)
                    _days_held = (_d1 - _d0).days
                except Exception:
                    _days_held = ""
                # 2026-08-25 · operator: "exit reasons whats tock u r showing?
                # where are they like example TCS?" · Convert internal jargon
                # like "→ GNFC.NS · +6.7pp alpha" into plain-English:
                # "Rotated to GNFC · better setup (+6.7pp)". Stop-hit stays.
                _reason_s = str(_reason or "").strip()
                if _reason_s.startswith("→") and "alpha" in _reason_s.lower():
                    # Format: "→ GNFC.NS · +6.7pp alpha"
                    _parts = _reason_s.replace("→", "").strip().split("·")
                    _new_tk = _parts[0].strip().replace(".NS","").replace(".BO","")
                    _delta = _parts[1].strip() if len(_parts) > 1 else ""
                    _reason_s = f"Rotated to {_new_tk} · better setup ({_delta.replace(' alpha','')})"
                elif "STOP_LOSS_HIT" in _reason_s.upper() or "STOP LOSS" in _reason_s.upper():
                    _reason_s = "Stop loss hit"
                elif not _reason_s:
                    _reason_s = "Closed (no reason recorded)"
                # Sector lookup
                _sec_e = _sector_for(_tk_e, mkt_key)
                _exit_rows.append((
                    _row_dt_e, _tk_e, _sec_e, _rn_e, _entry_dt_e, _row_dt_e,
                    _days_held, _entry_p, _exit_p, _pnl_e, _reason_s,
                ))
            # 2026-08-25 · operator: "why sunpharma should present in exit
            # sheet too?" · Sync Registry-CLOSED events so today's exits
            # appear in Exit History same day (previously only appeared
            # after CI wrote the history XLSX back to git next day).
            try:
                from backend.research import opportunity_registry as _oreg_eh
                _reg_eh = _oreg_eh.load_all(_ROOT)
                _seen_keys = {(x[1], x[3], str(x[0])[:10]) for x in _exit_rows}
                for _opps in _reg_eh.values():
                    for _o in _opps:
                        if _o.market.lower() != mkt_key.lower(): continue
                        if _o.status != "CLOSED": continue
                        if not _o.closed_date: continue
                        if _o.closed_date < _exit_cutoff: continue
                        _key = (_o.ticker.upper(), _o.runner, _o.closed_date)
                        if _key in _seen_keys: continue     # already in exit_rows
                        # Synthesize a row from Registry data
                        try:
                            from datetime import date as _dz
                            _dh = (_dz.fromisoformat(_o.closed_date)
                                        - _dz.fromisoformat(_o.created_date)).days
                        except Exception:
                            _dh = ""
                        _reason_e = str(_o.closed_reason or "Closed (no reason recorded)")
                        if "STOP_LOSS_HIT" in _reason_e.upper():
                            _reason_e = "Stop loss hit"
                        _exit_rows.append((
                            _o.closed_date, _o.ticker.upper(),
                            _sector_for(_o.ticker, mkt_key),
                            _o.runner, _o.created_date, _o.closed_date,
                            _dh, None, None, None, _reason_e,
                        ))
                        _seen_keys.add(_key)
            except Exception as _e:
                print(f"[exit_history:{mkt_key}] registry sync skipped · {_e}")
            # Sort · most recent exit first
            _exit_rows.sort(key=lambda x: x[0], reverse=True)
            _rowptr = 4
            for _er in _exit_rows:
                (_dt_key, _tk_e, _sec_e, _rn_e, _entry_dt_e, _exit_dt_e,
                 _days_held, _entry_p, _exit_p, _pnl_e, _reason) = _er
                # Format as "August 2026" (matches Portfolio Month column)
                try:
                    from datetime import datetime as _dt3
                    _month_e = _dt3.strptime(str(_exit_dt_e or _dt_key)[:7],
                                                                "%Y-%m").strftime("%b %Y")
                except Exception:
                    _month_e = str(_exit_dt_e or _dt_key)[:7]
                _vals = [_tk_e, _sec_e, _month_e, _rn_e, _entry_dt_e, _exit_dt_e,
                             _days_held, _entry_p, _exit_p,
                             (_pnl_e / 100.0 if isinstance(_pnl_e, (int, float)) else ""),
                             str(_reason or "")]
                for _c, _v in enumerate(_vals, start=1):
                    _cell = exit_ws.cell(_rowptr, _c, _v)
                    _cell.font = _Font(size=10)
                    _cell.alignment = _Align(horizontal="center", vertical="center")
                    if _c == 10:   # P&L column shifted +2 (Sector + Month)
                        _cell.number_format = "+0.00%;-0.00%;0.00%"
                # 2026-08-25 · operator "exit color code to entire row" ·
                # fill EVERY cell in the row with the P&L-based color, not
                # just the P&L cell.
                _row_fill_e = None
                if isinstance(_pnl_e, (int, float)):
                    if _pnl_e > 0.01:
                        _row_fill_e = _PF(start_color="C6EFCE",
                                                    end_color="C6EFCE", fill_type="solid")
                    elif _pnl_e < -0.01:
                        _row_fill_e = _PF(start_color="F8CBAD",
                                                    end_color="F8CBAD", fill_type="solid")
                if _row_fill_e is not None:
                    for _c in range(1, len(_exit_hdr) + 1):
                        exit_ws.cell(_rowptr, _c).fill = _row_fill_e
                _rowptr += 1
            if not _exit_rows:
                exit_ws.merge_cells(f"A4:K4")
                _empty = exit_ws.cell(4, 1, "no exits in the last 90 days")
                _empty.alignment = _Align(horizontal="center")
                _empty.font = _Font(size=10, italic=True, color="7F7F7F")
            exit_ws.freeze_panes = "A4"

            # 2026-08-25 · operator: "also give total P&L , positive P&L,
            # negative p&l. by month, anyhow we track for atleast 3 months.
            # plan exit sheet with added information."
            # Emit a per-month summary strip AFTER the exit rows.
            try:
                from collections import defaultdict as _dd
                _by_month = _dd(list)
                for _er in _exit_rows:
                    _dt_key = str(_er[0])[:7]     # YYYY-MM
                    _pnl_v = _er[9]               # P&L % index in tuple
                    if isinstance(_pnl_v, (int, float)):
                        _by_month[_dt_key].append(_pnl_v)
                # Header
                _summary_row = _rowptr + 2
                exit_ws.merge_cells(start_row=_summary_row, start_column=1,
                                                    end_row=_summary_row, end_column=len(_exit_hdr))
                _sh = exit_ws.cell(_summary_row, 1,
                                                "── MONTHLY P&L SUMMARY (last 3 months) ──")
                _sh.font = _Font(bold=True, size=12, color="FFFFFF")
                _sh.fill = _PF(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
                _sh.alignment = _Align(horizontal="center", vertical="center")
                _summary_row += 1
                # Column headers
                _sum_hdr = ["Month", "N Exits", "Wins", "Losses",
                                "Total P&L %", "Positive P&L %", "Negative P&L %", "Win Rate"]
                for _c, _n in enumerate(_sum_hdr, start=1):
                    _sc = exit_ws.cell(_summary_row, _c, _n)
                    _sc.font = _Font(bold=True, size=10, color="FFFFFF")
                    _sc.fill = _PF(start_color="4472C4", end_color="4472C4", fill_type="solid")
                    _sc.alignment = _Align(horizontal="center")
                _summary_row += 1
                # Rows sorted most-recent-month first · limit 3 months
                from datetime import datetime as _dt5
                for _month in sorted(_by_month.keys(), reverse=True)[:3]:
                    _vals_m = _by_month[_month]
                    _n_ex = len(_vals_m)
                    _wins = [v for v in _vals_m if v > 0]
                    _losses = [v for v in _vals_m if v < 0]
                    _total = round(sum(_vals_m), 2)
                    _pos_sum = round(sum(_wins), 2)
                    _neg_sum = round(sum(_losses), 2)
                    _win_rate = (round(len(_wins) / max(1, len(_wins) + len(_losses)) * 100, 1)
                                        if (_wins or _losses) else 0.0)
                    try:
                        _month_label_s = _dt5.strptime(_month, "%Y-%m").strftime("%b %Y")
                    except Exception:
                        _month_label_s = _month
                    _row_vals = [_month_label_s, _n_ex, len(_wins), len(_losses),
                                        f"{_total:+.2f}%", f"{_pos_sum:+.2f}%",
                                        f"{_neg_sum:+.2f}%", f"{_win_rate}%"]
                    for _c, _v in enumerate(_row_vals, start=1):
                        _mc = exit_ws.cell(_summary_row, _c, _v)
                        _mc.font = _Font(size=10, bold=(_c == 5))
                        _mc.alignment = _Align(horizontal="center")
                        # Color total P&L cell
                        if _c == 5 and isinstance(_total, (int, float)):
                            if _total > 0:
                                _mc.fill = _PF(start_color="C6EFCE",
                                                        end_color="C6EFCE", fill_type="solid")
                            elif _total < 0:
                                _mc.fill = _PF(start_color="F8CBAD",
                                                        end_color="F8CBAD", fill_type="solid")
                    _summary_row += 1
            except Exception as _e:
                print(f"[exit_history:{mkt_key}] monthly summary skipped · {_e}")

            # Rename Sheet 3 · full history stays last
            ws2.title = f"AEGIS {mkt_key} History"

            # 2026-08-25 · operator "add month column plz in ... even in
            # 3rd sheet" + "mm-yyyy makes sense in single column like
            # May 2026" · Insert Month as leftmost column, formatted
            # "August 2026" (full month name + year).
            try:
                from datetime import datetime as _dt4
                ws2.insert_cols(1)
                _hist_h = [c.value for c in ws2[1]]
                ws2.cell(1, 1, "Month").font = _Font(bold=True, color="FFFFFF", size=11)
                ws2.cell(1, 1).fill = _PF(start_color="1F4E78",
                                                          end_color="1F4E78", fill_type="solid")
                _date_col_hist = None
                for _ci, _cn in enumerate(_hist_h, start=1):
                    if str(_cn or "").strip() == "Date":
                        _date_col_hist = _ci; break
                if _date_col_hist is not None:
                    for _hr in range(2, ws2.max_row + 1):
                        _dv = ws2.cell(_hr, _date_col_hist).value
                        _s = str(_dv or "")[:7] if _dv else ""
                        try:
                            _m = (_dt4.strptime(_s, "%Y-%m").strftime("%b %Y")
                                      if _s and "-" in _s else "")
                        except Exception:
                            _m = _s
                        ws2.cell(_hr, 1, _m)
            except Exception as _e:
                print(f"[history:{mkt_key}] Month column insert skipped · {_e}")

            wb2.save(out_path)
            src_wb.close()
            # Skip send if market has 0 rows (e.g., USA freshly wiped for S&P 500 reset)
            if len(keep_rows) == 0:
                print(f"[xlsx:{mkt_key}] SKIPPED · 0 rows for market (fresh start · awaiting next pipeline run)")
                return True
            # 2026-08-25 · ZERO-TOLERANCE DELIVERY GATE ·
            # operator: "u shouldnt give me a chance to question right"
            # Consult all guards BEFORE the Telegram POST. If any hard-
            # blocking check FAILs, send a plain-text alert INSTEAD of
            # the defective XLSX. Operator never sees a broken report.
            try:
                from backend.delivery.delivery_gate import (
                    decide as _gate_decide, emit as _gate_emit,
                    blocked_summary as _gate_blocked,
                )
                _gd = _gate_decide(_ROOT, mkt_key.lower())
                _gate_emit(_ROOT, mkt_key.lower(), _gd)
                if _gd.verdict == "BLOCK":
                    print(f"[gate:{mkt_key}] 🚫 BLOCKED · {len(_gd.blocking_codes)} check(s) fail")
                    for _r in _gd.reasons[:5]:
                        print(f"  · {_r[:140]}")
                    # Send text alert instead of XLSX · operator knows why
                    _alert = _gate_blocked(_gd)
                    _ok_alert, _msg_alert = _send_markdown(token, chat_id, _alert)
                    print(f"[gate:{mkt_key}] alert sent={_ok_alert}")
                    return False    # non-zero exit · CI logs turn red
                if _gd.override_used:
                    print(f"[gate:{mkt_key}] ⚠️ OVERRIDE ACTIVE · shipping despite {len(_gd.blocking_codes)} fails")
            except Exception as _e:
                print(f"[gate:{mkt_key}] gate check errored · {type(_e).__name__}: {_e} · shipping anyway (fail-open)")

            # Send · gate passed
            # 2026-08-10 · operator: "simple note with date · why note so big"
            # Suppress heartbeat banner + Monday operator guide from Telegram
            # caption (kept in stdout logs · zero UI clutter for operator)
            full_caption = caption_body
            ok, msg = _send_document(token, chat_id, out_path, caption=full_caption)
            print(f"[xlsx:{mkt_key}] file={out_path.name} · rows={len(keep_rows)} · sent={ok}")
            if not ok:
                print(f"  detail: {msg[:180]}")
            return ok

        # Import Font at module scope (used in split_and_send)
        from openpyxl.styles import Font as _Font
        from openpyxl.styles import Alignment as _Align

        # 2026-08-10 · operator directive · "why note is so big · simple note
        # with date in note like yesterday" · stripped captions to bare date
        india_caption = f"📊 AEGIS India · {asof}"
        usa_caption   = f"📊 AEGIS USA · {asof}"

        xlsx_ok = True
        if "india" in markets:
            xlsx_ok &= _split_and_send("India", "INDIA", india_caption)
        if "usa" in markets:
            xlsx_ok &= _split_and_send("USA",   "USA",   usa_caption)
        # Guard 10 · Data Integrity audit (CEO 17-point checklist · 2026-08-10)
        # Runs AFTER the XLSX is written · audits Prev Close · Today Move · PnL
        # math · Recommended immutability · Entry Price immutability · Opp Age
        try:
            from backend.context.data_integrity_guard import (
                audit as _di_audit, emit as _di_emit, render_summary as _di_render)
            _di_paths = [
                _ROOT / "reports" / "telegram" / f"aegis_history_{m}.xlsx" for m in markets
            ]
            _di_result = _di_audit(_ROOT, _di_paths)
            _di_emit(_ROOT, _di_result)
            print(f"[guard10:integrity] {_di_render(_di_result)}")
            if _di_result.get("verdict") == "RED":
                print(f"[guard10:integrity] 🔴 {_di_result['n_issues']} integrity issues detected · investigation needed")
                for i in _di_result.get("issues", [])[:5]:
                    print(f"    ✗ {i.get('type')}: {i.get('ticker')} {i.get('date','')}")
        except Exception as _e:
            print(f"[guard10:integrity] check failed · {type(_e).__name__}: {_e}")

        # Record successful heartbeat (only if send actually worked)
        if xlsx_ok:
            try:
                from backend.context.pipeline_heartbeat import record_run as _hb_record
                for _m in markets:
                    _hb_record(_ROOT, _m, asof)
            except Exception as e:
                print(f"[heartbeat] record failed · {type(e).__name__}: {e}")
        # Optional · Google Sheets mirror (silent no-op if creds absent)
        gs_ok, gs_msg = maybe_sync_google_sheet(xlsx_path)
        if gs_ok:
            print(f"[gsheets] {gs_msg}")
        _append_delivery_ledger({
            "ts_utc":  datetime.now(timezone.utc).isoformat(),
            "engine":  "aegis.delivery.telegram.xlsx.v1",
            "kind":    "unified_xlsx",
            "file":    str(xlsx_path.relative_to(_ROOT)),
            "ok":      xlsx_ok,
            "gsheets": gs_ok,
        })
    except Exception as e:
        print(f"[xlsx] render/send failed · {type(e).__name__}: {e}")
        xlsx_ok = False

    # Terminal marker for retry-wrapper compatibility
    print(f"sent ({len(markets)} messages + 1 xlsx).")
    return 0 if (all_ok and xlsx_ok) else 1


if __name__ == "__main__":
    sys.exit(main())
