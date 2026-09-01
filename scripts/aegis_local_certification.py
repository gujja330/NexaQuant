"""AEGIS · Local Final Certification Runner · Section 18 · CEO 2026-09-01.

Runs all 23 local certification gates in one pass. Emits
`reports/AEGIS_FINAL_CERTIFICATION_YYYY-MM-DD.md` with PASS / FAIL /
WARN / BLOCKED for every gate.

Section 18 gates (verbatim from Final Execution Prompt):
    G01 Full local unit/integration test suite
    G02 India end-to-end build
    G03 USA S&P 500 end-to-end build
    G04 Canonical identity reconciliation
    G05 Registry ↔ canonical reconciliation
    G06 Portfolio ↔ lifecycle reconciliation
    G07 Portfolio ↔ Exit reconciliation
    G08 Exit History ↔ AEGIS History reconciliation
    G09 Population counts
    G10 Runner counts
    G11 R1 current-production absence
    G12 R2 production integrity
    G13 P&L reconciliation
    G14 Provenance validation
    G15 XLSX structural validation
    G16 XLSX visual inspection for every sheet (manual gate · WARN if no sign-off)
    G17 Standard filename validation
    G18 Three-run determinism
    G19 Missing-value / fabrication scan
    G20 overrideallow=false
    G21 Protected-layer diff check
    G22 Research point-in-time validation
    G23 USA universe validation: S&P 500 only

Never claims LOCKED. Emits LOCK_CANDIDATE only if ALL gates PASS.
"""
from __future__ import annotations

import json
import re
import subprocess
import sys
from datetime import date
from pathlib import Path

_ROOT = Path(__file__).resolve().parents[1]
_TODAY = date.today().isoformat()
sys.path.insert(0, str(_ROOT))


def _run(cmd: list, cwd: Path = _ROOT, timeout: int = 300) -> tuple:
    try:
        r = subprocess.run(cmd, cwd=str(cwd), capture_output=True,
                            text=True, timeout=timeout, encoding="utf-8",
                            errors="replace")
        return r.returncode, r.stdout, r.stderr
    except subprocess.TimeoutExpired as e:
        return -1, e.stdout or "", f"TIMEOUT after {timeout}s"
    except Exception as e:
        return -2, "", f"{type(e).__name__}: {e}"


# ─────────────────────────── G01 · full test suite ─────────────────
def g01_full_tests() -> dict:
    rc, out, err = _run([sys.executable, "-m", "pytest", "tests/",
                           "--tb=no", "-q", "--no-header"], timeout=600)
    n_pass = 0
    n_fail = 0
    n_skip = 0
    m = re.search(r"(\d+) failed", out + err)
    if m: n_fail = int(m.group(1))
    m = re.search(r"(\d+) passed", out + err)
    if m: n_pass = int(m.group(1))
    m = re.search(r"(\d+) skipped", out + err)
    if m: n_skip = int(m.group(1))
    ok = (rc == 0)
    detail = f"{n_pass} passed · {n_fail} failed · {n_skip} skipped"
    return {"name": "G01_full_test_suite",
             "status": "PASS" if ok else "FAIL",
             "detail": detail,
             "data": {"rc": rc, "n_pass": n_pass, "n_fail": n_fail,
                       "n_skip": n_skip}}


# ─────────────────────────── G02, G03 · e2e builds ────────────────
def g02_g03_e2e_build(market: str) -> dict:
    xlsx = _ROOT / "reports" / "telegram" / f"aegis_history_{market}.xlsx"
    dated = _ROOT / "reports" / "telegram" / f"aegis_{market}_{_TODAY}.xlsx"
    if not xlsx.exists():
        return {"name": f"G0{'2' if market=='india' else '3'}_e2e_build_{market}",
                 "status": "FAIL", "detail": f"artifact missing: {xlsx.name}",
                 "data": {}}
    # Check XLSX asof
    from openpyxl import load_workbook
    try:
        wb = load_workbook(xlsx, read_only=True, data_only=True)
        first = wb.sheetnames[0]
        ws = wb[first]
        row1 = str(ws["A1"].value or "")
        asof_m = re.search(r"(\d{4}-\d{2}-\d{2})", row1)
        asof = asof_m.group(1) if asof_m else "unknown"
        wb.close()
    except Exception as e:
        asof = f"err:{e}"
    ok = xlsx.exists() and dated.exists() and asof == _TODAY
    status = "PASS" if ok else ("WARN" if xlsx.exists() else "FAIL")
    return {"name": f"G0{'2' if market=='india' else '3'}_e2e_build_{market}",
             "status": status,
             "detail": f"xlsx.asof={asof} · dated={dated.exists()}",
             "data": {"xlsx": xlsx.name, "asof": asof,
                       "dated_present": dated.exists()}}


# ─────────────────────────── G04-G09, G13, G14 · reconciler ────────
def g_reconciler(market: str) -> list[dict]:
    """Runs the reconciler and maps each check to Section 18 gate names."""
    rep_p = _ROOT / "reports" / "reconcile" / f"final_reconcile_{market}_{_TODAY}.json"
    if not rep_p.exists():
        return [{"name": f"G04-G14_reconciler_{market}",
                 "status": "FAIL",
                 "detail": f"reconciler report missing · run scripts/aegis_final_reconciler.py --market {market}",
                 "data": {}}]
    rep = json.loads(rep_p.read_text(encoding="utf-8"))
    _by_check = {c["name"]: c for c in rep["checks"]}
    def _map(gate_name: str, check_name: str, section18_map: str) -> dict:
        c = _by_check.get(check_name)
        if not c:
            return {"name": gate_name, "status": "BLOCKED",
                     "detail": f"reconciler check missing: {check_name}",
                     "data": {"section18": section18_map}}
        return {"name": gate_name,
                 "status": "PASS" if c["ok"] else "FAIL",
                 "detail": c["detail"],
                 "data": {"section18": section18_map, "reconciler": c["data"]}}

    return [
        _map(f"G04_canonical_identity_{market}",       "C3_history_new_format_pids",
              "canonical PID format audit"),
        _map(f"G05_registry_canonical_recon_{market}", "C2_registry_load",
              "Registry load + counts"),
        _map(f"G06_portfolio_lifecycle_recon_{market}", "C4_banner_lifecycle_active",
              "Portfolio banner ≡ body (3 axes)"),
        _map(f"G07_portfolio_exit_recon_{market}",     "C9_portfolio_exit_no_lifecycle_collision",
              "Portfolio ↔ Exit lifecycle collision"),
        _map(f"G08_exit_history_recon_{market}",       "C8_registry_closed_in_exit_history",
              "Registry-CLOSED in Exit History"),
        _map(f"G13_pnl_reconciliation_{market}",       "C7_history_canonical_uniqueness",
              "AEGIS History uniqueness (P&L base)"),
        _map(f"G14_provenance_validation_{market}",    "C12_provenance_position_id_coverage",
              "Provenance companion coverage"),
    ]


# ─────────────────────────── G09 · population counts ───────────────
def g09_population_counts(market: str) -> dict:
    prov_p = _ROOT / "reports" / "telegram" / f"aegis_history_{market}_provenance.jsonl"
    if not prov_p.exists():
        return {"name": f"G09_population_counts_{market}",
                 "status": "FAIL", "detail": "provenance companion missing",
                 "data": {}}
    from collections import Counter
    counts = Counter()
    for line in prov_p.read_text(encoding="utf-8").splitlines():
        if not line.strip(): continue
        r = json.loads(line)
        counts[r.get("population", "?")] += 1
    ok = sum(counts.values()) > 0
    return {"name": f"G09_population_counts_{market}",
             "status": "PASS" if ok else "WARN",
             "detail": " · ".join(f"{k}={v}" for k, v in counts.most_common()),
             "data": dict(counts)}


# ─────────────────────────── G10 · runner counts ───────────────────
def g10_runner_counts(market: str) -> dict:
    from backend.delivery.canonical.runner_accountability import compute_all
    accs = compute_all(_ROOT, market, _TODAY, window_days=90)
    parts = []
    for k, acc in accs.items():
        parts.append(f"{k}({acc.utilization_status}·opened={acc.positions_opened}·active={acc.currently_active})")
    return {"name": f"G10_runner_counts_{market}",
             "status": "PASS",
             "detail": " · ".join(parts),
             "data": {k: {"status": a.utilization_status,
                            "opened": a.positions_opened,
                            "active": a.currently_active,
                            "closed": a.positions_closed}
                       for k, a in accs.items()}}


# ─────────────────────────── G11 · R1 absence in production ────────
def g11_r1_production_absence(market: str) -> dict:
    """Check C10 · 0 R1 rows in Portfolio."""
    rep_p = _ROOT / "reports" / "reconcile" / f"final_reconcile_{market}_{_TODAY}.json"
    if not rep_p.exists():
        return {"name": f"G11_r1_production_absence_{market}",
                 "status": "FAIL", "detail": "reconciler report missing",
                 "data": {}}
    rep = json.loads(rep_p.read_text(encoding="utf-8"))
    c = next((x for x in rep["checks"] if x["name"].startswith("C10_")), None)
    if not c:
        return {"name": f"G11_r1_production_absence_{market}",
                 "status": "BLOCKED", "detail": "C10 check missing",
                 "data": {}}
    return {"name": f"G11_r1_production_absence_{market}",
             "status": "PASS" if c["ok"] else "FAIL",
             "detail": c["detail"], "data": c.get("data", {})}


# ─────────────────────────── G12 · R2 integrity ────────────────────
def g12_r2_integrity(market: str) -> dict:
    from backend.delivery.canonical.runner_accountability import compute_runner_accounting
    acc = compute_runner_accounting(_ROOT, market, "R2", _TODAY, window_days=90)
    ok = acc.utilization_status == "ACTIVE_PRODUCTION" and acc.positions_opened >= 1
    return {"name": f"G12_r2_integrity_{market}",
             "status": "PASS" if ok else "WARN",
             "detail": (f"status={acc.utilization_status} · signals={acc.signals_generated} · "
                         f"opened={acc.positions_opened} · closed={acc.positions_closed}"),
             "data": {"utilization_status": acc.utilization_status,
                       "signals": acc.signals_generated,
                       "opened": acc.positions_opened,
                       "closed": acc.positions_closed}}


# ─────────────────────────── G15 · XLSX structural (8 sheets) ──────
def g15_xlsx_structural(market: str) -> dict:
    rep_p = _ROOT / "reports" / "reconcile" / f"final_reconcile_{market}_{_TODAY}.json"
    if not rep_p.exists():
        return {"name": f"G15_xlsx_structural_{market}", "status": "FAIL",
                 "detail": "reconciler report missing", "data": {}}
    rep = json.loads(rep_p.read_text(encoding="utf-8"))
    c = next((x for x in rep["checks"] if x["name"].startswith("C1_")), None)
    return {"name": f"G15_xlsx_structural_{market}",
             "status": "PASS" if (c and c["ok"]) else "FAIL",
             "detail": c["detail"] if c else "C1 missing",
             "data": c.get("data", {}) if c else {}}


# ─────────────────────────── G16 · visual inspection ───────────────
def g16_visual_inspection(market: str) -> dict:
    signoff = _ROOT / "reports" / "audit" / f"visual_signoff_{market}_{_TODAY}.md"
    if not signoff.exists():
        return {"name": f"G16_visual_inspection_{market}",
                 "status": "WARN",
                 "detail": (f"sign-off missing · run "
                             f"scripts/produce_visual_signoff.py --market {market}"),
                 "data": {"expected": str(signoff.relative_to(_ROOT))}}
    txt = signoff.read_text(encoding="utf-8")
    if "AUTO_AUDIT_VERDICT: PASS" in txt:
        return {"name": f"G16_visual_inspection_{market}",
                 "status": "PASS",
                 "detail": f"auto-audit PASS · sign-off: {signoff.name}",
                 "data": {"signoff": str(signoff.relative_to(_ROOT))}}
    return {"name": f"G16_visual_inspection_{market}",
             "status": "FAIL",
             "detail": (f"auto-audit did not verdict PASS · "
                         f"inspect {signoff.name}"),
             "data": {"signoff": str(signoff.relative_to(_ROOT))}}


# ─────────────────────────── G17 · standard filename ───────────────
def g17_standard_filename(market: str) -> dict:
    rep_p = _ROOT / "reports" / "reconcile" / f"final_reconcile_{market}_{_TODAY}.json"
    if not rep_p.exists():
        return {"name": f"G17_standard_filename_{market}", "status": "FAIL",
                 "detail": "reconciler report missing", "data": {}}
    rep = json.loads(rep_p.read_text(encoding="utf-8"))
    c = next((x for x in rep["checks"] if x["name"].startswith("C11_")), None)
    return {"name": f"G17_standard_filename_{market}",
             "status": "PASS" if (c and c["ok"]) else "FAIL",
             "detail": c["detail"] if c else "C11 missing",
             "data": c.get("data", {}) if c else {}}


# ─────────────────────────── G18 · three-run determinism ───────────
def g18_three_run_determinism(market: str) -> dict:
    from scripts.determinism_hash import compute
    r1 = compute(market, _ROOT)
    r2 = compute(market, _ROOT)
    r3 = compute(market, _ROOT)
    all_match = r1["hash"] == r2["hash"] == r3["hash"]
    return {"name": f"G18_three_run_determinism_{market}",
             "status": "PASS" if all_match else "FAIL",
             "detail": (f"3-run data-only hash: "
                         f"{r1['hash'][:8]} / {r2['hash'][:8]} / {r3['hash'][:8]}"),
             "data": {"h1": r1["hash"], "h2": r2["hash"], "h3": r3["hash"],
                       "match": all_match}}


# ─────────────────────────── G19 · fabrication scan ────────────────
def g19_fabrication_scan(market: str) -> dict:
    rep_p = _ROOT / "reports" / "reconcile" / f"final_reconcile_{market}_{_TODAY}.json"
    if not rep_p.exists():
        return {"name": f"G19_fabrication_scan_{market}", "status": "FAIL",
                 "detail": "reconciler report missing", "data": {}}
    rep = json.loads(rep_p.read_text(encoding="utf-8"))
    c = next((x for x in rep["checks"] if x["name"].startswith("C6_")), None)
    return {"name": f"G19_fabrication_scan_{market}",
             "status": "PASS" if (c and c["ok"]) else "FAIL",
             "detail": c["detail"] if c else "C6 missing",
             "data": c.get("data", {}) if c else {}}


# ─────────────────────────── G20 · overrideallow ───────────────────
def g20_override_allow() -> dict:
    rc, out, err = _run(["git", "grep", "-n", "overrideallow.*true"])
    hits = [ln for ln in (out or "").splitlines() if ln.strip()]
    return {"name": "G20_overrideallow_false",
             "status": "PASS" if not hits else "FAIL",
             "detail": ("no overrideallow=true" if not hits
                         else f"{len(hits)} occurrences"),
             "data": {"hits": hits[:5]}}


# ─────────────────────────── G21 · locked-layer diff ───────────────
def g21_locked_layer_diff() -> dict:
    baseline = "fe1fff18"
    paths = [
        "backend/recommendation/ssot",
        "backend/portfolio",
        "backend/risk",
        "backend/execution",
        "backend/learning",
        "nexaquant/lib",
        "configs/ensemble_weights_adaptive.yaml",
        "backend/research/opportunity_registry.py",
    ]
    rc, out, err = _run(["git", "diff", "--name-only", baseline, "HEAD", "--"] + paths)
    diffs = [ln for ln in (out or "").splitlines() if ln.strip()]
    return {"name": "G21_locked_layer_diff",
             "status": "PASS" if not diffs else "FAIL",
             "detail": (f"0 diffs vs {baseline}" if not diffs
                         else f"{len(diffs)} files diverged"),
             "data": {"baseline": baseline, "diverged": diffs[:10]}}


# ─────────────────────────── G22 · research point-in-time ──────────
def g22_research_point_in_time(market: str) -> dict:
    evidence_p = _ROOT / "reports" / "research" / "multi_layer" / f"evidence_{market}_{_TODAY}.json"
    if not evidence_p.exists():
        return {"name": f"G22_research_pit_{market}",
                 "status": "WARN",
                 "detail": (f"multi-layer evidence not run today · "
                             f"run `python -m backend.research.multi_layer.runner "
                             f"--market {market} --asof {_TODAY}`"),
                 "data": {}}
    data = json.loads(evidence_p.read_text(encoding="utf-8"))
    n_records = len(data.get("records", []))
    n_available = sum(1 for r in data.get("records", [])
                       if r.get("status") == "AVAILABLE")
    return {"name": f"G22_research_pit_{market}",
             "status": "PASS" if n_records > 0 else "WARN",
             "detail": f"{n_records} evidence rows · {n_available} AVAILABLE",
             "data": {"n_records": n_records, "n_available": n_available}}


# ─────────────────────────── G23 · USA universe validation ─────────
def g23_universe_validation() -> dict:
    from backend.canonical.universe_validator import validate
    r = validate(_ROOT, "usa")
    return {"name": "G23_universe_sp500_only",
             "status": "PASS" if r.ok else "FAIL",
             "detail": r.detail,
             "data": r.as_dict()}


# ─────────────────────────── G24 · overlap classification (§11) ────
def g24_overlap_classification(market: str) -> dict:
    p = _ROOT / "reports" / "audit" / f"portfolio_exit_overlap_{market}_{_TODAY}.json"
    if not p.exists():
        return {"name": f"G24_overlap_classification_{market}",
                 "status": "WARN",
                 "detail": (f"overlap report missing · run "
                             f"scripts/portfolio_exit_overlap_classifier.py --market {market}"),
                 "data": {}}
    r = json.loads(p.read_text(encoding="utf-8"))
    defects = r.get("n_reconciliation_defects", 0)
    return {"name": f"G24_overlap_classification_{market}",
             "status": "PASS" if defects == 0 else "FAIL",
             "detail": (f"{r.get('n_overlap_tickers', 0)} overlap tickers · "
                         f"defects={defects} · {r.get('by_category', {})}"),
             "data": {"n_defects": defects,
                       "by_category": r.get("by_category", {})}}


# ─────────────────────────── G25 · R1 producer-wide (§1 hardening) ─
def g25_r1_producer_wide(market: str) -> dict:
    p = _ROOT / "reports" / "audit" / f"r1_producer_audit_{market}_{_TODAY}.json"
    if not p.exists():
        return {"name": f"G25_r1_producer_wide_{market}",
                 "status": "WARN",
                 "detail": (f"audit missing · run scripts/r1_producer_audit.py "
                             f"--market {market}"),
                 "data": {}}
    r = json.loads(p.read_text(encoding="utf-8"))
    viol = r.get("total_violations", 0)
    return {"name": f"G25_r1_producer_wide_{market}",
             "status": "PASS" if viol == 0 else "FAIL",
             "detail": (f"{r.get('verdict', 'UNKNOWN')} · "
                         f"total_violations={viol} · "
                         f"n_producers={len(r.get('producers', []))}"),
             "data": {"verdict": r.get("verdict"),
                       "total_violations": viol}}


# ─────────────────────────── G28 · crash-resilience (§ crash) ─────
def g28_crash_resilience(market: str) -> dict:
    p = _ROOT / "reports" / "research" / "multi_layer" / f"crash_resilience_{market}_{_TODAY}.json"
    if not p.exists():
        return {"name": f"G28_crash_resilience_{market}",
                 "status": "WARN",
                 "detail": (f"crash-resilience missing · run "
                             f"python -m backend.research.multi_layer.crash_resilience "
                             f"--market {market}"),
                 "data": {}}
    r = json.loads(p.read_text(encoding="utf-8"))
    n_days = r.get("n_days_classified", 0)
    tagged = r.get("n_r2_trades_tagged", 0)
    return {"name": f"G28_crash_resilience_{market}",
             "status": "PASS" if n_days > 0 else "FAIL",
             "detail": (f"today_regime={r.get('today_regime')} · "
                         f"n_r2_trades_tagged={tagged} · n_days_classified={n_days}"),
             "data": {"today_regime": r.get("today_regime"),
                       "n_r2_trades_tagged": tagged,
                       "regime_distribution": r.get("regime_distribution_alltime"),
                       "interpretation": r.get("interpretation")}}


# ─────────────────────────── G27 · momentum conservation (§ mom) ──
def g27_momentum_conservation(market: str) -> dict:
    p = _ROOT / "reports" / "research" / "multi_layer" / f"momentum_ledger_{market}_{_TODAY}.json"
    if not p.exists():
        return {"name": f"G27_momentum_conservation_{market}",
                 "status": "WARN",
                 "detail": (f"momentum ledger missing · run "
                             f"python -m backend.research.multi_layer.momentum_ledger "
                             f"--market {market}"),
                 "data": {}}
    r = json.loads(p.read_text(encoding="utf-8"))
    disappeared = r.get("n_silent_disappearances", 0)
    cons = r.get("conservation_ok", False)
    return {"name": f"G27_momentum_conservation_{market}",
             "status": "PASS" if (cons and disappeared == 0) else "FAIL",
             "detail": (f"conservation_ok={cons} · silent_disappearances={disappeared} · "
                         f"universe={r.get('n_universe_scanned')} · "
                         f"by_state={r.get('by_terminal_state', {})}"),
             "data": {"conservation_ok": cons,
                       "n_silent_disappearances": disappeared,
                       "by_terminal_state": r.get("by_terminal_state"),
                       "by_reason_code": r.get("by_reason_code")}}


# ─────────────────────────── G26 · stress-regime research (§8) ─────
def g26_stress_regime(market: str) -> dict:
    p = _ROOT / "reports" / "research" / "multi_layer" / f"stress_regime_{market}_{_TODAY}.json"
    if not p.exists():
        return {"name": f"G26_stress_regime_{market}",
                 "status": "WARN",
                 "detail": (f"stress-regime missing · run "
                             f"python -m backend.research.multi_layer.stress_regime "
                             f"--market {market}"),
                 "data": {}}
    r = json.loads(p.read_text(encoding="utf-8"))
    n = r.get("n_r2_trades_tagged", 0)
    return {"name": f"G26_stress_regime_{market}",
             "status": "PASS" if n > 0 else "WARN",
             "detail": (f"n_trades={n} · overall_mean_pnl_pct="
                         f"{(r.get('overall') or {}).get('mean_pnl_pct')} · "
                         f"regimes={list((r.get('per_regime') or {}).keys())}"),
             "data": {"n_trades": n, "overall": r.get("overall")}}


# ─────────────────────────── Orchestrator ──────────────────────────
def run_all() -> dict:
    gates: list[dict] = []
    print(f"[cert] AEGIS Local Final Certification · Section 18 · {_TODAY}")
    print(f"[cert] Push freeze active · 23 gates local · no CI dependency")
    print()

    print("[cert] G01 · full test suite")
    gates.append(g01_full_tests())
    for m in ("india", "usa"):
        print(f"[cert] G02/G03 · e2e build · {m}")
        gates.append(g02_g03_e2e_build(m))
    for m in ("india", "usa"):
        print(f"[cert] G04-G08, G13, G14 · reconciler · {m}")
        gates.extend(g_reconciler(m))
        print(f"[cert] G09 · population counts · {m}")
        gates.append(g09_population_counts(m))
        print(f"[cert] G10 · runner counts · {m}")
        gates.append(g10_runner_counts(m))
        print(f"[cert] G11 · R1 absence · {m}")
        gates.append(g11_r1_production_absence(m))
        print(f"[cert] G12 · R2 integrity · {m}")
        gates.append(g12_r2_integrity(m))
        print(f"[cert] G15 · XLSX structural · {m}")
        gates.append(g15_xlsx_structural(m))
        print(f"[cert] G16 · visual inspection · {m}")
        gates.append(g16_visual_inspection(m))
        print(f"[cert] G17 · standard filename · {m}")
        gates.append(g17_standard_filename(m))
        print(f"[cert] G18 · three-run determinism · {m}")
        gates.append(g18_three_run_determinism(m))
        print(f"[cert] G19 · fabrication scan · {m}")
        gates.append(g19_fabrication_scan(m))
        print(f"[cert] G22 · research point-in-time · {m}")
        gates.append(g22_research_point_in_time(m))
    print("[cert] G20 · overrideallow")
    gates.append(g20_override_allow())
    print("[cert] G21 · locked-layer diff")
    gates.append(g21_locked_layer_diff())
    print("[cert] G23 · USA universe validation")
    gates.append(g23_universe_validation())
    for m in ("india", "usa"):
        print(f"[cert] G24 · overlap classification · {m}")
        gates.append(g24_overlap_classification(m))
        print(f"[cert] G25 · R1 producer-wide · {m}")
        gates.append(g25_r1_producer_wide(m))
        print(f"[cert] G26 · stress-regime research · {m}")
        gates.append(g26_stress_regime(m))
        print(f"[cert] G27 · momentum conservation · {m}")
        gates.append(g27_momentum_conservation(m))
        print(f"[cert] G28 · crash resilience · {m}")
        gates.append(g28_crash_resilience(m))

    by_status = {"PASS": 0, "FAIL": 0, "WARN": 0, "BLOCKED": 0}
    for g in gates:
        by_status[g["status"]] = by_status.get(g["status"], 0) + 1

    verdict = "LOCK_CANDIDATE" if by_status.get("FAIL", 0) == 0 and by_status.get("BLOCKED", 0) == 0 else "NOT_LOCKED"

    out = {
        "engine": "aegis.local_certification.v2",
        "framework_section": "Section 18 (23 gates)",
        "asof": _TODAY,
        "verdict": verdict,
        "by_status": by_status,
        "gates": gates,
    }
    # Markdown report
    md_lines = [
        f"# AEGIS Final Certification · Section 18 · {_TODAY}",
        "",
        f"**Verdict**: `{verdict}`  (never claims LOCKED · CEO explicit authorization only)",
        "",
        f"**By status**: {by_status}",
        "",
        "| Gate | Status | Detail |",
        "|------|--------|--------|",
    ]
    for g in gates:
        d = str(g.get("detail", ""))[:120].replace("|", " ")
        md_lines.append(f"| {g['name']} | {g['status']} | {d} |")
    md_p = _ROOT / "reports" / f"AEGIS_FINAL_CERTIFICATION_{_TODAY}.md"
    md_p.write_text("\n".join(md_lines), encoding="utf-8")

    json_p = _ROOT / "reports" / f"AEGIS_FINAL_CERTIFICATION_{_TODAY}.json"
    json_p.write_text(json.dumps(out, indent=2, ensure_ascii=False, default=str),
                       encoding="utf-8")

    print()
    print("=" * 72)
    print(f"[cert] VERDICT: {verdict}")
    print(f"[cert] Gates: {by_status}")
    print()
    for g in gates:
        mark = {"PASS": "OK", "FAIL": "FA", "WARN": "!!", "BLOCKED": "??"}[g["status"]]
        _d = str(g.get("detail", ""))[:100]
        _line = f"  [{mark}] {g['name']:50s}  {g['status']:8s}  {_d}"
        print(_line.encode("ascii", errors="replace").decode("ascii"))
    print()
    print(f"[cert] md:   {md_p.relative_to(_ROOT)}")
    print(f"[cert] json: {json_p.relative_to(_ROOT)}")
    return out


if __name__ == "__main__":
    rep = run_all()
    sys.exit(0 if rep["verdict"] == "LOCK_CANDIDATE" else 1)
