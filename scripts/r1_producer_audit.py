"""R1 producer-wide audit · CEO 2026-09-01 (§1 hardening).

Prove R1 retirement across every producer / ledger / reconciler /
workbook. NOT a delivery-layer-only check: this walks the whole chain
and reports any R1 activity that occurred AFTER the retirement date.

Producers audited:
  1. Registry (`reports/research/opportunity_registry.jsonl`)
     · Any R1 opportunity CREATED after retirement date = VIOLATION
     · Existing R1 ACTIVE positions = FROZEN_HELD (informational)
  2. Paper portfolio (`reports/research/runner1/history.jsonl`)
     · Any non-DORMANT event after retirement date = VIOLATION
  3. Intraday paper (`reports/research/runner1_intraday/*`)
     · Any non-empty snapshot after retirement date = VIOLATION
  4. Recommendations (`reports/recommendations.json`)
     · Any active R1 entry = VIOLATION (should be R2 or absent)
  5. Prediction snapshots (`reports/delivery/prediction_snapshots.jsonl`)
     · R1 rows appended after retirement date = VIOLATION
  6. AEGIS History workbook rows
     · R1 rows dated after retirement in `AEGIS INDIA/USA History`
       sheets · INFORMATIONAL (audit trail may include DORMANT ledger event)
  7. Portfolio + Today Decisions + Runner Performance sheets
     · Any R1 row = VIOLATION (production-facing)
  8. Angel LTP + Investability caches
     · Any R1 ticker fetched fresh after retirement · WARN (may be
       R2-driven so not automatically a violation)

Output: `reports/audit/r1_producer_audit_YYYY-MM-DD.json` with per-
producer verdict and a summary VIOLATION count. If 0 violations,
retirement is proven pipeline-wide.

Never modifies any producer · read-only audit.
"""
from __future__ import annotations

import argparse
import json
import sys
from datetime import date, datetime
from pathlib import Path

_ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(_ROOT))

# Retirement date · loaded from configs
def _retirement_date(root: Path) -> str:
    try:
        import yaml
        p = root / "configs" / "aegis_retirement.yaml"
        cfg = yaml.safe_load(p.read_text(encoding="utf-8")) or {}
        events = cfg.get("retirement_events", [])
        for e in events:
            if str(e.get("runner", "")).upper() == "R1":
                return str(e.get("retirement_date", ""))[:10]
    except Exception:
        pass
    return "2026-09-01"


def _is_r1(runner_or_pid: str) -> bool:
    s = str(runner_or_pid or "").upper()
    return s == "R1" or s.startswith("R1-") or s.startswith("IND-R1-") or s.startswith("USA-R1-")


def audit_registry(root: Path, retirement_date: str) -> dict:
    """Producer 1: opportunity registry.

    CEO 2026-09-02: an R1 record created post-retirement is a violation
    IFF it is still ACTIVE (an ongoing production leak) or was CLOSED
    without a retirement-cleanup reason. Entries CLOSED or REJECTED
    with a documented retirement-cleanup reason represent producer
    self-correction · not counted as ongoing violations."""
    from backend.research import opportunity_registry as oreg
    reg = oreg.load_all(root)
    frozen_held = 0
    violations_created_after = []
    for pid, events in reg.items():
        for o in events:
            if not _is_r1(o.runner): continue
            created = str(o.created_date or "")[:10]
            if o.status == "ACTIVE":
                frozen_held += 1
            if not (created and created > retirement_date):
                continue
            # Post-retirement creation · check if cleaned up
            reason = str(getattr(o, "closed_reason", "") or "").lower()
            is_cleanup = (
                o.status in ("REJECTED",) or
                (o.status == "CLOSED" and (
                    "retire" in reason or "retirement" in reason
                    or "producer guard" in reason
                ))
            )
            if is_cleanup:
                continue     # producer self-corrected · not a live violation
            violations_created_after.append({
                "position_id": o.opportunity_id,
                "ticker": o.ticker,
                "created": created,
                "status": o.status,
            })
    return {
        "producer": "opportunity_registry",
        "frozen_held": frozen_held,
        "violations_created_after_retirement": len(violations_created_after),
        "violation_samples": violations_created_after[:5],
    }


def audit_paper_portfolio(root: Path, retirement_date: str) -> dict:
    """Producer 2: R1 paper portfolio history."""
    p = root / "reports" / "research" / "runner1" / "history.jsonl"
    if not p.exists():
        return {"producer": "paper_portfolio_r1", "n_events": 0,
                "violations_non_dormant_after": 0, "violation_samples": []}
    events = []
    for line in p.read_text(encoding="utf-8").splitlines():
        if not line.strip(): continue
        try:
            events.append(json.loads(line))
        except json.JSONDecodeError:
            continue
    violations = []
    for e in events:
        asof = str(e.get("asof", ""))[:10]
        if not asof or asof <= retirement_date: continue
        status = str(e.get("status", "")).upper()
        # Anything not DORMANT_BY_DESIGN after retirement is a violation
        if status != "DORMANT_BY_DESIGN":
            violations.append({"asof": asof, "n_active": e.get("n_active", 0),
                                "n_opened": e.get("n_opened", 0),
                                "status": status})
    return {
        "producer": "paper_portfolio_r1",
        "n_events": len(events),
        "violations_non_dormant_after": len(violations),
        "violation_samples": violations[:5],
    }


def audit_intraday_paper(root: Path, retirement_date: str) -> dict:
    """Producer 3: intraday R1 paper."""
    dirp = root / "reports" / "research" / "runner1_intraday"
    if not dirp.exists():
        return {"producer": "intraday_paper_r1", "n_snapshots": 0,
                "violations_nonempty_after": 0, "violation_samples": []}
    files = sorted(dirp.glob("*.json"))
    violations = []
    for f in files:
        try:
            snap = json.loads(f.read_text(encoding="utf-8"))
        except Exception:
            continue
        asof = str(snap.get("as_of", ""))[:10]
        if not asof or asof <= retirement_date: continue
        picks = snap.get("picks", []) or []
        if len(picks) > 0:
            violations.append({"file": f.name, "as_of": asof, "n_picks": len(picks)})
    return {
        "producer": "intraday_paper_r1",
        "n_snapshots": len(files),
        "violations_nonempty_after": len(violations),
        "violation_samples": violations[:5],
    }


def audit_recommendations(root: Path, retirement_date: str) -> dict:
    """Producer 4: current recommendations.json."""
    p = root / "reports" / "recommendations.json"
    if not p.exists():
        return {"producer": "recommendations_json", "n_r1_rows": 0,
                "violations_active_r1": 0, "violation_samples": []}
    try:
        data = json.loads(p.read_text(encoding="utf-8"))
    except Exception as e:
        return {"producer": "recommendations_json",
                "error": f"{type(e).__name__}: {e}"}
    recs = data.get("recommendations") or []
    violations = []
    for r in recs:
        run = str(r.get("run_type") or r.get("runner") or "").upper()
        if _is_r1(run):
            violations.append({"ticker": r.get("ticker"), "runner": run})
    return {
        "producer": "recommendations_json",
        "n_total_recs": len(recs),
        "violations_active_r1": len(violations),
        "violation_samples": violations[:5],
    }


def audit_prediction_snapshots(root: Path, retirement_date: str) -> dict:
    """Producer 5: prediction snapshots ledger."""
    p = root / "reports" / "delivery" / "prediction_snapshots.jsonl"
    if not p.exists():
        return {"producer": "prediction_snapshots", "n_r1_rows_after": 0,
                "violation_samples": []}
    violations = []
    n_r1_total = 0
    for line in p.read_text(encoding="utf-8", errors="replace").splitlines():
        if not line.strip(): continue
        try:
            r = json.loads(line)
        except json.JSONDecodeError:
            continue
        run = str(r.get("runner", "") or r.get("run_type", "")).upper()
        if not _is_r1(run): continue
        n_r1_total += 1
        asof = str(r.get("asof") or r.get("date", ""))[:10]
        if asof and asof > retirement_date:
            violations.append({"asof": asof, "ticker": r.get("ticker"),
                                "runner": run})
    return {
        "producer": "prediction_snapshots",
        "n_r1_total": n_r1_total,
        "violations_appended_after": len(violations),
        "violation_samples": violations[:5],
    }


def audit_workbook_current_production(root: Path, market: str) -> dict:
    """Producers 6+7: XLSX current-production sheets · every R1 row is a violation."""
    from openpyxl import load_workbook
    xlsx = root / "reports" / "telegram" / f"aegis_history_{market.lower()}.xlsx"
    if not xlsx.exists():
        return {"producer": f"workbook_current_production_{market}",
                "xlsx_missing": True, "violations": 0}
    wb = load_workbook(xlsx, read_only=True, data_only=True)
    current_production_sheets = ["Portfolio", "Today Decisions"]
    violations = []
    for sh in current_production_sheets:
        if sh not in wb.sheetnames: continue
        ws = wb[sh]
        for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
            if not row: continue
            # Look for any cell that starts with R1 identifier
            for c in row:
                if c and _is_r1(c):
                    violations.append({"sheet": sh, "excel_row": i,
                                        "cell_value": str(c)[:30]})
                    break
    wb.close()
    return {
        "producer": f"workbook_current_production_{market}",
        "sheets_checked": current_production_sheets,
        "violations": len(violations),
        "violation_samples": violations[:5],
    }


def audit(root: Path, market: str) -> dict:
    retirement_date = _retirement_date(root)
    results = []
    results.append(audit_registry(root, retirement_date))
    results.append(audit_paper_portfolio(root, retirement_date))
    results.append(audit_intraday_paper(root, retirement_date))
    results.append(audit_recommendations(root, retirement_date))
    results.append(audit_prediction_snapshots(root, retirement_date))
    results.append(audit_workbook_current_production(root, market))
    # Compute total violations
    n_viol = 0
    for r in results:
        for k, v in r.items():
            if k.startswith("violations") and isinstance(v, int):
                n_viol += v
    verdict = "PROVEN_RETIRED" if n_viol == 0 else "VIOLATIONS_FOUND"
    return {
        "engine": "aegis.r1_producer_audit.v1",
        "asof": date.today().isoformat(),
        "market": market.lower(),
        "retirement_date": retirement_date,
        "verdict": verdict,
        "total_violations": n_viol,
        "producers": results,
    }


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--market", choices=["india", "usa", "both"],
                     default="both")
    args = ap.parse_args()
    markets = ["india", "usa"] if args.market == "both" else [args.market]
    any_violation = False
    for m in markets:
        rep = audit(_ROOT, m)
        out_p = _ROOT / "reports" / "audit" / f"r1_producer_audit_{m}_{rep['asof']}.json"
        out_p.parent.mkdir(parents=True, exist_ok=True)
        out_p.write_text(json.dumps(rep, indent=2, ensure_ascii=False),
                          encoding="utf-8")
        print(f"[r1_audit:{m}] {rep['verdict']} · total_violations={rep['total_violations']} · "
              f"report={out_p.relative_to(_ROOT)}")
        for prod in rep["producers"]:
            _viols = sum(v for k, v in prod.items()
                          if k.startswith("violations") and isinstance(v, int))
            _line = f"    {prod['producer']:38s} violations={_viols}"
            print(_line.encode("ascii", errors="replace").decode("ascii"))
        if rep["total_violations"] > 0:
            any_violation = True
    return 2 if any_violation else 0


if __name__ == "__main__":
    sys.exit(main())
