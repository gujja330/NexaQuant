"""One-shot Opportunity Registry bootstrap · seeds from workbook history.

The Registry (backend/research/opportunity_registry.py) was added
2026-08-18 with an EMPTY starting state. Every daily CI run since has
called `get_or_create` with today's asof as `created_date` — because
the Registry file wasn't yet populated with true first-appearance
dates for tickers that were already active before Aug 18.

Symptom: ONGC recommended Aug 12, HINDUNILVR Aug 12, ZYDUS Aug 11 all
kept showing as NEW because the Registry had no prior record.

Fix: walk the current unified `aegis_history.xlsx` · for each unique
(market, runner, ticker) find the earliest Date it appears · CREATE
a Registry entry with that Date as created_date. Then re-runs use
those correct historical dates instead of restamping to today.

Idempotent · safe to run multiple times. Skips (market, runner, ticker)
tuples that already have an ACTIVE Registry entry (won't overwrite).
Any subsequent CLOSED events are NOT reconstructed here (that's manual
per-position analysis · out of scope). This bootstrap only backfills
the CREATION event.

Run:  python scripts/bootstrap_opportunity_registry.py
      python scripts/bootstrap_opportunity_registry.py --dry-run   (preview only)
"""
from __future__ import annotations

import argparse
import io
import json
import sys
from datetime import datetime, timezone
from pathlib import Path

try:
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
except Exception:
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")

_ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(_ROOT))

from backend.research import opportunity_registry as _oreg   # noqa: E402


def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--dry-run", action="store_true",
                       help="show what would be created · do not write")
    args = ap.parse_args()

    xlsx = _ROOT / "reports" / "telegram" / "aegis_history.xlsx"
    if not xlsx.exists():
        print(f"[bootstrap] source workbook missing: {xlsx}")
        return 1

    try:
        from openpyxl import load_workbook
    except ImportError:
        print("[bootstrap] openpyxl not installed")
        return 1

    print("=" * 70)
    print("  Opportunity Registry bootstrap · seed from workbook history")
    print("=" * 70)
    print(f"  Source: {xlsx.relative_to(_ROOT)}")
    print(f"  Mode:   {'DRY RUN · no writes' if args.dry_run else 'LIVE · will append to registry'}")
    print()

    wb = load_workbook(xlsx, read_only=True)
    ws = wb["AEGIS Daily"] if "AEGIS Daily" in wb.sheetnames else wb.active
    h = [c.value for c in ws[1]]
    def _col(name):
        try: return h.index(name) + 1
        except ValueError: return None
    c_ctry  = _col("Country")
    c_run   = _col("Run_Type")
    c_tk    = _col("Ticker")
    c_date  = _col("Date")
    c_rec   = _col("Recommended")   # may be restamped · use for status hint only
    c_st    = _col("Status")

    if not all([c_ctry, c_run, c_tk, c_date]):
        print("[bootstrap] required columns missing · abort")
        wb.close(); return 1

    # Walk every row · for each (market, runner, ticker):
    #   · earliest[key]        = (earliest Date, Status on that earliest date)
    #   · latest_exit[key]     = (EXIT Date, exit reason) if ever exited · else None
    # 2026-08-20 · enhanced from creation-only seed to also close opportunities
    # whose most-recent status in history is EXIT · fixes 'all 596 records
    # marked ACTIVE despite dozens of visible ⚪ CLOSED rows in Aug 20 output'.
    earliest: dict = {}
    latest_exit: dict = {}   # key -> (exit_date, exit_reason)
    c_exit_reason = _col("Exit Reason")
    for r in range(2, ws.max_row + 1):
        mk = str(ws.cell(r, c_ctry).value or "").lower()
        rn = str(ws.cell(r, c_run).value or "").upper().replace("_NEW", "")
        tk = str(ws.cell(r, c_tk).value or "").upper().replace(".NS","").replace(".BO","")
        dt = str(ws.cell(r, c_date).value or "")[:10]
        st = str(ws.cell(r, c_st).value or "") if c_st else ""
        if not (mk and rn and tk and dt):
            continue
        key = (mk, rn, tk)
        prev = earliest.get(key)
        if prev is None or dt < prev[0]:
            earliest[key] = (dt, st)
        # Track EXIT events · keep latest exit date per key
        if str(st).upper() == "EXIT":
            reason = str(ws.cell(r, c_exit_reason).value or "").strip() \
                          if c_exit_reason else ""
            prior_exit = latest_exit.get(key)
            if prior_exit is None or dt > prior_exit[0]:
                latest_exit[key] = (dt, reason or "closed in history")
    wb.close()

    print(f"[bootstrap] found {len(earliest)} unique (market, runner, ticker) tuples")
    print(f"[bootstrap] detected {len(latest_exit)} tuples that were EXITed at some point")

    # Load current registry · skip tuples that already have ACTIVE entry
    reg = _oreg.load_all(_ROOT)
    n_skipped = n_seeded = 0
    seeded_samples: list = []
    for key, (first_dt, first_status) in sorted(earliest.items()):
        mk, rn, tk = key
        existing = reg.get(key, [])
        has_active = any(o.is_active() for o in existing)
        if has_active:
            n_skipped += 1
            continue
        if args.dry_run:
            n_seeded += 1
            if len(seeded_samples) < 15:
                seeded_samples.append((key, first_dt, first_status))
            continue
        # Live · create the seed
        try:
            _oreg.get_or_create(
                _ROOT, mk, rn, tk, first_dt,
                initial_signal=first_status, initial_rank=None,
            )
            n_seeded += 1
            if len(seeded_samples) < 15:
                seeded_samples.append((key, first_dt, first_status))
        except Exception as e:
            print(f"  ! failed to seed {key}: {type(e).__name__}: {e}")

    # 2026-08-20 · close-transition pass · walk latest_exit and mark those
    # opportunities EXITed. Operator directive "exit or close one word is
    # enough · prefer exit default" · so we call registry.close() (which
    # sets status=CLOSED internally · display layer renders as EXIT).
    n_closed = 0
    if not args.dry_run:
        reg = _oreg.load_all(_ROOT)   # refresh · include just-seeded records
        for key, (exit_dt, reason) in sorted(latest_exit.items()):
            opps = reg.get(key, [])
            active = [o for o in opps if o.is_active()]
            if not active:
                continue
            opp = active[0]
            try:
                _oreg.close(_ROOT, opp.opportunity_id, exit_dt, reason,
                                  registry=reg)
                n_closed += 1
            except Exception as e:
                print(f"  ! failed to close {key}: {type(e).__name__}: {e}")

    print()
    print(f"[bootstrap] {'would seed' if args.dry_run else 'seeded'}: {n_seeded}")
    print(f"[bootstrap] skipped (already active in registry): {n_skipped}")
    if not args.dry_run:
        print(f"[bootstrap] closed (Exit detected in history): {n_closed}")
    if seeded_samples:
        print()
        print(f"  sample of {'planned' if args.dry_run else 'seeded'} entries:")
        for (mk, rn, tk), dt, st in seeded_samples:
            print(f"    {mk:5} {rn:3} {tk:12}  first-seen={dt}  status={st}")

    print()
    if args.dry_run:
        print("  DRY RUN complete · re-run without --dry-run to persist")
    else:
        print(f"  ✓ registry file: {_oreg._registry_path(_ROOT).relative_to(_ROOT)}")
        # Verify by reading back
        reg2 = _oreg.load_all(_ROOT)
        actives = _oreg.active_opportunities(reg2)
        closes  = _oreg.count_by_status(reg2).get("CLOSED", 0)
        print(f"  ✓ verified · ACTIVE={len(actives)}  CLOSED={closes}  (CLOSED renders as EXIT in Portfolio)")
    return 0


if __name__ == "__main__":
    sys.exit(main())
