"""AEGIS Daily Scorecard · passive observation collector.

Not a feature. Reuses existing data (aegis_history_*.xlsx +
investability_*.json). Emits reports/research/daily_scorecard_{date}.json
for Sprint K Part 25 attribution to consume.

Runs nightly · zero cost · zero side effects.

Answers three questions each day:
1. What's the Priority bucket distribution across the portfolio?
2. Day-over-day P&L movement per bucket?
3. Cumulative outcomes since lock date (2026-08-08)?
"""
from __future__ import annotations

import json
import sys
from collections import defaultdict
from datetime import date
from pathlib import Path


_ROOT = Path(__file__).resolve().parents[1]
_LOCK_DATE = "2026-08-08"


def _load_portfolio(market: str) -> list:
    """Load Portfolio sheet as list of dicts."""
    xlsx = _ROOT / "reports" / "telegram" / f"aegis_history_{market}.xlsx"
    if not xlsx.exists(): return []
    try:
        from openpyxl import load_workbook
        wb = load_workbook(xlsx, read_only=True)
        if "Portfolio" not in wb.sheetnames:
            wb.close(); return []
        ws = wb["Portfolio"]
        header = [ws.cell(7, c).value for c in range(1, ws.max_column + 1)]
        rows = []
        for r in range(8, ws.max_row + 1):
            row = {header[c-1]: ws.cell(r, c).value
                       for c in range(1, ws.max_column + 1)}
            if row.get("Ticker"):
                rows.append(row)
        wb.close()
        return rows
    except Exception as e:
        print(f"[scorecard] load {market} failed: {e}")
        return []


def _bucket_from_urgency_reason(urgency: str, reason: str) -> str:
    """Reverse-lookup bucket letter from Urgency + Reason combo."""
    mapping = {
        ("Conviction Buy",       "🔴 HIGH"):   "A",
        ("Confirmed Buy",        "🟠 HIGH"):   "B",
        ("Quality Dip",          "🟠 HIGH"):   "C",
        ("Compounder",           "🟢 LOW"):    "D",
        ("Watch",                "🟡 MEDIUM"): "E",
        ("Signal Warning",       "🟠 HIGH"):   "F",
        ("Structural Failure",   "🔴 HIGH"):   "G",
        ("Premature Exit?",      "🟡 MEDIUM"): "H",
        ("Clean Exit",           "⚪ CLOSED"): "I",
        ("Artifact",             "⚪ CLOSED"): "J",
    }
    return mapping.get((str(reason or "").strip(), str(urgency or "").strip()), "?")


def generate(asof: str | None = None) -> dict:
    asof = asof or date.today().isoformat()
    scorecard = {
        "engine":     "daily_scorecard.v1",
        "asof":       asof,
        "since_lock": _LOCK_DATE,
        "markets":    {},
    }
    for market in ("india", "usa"):
        rows = _load_portfolio(market)
        if not rows:
            scorecard["markets"][market] = {"n": 0, "note": "no data"}
            continue

        bucket_counts = defaultdict(int)
        bucket_pnl = defaultdict(list)
        for row in rows:
            bucket = _bucket_from_urgency_reason(row.get("🎯 Urgency"), row.get("Reason"))
            bucket_counts[bucket] += 1
            pnl = row.get("P&L %")
            if isinstance(pnl, (int, float)):
                bucket_pnl[bucket].append(pnl * 100)

        # Aggregate per bucket
        by_bucket = {}
        for b, cnt in bucket_counts.items():
            pnls = bucket_pnl.get(b, [])
            by_bucket[b] = {
                "n":          cnt,
                "avg_pnl":    round(sum(pnls) / len(pnls), 2) if pnls else None,
                "n_winners":  sum(1 for p in pnls if p > 0),
                "n_losers":   sum(1 for p in pnls if p < 0),
                "n_flat":     sum(1 for p in pnls if abs(p) < 0.01),
            }

        # Total portfolio
        all_pnls = [p for lst in bucket_pnl.values() for p in lst]
        total_pnl = round(sum(all_pnls), 2) if all_pnls else 0
        avg_pnl = round(total_pnl / len(all_pnls), 2) if all_pnls else 0
        n_win = sum(1 for p in all_pnls if p > 0)
        n_loss = sum(1 for p in all_pnls if p < 0)

        scorecard["markets"][market] = {
            "n_positions":    len(rows),
            "total_pnl_pct":  total_pnl,
            "avg_pnl_pct":    avg_pnl,
            "win_rate_pct":   round(n_win / max(1, n_win + n_loss) * 100, 1),
            "by_bucket":      by_bucket,
        }

    return scorecard


def main() -> int:
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    from datetime import date as _d
    asof = _d.today().isoformat()
    sc = generate(asof)
    out_dir = _ROOT / "reports" / "research"
    out_dir.mkdir(parents=True, exist_ok=True)
    out = out_dir / f"daily_scorecard_{asof}.json"
    out.write_text(json.dumps(sc, indent=2, default=str, ensure_ascii=False),
                       encoding="utf-8")
    print(f"[scorecard] wrote {out}")
    for market, data in sc["markets"].items():
        if data.get("n_positions"):
            print(f"  {market}: {data['n_positions']} positions · "
                      f"total P&L {data['total_pnl_pct']:+.2f}% · "
                      f"win rate {data['win_rate_pct']}%")
    return 0


if __name__ == "__main__":
    sys.exit(main())
