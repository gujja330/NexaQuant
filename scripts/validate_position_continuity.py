"""Operator directive 2026-08-15 · Section 22 · 10-metric acceptance report.

Reads the freshly-built XLSX + P0 outcome dataset + config files ·
computes all 10 acceptance metrics + prints PASS/FAIL per criterion.
Non-zero exit if any criterion fails.

10 metrics (Section 22):
   1. number of active positions
   2. number of new opportunities
   3. number of closed positions
   4. number of skipped candidates in research dataset
   5. portfolio realized P&L
   6. portfolio unrealized P&L
   7. new-opportunity count by runner
   8. duplicate Position ID count               (must be 0)
   9. repeated-NEW violations                    (must be 0)
  10. stop-loss/Decision conflicts               (must be 0)

Usage:  python scripts/validate_position_continuity.py
"""
from __future__ import annotations

import io
import json
import sys
from collections import Counter, defaultdict
from pathlib import Path

try:
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
except Exception:
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")

_ROOT = Path(__file__).resolve().parents[1]
_BINDING_SIGNALS = (
    "EMERGENCY_EXIT", "PORTFOLIO_MAX_DD", "HARD_STOP",
    "STOP_LOSS_HIT", "GAP_EXIT", "TRAILING_STOP_HIT",
    "CRITICAL_DEEP_LOSS",
)


def _read_source_history() -> list[dict]:
    """Read every row from aegis_history.xlsx as list of dicts."""
    from openpyxl import load_workbook
    p = _ROOT / "reports" / "telegram" / "aegis_history.xlsx"
    if not p.exists():
        return []
    wb = load_workbook(p, read_only=True)
    ws = wb["AEGIS Daily"] if "AEGIS Daily" in wb.sheetnames else wb.active
    header = None
    rows = []
    for row in ws.iter_rows(values_only=True):
        if header is None:
            header = list(row)
            continue
        if row and any(row):
            rows.append(dict(zip(header, row)))
    wb.close()
    return rows


def _read_per_market_portfolio(market: str) -> list[dict]:
    """Read the Portfolio sheet of a market's XLSX."""
    from openpyxl import load_workbook
    p = _ROOT / "reports" / "telegram" / f"aegis_history_{market}.xlsx"
    if not p.exists():
        return []
    wb = load_workbook(p, read_only=True)
    if "Portfolio" not in wb.sheetnames:
        wb.close(); return []
    ws = wb["Portfolio"]
    header = None
    rows = []
    for row in ws.iter_rows(values_only=True):
        if header is None:
            if row and row[0] == "Ticker":
                header = list(row)
            continue
        if row and row[0]:
            rows.append(dict(zip(header, row)))
    wb.close()
    return rows


def main() -> int:
    print("=" * 70)
    print("  Operator Section 22 · Position Continuity Acceptance Report")
    print("=" * 70)
    print()

    all_hist = _read_source_history()
    india_pf = _read_per_market_portfolio("india")
    usa_pf   = _read_per_market_portfolio("usa")

    print(f"  Loaded {len(all_hist)} history rows · "
              f"{len(india_pf)} India Portfolio rows · "
              f"{len(usa_pf)} USA Portfolio rows")
    print()

    failures = 0

    # ── #1 · Active positions ──
    active_india = sum(1 for r in india_pf
                                if str(r.get("Status") or "").upper() != "EXIT"
                                and "ARTIFACT" not in str(r.get("🎯 DECISION") or "").upper())
    active_usa = sum(1 for r in usa_pf
                              if str(r.get("Status") or "").upper() != "EXIT"
                              and "ARTIFACT" not in str(r.get("🎯 DECISION") or "").upper())
    print(f"  #1  Active positions           India={active_india}  USA={active_usa}")

    # ── #2 · New opportunities ──
    new_india = sum(1 for r in india_pf if str(r.get("Lifecycle") or "").upper().find("NEW") >= 0)
    new_usa   = sum(1 for r in usa_pf   if str(r.get("Lifecycle") or "").upper().find("NEW") >= 0)
    print(f"  #2  New opportunities          India={new_india}  USA={new_usa}")

    # ── #3 · Closed positions ──
    closed_india = sum(1 for r in india_pf
                                 if "CLOSED" in str(r.get("🎯 DECISION") or "").upper())
    closed_usa = sum(1 for r in usa_pf
                                if "CLOSED" in str(r.get("🎯 DECISION") or "").upper())
    print(f"  #3  Closed positions           India={closed_india}  USA={closed_usa}")

    # ── #4 · Skipped candidates in research dataset ──
    for mk in ("india", "usa"):
        skip_p = _ROOT / "reports" / "research" / f"skip_candidates_{mk}.jsonl"
        n_skip = 0
        if skip_p.exists():
            n_skip = sum(1 for l in skip_p.read_text(encoding="utf-8").splitlines() if l.strip())
        print(f"  #4  Skip candidates ({mk:5}) tracked: {n_skip}")

    # ── #5 + #6 · Realized / Unrealized P&L ──
    # Extract from KPI banner rows (top of Portfolio sheet · read raw workbook)
    for mk, tag in (("india", "India"), ("usa", "USA")):
        from openpyxl import load_workbook
        p = _ROOT / "reports" / "telegram" / f"aegis_history_{mk}.xlsx"
        if not p.exists(): continue
        wb = load_workbook(p, read_only=True)
        if "Portfolio" not in wb.sheetnames:
            wb.close(); continue
        ws = wb["Portfolio"]
        vals = {"Realized": None, "Unrealized": None}
        for row in ws.iter_rows(min_row=1, max_row=10, values_only=True):
            if not row: continue
            for k in vals:
                if row[0] and k in str(row[0]):
                    v = row[1] if len(row) > 1 else None
                    if isinstance(v, (int, float)): vals[k] = v * 100
        wb.close()
        r_pct = f"{vals['Realized']:+.2f}%" if vals['Realized'] is not None else "n/a"
        u_pct = f"{vals['Unrealized']:+.2f}%" if vals['Unrealized'] is not None else "n/a"
        print(f"  #5/6  {tag} Portfolio · Realized={r_pct}  Unrealized={u_pct}")

    # ── #7 · New opportunities by runner ──
    print("  #7  New opportunities by runner")
    by_runner_new = defaultdict(int)
    for r in india_pf + usa_pf:
        runner = str(r.get("Runner") or "?")
        lifecycle = str(r.get("Lifecycle") or "")
        if "NEW" in lifecycle.upper():
            by_runner_new[runner] += 1
    for r, n in sorted(by_runner_new.items()):
        print(f"      · {r}  =  {n}")

    # ── #8 · Duplicate Position ID (must be 0) ──
    pos_ids = Counter()
    for r in all_hist:
        pid = r.get("Position ID")
        if pid:
            pos_ids[str(pid)] += 1
    # Duplicates within same (Position ID, Date) = OK (multi-runner append)
    # Duplicates across DIFFERENT Position IDs pointing at same (ticker, runner,
    # first_seen) triple = FAIL. Check by grouping.
    unique_pids = set(pos_ids.keys())
    # Extract (runner, ticker, first_seen_date) from Position ID
    triples = defaultdict(set)
    for pid in unique_pids:
        # Format 1: R1-LUPIN-IND-20260731-abc123  (new)
        # Format 2: LUPIN_IND_20260731             (legacy · pre-refactor)
        if pid.startswith("R") and "-" in pid:
            parts = pid.split("-")
            if len(parts) >= 5:
                key = (parts[0], parts[1], parts[3])
                triples[key].add(pid)
        else:
            parts = pid.split("_")
            if len(parts) >= 3:
                key = ("?", parts[0], parts[2])
                triples[key].add(pid)
    n_collisions = sum(1 for pids in triples.values() if len(pids) > 1)
    ok8 = n_collisions == 0
    print(f"  {'✅' if ok8 else '❌'} #8  Duplicate Position IDs (semantic): {n_collisions}  · target 0")
    if not ok8:
        for key, pids in triples.items():
            if len(pids) > 1:
                print(f"       ✗ {key}: {pids}")
        failures += 1

    # ── #9 · Repeated-NEW violations (same ticker+runner NEW on 2+ days) ──
    hist_appearances = defaultdict(list)   # (mk, runner, ticker) -> list of Dates
    for r in all_hist:
        mk = str(r.get("Country") or "").lower()
        runner = str(r.get("Run_Type") or "").upper().replace("_NEW", "")
        tk = str(r.get("Ticker") or "").upper().replace(".NS","").replace(".BO","")
        dt = str(r.get("Date") or "")[:10]
        rec = str(r.get("Recommended") or "")[:10]
        if mk and runner and tk and dt:
            hist_appearances[(mk, runner, tk)].append((dt, rec))
    n_repeated_new = 0
    repeated_examples = []
    for key, appears in hist_appearances.items():
        if len(appears) < 2: continue
        # Sort chronologically
        appears.sort()
        # count "day where recommended==date" · that's a NEW-day marker
        n_new_days = sum(1 for dt, rec in appears if dt == rec)
        if n_new_days > 1:
            n_repeated_new += 1
            if len(repeated_examples) < 5:
                repeated_examples.append((key, appears[:5]))
    ok9 = n_repeated_new == 0
    print(f"  {'✅' if ok9 else '❌'} #9  Repeated-NEW violations (same tkr+runner NEW twice): {n_repeated_new}  · target 0")
    if not ok9:
        for key, appears in repeated_examples:
            print(f"       ✗ {key}: {appears}")
        failures += 1

    # ── #10 · Stop-loss / Decision conflicts (any risk signal + non-EXIT decision) ──
    conflicts = []
    for r in india_pf + usa_pf:
        alerts_up = str(r.get("Alerts") or "").upper()
        dec_up = str(r.get("🎯 DECISION") or "").upper()
        if any(sig in alerts_up for sig in _BINDING_SIGNALS):
            if "EXIT" not in dec_up:
                conflicts.append((r.get("Ticker"), alerts_up[:60], dec_up[:60]))
    ok10 = len(conflicts) == 0
    print(f"  {'✅' if ok10 else '❌'} #10 Stop-loss/Decision conflicts: {len(conflicts)}  · target 0")
    if not ok10:
        for tk, al, dec in conflicts[:5]:
            print(f"       ✗ {tk} · Alerts={al!r} · Decision={dec!r}")
        failures += 1

    print()
    print("=" * 70)
    if failures == 0:
        print("  ✅  ALL ACCEPTANCE METRICS · WITHIN THRESHOLD · production-ready")
        return 0
    print(f"  ❌  {failures} metric(s) VIOLATED · address above · not production-ready")
    return 1


if __name__ == "__main__":
    sys.exit(main())
