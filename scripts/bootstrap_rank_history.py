"""Seed rank_history.jsonl from prior-day rows in the existing XLSX.

One-time bootstrap so operator's day-over-day delta (Prior Rank + Rank Δ +
Alert columns) works from day 1 without waiting for a week of CI runs to
build up rank history.

Reads reports/telegram/aegis_history.xlsx · for every (Date, Country,
Run_Type, Ticker) row it hasn't already stamped, appends a snapshot to
rank_history.jsonl with (rank, confidence, model_score, status).

Idempotent · safe to re-run.

Usage:
    python scripts/bootstrap_rank_history.py
"""
from __future__ import annotations

import sys
from datetime import datetime, timezone
from pathlib import Path

_ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(_ROOT))

from openpyxl import load_workbook

from backend.portfolio.rank_history import append_snapshot, load_all


def main() -> int:
    xlsx = _ROOT / "reports" / "telegram" / "aegis_history.xlsx"
    if not xlsx.exists():
        print("[bootstrap_rank] no XLSX at reports/telegram/aegis_history.xlsx · nothing to seed")
        return 0

    wb = load_workbook(xlsx, read_only=True)
    ws = wb.active
    h = [c.value for c in ws[1]]

    def idx(name: str) -> int | None:
        try:  return h.index(name) + 1
        except ValueError: return None

    ic = {name: idx(name) for name in
              ["Date", "Country", "Run_Type", "Ticker",
               "Rank", "Confidence %", "Model Score", "Status"]}
    if not all(ic[k] for k in ("Date", "Country", "Run_Type", "Ticker", "Rank")):
        print(f"[bootstrap_rank] XLSX missing required columns · headers: {h}")
        return 1

    existing = {(r["asof"], r["market"], r["runner"], r["ticker"])
                     for r in load_all(_ROOT)}
    print(f"[bootstrap_rank] existing rank_history entries: {len(existing)}")

    seeded = 0
    skipped = 0
    for row in range(2, ws.max_row + 1):
        d = ws.cell(row=row, column=ic["Date"]).value
        country = ws.cell(row=row, column=ic["Country"]).value
        run_type = ws.cell(row=row, column=ic["Run_Type"]).value
        ticker = ws.cell(row=row, column=ic["Ticker"]).value
        rank = ws.cell(row=row, column=ic["Rank"]).value
        conf = ws.cell(row=row, column=ic["Confidence %"]).value if ic["Confidence %"] else None
        model = ws.cell(row=row, column=ic["Model Score"]).value if ic["Model Score"] else None
        status = ws.cell(row=row, column=ic["Status"]).value if ic["Status"] else None
        if not d or not country or not run_type or not ticker:
            continue
        asof = str(d)[:10] if not isinstance(d, str) else d[:10]
        market = str(country).lower()
        runner = "runner2" if run_type == "R2" else "runner1"
        key = (asof, market, runner, str(ticker))
        if key in existing:
            skipped += 1
            continue
        # Normalize confidence · XLSX stores as % (43.1) · rank_history as [0..1]
        conf_frac = None
        if isinstance(conf, (int, float)):
            conf_frac = conf / 100.0 if conf > 1 else conf
        rank_int = int(rank) if isinstance(rank, (int, float)) else None
        append_snapshot(_ROOT, asof=asof, market=market, runner=runner,
                              ticker=str(ticker), rank=rank_int,
                              confidence=conf_frac,
                              model_score=float(model) if isinstance(model, (int, float)) else None,
                              status=str(status) if status else None)
        seeded += 1

    wb.close()
    print(f"[bootstrap_rank] seeded={seeded} skipped_already_present={skipped}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
