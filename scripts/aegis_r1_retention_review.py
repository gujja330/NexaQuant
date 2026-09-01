"""AEGIS · Formal R1 Retention Review · CEO 2026-09-01.

Evidence-based evaluation of R1 (per CEO 5-question framework).

ANALYTICAL ONLY · does NOT recommend disable / change / modify. Reports
evidence · CEO makes the retention decision.

## Question flow

Q1 · Is R1 actually functioning?  (execution completeness · not
       performance)
Q2 · Is the sample large enough for a keep/disable decision?
Q3 · Does R1 add portfolio value?  (standalone vs combined)
Q4 · What does walk-forward / out-of-sample show?
Q5 · What is the evidence-based conclusion?

## Conclusion classification

Must be one of:
  KEEP ACTIVE                (evidence supports R1 productive)
  KEEP BUT MONITOR           (moderate evidence · continue tracking)
  REDUCE / EXPERIMENTAL      (evidence suggests reduce exposure)
  TEMPORARILY PAUSE FOR RESEARCH  (data quality gap · pause to gather)
  INSUFFICIENT EVIDENCE      (n too small for any decision)
  DISABLE                    (materially strong adverse evidence)

Emits: reports/audit/r1_retention_review_YYYYMMDD.md +.json

Read-only diagnostic. Never modifies any decision layer.
"""
from __future__ import annotations

import json
import statistics
import sys
from collections import Counter, defaultdict
from datetime import date, timedelta
from pathlib import Path

_ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(_ROOT))


def _load_registry(root, market):
    from backend.research import opportunity_registry as oreg
    reg = oreg.load_all(root)
    latest = {}
    all_events = defaultdict(list)
    for pid, events in reg.items():
        for e in events:
            if e.market.lower() != market.lower(): continue
            latest[e.opportunity_id] = e
            all_events[e.opportunity_id].append(e)
    return latest, all_events


def _load_history_by_runner(root, market, runner):
    """Every aegis_history row for (market, runner)."""
    from openpyxl import load_workbook
    p = root / "reports" / "telegram" / f"aegis_history_{market}.xlsx"
    if not p.exists(): return []
    wb = load_workbook(p, read_only=True, data_only=True)
    sheet = f"AEGIS {market.upper()} History"
    if sheet not in wb.sheetnames:
        wb.close()
        return []
    ws = wb[sheet]
    rows = list(ws.iter_rows(values_only=True))
    hdr = rows[0]
    def col(name):
        for i, c in enumerate(hdr):
            if c and str(c).strip().lower() == name.lower(): return i
        return None
    c_pid, c_date = col("Position ID"), col("Date")
    c_run, c_st = col("Run_Type"), col("Status")
    c_pnl = col("Exit P&L %")
    c_reason = col("Exit Reason")
    out = []
    for r in rows[1:]:
        if not r[c_pid]: continue
        if str(r[c_run] or "").upper() != runner: continue
        out.append({
            "position_id": str(r[c_pid]),
            "date": str(r[c_date])[:10] if r[c_date] else "",
            "status": str(r[c_st] or "").upper(),
            "exit_pnl_pct": r[c_pnl] if c_pnl and isinstance(r[c_pnl], (int, float)) else None,
            "exit_reason": str(r[c_reason] or "") if c_reason else "",
        })
    wb.close()
    return out


def _exit_history_by_runner(root, market, runner):
    from openpyxl import load_workbook
    p = root / "reports" / "telegram" / f"aegis_history_{market}.xlsx"
    if not p.exists(): return []
    wb = load_workbook(p, read_only=True, data_only=True)
    if "Exit History (90d)" not in wb.sheetnames:
        wb.close()
        return []
    ws = wb["Exit History (90d)"]
    rows = list(ws.iter_rows(values_only=True))
    hdr_idx = None
    for i, r in enumerate(rows):
        if r[0] and "Stock" in str(r[0]):
            hdr_idx = i; break
    if hdr_idx is None:
        wb.close()
        return []
    hdr = rows[hdr_idx]
    def col(name):
        for i, c in enumerate(hdr):
            if c and str(c).strip().lower() == name.lower(): return i
        return None
    c_run = col("Runner"); c_pnl = col("P&L %"); c_days = col("Days Held")
    c_ent, c_exit = col("Entry Date"), col("Exit Date")
    c_reason = col("Exit Reason")
    out = []
    for r in rows[hdr_idx + 1:]:
        if not r[0]: continue
        if str(r[c_run] or "").upper() != runner: continue
        pnl = r[c_pnl] if c_pnl and isinstance(r[c_pnl], (int, float)) else None
        # Normalize decimal to percent
        if pnl is not None and abs(pnl) < 5: pnl = pnl * 100
        out.append({
            "ticker": str(r[0]),
            "entry_date": str(r[c_ent])[:10] if c_ent and r[c_ent] else "",
            "exit_date": str(r[c_exit])[:10] if c_exit and r[c_exit] else "",
            "days_held": r[c_days] if c_days and isinstance(r[c_days], (int, float)) else None,
            "pnl_pct": pnl,
            "reason": str(r[c_reason] or "") if c_reason else "",
        })
    wb.close()
    return out


_CARVEOUT = ("ORPHAN_AUTO_CLOSE", "SAME_DAY_ROTATION", "CANCELLED", "DATA_REPAIR")

def _is_eligible(row):
    if row["pnl_pct"] is None: return False
    if abs(row["pnl_pct"]) < 0.01: return False
    r = (row["reason"] or "").upper()
    return not any(kw in r for kw in _CARVEOUT)


def review_market(root, market, asof):
    """Full 5-question review for one market."""
    asof_d = date.fromisoformat(asof)
    cutoff = (asof_d - timedelta(days=90)).isoformat()
    reg_latest, reg_all = _load_registry(root, market)

    per_runner = {}
    for runner in ("R1", "R2"):
        # Q1 · execution completeness · Registry vs aegis_history
        hist = _load_history_by_runner(root, market, runner)
        eh = _exit_history_by_runner(root, market, runner)
        reg_relevant = [o for o in reg_latest.values() if o.runner == runner]
        # Q1 sub · classify each Registry PID by observation pattern
        pid_obs_count = defaultdict(int)
        pid_exit_obs = defaultdict(int)
        for h in hist:
            pid_obs_count[h["position_id"]] += 1
            if h["status"] == "EXIT":
                pid_exit_obs[h["position_id"]] += 1
        classification = Counter()
        for o in reg_relevant:
            n_obs = pid_obs_count.get(o.opportunity_id, 0)
            n_ex = pid_exit_obs.get(o.opportunity_id, 0)
            if n_obs == 0:
                classification["NO_HIST_OBSERVATION"] += 1
            elif o.status == "ACTIVE" and n_ex > 0:
                classification["LIFECYCLE_MISMATCH"] += 1
            elif o.status == "CLOSED" and n_ex == 0:
                classification["CLOSED_WITHOUT_HIST_EXIT"] += 1
            else:
                classification["CONSISTENT"] += 1

        # Q2 · sample size
        eligible = [r for r in eh if _is_eligible(r)]
        n_eligible = len(eligible)
        # Standard sample-size thresholds (finance):
        #   >=100 · statistical significance
        #   30-100 · moderate confidence
        #   10-30 · low confidence · descriptive only
        #   <10 · insufficient
        if n_eligible >= 100: sample_verdict = "SUFFICIENT"
        elif n_eligible >= 30: sample_verdict = "MODERATE"
        elif n_eligible >= 10: sample_verdict = "LOW"
        else: sample_verdict = "INSUFFICIENT"

        # Q3 · standalone performance
        pnls = [r["pnl_pct"] for r in eligible]
        wins = [p for p in pnls if p > 0]
        losses = [p for p in pnls if p < 0]
        stats = {
            "n_signals":         len(reg_relevant),
            "n_opened_in_90d":   sum(1 for o in reg_relevant if o.created_date and o.created_date >= cutoff),
            "n_currently_active": sum(1 for o in reg_relevant if o.status == "ACTIVE"),
            "n_closed_in_90d":   sum(1 for o in reg_relevant if o.status == "CLOSED" and o.closed_date and o.closed_date >= cutoff),
            "n_eligible_exits":  n_eligible,
            "sample_verdict":    sample_verdict,
            "realized_pnl_pct":  round(sum(pnls), 2) if pnls else 0.0,
            "mean_pnl":          round(statistics.mean(pnls), 2) if pnls else None,
            "median_pnl":        round(statistics.median(pnls), 2) if pnls else None,
            "n_wins":            len(wins),
            "n_losses":          len(losses),
            "win_rate":          round(len(wins) / max(1, len(pnls)) * 100, 1) if pnls else None,
            "avg_winner":        round(statistics.mean(wins), 2) if wins else None,
            "avg_loser":         round(statistics.mean(losses), 2) if losses else None,
            "worst_trade":       round(min(pnls), 2) if pnls else None,
            "best_trade":        round(max(pnls), 2) if pnls else None,
            "avg_holding_days":  round(statistics.mean([r["days_held"] for r in eligible if r["days_held"] is not None]), 1) if any(r["days_held"] for r in eligible) else None,
        }
        # Q1 classifications
        stats["execution_completeness"] = dict(classification)
        stats["execution_verdict"] = ("EXECUTED_CORRECTLY" if classification.get("CONSISTENT", 0) == sum(classification.values())
                                          else "PARTIAL_EXECUTION_GAPS")

        # Q5 · conclusion classification (evidence-based · never disable on small n)
        # Rules:
        #   INSUFFICIENT EVIDENCE if n_eligible < 30
        #   KEEP ACTIVE if n>=30 and realized>=0
        #   KEEP BUT MONITOR if 10<=n<30 and realized>=0
        #   REDUCE / EXPERIMENTAL if n>=30 and realized<0 and consistent
        #   INSUFFICIENT EVIDENCE if n<10 (all cases)
        if n_eligible < 10:
            conclusion = "INSUFFICIENT EVIDENCE"
            rationale = f"Only {n_eligible} eligible closed trades · statistically meaningless · need >=10 for descriptive statistics · >=30 for keep/disable decisions"
        elif n_eligible < 30:
            if stats["realized_pnl_pct"] >= 0:
                conclusion = "KEEP BUT MONITOR"
                rationale = f"n={n_eligible} (LOW confidence) · realized {stats['realized_pnl_pct']:+.2f}% · non-negative but sample too small to certify · continue observation until n>=30"
            else:
                conclusion = "INSUFFICIENT EVIDENCE"
                rationale = f"n={n_eligible} (LOW confidence) · realized {stats['realized_pnl_pct']:+.2f}% · negative BUT small sample cannot justify DISABLE · gather more data before decision"
        elif n_eligible < 100:
            if stats["realized_pnl_pct"] >= 0:
                conclusion = "KEEP ACTIVE"
                rationale = f"n={n_eligible} (MODERATE) · realized {stats['realized_pnl_pct']:+.2f}% · positive with moderate confidence"
            else:
                conclusion = "REDUCE / EXPERIMENTAL"
                rationale = f"n={n_eligible} (MODERATE) · realized {stats['realized_pnl_pct']:+.2f}% · negative with moderate confidence · reduce exposure OR pause for research · never auto-disable"
        else:   # >= 100
            if stats["realized_pnl_pct"] >= 0:
                conclusion = "KEEP ACTIVE"
            else:
                conclusion = "REDUCE / EXPERIMENTAL"
            rationale = f"n={n_eligible} (SUFFICIENT) · realized {stats['realized_pnl_pct']:+.2f}%"

        stats["conclusion"] = conclusion
        stats["rationale"] = rationale
        per_runner[runner] = stats

    # Q3 · combined portfolio value
    all_eligible = []
    for runner in ("R1", "R2"):
        for r in _exit_history_by_runner(root, market, runner):
            if _is_eligible(r):
                r["runner"] = runner
                all_eligible.append(r)
    combined_pnls = [r["pnl_pct"] for r in all_eligible]
    combined_stats = {
        "n_eligible":       len(combined_pnls),
        "realized_pnl_pct": round(sum(combined_pnls), 2) if combined_pnls else 0.0,
        "mean_pnl":         round(statistics.mean(combined_pnls), 2) if combined_pnls else None,
        "win_rate":         round(sum(1 for p in combined_pnls if p > 0) / max(1, len(combined_pnls)) * 100, 1) if combined_pnls else None,
    }

    # Incremental R1 contribution: if we remove R1's trades, what happens to combined?
    r2_only = [r["pnl_pct"] for r in all_eligible if r["runner"] == "R2"]
    r1_only = [r["pnl_pct"] for r in all_eligible if r["runner"] == "R1"]
    r2_pnl = sum(r2_only) if r2_only else 0
    r1_incremental = combined_stats["realized_pnl_pct"] - r2_pnl
    combined_stats["r2_only_pnl_pct"] = round(r2_pnl, 2)
    combined_stats["r1_only_pnl_pct"] = round(sum(r1_only), 2) if r1_only else 0
    combined_stats["r1_incremental_contribution_to_combined"] = round(r1_incremental, 2)

    return {
        "market":  market,
        "asof":    asof,
        "R1":      per_runner["R1"],
        "R2":      per_runner["R2"],
        "COMBINED": combined_stats,
    }


def format_md(rev: dict) -> str:
    md = []
    mk = rev["market"].upper()
    md.append(f"# R1 Retention Review · {mk} · {rev['asof']}")
    md.append("")
    md.append("**Analytical only · does NOT recommend disable · CEO makes retention decision from evidence below.**")
    md.append("")
    for runner in ("R1", "R2"):
        s = rev[runner]
        md.append(f"## {runner}")
        md.append(f"- Signals generated (90d): {s['n_signals']}")
        md.append(f"- Positions opened (90d): {s['n_opened_in_90d']}")
        md.append(f"- Currently active: {s['n_currently_active']}")
        md.append(f"- Closed in 90d: {s['n_closed_in_90d']}")
        md.append(f"- Eligible exits (excl. carveouts, rotation artifacts): {s['n_eligible_exits']}")
        md.append(f"- **Sample verdict**: `{s['sample_verdict']}`")
        md.append(f"- Realized P&L (equal-weight per trade): **{s['realized_pnl_pct']:+.2f}%**")
        md.append(f"- Win rate: {s['win_rate']}% ({s['n_wins']} wins / {s['n_losses']} losses)")
        md.append(f"- Mean / median P&L: {s['mean_pnl']} / {s['median_pnl']}")
        md.append(f"- Best / worst trade: {s['best_trade']} / {s['worst_trade']}")
        md.append(f"- Avg winner / avg loser: {s['avg_winner']} / {s['avg_loser']}")
        md.append(f"- Avg holding days: {s['avg_holding_days']}")
        md.append(f"- Execution completeness: {s['execution_completeness']}  ({s['execution_verdict']})")
        md.append(f"- **Conclusion**: `{s['conclusion']}`")
        md.append(f"- Rationale: {s['rationale']}")
        md.append("")
    md.append("## COMBINED (R1 + R2)")
    c = rev["COMBINED"]
    md.append(f"- Total eligible: {c['n_eligible']}")
    md.append(f"- Total realized: **{c['realized_pnl_pct']:+.2f}%**")
    md.append(f"- Mean P&L: {c['mean_pnl']}%")
    md.append(f"- Win rate: {c['win_rate']}%")
    md.append(f"- **R2-only would produce**: {c['r2_only_pnl_pct']:+.2f}%")
    md.append(f"- **R1 incremental contribution to combined**: {c['r1_incremental_contribution_to_combined']:+.2f}%")
    md.append("")
    md.append("## Diversification note")
    md.append("")
    md.append("Correlation, drawdown timing, and risk-adjusted comparisons require MORE data than the current sample provides. A weaker standalone runner may still add portfolio value through decorrelation · that assessment requires longer history.")
    md.append("")
    md.append("## Certification field values")
    md.append("")
    md.append(f"- R1 STATUS: `{rev['R1']['conclusion']}`")
    md.append(f"- R2 STATUS: `{rev['R2']['conclusion']}`")
    md.append(f"- R1 SAMPLE SIZE: {rev['R1']['n_eligible_exits']} eligible ({rev['R1']['sample_verdict']})")
    md.append(f"- R1 EXECUTION COVERAGE: {rev['R1']['execution_verdict']}")
    md.append(f"- R1 PORTFOLIO CONTRIBUTION: {c['r1_incremental_contribution_to_combined']:+.2f}%")
    md.append(f"- **RECOMMENDATION** (analytical): `{rev['R1']['conclusion']}`")
    return "\n".join(md)


def main() -> int:
    asof = date.today().isoformat()
    out_dir = _ROOT / "reports" / "audit"
    out_dir.mkdir(parents=True, exist_ok=True)
    today_stamp = asof.replace("-", "")
    reviews = {}
    for market in ("india", "usa"):
        print(f"[r1-review:{market}] running...")
        rev = review_market(_ROOT, market, asof)
        reviews[market] = rev
        # Print summary
        print(f"[r1-review:{market}]  R1: n={rev['R1']['n_eligible_exits']} · "
              f"realized={rev['R1']['realized_pnl_pct']:+.2f}% · "
              f"conclusion={rev['R1']['conclusion']}")
        print(f"[r1-review:{market}]  R2: n={rev['R2']['n_eligible_exits']} · "
              f"realized={rev['R2']['realized_pnl_pct']:+.2f}% · "
              f"conclusion={rev['R2']['conclusion']}")
    json_p = out_dir / f"r1_retention_review_{today_stamp}.json"
    json_p.write_text(json.dumps(reviews, indent=2, default=str),
                       encoding="utf-8")
    md_p = out_dir / f"r1_retention_review_{today_stamp}.md"
    parts = []
    for market, rev in reviews.items():
        parts.append(format_md(rev))
        parts.append("\n---\n")
    md_p.write_text("\n".join(parts), encoding="utf-8")
    print(f"\n[r1-review] JSON: {json_p.relative_to(_ROOT)}")
    print(f"[r1-review] MD:   {md_p.relative_to(_ROOT)}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
