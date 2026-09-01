"""Emit a provenance companion JSONL for each market's XLSX.

For every Portfolio row + Exit-History row currently visible in the
production XLSX, resolve the (Ticker, Runner, entry_date) tuple to a
canonical Position ID by joining against the AEGIS History sheet and
the Registry. Emit one record per row with the full provenance:

    { position_id, legacy_position_id, ticker, runner, entry_date,
      exit_date, lifecycle, population, asof, source, engine, sheet }

Output:
    reports/telegram/aegis_history_{market}_provenance.jsonl

The reconciler can consume this file to enforce provenance without
requiring the XLSX itself to change layout. CEO 2026-09-01.
"""
from __future__ import annotations
import argparse
import json
import sys
from datetime import date
from pathlib import Path
from openpyxl import load_workbook

_ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(_ROOT))

from backend.research import opportunity_registry as oreg


def _col(hdr, name):
    for i, c in enumerate(hdr):
        if c and name.lower() == str(c).lower(): return i
    return None


def _col_any(hdr, *names):
    """Return the first matching column index for any of the given header
    names. Handles the c=0 case correctly (0 is a valid index · falsy in
    Python · so `_col(...) or _col(...)` would silently miss column A)."""
    for n in names:
        i = _col(hdr, n)
        if i is not None:
            return i
    return None


def _find_hdr_row(rows):
    for i, r in enumerate(rows[:8]):
        if r and sum(1 for c in r if c is not None) >= 5:
            return i
    return 0


def emit(market: str, root: Path) -> dict:
    market_l = market.lower()
    xlsx = root / "reports" / "telegram" / f"aegis_history_{market_l}.xlsx"
    if not xlsx.exists():
        return {"error": f"artifact not present · {xlsx}", "n": 0}
    wb = load_workbook(xlsx, read_only=True, data_only=True)
    asof = date.today().isoformat()
    out_records: list[dict] = []

    # Load Registry index (ticker, runner, entry_date) -> Position ID
    reg = oreg.load_all(root)
    reg_by_key: dict[tuple, str] = {}
    reg_legacy_by_key: dict[tuple, str] = {}
    for _pid, opps in reg.items():
        for o in opps:
            if o.market.lower() != market_l: continue
            tk = str(o.ticker or "").split(".", 1)[0].upper()
            rn = str(o.runner or "").upper()
            ed = str(o.created_date or "")[:10]
            key = (tk, rn, ed)
            reg_by_key[key] = o.opportunity_id
            reg_legacy_by_key[key] = getattr(o, "legacy_position_id", "") or ""

    # Also build a lookup from AEGIS History (has PID column already)
    hist_sheet = f"AEGIS {market.upper()} History"
    hist_by_key: dict[tuple, str] = {}
    hist_legacy_by_key: dict[tuple, str] = {}
    if hist_sheet in wb.sheetnames:
        ws_h = wb[hist_sheet]
        rows_h = list(ws_h.iter_rows(values_only=True))
        hdr_h = rows_h[0]
        c_pid = _col(hdr_h, "Position ID")
        c_legacy = _col(hdr_h, "Legacy Position ID")
        c_tk = _col(hdr_h, "Ticker")
        c_run = _col_any(hdr_h, "Run_Type", "Runner")
        c_rec = _col(hdr_h, "Recommended")
        for r in rows_h[1:]:
            if not r or c_pid is None or not r[c_pid]: continue
            tk = str(r[c_tk] or "").split(".", 1)[0].upper() if c_tk is not None else ""
            rn = str(r[c_run] or "").upper() if c_run is not None else ""
            ed = str(r[c_rec] or "")[:10] if c_rec is not None else ""
            key = (tk, rn, ed)
            if key not in hist_by_key:
                hist_by_key[key] = str(r[c_pid])
            if c_legacy is not None and r[c_legacy] and key not in hist_legacy_by_key:
                hist_legacy_by_key[key] = str(r[c_legacy])

    def _resolve_pid(tk: str, rn: str, ed: str) -> tuple[str, str]:
        key = (tk, rn, ed)
        pid = reg_by_key.get(key) or hist_by_key.get(key) or ""
        legacy = reg_legacy_by_key.get(key) or hist_legacy_by_key.get(key) or ""
        return pid, legacy

    # ── Portfolio rows ─────────────────────────────────────────────
    if "01_Portfolio" in wb.sheetnames:
        ws_p = wb["01_Portfolio"]
        rows_p = list(ws_p.iter_rows(values_only=True))
        hr = _find_hdr_row(rows_p)
        hdr_p = rows_p[hr]
        c_tk = _col(hdr_p, "Ticker")
        c_run = _col(hdr_p, "Runner")
        c_life = _col(hdr_p, "Lifecycle")
        c_dec = _col_any(hdr_p, "🎯 DECISION", "DECISION")
        c_ent = _col(hdr_p, "Entry Date")
        c_ext = _col(hdr_p, "Exit Date")
        for r in rows_p[hr + 1:]:
            if not r or not r[c_tk if c_tk is not None else 0]: continue
            tk = str(r[c_tk] or "").split(".", 1)[0].upper() if c_tk is not None else ""
            rn = str(r[c_run] or "").upper() if c_run is not None else ""
            ed = str(r[c_ent] or "")[:10] if c_ent is not None else ""
            xd = str(r[c_ext] or "")[:10] if c_ext is not None else ""
            life = str(r[c_life] or "") if c_life is not None else ""
            dec = str(r[c_dec] or "") if c_dec is not None else ""
            pid, legacy = _resolve_pid(tk, rn, ed)
            # Population classification per Population enum
            if "SUGGESTED" in dec.upper():
                pop = "FRESH_RECOMMENDATION"
            elif "SHADOW" in rn:
                pop = "SHADOW"
            elif "ACTIVE" in life.upper():
                pop = "CURRENT_HOLDING"
            elif "NEW" in life.upper():
                pop = "CURRENT_SIGNAL"
            else:
                pop = "CURRENT_HOLDING"
            out_records.append({
                "sheet": "01_Portfolio",
                "position_id": pid,
                "legacy_position_id": legacy,
                "ticker": tk,
                "runner": rn,
                "entry_date": ed,
                "exit_date": xd,
                "lifecycle": life,
                "population": pop,
                "asof": asof,
                "source": "build_aegis_3sheet_workbook",
                "engine": "aegis_canonical_v3",
            })

    # ── 03_Exit_History body rows (canonical PID in col 0) ──────────
    eh_sheet = "03_Exit_History"
    if eh_sheet in wb.sheetnames:
        ws_e = wb[eh_sheet]
        rows_e = list(ws_e.iter_rows(values_only=True))
        for r in rows_e:
            if not r or not r[0]: continue
            pid = str(r[0])
            if not (pid.upper().startswith("USA-") or pid.upper().startswith("IND-")):
                continue
            # cols: 0=PID 1=Ticker 2=Runner 3=Market 4=EntryDate 5=ExitDate
            tk = str(r[1] or "").split(".", 1)[0].upper() if len(r) > 1 else ""
            rn = str(r[2] or "").upper() if len(r) > 2 else ""
            ed = str(r[4] or "")[:10] if len(r) > 4 and str(r[4] or "") != "—" else ""
            xd = str(r[5] or "")[:10] if len(r) > 5 and str(r[5] or "") != "—" else ""
            legacy = reg_legacy_by_key.get((tk, rn, ed), "")
            out_records.append({
                "sheet": eh_sheet,
                "position_id": pid,
                "legacy_position_id": legacy,
                "ticker": tk,
                "runner": rn,
                "entry_date": ed,
                "exit_date": xd,
                "lifecycle": "EXIT",
                "population": "HISTORICAL_CLOSED",
                "asof": asof,
                "source": "build_aegis_3sheet_workbook",
                "engine": "aegis_daily_v2",
            })

    wb.close()
    out_path = root / "reports" / "telegram" / f"aegis_history_{market_l}_provenance.jsonl"
    with out_path.open("w", encoding="utf-8") as f:
        for rec in out_records:
            f.write(json.dumps(rec, ensure_ascii=False) + "\n")

    # Coverage summary
    n = len(out_records)
    n_pid = sum(1 for r in out_records if r["position_id"])
    n_no_pid = n - n_pid
    return {
        "market": market_l,
        "n_records": n,
        "n_with_position_id": n_pid,
        "n_missing_position_id": n_no_pid,
        "coverage_pct": round(n_pid / max(1, n) * 100, 1),
        "out_path": str(out_path.relative_to(root)),
    }


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--market", choices=["india", "usa", "both"],
                     default="both")
    args = ap.parse_args()
    markets = ["india", "usa"] if args.market == "both" else [args.market]
    for m in markets:
        rep = emit(m, _ROOT)
        print(json.dumps(rep, indent=2, ensure_ascii=False))
    return 0


if __name__ == "__main__":
    sys.exit(main())
